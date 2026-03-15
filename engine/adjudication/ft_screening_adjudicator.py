"""Full-text screening adjudication pipeline — export FT_FLAGGED papers for
human review, import decisions back into the database.

Mirrors the screening_adjudicator pattern but for full-text screening results.
Supports both xlsx (workbook) and JSON (HTML tool) import formats.
"""

import json
import logging
from datetime import datetime, timezone
from pathlib import Path

from engine.adjudication.schema import ensure_adjudication_table
from engine.adjudication.workflow import complete_stage
from engine.core.database import ReviewDatabase

logger = logging.getLogger(__name__)

# FT reason code descriptions for reference sheet
_REASON_CODE_DESCRIPTIONS = {
    "eligible": "Paper meets all inclusion criteria based on full text",
    "wrong_specialty": "Paper's specialty falls outside the included specialty scope",
    "no_autonomy_content": "Full text reveals no autonomous or semi-autonomous robot execution",
    "wrong_intervention": "Intervention does not involve autonomous surgical robot control",
    "protocol_only": "Paper describes a study protocol without results",
    "duplicate_cohort": "Same cohort/data as another included paper",
    "insufficient_data": "Insufficient methodological detail to assess eligibility",
}


# ── Data Collection ─────────────────────────────────────────────────


def _collect_ft_flagged(db: ReviewDatabase) -> list[dict]:
    """Collect FT_FLAGGED papers from review.db with screening rationale."""
    papers = db.get_papers_by_status("FT_FLAGGED")
    results = []

    for p in papers:
        pid = p["id"]

        # Get FT screening decision (primary)
        ft_row = db._conn.execute(
            "SELECT model, decision, reason_code, rationale, confidence "
            "FROM ft_screening_decisions WHERE paper_id = ? ORDER BY id DESC LIMIT 1",
            (pid,),
        ).fetchone()

        primary_decision = ft_row["decision"] if ft_row else ""
        primary_rationale = ft_row["rationale"] if ft_row else ""
        reason_code = ft_row["reason_code"] if ft_row else ""

        # Get FT verification decision
        vrow = db._conn.execute(
            "SELECT model, decision, rationale, confidence "
            "FROM ft_verification_decisions WHERE paper_id = ? "
            "ORDER BY id DESC LIMIT 1",
            (pid,),
        ).fetchone()

        verifier_decision = vrow["decision"] if vrow else ""
        verifier_rationale = vrow["rationale"] if vrow else ""

        # Try to get parsed text excerpt (intro/methods)
        text_excerpt = ""
        ft_asset = db._conn.execute(
            "SELECT parsed_text_path FROM full_text_assets WHERE paper_id = ? "
            "ORDER BY id DESC LIMIT 1",
            (pid,),
        ).fetchone()
        if ft_asset and ft_asset["parsed_text_path"]:
            try:
                md_path = Path(ft_asset["parsed_text_path"])
                if md_path.exists():
                    text_excerpt = md_path.read_text()[:500]
            except Exception:
                pass

        results.append({
            "paper_id": pid,
            "ee_identifier": p.get("ee_identifier") or "",
            "title": p["title"],
            "abstract": p.get("abstract") or "",
            "doi": p.get("doi") or "",
            "pmid": p.get("pmid") or "",
            "year": p.get("year") or "",
            "journal": p.get("journal") or "",
            "reason_code": reason_code,
            "primary_decision": primary_decision,
            "primary_rationale": primary_rationale,
            "verifier_decision": verifier_decision,
            "verifier_rationale": verifier_rationale,
            "text_excerpt": text_excerpt,
        })

    return results


# ── Reference Content Builder ─────────────────────────────────────


def _build_ft_reference_content(spec=None) -> str:
    """Build reference sheet content for FT screening adjudication."""
    lines = []

    lines.append("FULL-TEXT SCREENING REASON CODES")
    lines.append("")
    for code, desc in _REASON_CODE_DESCRIPTIONS.items():
        lines.append(f"  {code}: {desc}")
    lines.append("")

    if spec:
        if hasattr(spec, "screening_criteria") and spec.screening_criteria:
            lines.append("INCLUSION CRITERIA:")
            for criterion in spec.screening_criteria.inclusion:
                lines.append(f"  + {criterion}")
            lines.append("")
            lines.append("EXCLUSION CRITERIA:")
            for criterion in spec.screening_criteria.exclusion:
                lines.append(f"  - {criterion}")
            lines.append("")

        if hasattr(spec, "pico") and spec.pico:
            lines.append("PICO FRAMEWORK:")
            lines.append(f"  Population:   {spec.pico.population}")
            lines.append(f"  Intervention: {spec.pico.intervention}")
            lines.append(f"  Comparator:   {spec.pico.comparator}")
            if isinstance(spec.pico.outcomes, list):
                lines.append(f"  Outcomes:     {'; '.join(spec.pico.outcomes)}")
            else:
                lines.append(f"  Outcomes:     {spec.pico.outcomes}")
            lines.append("")

        if hasattr(spec, "specialty_scope") and spec.specialty_scope:
            lines.append("SPECIALTY SCOPE:")
            lines.append("  Included specialties:")
            for s in spec.specialty_scope.included:
                lines.append(f"    + {s}")
            lines.append("  Excluded specialties:")
            for s in spec.specialty_scope.excluded:
                lines.append(f"    - {s}")
            if spec.specialty_scope.notes:
                lines.append(f"  Notes: {spec.specialty_scope.notes}")

    return "\n".join(lines)


def _build_ft_decision_criteria(spec=None) -> list[str]:
    """Build decision criteria for FT screening adjudication."""
    criteria = [
        "FT_ELIGIBLE: The full text confirms the paper describes autonomous or "
        "semi-autonomous surgical robot execution of a physical task.",
        "FT_SCREENED_OUT: The full text reveals the paper does not meet inclusion "
        "criteria (see reason code for likely cause).",
    ]

    if spec and hasattr(spec, "specialty_scope") and spec.specialty_scope:
        included = ", ".join(spec.specialty_scope.included)
        excluded = ", ".join(spec.specialty_scope.excluded)
        criteria.append(f"SPECIALTY SCOPE — Included: {included}")
        criteria.append(f"SPECIALTY SCOPE — Excluded: {excluded}")

    return criteria


def _build_ft_edge_case_guidance(spec=None) -> str:
    """Build edge case guidance for FT screening."""
    parts = []
    if spec and hasattr(spec, "specialty_scope") and spec.specialty_scope and spec.specialty_scope.notes:
        parts.append(spec.specialty_scope.notes.strip())
    parts.append(
        "These papers were flagged because the primary screener and verifier disagreed. "
        "Review the full-text reason code and both rationales to make your decision."
    )
    return " ".join(parts)


# ── Export ──────────────────────────────────────────────────────────


def export_ft_adjudication_queue(
    review_db: ReviewDatabase,
    output_path: str | Path,
    *,
    format: str = "xlsx",
    review_name: str | None = None,
    review_spec=None,
) -> dict:
    """Export all FT_FLAGGED papers as a human-review Excel queue.

    Returns summary dict with counts.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    all_flagged = _collect_ft_flagged(review_db)

    if not all_flagged:
        logger.warning("No FT_FLAGGED papers found — nothing to export")
        return {"total": 0, "by_reason": {}}

    # Count by reason code
    reason_counts: dict[str, int] = {}
    for p in all_flagged:
        rc = p["reason_code"] or "unknown"
        reason_counts[rc] = reason_counts.get(rc, 0) + 1

    # Sort by reason code then title
    all_flagged.sort(key=lambda p: (p["reason_code"] or "zzz", p["title"]))

    if format == "html":
        from engine.adjudication.ft_adjudication_html import (
            generate_ft_adjudication_html,
        )
        out, html_stats = generate_ft_adjudication_html(
            review_name=review_name or "unknown",
            output_path=str(output_path),
        )
        output_path = out
    elif format == "xlsx":
        import sys
        print(
            "Warning: xlsx export is deprecated for interactive review. "
            "Use --format html (default). xlsx retained for reference exports.",
            file=sys.stderr,
        )
        _write_ft_xlsx(
            all_flagged, output_path, reason_counts,
            review_name=review_name or "unknown",
            review_spec=review_spec,
            db_path=str(review_db.db_path),
        )
    else:
        raise ValueError(f"Unsupported format: {format}")

    logger.info(
        "Exported FT adjudication queue: %d papers to %s",
        len(all_flagged), output_path,
    )
    for rc, count in sorted(reason_counts.items()):
        logger.info("  %s: %d", rc, count)

    return {
        "total": len(all_flagged),
        "by_reason": reason_counts,
        "output_path": str(output_path),
    }


def _write_ft_xlsx(
    papers: list[dict],
    output_path: Path,
    reason_counts: dict,
    review_name: str = "unknown",
    review_spec=None,
    db_path: str = "",
) -> None:
    """Write the FT adjudication queue as a self-documenting Excel workbook."""
    from engine.exporters.review_workbook import (
        ColumnDef,
        DecisionColumnDef,
        FreeTextColumnDef,
        InstructionsConfig,
        create_review_workbook,
    )

    rows = []
    for i, paper in enumerate(papers, 1):
        rows.append({
            "row_num": i,
            "paper_id": paper["paper_id"],
            "ee_identifier": paper["ee_identifier"],
            "reason_code": paper["reason_code"],
            "title": paper["title"],
            "abstract": paper["abstract"][:2000] if paper["abstract"] else "",
            "doi": paper["doi"],
            "pmid": paper["pmid"],
            "year": paper.get("year", ""),
            "journal": paper["journal"],
            "primary_decision": paper["primary_decision"],
            "primary_rationale": paper["primary_rationale"][:1000] if paper["primary_rationale"] else "",
            "verifier_decision": paper["verifier_decision"],
            "verifier_rationale": paper["verifier_rationale"][:1000] if paper["verifier_rationale"] else "",
            "text_excerpt": paper.get("text_excerpt", "")[:500],
        })

    columns = [
        ColumnDef(key="row_num", header="Row #", width=6),
        ColumnDef(key="paper_id", header="Paper ID", width=10),
        ColumnDef(key="ee_identifier", header="EE-ID", width=10),
        ColumnDef(key="reason_code", header="Reason Code", width=20),
        ColumnDef(key="title", header="Title", width=60, wrap=True),
        ColumnDef(key="abstract", header="Abstract", width=80, wrap=True),
        ColumnDef(key="doi", header="DOI", width=25),
        ColumnDef(key="pmid", header="PMID", width=12),
        ColumnDef(key="year", header="Year", width=6),
        ColumnDef(key="journal", header="Journal", width=30),
        ColumnDef(key="primary_decision", header="Primary Decision", width=15),
        ColumnDef(key="primary_rationale", header="Primary Rationale", width=50, wrap=True),
        ColumnDef(key="verifier_decision", header="Verifier Decision", width=15),
        ColumnDef(key="verifier_rationale", header="Verifier Rationale", width=50, wrap=True),
        ColumnDef(key="text_excerpt", header="Text Excerpt (~500 chars)", width=60, wrap=True),
    ]

    decision_columns = [
        DecisionColumnDef(
            key="PI_decision",
            header="PI_decision (FT_ELIGIBLE/FT_SCREENED_OUT)",
            valid_values=["FT_ELIGIBLE", "FT_SCREENED_OUT"],
            width=30,
        ),
    ]

    free_text_columns = [
        FreeTextColumnDef(key="PI_notes", header="PI_notes (optional)", width=30),
    ]

    import_cmd = (
        f"python -c \"\n"
        f"from engine.core.database import ReviewDatabase\n"
        f"from engine.adjudication import import_ft_adjudication_decisions\n"
        f"db = ReviewDatabase('{review_name}')\n"
        f"import_ft_adjudication_decisions(db, '{output_path}')\n"
        f"\""
    )

    decision_criteria = _build_ft_decision_criteria(review_spec)
    edge_case = _build_ft_edge_case_guidance(review_spec)

    instr = InstructionsConfig(
        review_name=review_name,
        review_spec_id=f"{review_spec.title} v{review_spec.version}" if review_spec else "",
        db_path=db_path,
        export_trigger=(
            f"{len(papers)} papers flagged during full-text screening where "
            f"primary and verification models disagreed"
        ),
        row_count=len(papers),
        decision_column_name="PI_decision (FT_ELIGIBLE/FT_SCREENED_OUT)",
        valid_values=["FT_ELIGIBLE", "FT_SCREENED_OUT"],
        decision_criteria=decision_criteria,
        edge_case_guidance=edge_case,
        import_command=import_cmd,
        columns_importer_reads=[
            "Paper ID (B)", "Title (E)", "Reason Code (D)",
            "PI_decision (P)", "PI_notes (Q)",
        ],
        columns_importer_ignores=(
            "All other columns (Row #, EE-ID, Abstract, DOI, PMID, Journal, "
            "Primary/Verifier Decision/Rationale, Text Excerpt) are read-only context."
        ),
    )

    reference_content = _build_ft_reference_content(review_spec)

    create_review_workbook(
        output_path=output_path,
        rows=rows,
        columns=columns,
        decision_columns=decision_columns,
        free_text_columns=free_text_columns,
        instructions=instr,
        reference_content=reference_content,
        reference_sheet_title="FT Screening Criteria",
    )


# ── Import ─────────────────────────────────────────────────────────


def import_ft_adjudication_decisions(
    review_db: ReviewDatabase,
    input_path: str | Path | None = None,
) -> dict:
    """Read completed FT adjudication decisions and write to database.

    If input_path is None, auto-discovers the decisions file using the
    naming convention: {review}_ft_adjudication_decisions.json

    Supports two input formats (detected by file extension):
      - .xlsx  — Excel workbook (from export_ft_adjudication_queue)
      - .json  — JSON array (from ft_adjudication_html.py HTML tool)

    JSON schema: [{paper_id: int, decision: "FT_ELIGIBLE"|"FT_SCREENED_OUT", note: str|null}]

    Validates the entire file before making any changes:
      - Rejects if any decision is missing or invalid

    Records decisions in ft_screening_adjudication table.
    Auto-advances FULL_TEXT_ADJUDICATION_COMPLETE if zero unresolved.

    Returns summary dict.
    """
    if input_path is None:
        from engine.core.naming import review_artifact_path
        review_name = Path(review_db.db_path).parent.name
        data_dir = Path(review_db.db_path).parent
        input_path = review_artifact_path(
            data_dir, review_name, "ft_adjudication", "decisions", "json",
        )
        if not input_path.exists():
            raise FileNotFoundError(
                f"Expected decisions file at: {input_path}\n"
                "Export from the HTML review tool first."
            )

    input_path = Path(input_path)

    if input_path.suffix.lower() == ".json":
        return _import_ft_json(review_db, input_path)

    return _import_ft_xlsx(review_db, input_path)


def _import_ft_json(
    review_db: ReviewDatabase,
    input_path: Path,
) -> dict:
    """Import FT adjudication decisions from JSON (HTML tool output).

    Expected schema: [{paper_id: int, decision: "FT_ELIGIBLE"|"FT_SCREENED_OUT", note: str|null}]
    """
    try:
        with open(input_path) as f:
            records = json.load(f)
    except (json.JSONDecodeError, OSError) as e:
        error_msg = f"\nIMPORT REJECTED — Cannot read JSON file: {e}\n"
        print(error_msg)
        logger.error(error_msg)
        return {
            "stats": {"ft_eligible": 0, "ft_screened_out": 0, "missing": 0, "invalid": 0, "total": 0},
            "warnings": [error_msg],
        }

    if not isinstance(records, list):
        error_msg = "\nIMPORT REJECTED — JSON must be an array of decision objects.\n"
        print(error_msg)
        logger.error(error_msg)
        return {
            "stats": {"ft_eligible": 0, "ft_screened_out": 0, "missing": 0, "invalid": 0, "total": 0},
            "warnings": [error_msg],
        }

    # ── Pass 1: Validate ─────────────────────────────────────────
    parsed_rows = []
    invalid_rows = []

    for i, rec in enumerate(records):
        paper_id = rec.get("paper_id")
        decision_raw = rec.get("decision", "")
        note = rec.get("note") or ""

        if not paper_id:
            invalid_rows.append(f"  Record {i}: missing paper_id")
            continue

        decision = str(decision_raw).strip().upper()
        if decision not in ("FT_ELIGIBLE", "FT_SCREENED_OUT"):
            invalid_rows.append(
                f"  Record {i} (paper {paper_id}): '{decision_raw}' "
                f"(must be FT_ELIGIBLE or FT_SCREENED_OUT)"
            )
            continue

        # Look up title and reason_code from the DB for the adjudication record
        row = review_db._conn.execute(
            "SELECT title FROM papers WHERE id = ?", (paper_id,)
        ).fetchone()
        title = row["title"] if row else ""

        ft_row = review_db._conn.execute(
            "SELECT reason_code FROM ft_screening_decisions "
            "WHERE paper_id = ? ORDER BY id DESC LIMIT 1",
            (paper_id,),
        ).fetchone()
        reason_code = ft_row["reason_code"] if ft_row else ""

        parsed_rows.append({
            "paper_id": paper_id,
            "title": title,
            "reason_code": reason_code,
            "decision": decision,
            "notes": note,
        })

    # ── Reject on validation failure ──────────────────────────────
    if invalid_rows:
        msg_parts = [
            f"\nIMPORT REJECTED — {len(invalid_rows)} validation error(s) found.",
            "No database changes were made.\n",
            f"INVALID RECORDS ({len(invalid_rows)}):",
        ]
        msg_parts.extend(invalid_rows)
        msg_parts.append("\nFix the JSON file and re-run.")

        error_msg = "\n".join(msg_parts)
        print(error_msg)
        logger.error(error_msg)
        return {
            "stats": {
                "ft_eligible": 0, "ft_screened_out": 0,
                "missing": 0, "invalid": len(invalid_rows),
                "total": len(parsed_rows) + len(invalid_rows),
            },
            "warnings": invalid_rows,
        }

    # ── Pass 2: Apply ─────────────────────────────────────────────
    return _apply_ft_decisions(review_db, parsed_rows)


def _apply_ft_decisions(
    review_db: ReviewDatabase,
    parsed_rows: list[dict],
) -> dict:
    """Shared logic: write validated decisions to DB and advance workflow."""
    ensure_adjudication_table(review_db._conn)

    now = datetime.now(timezone.utc).isoformat()
    stats = {
        "ft_eligible": 0, "ft_screened_out": 0,
        "missing": 0, "invalid": 0,
        "total": len(parsed_rows),
    }

    for pr in parsed_rows:
        decision = pr["decision"]
        paper_id = pr["paper_id"]
        title = pr["title"]
        reason_code = pr["reason_code"]
        notes = pr["notes"]

        # Record in ft_screening_adjudication table
        review_db._conn.execute(
            """INSERT INTO ft_screening_adjudication
               (paper_id, title, reason_code, adjudication_decision,
                adjudication_reason, adjudication_timestamp, created_at)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (paper_id, title, reason_code, decision, notes, now, now),
        )

        # Update paper status (FT_FLAGGED → FT_ELIGIBLE or FT_SCREENED_OUT)
        if paper_id:
            try:
                review_db.update_status(int(paper_id), decision)
            except ValueError as e:
                logger.warning("Paper %s: status update failed — %s", paper_id, e)

        if decision == "FT_ELIGIBLE":
            stats["ft_eligible"] += 1
        else:
            stats["ft_screened_out"] += 1

    review_db._conn.commit()

    # Success summary
    print(
        f"\nIMPORT SUCCESSFUL — {stats['total']} decisions processed.\n"
        f"  FT_ELIGIBLE:     {stats['ft_eligible']}\n"
        f"  FT_SCREENED_OUT: {stats['ft_screened_out']}\n"
        f"  Database updated."
    )

    logger.info(
        "FT adjudication import: %d eligible, %d screened out (of %d total)",
        stats["ft_eligible"], stats["ft_screened_out"], stats["total"],
    )

    # Auto-advance workflow: FULL_TEXT_ADJUDICATION_COMPLETE
    complete_stage(
        review_db._conn, "FULL_TEXT_ADJUDICATION_COMPLETE",
        metadata=(
            f"{stats['ft_eligible']} eligible, {stats['ft_screened_out']} screened out "
            f"(of {stats['total']} total)"
        ),
    )

    return {"stats": stats, "warnings": []}


def _import_ft_xlsx(
    review_db: ReviewDatabase,
    input_path: Path,
) -> dict:
    """Import FT adjudication decisions from Excel workbook."""
    from openpyxl import load_workbook

    wb = load_workbook(input_path)

    # Find the review queue sheet (support both old and new naming)
    sheet_name = None
    for name in ["Review Queue", "FT Review Queue"]:
        if name in wb.sheetnames:
            sheet_name = name
            break
    if sheet_name is None:
        error_msg = f"\nIMPORT REJECTED — No 'Review Queue' sheet found in {input_path}\n"
        print(error_msg)
        logger.error(error_msg)
        return {
            "stats": {"ft_eligible": 0, "ft_screened_out": 0, "missing": 0, "invalid": 0, "total": 0},
            "warnings": [error_msg],
        }

    ws = wb[sheet_name]

    # ── Build header index for robust column lookup ───────────
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_index = {}
    for idx, val in enumerate(header_row):
        if val:
            col_index[str(val).strip()] = idx

    def _find_col(candidates: list[str]) -> int | None:
        for c in candidates:
            for header, idx in col_index.items():
                if c.lower() in header.lower():
                    return idx
        return None

    title_col = _find_col(["Title"])
    paper_id_col = _find_col(["Paper ID"])
    reason_code_col = _find_col(["Reason Code", "Reason"])
    decision_col = _find_col(["PI_decision", "DECISION"])
    notes_col = _find_col(["PI_notes", "Notes"])
    row_num_col = _find_col(["Row #", "Row"])

    if title_col is None or decision_col is None:
        error_msg = (
            "\nIMPORT REJECTED — Required columns not found.\n"
            f"  Found headers: {list(col_index.keys())}\n"
            "  Required: Title, PI_decision (or DECISION)\n"
        )
        print(error_msg)
        logger.error(error_msg)
        return {
            "stats": {"ft_eligible": 0, "ft_screened_out": 0, "missing": 0, "invalid": 0, "total": 0},
            "warnings": [error_msg],
        }

    # ── Pass 1: Validate all rows before any DB writes ──────────
    parsed_rows = []
    blank_rows = []
    invalid_rows = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        title_val = row[title_col].value
        if not title_val:
            continue

        row_num = row[row_num_col].value if row_num_col is not None else row[0].row - 1
        paper_id = row[paper_id_col].value if paper_id_col is not None else None
        reason_code = row[reason_code_col].value if reason_code_col is not None else ""
        title = row[title_col].value
        decision_raw = row[decision_col].value
        notes = (row[notes_col].value or "") if notes_col is not None else ""

        if not decision_raw or str(decision_raw).strip() == "":
            blank_rows.append(
                f"  Row {row_num}: '{title[:60]}...'" if len(str(title)) > 60
                else f"  Row {row_num}: '{title}'"
            )
            continue

        decision = str(decision_raw).strip().upper()
        if decision not in ("FT_ELIGIBLE", "FT_SCREENED_OUT"):
            invalid_rows.append(
                f"  Row {row_num}: '{decision_raw}' (must be FT_ELIGIBLE or FT_SCREENED_OUT)"
            )
            continue

        parsed_rows.append({
            "row_num": row_num,
            "paper_id": paper_id,
            "reason_code": reason_code,
            "title": title,
            "decision": decision,
            "notes": notes,
        })

    # ── Reject on validation failure ────────────────────────────
    if blank_rows or invalid_rows:
        total_issues = len(blank_rows) + len(invalid_rows)
        msg_parts = [
            f"\nIMPORT REJECTED — {total_issues} validation error(s) found.",
            "No database changes were made.\n",
        ]
        if blank_rows:
            msg_parts.append(f"BLANK DECISION CELLS ({len(blank_rows)} rows):")
            msg_parts.extend(blank_rows)
            msg_parts.append("")
        if invalid_rows:
            msg_parts.append(f"INVALID DECISION VALUES ({len(invalid_rows)} rows):")
            msg_parts.extend(invalid_rows)
            msg_parts.append("")
        msg_parts.append("Fix the workbook and re-run the import command.")

        error_msg = "\n".join(msg_parts)
        print(error_msg)
        logger.error(error_msg)

        return {
            "stats": {
                "ft_eligible": 0,
                "ft_screened_out": 0,
                "missing": len(blank_rows),
                "invalid": len(invalid_rows),
                "total": len(parsed_rows) + len(blank_rows) + len(invalid_rows),
            },
            "warnings": blank_rows + invalid_rows,
        }

    # ── Pass 2: Apply all validated decisions ───────────────────
    return _apply_ft_decisions(review_db, parsed_rows)


# ── Pipeline Gate ──────────────────────────────────────────────────


def check_ft_adjudication_gate(review_db: ReviewDatabase) -> int:
    """Check for unresolved FT_FLAGGED papers. Returns count."""
    flagged = review_db.get_papers_by_status("FT_FLAGGED")
    return len(flagged)
