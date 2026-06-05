"""PI audit workbook generator v2 — verdict-transition + intra-rater audit.

Where v1 (``pi_audit_sampler.py``) sampled a single Pass 2 run stratified by
verdict, v2 audits the *delta* between two Pass 2 runs:

  - old run: ``surgical_autonomy_pass2_full_20260421T174729Z`` (original prompt)
  - new run: ``surgical_autonomy_pass2_codebook_v2_20260604T042317Z`` (codebook-aware)

joined at arm-row level over the 1,211 shared triples. Four strata, 200 rows:

  1. ``UNSUPPORTED_to_SUPPORTED`` — all arm-rows that flipped UNSUPPORTED→SUPPORTED
     (the largest leniency swing; expect 97).
  2. ``UNSUPPORTED_to_PARTIALLY`` — 60 seeded-random of the 100 that flipped
     UNSUPPORTED→PARTIALLY_SUPPORTED.
  3. ``SUPPORTED_to_UNSUPPORTED`` — all arm-rows that flipped the other way
     (new run *more* skeptical; expect 27).
  4. ``intra_rater_overlap`` — 16 rows re-drawn from the original (v1) n=100
     audit, ordinal-field-weighted, re-presented under the new ergonomics so
     within-rater consistency can be measured against the original adjudication.

Two xlsx files, fully blinded (mirrors v1): a blinded adjudication workbook and
a separate held-back key. The blinded sheet shows ONLY field_name + codebook
definition, arm_value, source context, a 4-state adjudication cell, and notes.
It withholds both judge verdicts, the judge reasoning, the arm identity, the
sampling stratum, and the model's cited evidence span.

Source-text policy (v2):
  - Full parsed text when it fits under a 30,000-char safe cell cap.
  - Otherwise a large span-centered window (±8,000 chars). NEVER a silent
    truncation — ``source_truncated`` is an explicit visible flag whenever
    windowed (the v1 truncation lesson).
  - A neutral locator (``»…«``) marks where arm_value occurs verbatim in the
    source, as a finding aid only. Most categorical values do not occur
    verbatim and get no locator. The model's evidence span is NEVER surfaced
    as "the evidence" — it only positions the window when arm_value is not
    locatable, and is never marked.

Deterministic. No DB writes.

Usage:
  PYTHONPATH=. python -m analysis.paper1.pi_audit_sampler_v2 \\
      --review surgical_autonomy \\
      --codebook data/surgical_autonomy/extraction_codebook.yaml \\
      --out-dir artifacts/paper1/pi_audit_v2
"""

from __future__ import annotations

import argparse
import copy
import hashlib
import logging
import random
import re
import sqlite3
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from analysis.paper1.judge_loader import _coerce_value
from analysis.paper1.judge_prompts import is_absence_claim
from analysis.paper1.pi_audit_sampler import (
    ARMS,
    BRAND_CHARCOAL,
    BRAND_MIST,
    BRAND_MIST_DARK,
    BRAND_TEAL,
    BRAND_WHITE,
    _brand_header_fill,
    _brand_header_font,
    _fetch_all_arm_snippets_for_field,
    _fetch_arm_value_and_snippet,
    _fetch_ee_identifier,
    _locate_span_char_offset,
    _read_paper_text,
    _set_column_widths,
    _thin_border,
    _wrap_align,
    _write_header,
)
from engine.core.database import ReviewDatabase

logger = logging.getLogger(__name__)


# ═════════════════════════════════════════════════════════════════════
# Fixed parameters — recorded verbatim in the Metadata sheets.
# ═════════════════════════════════════════════════════════════════════

PI_AUDIT_V2_CONFIG: dict = {
    "new_run_id": "surgical_autonomy_pass2_codebook_v2_20260604T042317Z",
    "old_run_id": "surgical_autonomy_pass2_full_20260421T174729Z",
    "master_seed": 20260605,  # v2 audit seed; distinct from v1 (20260422).
    "strata": {
        # mode "all": take every matching arm-row.
        # mode "sample": seeded random k of the matching arm-rows.
        # mode "overlap": re-draw from the original v1 audit (see below).
        "UNSUPPORTED_to_SUPPORTED": {
            "mode": "all", "old": "UNSUPPORTED", "new": "SUPPORTED",
        },
        "UNSUPPORTED_to_PARTIALLY": {
            "mode": "sample", "k": 60,
            "old": "UNSUPPORTED", "new": "PARTIALLY_SUPPORTED",
        },
        "SUPPORTED_to_UNSUPPORTED": {
            "mode": "all", "old": "SUPPORTED", "new": "UNSUPPORTED",
        },
        "intra_rater_overlap": {"mode": "overlap", "k": 16},
    },
    # Source of the original audit's per-row adjudication + identifiers.
    "original_audit_results_xlsx":
        "artifacts/paper1/pi_audit/pi_audit_results_2026-04-22T19-05-36Z.xlsx",
    # Authoritative arm values exactly as Pass 2 saw them. An empty cell here
    # means the arm extracted no value — i.e. it claims the field is absent.
    "pairs_csv":
        "data/surgical_autonomy/exports/disagreement_pairs_3arm.csv",
}

# Canonical display string for an arm that extracted no value (empty pairs-CSV
# cell). is_absence_claim() treats this as an absence assertion, so it gets no
# locator and a head/full-text window — matching how Pass 2 judged it.
ABSENCE_DISPLAY = "NOT REPORTED"

VERDICTS = ("UNSUPPORTED", "PARTIALLY_SUPPORTED", "SUPPORTED")
ADJUDICATION_STATES = ("SUPPORTED", "PARTIALLY_SUPPORTED", "UNSUPPORTED", "UNCLEAR")

# Excel per-cell hard limit. The safe cap leaves headroom under it for the
# locator markers + context preamble so a full-text cell can never spill.
EXCEL_HARD_CAP = 32_767
SAFE_CELL_CAP = 30_000
WINDOW_RADIUS_CHARS = 8_000

# A neutral finding aid — wrap the verbatim arm_value occurrence. Guillemets
# are visually distinct from any character likely to appear in clinical prose
# and carry no "this is the evidence" semantics.
LOCATOR_OPEN = "»"   # »
LOCATOR_CLOSE = "«"  # «
# Don't locate very short values — a bare "No"/"RCT" matches arbitrary prose
# and the mark would mislead rather than aid. Categorical values are expected
# to fall below this floor and therefore (correctly) get no locator.
MIN_LOCATOR_CHARS = 12

_WINDOW_MARKER = (
    "[Context: windowed view — a {radius:,}-char window centered on the "
    "relevant region of a {full:,}-char paper. source_truncated=TRUE; if the "
    "context shown is insufficient to decide, choose UNCLEAR.]\n\n"
)

# Inline accent (brand terracotta) — used ONLY for the source_truncated=TRUE
# flag text, never as a fill. One accent per surface.
BRAND_TERRACOTTA = "B85D3A"


# openpyxl's ILLEGAL_CHARACTERS_RE only strips C0 controls; parsed PDF/markdown
# also carries DEL + C1 controls (0x7f–0x9f), surrogates, and noncharacters
# that lxml rejects on save. Strip the full XML-1.0-illegal set here.
_XML_ILLEGAL_RE = re.compile(
    "[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f\ud800-\udfff￾￿]"
)


def _xml_safe(value):
    if isinstance(value, str):
        return _XML_ILLEGAL_RE.sub("", value)
    return value


class WindowAnchor:
    """How the window (when truncated) was centered. Key-only diagnostic."""
    FULL = "none_full_text"
    ARM_VALUE = "arm_value"      # centered on the verbatim arm_value occurrence
    MODEL_SPAN = "model_span"    # arm_value not locatable; centered on this
                                 # arm's span — positioning only, never shown
    COFIELD_SPAN = "cofield_span"  # this arm's span unlocatable; centered on
                                   # another arm's span for the SAME field (the
                                   # field's evidence region) — positioning only
    HEAD = "head"                # nothing locatable; head of paper


# ── Provenance ───────────────────────────────────────────────────────


def _sha256_file(path: Path) -> Optional[str]:
    try:
        h = hashlib.sha256()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(1 << 16), b""):
                h.update(chunk)
        return h.hexdigest()
    except OSError:
        return None


def _stratum_seed(master_seed: int, stratum: str) -> int:
    """SHA-256(master_seed || stratum), first 4 bytes as int."""
    key = f"{master_seed}\x1f{stratum}".encode("utf-8")
    return int.from_bytes(hashlib.sha256(key).digest()[:4], "big") % (2**31)


# ── Codebook field definitions ───────────────────────────────────────


def load_codebook_fields(codebook_path: Path) -> tuple[dict, dict]:
    """Return (family_by_field, definition_by_field).

    family_by_field maps field name -> judge_rubric_family.
    definition_by_field maps field name -> a human-readable rubric block
    (definition + instruction + value enumeration) for the blinded sheet.
    """
    cb = yaml.safe_load(codebook_path.read_text())
    family: dict[str, str] = {}
    definition: dict[str, str] = {}
    for node in cb.get("fields", []):
        if not isinstance(node, dict):
            continue
        name = node.get("name")
        if not name:
            continue
        if "judge_rubric_family" in node:
            family[name] = node["judge_rubric_family"]
        definition[name] = _compose_field_definition(node)
    return family, definition


def _compose_field_definition(node: dict) -> str:
    parts: list[str] = []
    d = node.get("definition")
    if d:
        parts.append(str(d).strip())
    instr = node.get("instruction")
    if instr:
        parts.append("Instruction: " + str(instr).strip())
    if node.get("ordered_values"):
        levels = " < ".join(str(v) for v in node["ordered_values"])
        parts.append("Ordered levels (low→high): " + levels)
    if node.get("valid_values"):
        lines = []
        for v in node["valid_values"]:
            if isinstance(v, dict):
                val = v.get("value", "")
                vd = v.get("definition", "")
                lines.append(f"  - {val}: {vd}".rstrip())
            else:
                lines.append(f"  - {v}")
        parts.append("Allowed values:\n" + "\n".join(lines))
    return "\n\n".join(parts)


# ── Cross-run join ───────────────────────────────────────────────────


def _load_run_arm_rows(
    conn: sqlite3.Connection, run_id: str,
) -> dict[tuple[str, str, str], tuple[int, str]]:
    """Map (paper_id, field_name, arm_name) -> (verification_id, verdict)."""
    out: dict[tuple[str, str, str], tuple[int, str]] = {}
    for vid, pid, fn, arm, verdict in conn.execute(
        """SELECT verification_id, paper_id, field_name, arm_name, verdict
           FROM fabrication_verifications WHERE judge_run_id = ?""",
        (run_id,),
    ):
        out[(pid, fn, arm)] = (vid, verdict)
    return out


def build_transitions(
    new_rows: dict[tuple[str, str, str], tuple[int, str]],
    old_rows: dict[tuple[str, str, str], tuple[int, str]],
) -> list[dict]:
    """Return one record per shared arm-row with old/new verdict + ids.

    Records are sorted by new verification_id for stable, reproducible
    downstream sampling.
    """
    shared = sorted(
        set(new_rows) & set(old_rows),
        key=lambda k: new_rows[k][0],
    )
    records = []
    for k in shared:
        pid, fn, arm = k
        new_vid, new_verdict = new_rows[k]
        old_vid, old_verdict = old_rows[k]
        records.append({
            "paper_id": pid, "field_name": fn, "arm_name": arm,
            "new_verification_id": new_vid, "old_verification_id": old_vid,
            "old_verdict": old_verdict, "new_verdict": new_verdict,
        })
    return records


# ── Original (v1) audit overlap source ───────────────────────────────


def load_original_audit_joined(results_xlsx: Path) -> list[dict]:
    """Read the JoinedRows sheet of the v1 unblinding results workbook.

    Returns one dict per original audit row with the identifiers and the
    PI's original adjudication needed for the intra-rater overlap stratum.
    """
    wb = load_workbook(results_xlsx, read_only=True)
    if "JoinedRows" not in wb.sheetnames:
        wb.close()
        raise RuntimeError(
            f"{results_xlsx} has no JoinedRows sheet; cannot build overlap "
            "stratum. Re-run the v1 unblinding first."
        )
    ws = wb["JoinedRows"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = list(rows[0])
    idx = {h: i for i, h in enumerate(header)}
    needed = ("paper_id", "field_name", "arm_name", "verification_id",
              "gemma_verdict", "pi_adjudication", "pi_notes")
    for col in needed:
        if col not in idx:
            raise RuntimeError(f"JoinedRows missing column {col!r}")
    out = []
    for r in rows[1:]:
        out.append({
            "paper_id": str(r[idx["paper_id"]]),
            "field_name": r[idx["field_name"]],
            "arm_name": r[idx["arm_name"]],
            "orig_verification_id": r[idx["verification_id"]],
            "orig_gemma_verdict": r[idx["gemma_verdict"]],
            "orig_pi_adjudication": r[idx["pi_adjudication"]],
            "orig_pi_notes": r[idx["pi_notes"]],
        })
    return out


# ── Selection ────────────────────────────────────────────────────────


def select_records(
    transitions: list[dict],
    original_joined: list[dict],
    new_rows: dict[tuple[str, str, str], tuple[int, str]],
    old_rows: dict[tuple[str, str, str], tuple[int, str]],
    family_by_field: dict[str, str],
    config: dict,
) -> tuple[list[dict], dict]:
    """Return (selected_records, selection_report).

    Each selected record carries: paper_id, field_name, arm_name,
    new/old verification_id, old/new verdict, stratum, judge_rubric_family,
    and (overlap only) orig_pi_adjudication / orig_pi_notes.
    """
    master_seed = config["master_seed"]
    by_transition: dict[tuple[str, str], list[dict]] = {}
    for rec in transitions:
        by_transition.setdefault(
            (rec["old_verdict"], rec["new_verdict"]), []
        ).append(rec)

    selected: list[dict] = []
    report: dict = {"strata": {}}

    # Strata 1–3: transition-defined.
    for name, spec in config["strata"].items():
        if spec["mode"] not in ("all", "sample"):
            continue
        pool = by_transition.get((spec["old"], spec["new"]), [])
        pool = sorted(pool, key=lambda r: r["new_verification_id"])
        if spec["mode"] == "all":
            chosen = pool
        else:
            k = spec["k"]
            if len(pool) < k:
                raise RuntimeError(
                    f"stratum {name}: pool has {len(pool)} < requested {k}"
                )
            rng = random.Random(_stratum_seed(master_seed, name))
            chosen = rng.sample(pool, k)
        for rec in chosen:
            r = dict(rec)
            r["stratum"] = name
            r["judge_rubric_family"] = family_by_field.get(
                rec["field_name"], "UNKNOWN")
            r["orig_pi_adjudication"] = None
            r["orig_pi_notes"] = None
            selected.append(r)
        report["strata"][name] = {
            "pool_size": len(pool), "selected": len(chosen),
        }

    # Stratum 4: intra-rater overlap, ordinal-weighted, dedup vs 1–3.
    overlap_spec = config["strata"]["intra_rater_overlap"]
    already = {
        (r["paper_id"], r["field_name"], r["arm_name"]) for r in selected
    }
    eligible = []
    for orow in original_joined:
        key = (orow["paper_id"], orow["field_name"], orow["arm_name"])
        if key in already:
            continue
        if key not in new_rows or key not in old_rows:
            # original-audit row not present in the shared join (e.g. the
            # one new-only triple has no old counterpart) — skip.
            continue
        eligible.append(orow)

    def is_ordinal(o):
        return family_by_field.get(o["field_name"]) == "ordinal"

    # Stable base ordering before the seeded shuffle.
    eligible.sort(key=lambda o: (
        int(o["paper_id"]) if str(o["paper_id"]).isdigit() else 0,
        o["field_name"], o["arm_name"],
    ))
    rng = random.Random(_stratum_seed(master_seed, "intra_rater_overlap"))
    ordinal_pool = [o for o in eligible if is_ordinal(o)]
    other_pool = [o for o in eligible if not is_ordinal(o)]
    rng.shuffle(ordinal_pool)
    rng.shuffle(other_pool)
    # Ordinal first, then fill.
    pool = ordinal_pool + other_pool
    k = overlap_spec["k"]
    if len(pool) < k:
        raise RuntimeError(
            f"intra_rater_overlap: only {len(pool)} eligible rows, need {k}"
        )
    chosen_overlap = pool[:k]
    n_ordinal_in_overlap = sum(1 for o in chosen_overlap if is_ordinal(o))

    for orow in chosen_overlap:
        key = (orow["paper_id"], orow["field_name"], orow["arm_name"])
        new_vid, new_verdict = new_rows[key]
        old_vid, old_verdict = old_rows[key]
        selected.append({
            "paper_id": orow["paper_id"],
            "field_name": orow["field_name"],
            "arm_name": orow["arm_name"],
            "new_verification_id": new_vid,
            "old_verification_id": old_vid,
            "old_verdict": old_verdict,
            "new_verdict": new_verdict,
            "stratum": "intra_rater_overlap",
            "judge_rubric_family": family_by_field.get(
                orow["field_name"], "UNKNOWN"),
            "orig_pi_adjudication": orow["orig_pi_adjudication"],
            "orig_pi_notes": orow["orig_pi_notes"],
        })
    report["strata"]["intra_rater_overlap"] = {
        "eligible": len(eligible),
        "ordinal_eligible": len(ordinal_pool),
        "selected": len(chosen_overlap),
        "ordinal_in_selection": n_ordinal_in_overlap,
        "filled_non_ordinal": len(chosen_overlap) - n_ordinal_in_overlap,
    }

    return selected, report


# ── Source-text construction (v2 windowing) ──────────────────────────


def _snap_to_boundary(text: str, start: int, end: int) -> tuple[int, int]:
    """Nudge window edges to nearby whitespace to avoid mid-word cuts."""
    if start > 0:
        sp = text.find(" ", start, min(len(text), start + 200))
        if sp != -1:
            start = sp + 1
    if end < len(text):
        sp = text.rfind(" ", max(0, end - 200), end)
        if sp != -1:
            end = sp
    return start, end


def _find_verbatim(text: str, value: str) -> Optional[int]:
    """Return first offset of value in text (exact, then case-insensitive),
    or None. Only meaningful for values >= MIN_LOCATOR_CHARS."""
    v = (value or "").strip()
    if len(v) < MIN_LOCATOR_CHARS:
        return None
    idx = text.find(v)
    if idx != -1:
        return idx
    low = text.lower().find(v.lower())
    return low if low != -1 else None


def build_source_text(
    full_text: str,
    arm_value: Optional[str],
    arm_snippet: Optional[str],
    cofield_snippets: Optional[list[Optional[str]]] = None,
) -> dict:
    """Return a dict describing the source cell for one row.

    Keys: source_text, source_truncated (bool), window_anchor,
    locator_applied (bool), locator_char_offset (int|None),
    window_start, window_end, full_text_chars, source_text_chars.

    cofield_snippets are the other arms' evidence spans for the SAME field,
    used only to position the window when this arm's value and span cannot
    be located. They are never shown or marked.
    """
    full_chars = len(full_text)
    value = (arm_value or "").strip()
    locatable_value = (
        not is_absence_claim(arm_value)
        and _find_verbatim(full_text, value) is not None
    )
    value_offset = _find_verbatim(full_text, value) if locatable_value else None

    if full_chars <= SAFE_CELL_CAP:
        window_start, window_end = 0, full_chars
        body = full_text
        truncated = False
        anchor = WindowAnchor.FULL
    else:
        truncated = True
        if value_offset is not None:
            center = value_offset + len(value) // 2
            anchor = WindowAnchor.ARM_VALUE
        else:
            span_offset = _locate_span_char_offset(full_text, arm_snippet)
            if span_offset is not None:
                center = span_offset
                anchor = WindowAnchor.MODEL_SPAN
            else:
                # This arm's span is unlocatable (often a paraphrased
                # snippet). Center on any co-field arm span that DOES
                # locate — the field's evidence lives in the same region.
                center = None
                for other in (cofield_snippets or []):
                    if other is None or other is arm_snippet:
                        continue
                    off = _locate_span_char_offset(full_text, other)
                    if off is not None:
                        center = off
                        anchor = WindowAnchor.COFIELD_SPAN
                        break
                if center is None:
                    center = 0
                    anchor = WindowAnchor.HEAD
        raw_start = max(0, center - WINDOW_RADIUS_CHARS)
        raw_end = min(full_chars, center + WINDOW_RADIUS_CHARS)
        window_start, window_end = _snap_to_boundary(
            full_text, raw_start, raw_end)
        body = full_text[window_start:window_end]

    # Locator: wrap the first verbatim occurrence of arm_value in the body
    # (full text or window). Only when the value is locatable in the body.
    locator_applied = False
    locator_char_offset = None
    if not is_absence_claim(arm_value):
        body_offset = _find_verbatim(body, value)
        if body_offset is not None:
            body = (
                body[:body_offset]
                + LOCATOR_OPEN + body[body_offset:body_offset + len(value)]
                + LOCATOR_CLOSE + body[body_offset + len(value):]
            )
            locator_applied = True
            locator_char_offset = window_start + body_offset

    source_text = body
    if truncated:
        marker = _WINDOW_MARKER.format(
            radius=2 * WINDOW_RADIUS_CHARS, full=full_chars)
        source_text = marker + body

    if len(source_text) > EXCEL_HARD_CAP:
        # Should be unreachable: full path is capped at SAFE_CELL_CAP and the
        # window path is bounded at ~16K + marker + 2 locator chars. Guard so
        # the v1 silent-truncation failure can never recur.
        raise RuntimeError(
            f"source_text {len(source_text)} chars exceeds Excel hard cap "
            f"{EXCEL_HARD_CAP} (anchor={anchor})"
        )

    return {
        "source_text": source_text,
        "source_truncated": truncated,
        "window_anchor": anchor,
        "locator_applied": locator_applied,
        "locator_char_offset": locator_char_offset,
        "window_start": window_start,
        "window_end": window_end,
        "full_text_chars": full_chars,
        "source_text_chars": len(source_text),
    }


# ── Enriched row ─────────────────────────────────────────────────────


@dataclass
class EnrichedRowV2:
    # Key-only identifiers / provenance
    paper_id: str
    ee_identifier: Optional[str]
    arm_name: str
    new_verification_id: int
    old_verification_id: int
    old_verdict: str
    new_verdict: str
    stratum: str
    judge_rubric_family: str
    orig_pi_adjudication: Optional[str]
    orig_pi_notes: Optional[str]
    # Key-only source diagnostics
    window_anchor: str
    locator_applied: bool
    locator_char_offset: Optional[int]
    window_start: int
    window_end: int
    full_text_chars: int
    source_text_chars: int
    # Blinded-sheet-visible
    field_name: str
    field_definition: str
    arm_value: str
    source_text: str
    source_truncated: bool
    # Assigned after shuffle
    row_id: int = 0


def load_pairs_values(
    csv_path: Path,
) -> dict[tuple[str, str], dict[str, Optional[str]]]:
    """Map (paper_id, field_name) -> {arm_name: value-or-None}.

    Mirrors judge_loader: arm values come from the disagreement-pairs CSV
    (the exact strings Pass 2 evaluated), coerced so empty/none/null → None
    (an absence claim).
    """
    import csv as _csv
    out: dict[tuple[str, str], dict[str, Optional[str]]] = {}
    with Path(csv_path).open(newline="") as f:
        for row in _csv.DictReader(f):
            key = (str(int(row["paper_id"])), row["field_name"])
            out[key] = {
                "local": _coerce_value(row.get("local_value")),
                "openai_o4_mini_high": _coerce_value(row.get("o4mini_value")),
                "anthropic_sonnet_4_6": _coerce_value(row.get("sonnet_value")),
            }
    return out


def enrich_rows(
    conn: sqlite3.Connection,
    review_dir: Path,
    selected: list[dict],
    definition_by_field: dict[str, str],
    pairs_values: dict[tuple[str, str], dict[str, Optional[str]]],
) -> list[EnrichedRowV2]:
    out: list[EnrichedRowV2] = []
    paper_cache: dict[str, Optional[str]] = {}
    for rec in selected:
        pid = rec["paper_id"]
        fname = rec["field_name"]
        aname = rec["arm_name"]
        if pid not in paper_cache:
            paper_cache[pid] = _read_paper_text(review_dir, pid)
        full_text = paper_cache[pid] or ""
        if not full_text:
            raise RuntimeError(
                f"No parsed text for paper_id={pid} "
                f"(field={fname}, arm={aname}); cannot build source_text"
            )
        cell = pairs_values.get((pid, fname))
        if cell is None or aname not in cell:
            raise RuntimeError(
                f"(paper_id={pid}, field={fname}) absent from pairs CSV; "
                "cannot recover the arm value Pass 2 evaluated"
            )
        raw_value = cell[aname]
        # Empty pairs cell → arm extracted nothing → absence claim. Render
        # the canonical absence label so the PI judges whether the paper
        # truly omits the field, exactly as Pass 2 framed it.
        arm_value = raw_value if raw_value is not None else ABSENCE_DISPLAY
        # Snippets are only used to position a window when the value is not
        # locatable; pull from the extraction store (None for absence). The
        # co-field snippets rescue rows whose own span is unlocatable.
        _, arm_snippet = _fetch_arm_value_and_snippet(conn, pid, fname, aname)
        cofield_snippets = _fetch_all_arm_snippets_for_field(conn, pid, fname)
        src = build_source_text(
            full_text, str(arm_value), arm_snippet, cofield_snippets)
        out.append(EnrichedRowV2(
            paper_id=pid,
            ee_identifier=_fetch_ee_identifier(conn, pid),
            arm_name=aname,
            new_verification_id=rec["new_verification_id"],
            old_verification_id=rec["old_verification_id"],
            old_verdict=rec["old_verdict"],
            new_verdict=rec["new_verdict"],
            stratum=rec["stratum"],
            judge_rubric_family=rec["judge_rubric_family"],
            orig_pi_adjudication=rec["orig_pi_adjudication"],
            orig_pi_notes=rec["orig_pi_notes"],
            window_anchor=src["window_anchor"],
            locator_applied=src["locator_applied"],
            locator_char_offset=src["locator_char_offset"],
            window_start=src["window_start"],
            window_end=src["window_end"],
            full_text_chars=src["full_text_chars"],
            source_text_chars=src["source_text_chars"],
            field_name=fname,
            field_definition=definition_by_field.get(fname, ""),
            arm_value=str(arm_value),
            source_text=src["source_text"],
            source_truncated=src["source_truncated"],
        ))
    return out


def randomize(rows: list[EnrichedRowV2], master_seed: int) -> list[EnrichedRowV2]:
    """Shuffle with seed=(master_seed + 1) and assign row_id 1..N."""
    rng = random.Random(master_seed + 1)
    shuffled = list(rows)
    rng.shuffle(shuffled)
    for i, r in enumerate(shuffled, 1):
        r.row_id = i
    return shuffled


# ── Workbook writers ─────────────────────────────────────────────────

BLINDED_HEADER = [
    "row_id", "field_name", "field_definition", "arm_value",
    "source_truncated", "source_text", "adjudication", "notes",
    "adjudicated_at",
]
# 1-based column indices that legitimately hold free prose / data and are
# therefore excluded from the forbidden-substring blinding scan.
_CONTENT_COLS = {3, 4, 6}  # field_definition, arm_value, source_text


def _write_instructions_sheet(ws) -> None:
    ws.title = "Instructions"
    heading = Font(name="Calibri", size=14, bold=True, color=BRAND_TEAL)
    sub = Font(name="Calibri", size=12, bold=True, color=BRAND_TEAL)
    body = Font(name="Calibri", size=11, color=BRAND_CHARCOAL)
    lines: list[tuple[str, Font]] = [
        ("PI Audit v2 — Blinded Adjudication", heading),
        ("", body),
        ("Purpose", sub),
        ("This workbook holds 200 arm-rows. For each, judge whether arm_value "
         "is grounded in the source_text and record your own adjudication. You "
         "are blinded to which model produced each row and to any machine "
         "verdict.",
         body),
        ("", body),
        ("Verdict definitions", sub),
        ("SUPPORTED — arm_value is directly and fully grounded in the "
         "source_text (exact match or trivial paraphrase).", body),
        ("PARTIALLY_SUPPORTED — the source_text partially grounds arm_value: "
         "overlapping content, but one is less complete, less specific, or "
         "adds unverified detail.", body),
        ("UNSUPPORTED — the source_text does not ground arm_value; it "
         "contradicts, exceeds, or is absent from the source.", body),
        ("UNCLEAR — the context is ambiguous, the field is not present in the "
         "visible text, or you cannot decide from what is shown. Prefer "
         "UNCLEAR over a forced call.", body),
        ("", body),
        ("Absence values", sub),
        ("If arm_value is an absence sentinel (NR, N/A, NA, NOT_FOUND, "
         "NOT FOUND, NOT REPORTED, or empty), judge whether the paper truly "
         "omits this field: SUPPORTED = the paper indeed does not report it; "
         "UNSUPPORTED = the paper does report it and the value missed it.",
         body),
        ("", body),
        ("Source text & the locator", sub),
        ("source_text is the parsed paper. When the paper is longer than the "
         "safe cell limit, you see a large window centered on the relevant "
         "region — in that case source_truncated is TRUE and a context note "
         "appears at the top of the cell. Text is NEVER silently cut.",
         body),
        ("Where the exact arm_value text appears in the source, it is wrapped "
         "in » guillemets « purely as a finding aid so you don't have "
         "to scan the whole passage. Many values (especially categorical "
         "codes) do not appear verbatim and so carry no locator — that is "
         "expected and is NOT evidence of absence. Judge the substance, not "
         "the presence of a marker.",
         body),
        ("", body),
        ("Working columns", sub),
        ("adjudication (dropdown) and notes are yours to fill; "
         "adjudicated_at is an optional date. All other columns are reference.",
         body),
        ("", body),
        ("Sign-off", sub),
        ("Completed on (YYYY-MM-DD):", body),
        ("Signed:", body),
    ]
    for i, (text, fnt) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = fnt
        c.alignment = _wrap_align()
    _set_column_widths(ws, [("A", 120)])


def _write_adjudication_sheet(ws, rows: list[EnrichedRowV2]) -> None:
    ws.title = "Adjudication"
    _write_header(ws, BLINDED_HEADER)
    ws.freeze_panes = "A2"

    locked = Font(name="Calibri", size=11, color=BRAND_CHARCOAL)
    accent = Font(name="Calibri", size=11, bold=True, color=BRAND_TERRACOTTA)
    mist = PatternFill(start_color=BRAND_MIST, end_color=BRAND_MIST,
                       fill_type="solid")
    mist_dark = PatternFill(start_color=BRAND_MIST_DARK,
                            end_color=BRAND_MIST_DARK, fill_type="solid")
    border = _thin_border()

    for i, r in enumerate(rows, 2):
        alt = mist if i % 2 == 0 else mist_dark
        values = [
            r.row_id, r.field_name, r.field_definition, r.arm_value,
            "TRUE" if r.source_truncated else "FALSE",
            r.source_text, "", "", "",
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=i, column=col, value=_xml_safe(val))
            c.alignment = _wrap_align()
            # source_truncated flag in brand accent when TRUE.
            if col == 5 and r.source_truncated:
                c.font = accent
            else:
                c.font = locked
            c.fill = alt
            c.border = border

    dv = DataValidation(
        type="list",
        formula1='"SUPPORTED,PARTIALLY_SUPPORTED,UNSUPPORTED,UNCLEAR"',
        allow_blank=True, showErrorMessage=True,
        errorTitle="Invalid adjudication",
        error="Choose one of: SUPPORTED, PARTIALLY_SUPPORTED, UNSUPPORTED, "
              "UNCLEAR.",
    )
    ws.add_data_validation(dv)
    adj_col = get_column_letter(BLINDED_HEADER.index("adjudication") + 1)
    dv.add(f"{adj_col}2:{adj_col}{len(rows) + 1}")

    _set_column_widths(ws, [
        ("A", 7), ("B", 22), ("C", 60), ("D", 36),
        ("E", 14), ("F", 96), ("G", 22), ("H", 40), ("I", 18),
    ])
    for i in range(2, len(rows) + 2):
        ws.row_dimensions[i].height = 64


KEY_HEADER = [
    "row_id", "stratum", "old_verdict", "new_verdict", "arm_name",
    "judge_rubric_family", "paper_id", "ee_identifier", "field_name",
    "arm_value", "new_verification_id", "old_verification_id",
    "orig_pi_adjudication", "orig_pi_notes",
    "source_truncated", "window_anchor", "locator_applied",
    "locator_char_offset", "window_start", "window_end",
    "full_text_chars", "source_text_chars",
]


def _write_key_sheet(ws, rows: list[EnrichedRowV2]) -> None:
    ws.title = "Key"
    _write_header(ws, KEY_HEADER)
    ws.freeze_panes = "A2"
    locked = Font(name="Calibri", size=11, color=BRAND_CHARCOAL)
    border = _thin_border()
    for i, r in enumerate(rows, 2):
        values = [
            r.row_id, r.stratum, r.old_verdict, r.new_verdict, r.arm_name,
            r.judge_rubric_family, r.paper_id, r.ee_identifier, r.field_name,
            r.arm_value, r.new_verification_id, r.old_verification_id,
            r.orig_pi_adjudication, r.orig_pi_notes,
            "TRUE" if r.source_truncated else "FALSE",
            r.window_anchor, "TRUE" if r.locator_applied else "FALSE",
            r.locator_char_offset, r.window_start, r.window_end,
            r.full_text_chars, r.source_text_chars,
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=i, column=col, value=_xml_safe(val))
            c.alignment = _wrap_align()
            c.font = locked
            c.border = border
    _set_column_widths(ws, [
        ("A", 7), ("B", 26), ("C", 20), ("D", 20), ("E", 22),
        ("F", 20), ("G", 9), ("H", 12), ("I", 22), ("J", 34),
        ("K", 14), ("L", 14), ("M", 18), ("N", 40),
        ("O", 14), ("P", 16), ("Q", 14), ("R", 16), ("S", 12),
        ("T", 12), ("U", 14), ("V", 14),
    ])


def _write_metadata_sheet(ws, config: dict, extra: dict) -> None:
    ws.title = "Metadata"
    _write_header(ws, ["field", "value"])
    rows: list[tuple[str, str]] = [
        ("new_run_id", config["new_run_id"]),
        ("old_run_id", config["old_run_id"]),
        ("master_seed", str(config["master_seed"])),
        ("safe_cell_cap_chars", str(SAFE_CELL_CAP)),
        ("window_radius_chars", str(WINDOW_RADIUS_CHARS)),
        ("excel_hard_cap_chars", str(EXCEL_HARD_CAP)),
        ("min_locator_chars", str(MIN_LOCATOR_CHARS)),
    ]
    for k, v in extra.items():
        rows.append((k, str(v)))
    for i, (k, v) in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=k).font = Font(
            bold=True, color=BRAND_CHARCOAL)
        ws.cell(row=i, column=2, value=_xml_safe(v)).alignment = \
            _wrap_align()
    _set_column_widths(ws, [("A", 40), ("B", 90)])


# ── Validation ───────────────────────────────────────────────────────

_FORBIDDEN_STRINGS = (
    "openai", "anthropic", "gemma", "sonnet", "o4_mini", "o4-mini",
    "verdict", "reasoning", "fabrication_hypothesis", "stratum",
    "sampling_stratum",
)


def _check_not_sorted(values: list) -> bool:
    """True if the sequence is NOT monotonic (catches ordering leakage)."""
    vals = [v for v in values if v is not None]
    if len(vals) < 2:
        return True
    asc = all(a <= b for a, b in zip(vals, vals[1:]))
    desc = all(a >= b for a, b in zip(vals, vals[1:]))
    return not (asc or desc)


def validate_pre_write(rows: list[EnrichedRowV2], config: dict) -> dict:
    n = len(rows)
    # Unique (paper, field, arm) — no arm-row presented twice.
    keys = [(r.paper_id, r.field_name, r.arm_name) for r in rows]
    assert len(set(keys)) == n, "duplicate (paper, field, arm) in selection"
    # row_id 1..n
    rids = sorted(r.row_id for r in rows)
    assert rids == list(range(1, n + 1)), "row_id not 1..n"
    # Non-empty visible fields
    for r in rows:
        assert r.arm_value.strip(), f"row_id={r.row_id} empty arm_value"
        assert r.source_text.strip(), f"row_id={r.row_id} empty source_text"
        assert len(r.source_text) <= EXCEL_HARD_CAP, (
            f"row_id={r.row_id} source_text over Excel cap")
        # source_truncated flag must match presence of windowing.
        assert r.source_truncated == (r.window_anchor != WindowAnchor.FULL), (
            f"row_id={r.row_id} source_truncated flag inconsistent with anchor")
    # Anti-sort: row order must not leak the design.
    ordered = sorted(rows, key=lambda r: r.row_id)
    for label, key in [
        ("paper_id", lambda r: int(r.paper_id) if r.paper_id.isdigit() else 0),
        ("field_name", lambda r: r.field_name),
        ("stratum", lambda r: r.stratum),
        ("arm_name", lambda r: r.arm_name),
        ("new_verdict", lambda r: r.new_verdict),
    ]:
        assert _check_not_sorted([key(r) for r in ordered]), (
            f"row order monotonic in {label} — blinding compromised")

    stratum_totals: dict[str, int] = {}
    arm_totals: dict[str, int] = {}
    family_by_stratum: dict[str, dict[str, int]] = {}
    arm_by_stratum: dict[str, dict[str, int]] = {}
    for r in rows:
        stratum_totals[r.stratum] = stratum_totals.get(r.stratum, 0) + 1
        arm_totals[r.arm_name] = arm_totals.get(r.arm_name, 0) + 1
        family_by_stratum.setdefault(r.stratum, {})
        family_by_stratum[r.stratum][r.judge_rubric_family] = (
            family_by_stratum[r.stratum].get(r.judge_rubric_family, 0) + 1)
        arm_by_stratum.setdefault(r.stratum, {})
        arm_by_stratum[r.stratum][r.arm_name] = (
            arm_by_stratum[r.stratum].get(r.arm_name, 0) + 1)

    return {
        "n_rows": n,
        "stratum_totals": stratum_totals,
        "arm_totals": arm_totals,
        "family_by_stratum": family_by_stratum,
        "arm_by_stratum": arm_by_stratum,
        "truncated_rows": sum(1 for r in rows if r.source_truncated),
        "full_text_rows": sum(1 for r in rows if not r.source_truncated),
        "locator_rows": sum(1 for r in rows if r.locator_applied),
        "anchor_totals": {
            a: sum(1 for r in rows if r.window_anchor == a)
            for a in (WindowAnchor.FULL, WindowAnchor.ARM_VALUE,
                      WindowAnchor.MODEL_SPAN, WindowAnchor.COFIELD_SPAN,
                      WindowAnchor.HEAD)
        },
    }


def _forbidden_hits(blinded_path: Path) -> list[str]:
    wb = load_workbook(blinded_path)
    ws = wb["Adjudication"]
    pats = {s: re.compile(re.escape(s), re.IGNORECASE)
            for s in _FORBIDDEN_STRINGS}
    hits: list[str] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.column in _CONTENT_COLS:
                continue
            v = cell.value
            if not isinstance(v, str):
                continue
            for s, pat in pats.items():
                if pat.search(v):
                    hits.append(f"{s!r} @ {cell.coordinate}: {v[:60]!r}")
    wb.close()
    return hits


def validate_post_write(
    blinded_path: Path, key_path: Path, n: int,
) -> dict:
    expected_max = n + 1
    wb_b = load_workbook(blinded_path)
    assert set(wb_b.sheetnames) == {"Instructions", "Adjudication", "Metadata"}, (
        f"blinded sheets: {wb_b.sheetnames}")
    ws = wb_b["Adjudication"]
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    assert header == BLINDED_HEADER, f"blinded header: {header}"
    assert ws.max_row == expected_max, (
        f"Adjudication rows {ws.max_row} != {expected_max}")
    # Hard structural guarantee: none of the key-only column names exist here.
    leaked = set(header) & {
        "arm_name", "stratum", "old_verdict", "new_verdict",
        "judge_rubric_family", "paper_id", "verification_id",
        "new_verification_id", "old_verification_id", "orig_pi_adjudication",
    }
    assert not leaked, f"blinded sheet leaks key columns: {leaked}"
    wb_b.close()

    wb_k = load_workbook(key_path)
    assert set(wb_k.sheetnames) == {"Key", "Metadata"}, (
        f"key sheets: {wb_k.sheetnames}")
    assert wb_k["Key"].max_row == expected_max, "key row count mismatch"
    wb_k.close()

    hits = _forbidden_hits(blinded_path)
    assert hits == [], "blinded sheet forbidden strings:\n" + "\n".join(hits[:20])

    # row_id pairing
    wb_b = load_workbook(blinded_path)
    wb_k = load_workbook(key_path)
    bids = {r[0].value for r in wb_b["Adjudication"].iter_rows(
        min_row=2, max_row=expected_max, max_col=1)}
    kids = {r[0].value for r in wb_k["Key"].iter_rows(
        min_row=2, max_row=expected_max, max_col=1)}
    wb_b.close()
    wb_k.close()
    assert bids == kids == set(range(1, n + 1)), "row_id sets differ"
    return {"forbidden_hits": len(hits)}


# ── Driver ───────────────────────────────────────────────────────────


def generate(
    conn: sqlite3.Connection,
    review_dir: Path,
    codebook_path: Path,
    out_dir: Path,
    config: Optional[dict] = None,
    *,
    run_metadata_extras: Optional[dict] = None,
    timestamp: Optional[str] = None,
) -> tuple[Path, Path, dict]:
    cfg = copy.deepcopy(config if config is not None else PI_AUDIT_V2_CONFIG)
    family_by_field, definition_by_field = load_codebook_fields(codebook_path)

    new_rows = _load_run_arm_rows(conn, cfg["new_run_id"])
    old_rows = _load_run_arm_rows(conn, cfg["old_run_id"])
    transitions = build_transitions(new_rows, old_rows)

    results_xlsx = Path(cfg["original_audit_results_xlsx"])
    original_joined = load_original_audit_joined(results_xlsx)

    pairs_csv = Path(cfg["pairs_csv"])
    pairs_values = load_pairs_values(pairs_csv)

    selected, sel_report = select_records(
        transitions, original_joined, new_rows, old_rows,
        family_by_field, cfg)
    enriched = enrich_rows(
        conn, review_dir, selected, definition_by_field, pairs_values)
    enriched = randomize(enriched, cfg["master_seed"])
    summary = validate_pre_write(enriched, cfg)
    summary["selection_report"] = sel_report

    out_dir.mkdir(parents=True, exist_ok=True)
    ts = timestamp or datetime.now(timezone.utc).strftime("%Y-%m-%dT%H-%M-%SZ")
    blinded_path = out_dir / f"pi_audit_v2_workbook_{ts}.xlsx"
    key_path = out_dir / f"pi_audit_v2_key_{ts}.xlsx"

    extras = dict(run_metadata_extras or {})
    extras.update({
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "shared_arm_rows": len(transitions),
        "n_rows": summary["n_rows"],
        "input_sha256.codebook": _sha256_file(codebook_path),
        "input_sha256.original_audit_results_xlsx": _sha256_file(results_xlsx),
        "input_sha256.pairs_csv": _sha256_file(pairs_csv),
        "truncated_rows": summary["truncated_rows"],
        "full_text_rows": summary["full_text_rows"],
        "locator_rows": summary["locator_rows"],
    })
    for s, n in summary["stratum_totals"].items():
        extras[f"stratum.{s}"] = n
    for s in cfg["strata"]:
        extras[f"stratum_seed.{s}"] = _stratum_seed(cfg["master_seed"], s)
    for a, n in summary["anchor_totals"].items():
        extras[f"anchor.{a}"] = n

    wb_b = Workbook()
    _write_instructions_sheet(wb_b.active)
    _write_adjudication_sheet(wb_b.create_sheet("Adjudication"), enriched)
    _write_metadata_sheet(wb_b.create_sheet("Metadata"), cfg, extras)
    wb_b.save(blinded_path)

    wb_k = Workbook()
    _write_key_sheet(wb_k.active, enriched)
    _write_metadata_sheet(wb_k.create_sheet("Metadata"), cfg, extras)
    wb_k.save(key_path)

    post = validate_post_write(blinded_path, key_path, summary["n_rows"])
    summary.update(post)
    summary["blinded_path"] = str(blinded_path)
    summary["key_path"] = str(key_path)
    return blinded_path, key_path, summary


# ── CLI ──────────────────────────────────────────────────────────────


def _run_exists(conn: sqlite3.Connection, run_id: str) -> bool:
    return conn.execute(
        "SELECT 1 FROM judge_runs WHERE run_id = ?", (run_id,)
    ).fetchone() is not None


def _build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="analysis.paper1.pi_audit_sampler_v2")
    p.add_argument("--review", required=True)
    p.add_argument("--codebook", type=Path, required=True)
    p.add_argument("--pairs-csv", type=Path, default=None,
                   help="Override the disagreement-pairs CSV (defaults to the "
                        "path baked into PI_AUDIT_V2_CONFIG).")
    p.add_argument("--out-dir", type=Path,
                   default=Path("artifacts/paper1/pi_audit_v2"))
    p.add_argument("--data-root", type=Path, default=None)
    return p


def _fmt_report(summary: dict) -> str:
    lines = []
    sr = summary["selection_report"]["strata"]
    lines.append(f"n_rows: {summary['n_rows']}")
    lines.append("")
    lines.append("Stratum composition:")
    for s, n in summary["stratum_totals"].items():
        info = sr.get(s, {})
        extra = ""
        if "pool_size" in info:
            extra = f" (pool={info['pool_size']}, selected={info['selected']})"
        elif "eligible" in info:
            extra = (f" (eligible={info['eligible']}, "
                     f"ordinal_eligible={info['ordinal_eligible']}, "
                     f"ordinal_in_selection={info['ordinal_in_selection']}, "
                     f"filled_non_ordinal={info['filled_non_ordinal']})")
        lines.append(f"  {s}: {n}{extra}")
    lines.append("")
    lines.append("Family composition by stratum:")
    for s, fam in summary["family_by_stratum"].items():
        fam_str = ", ".join(f"{k}={v}" for k, v in sorted(fam.items()))
        lines.append(f"  {s}: {fam_str}")
    lines.append("")
    lines.append("Arm composition by stratum:")
    for s, arm in summary["arm_by_stratum"].items():
        arm_str = ", ".join(f"{k}={v}" for k, v in sorted(arm.items()))
        lines.append(f"  {s}: {arm_str}")
    lines.append("")
    lines.append(f"Arm totals: {summary['arm_totals']}")
    lines.append(f"Source: full_text={summary['full_text_rows']}, "
                 f"truncated/windowed={summary['truncated_rows']}, "
                 f"locator_applied={summary['locator_rows']}")
    lines.append(f"Window anchors: {summary['anchor_totals']}")
    lines.append(f"Forbidden-string hits in blinded workbook: "
                 f"{summary['forbidden_hits']}")
    return "\n".join(lines)


def run(argv: Optional[list[str]] = None) -> int:
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    args = _build_arg_parser().parse_args(argv)

    db = (ReviewDatabase(args.review, data_root=args.data_root)
          if args.data_root else ReviewDatabase(args.review))
    try:
        for key in ("new_run_id", "old_run_id"):
            rid = PI_AUDIT_V2_CONFIG[key]
            if not _run_exists(db._conn, rid):
                raise SystemExit(f"{key} not found in judge_runs: {rid}")
        cfg = copy.deepcopy(PI_AUDIT_V2_CONFIG)
        if args.pairs_csv is not None:
            cfg["pairs_csv"] = str(args.pairs_csv)
        blinded, key_path, summary = generate(
            db._conn, db.db_path.parent, args.codebook, args.out_dir,
            config=cfg)
    finally:
        db.close()

    print(f"blinded workbook: {blinded.resolve()}")
    print(f"key workbook:     {key_path.resolve()}")
    print()
    print(_fmt_report(summary))
    return 0


if __name__ == "__main__":
    raise SystemExit(run())
