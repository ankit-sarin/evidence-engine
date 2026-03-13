"""Tests for the screening adjudication pipeline."""

import csv
import json
import tempfile
from datetime import datetime, timezone
from pathlib import Path

import pytest

from engine.adjudication.categorizer import (
    CategoryConfig,
    categorize_paper,
    generate_starter_config,
    get_category_descriptions,
    load_config,
)
from engine.adjudication.schema import ensure_adjudication_table
from engine.adjudication.screening_adjudicator import (
    _collect_db_flagged,
    _collect_expanded_flagged,
    check_adjudication_gate,
    export_adjudication_queue,
    import_adjudication_decisions,
)
from engine.core.database import ReviewDatabase
from engine.search.models import Citation


# ── Fixtures ────────────────────────────────────────────────────────


@pytest.fixture
def tmp_db(tmp_path):
    """Create a ReviewDatabase in a temp directory."""
    db = ReviewDatabase("test_review", data_root=tmp_path)
    yield db
    db.close()


@pytest.fixture
def flagged_db(tmp_db):
    """DB with some papers in ABSTRACT_SCREEN_FLAGGED status."""
    cits = [
        Citation(
            title="Deep Learning Surgical Tool Segmentation",
            abstract="We propose a CNN for detecting instruments in endoscopic images.",
            pmid="11111111", doi="10.1000/test1", source="pubmed",
            authors=["A"], journal="J Test", year=2024,
        ),
        Citation(
            title="Autonomous Suturing Robot for Tissue Closure",
            abstract="A robot that autonomously performs running sutures on porcine tissue.",
            pmid="22222222", doi="10.1000/test2", source="pubmed",
            authors=["B"], journal="J Surg", year=2024,
        ),
        Citation(
            title="Systematic Review of Robotic Surgery Autonomy",
            abstract="We conducted a systematic review and meta-analysis of autonomous surgical systems.",
            pmid="33333333", doi="10.1000/test3", source="pubmed",
            authors=["C"], journal="J Rev", year=2024,
        ),
    ]
    tmp_db.add_papers(cits)
    for pid in [1, 2, 3]:
        tmp_db.add_screening_decision(pid, 1, "include", "maybe relevant", "qwen3:8b")
        tmp_db.add_screening_decision(pid, 2, "exclude", "not sure", "qwen3:8b")
        tmp_db.update_status(pid, "ABSTRACT_SCREEN_FLAGGED")
    return tmp_db


@pytest.fixture
def expanded_dir(tmp_path):
    """Create mock expanded search CSV files."""
    edir = tmp_path / "expanded_search"
    edir.mkdir()

    # abstracts.jsonl
    abstracts = [
        {"key": "10.1000/exp1", "doi": "10.1000/exp1", "pmid": "", "title": "Force Sensor Design for Surgery",
         "abstract": "We present a novel piezoelectric force sensor for robotic instruments.",
         "year": "2024", "journal": "Sensors", "source": "openalex"},
        {"key": "10.1000/exp2", "doi": "10.1000/exp2", "pmid": "44444444", "title": "Robot-Assisted Autonomous Bone Cutting",
         "abstract": "An autonomous robotic system that performs femoral osteotomies.",
         "year": "2024", "journal": "J Ortho", "source": "pubmed"},
    ]
    with open(edir / "abstracts.jsonl", "w") as f:
        for a in abstracts:
            f.write(json.dumps(a) + "\n")

    # screening_results.csv — one flagged
    with open(edir / "screening_results.csv", "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["title", "doi", "pmid", "year", "journal", "source",
                     "has_abstract", "screening_decision",
                     "pass1_decision", "pass1_rationale", "pass1_confidence",
                     "pass2_decision", "pass2_rationale", "pass2_confidence"])
        w.writerow(["Force Sensor Design for Surgery", "10.1000/exp1", "", "2024",
                     "Sensors", "openalex", "yes", "flagged",
                     "exclude", "hardware only", "0.9",
                     "include", "might be relevant", "0.6"])
        w.writerow(["Some Excluded Paper", "10.1000/exp_ex", "", "2024",
                     "Other", "pubmed", "yes", "exclude",
                     "exclude", "not relevant", "0.95",
                     "exclude", "confirmed", "0.95"])

    # verification_results.csv — one flagged
    with open(edir / "verification_results.csv", "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["title", "doi", "pmid", "year", "journal", "source",
                     "primary_decision", "verification_decision", "final_decision",
                     "verification_rationale", "verification_confidence"])
        w.writerow(["Robot-Assisted Autonomous Bone Cutting", "10.1000/exp2", "44444444",
                     "2024", "J Ortho", "pubmed",
                     "include", "exclude", "flagged",
                     "No clear autonomous execution", "0.8"])
        w.writerow(["Confirmed Paper", "10.1000/exp_ok", "", "2024",
                     "Good", "pubmed", "include", "include", "include",
                     "Clearly autonomous", "0.95"])

    return edir


@pytest.fixture
def surgical_autonomy_config():
    """Load the actual surgical_autonomy config."""
    return load_config(review_name="surgical_autonomy")


@pytest.fixture
def keyword_config_path(tmp_path):
    """Create a keyword-based YAML config for testing."""
    import yaml
    config = {
        "categories": {
            "cv_perception": {
                "description": "CV without robot control",
                "title_keywords": ["segmentation", "detection", "tracking"],
                "abstract_keywords": ["deep learning", "convolutional"],
                "exclude_if_also": ["autonomous"],
            },
            "review_editorial": {
                "description": "Reviews and surveys",
                "title_keywords": ["review", "survey", "editorial"],
                "abstract_keywords": ["systematic review"],
                "exclude_if_also": [],
            },
        }
    }
    path = tmp_path / "test_review" / "adjudication_categories.yaml"
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w") as f:
        yaml.dump(config, f)
    return path


# ── Config Loading Tests ────────────────────────────────────────────


def test_load_config_from_yaml(keyword_config_path):
    config = CategoryConfig.load(keyword_config_path)
    assert len(config.categories) == 2
    assert config.categories[0]["name"] == "cv_perception"
    assert "segmentation" in config.categories[0]["title_keywords"]


def test_load_config_missing_file(tmp_path):
    config = CategoryConfig.load(tmp_path / "nonexistent.yaml")
    assert len(config.categories) == 0


def test_load_config_by_review_name():
    config = load_config(review_name="surgical_autonomy")
    assert len(config.categories) == 7
    names = [c["name"] for c in config.categories]
    assert "cv_perception" in names
    assert "teleoperation_only" in names


def test_load_config_no_review_returns_empty():
    config = load_config()
    assert len(config.categories) == 0


def test_load_config_by_review_name_missing(tmp_path):
    config = load_config(review_name="nonexistent_review", data_root=tmp_path)
    assert len(config.categories) == 0


# ── Categorizer Tests (with surgical_autonomy regex config) ─────────


def test_categorize_cv_perception(surgical_autonomy_config):
    assert categorize_paper(
        "Deep Learning for Surgical Tool Segmentation",
        "CNN-based object detection of instruments in endoscopic images",
        config=surgical_autonomy_config,
    ) == "cv_perception"


def test_categorize_review(surgical_autonomy_config):
    assert categorize_paper(
        "A Systematic Review of Autonomous Surgery",
        "We performed a systematic review and meta-analysis",
        config=surgical_autonomy_config,
    ) == "review_editorial"


def test_categorize_hardware(surgical_autonomy_config):
    assert categorize_paper(
        "Piezoelectric Force Sensor for MIS",
        "Novel force sensor design using piezoelectric materials",
        config=surgical_autonomy_config,
    ) == "hardware_sensing"


def test_categorize_planning(surgical_autonomy_config):
    assert categorize_paper(
        "3D Printed Patient-Specific Surgical Guides",
        "preoperative planning with anatomical model",
        config=surgical_autonomy_config,
    ) == "planning_only"


def test_categorize_rehab(surgical_autonomy_config):
    assert categorize_paper(
        "Exoskeleton for Gait Rehabilitation",
        "A lower-limb exoskeleton for post-stroke rehabilitation",
        config=surgical_autonomy_config,
    ) == "rehabilitation_prosthetics"


def test_categorize_ambiguous(surgical_autonomy_config):
    assert categorize_paper(
        "Autonomous Suturing Robot",
        "The robot performed running sutures on porcine tissue",
        config=surgical_autonomy_config,
    ) == "ambiguous"


def test_category_descriptions(surgical_autonomy_config):
    descs = get_category_descriptions(config=surgical_autonomy_config)
    assert "ambiguous" in descs
    assert "cv_perception" in descs
    assert len(descs) == 8  # 7 categories + ambiguous


# ── Keyword-Based Config Tests ──────────────────────────────────────


def test_keyword_matching(keyword_config_path):
    config = CategoryConfig.load(keyword_config_path)
    assert categorize_paper(
        "Tool Segmentation in Surgery", "We use deep learning",
        config=config,
    ) == "cv_perception"


def test_keyword_exclude_if_also(keyword_config_path):
    config = CategoryConfig.load(keyword_config_path)
    # "segmentation" matches cv_perception, but "autonomous" in exclude_if_also blocks it
    assert categorize_paper(
        "Autonomous Segmentation and Execution",
        "An autonomous system for surgical segmentation and cutting",
        config=config,
    ) != "cv_perception"


def test_keyword_review_match(keyword_config_path):
    config = CategoryConfig.load(keyword_config_path)
    assert categorize_paper(
        "A Review of Surgical Robotics", "",
        config=config,
    ) == "review_editorial"


# ── Fallback Behavior Tests ─────────────────────────────────────────


def test_no_config_all_ambiguous():
    """Without a config, all papers should be 'ambiguous'."""
    assert categorize_paper("Any Title", "Any abstract") == "ambiguous"
    assert categorize_paper("Systematic Review", "meta-analysis") == "ambiguous"


def test_empty_config_all_ambiguous():
    config = CategoryConfig.empty()
    assert categorize_paper("Tool Segmentation", "deep learning", config=config) == "ambiguous"


def test_descriptions_without_config():
    descs = get_category_descriptions()
    assert descs == {"ambiguous": "No clear FP pattern — needs careful human review"}


# ── Starter Config Generation ──────────────────────────────────────


def test_generate_starter_config(tmp_path):
    out = tmp_path / "starter.yaml"
    result = generate_starter_config(out)
    assert result == out
    assert out.exists()

    config = CategoryConfig.load(out)
    assert len(config.categories) >= 5
    names = [c["name"] for c in config.categories]
    assert "cv_perception" in names


def test_generate_starter_config_with_samples(tmp_path):
    out = tmp_path / "starter.yaml"
    generate_starter_config(out, sample_titles=["Paper A", "Paper B"])
    content = out.read_text()
    assert "Paper A" in content
    assert "Paper B" in content


# ── Schema Tests ────────────────────────────────────────────────────


def test_adjudication_table_created(tmp_db):
    """The adjudication table should be created by ReviewDatabase init."""
    row = tmp_db._conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='abstract_screening_adjudication'"
    ).fetchone()
    assert row is not None


def test_adjudication_table_idempotent(tmp_db):
    """Calling ensure_adjudication_table twice should not error."""
    ensure_adjudication_table(tmp_db._conn)
    ensure_adjudication_table(tmp_db._conn)


# ── DB Collection Tests ─────────────────────────────────────────────


def test_collect_db_flagged(flagged_db):
    results = _collect_db_flagged(flagged_db)
    assert len(results) == 3
    assert all(r["source_type"] == "db" for r in results)
    assert results[0]["paper_id"] in [1, 2, 3]
    assert results[0]["primary_decision"] == "include"


# ── Expanded Collection Tests ───────────────────────────────────────


def test_collect_expanded_flagged(expanded_dir):
    results = _collect_expanded_flagged(expanded_dir)
    assert len(results) == 2
    assert all(r["source_type"] == "expanded_csv" for r in results)
    titles = {r["title"] for r in results}
    assert "Force Sensor Design for Surgery" in titles
    assert "Robot-Assisted Autonomous Bone Cutting" in titles


def test_expanded_flagged_has_abstracts(expanded_dir):
    results = _collect_expanded_flagged(expanded_dir)
    for r in results:
        assert r["abstract"], f"Missing abstract for {r['title']}"


# ── Export Tests ────────────────────────────────────────────────────


def test_export_db_only(flagged_db, tmp_path):
    out = tmp_path / "queue.xlsx"
    result = export_adjudication_queue(flagged_db, out)
    assert result["total"] == 3
    assert out.exists()
    assert "categories" in result


def test_export_with_expanded(flagged_db, expanded_dir, tmp_path):
    out = tmp_path / "queue.xlsx"
    result = export_adjudication_queue(
        flagged_db, out, expanded_search_dir=expanded_dir,
    )
    assert result["total"] == 5  # 3 from DB + 2 from expanded
    assert out.exists()


def test_export_categorization(flagged_db, tmp_path, surgical_autonomy_config):
    out = tmp_path / "queue.xlsx"
    result = export_adjudication_queue(
        flagged_db, out, category_config=surgical_autonomy_config,
    )
    cats = result["categories"]
    # At least some papers should be categorized
    assert sum(cats.values()) == 3


def test_export_empty_db(tmp_db, tmp_path):
    out = tmp_path / "queue.xlsx"
    result = export_adjudication_queue(tmp_db, out)
    assert result["total"] == 0


def test_export_xlsx_structure(flagged_db, tmp_path):
    from openpyxl import load_workbook
    out = tmp_path / "queue.xlsx"
    export_adjudication_queue(flagged_db, out)
    wb = load_workbook(out)
    assert "Review Queue" in wb.sheetnames
    assert "Instructions" in wb.sheetnames
    # Instructions is the first sheet (opens by default)
    assert wb.sheetnames[0] == "Instructions"
    ws = wb["Review Queue"]
    # Decision column header includes valid values
    assert "PI_decision" in ws.cell(row=1, column=15).value
    assert "INCLUDE" in ws.cell(row=1, column=15).value
    # Notes column follows
    assert "PI_notes" in ws.cell(row=1, column=16).value


def test_export_with_review_name_loads_config(flagged_db, tmp_path):
    """Passing review_name should load categories from YAML config."""
    out = tmp_path / "queue.xlsx"
    result = export_adjudication_queue(
        flagged_db, out, review_name="surgical_autonomy",
    )
    cats = result["categories"]
    # With the surgical_autonomy config, the CV paper should NOT be ambiguous
    assert "cv_perception" in cats or "review_editorial" in cats


# ── Import Tests ────────────────────────────────────────────────────


def test_import_decisions(flagged_db, tmp_path):
    """Export, fill in decisions, import back."""
    out = tmp_path / "queue.xlsx"
    export_adjudication_queue(flagged_db, out)

    # Fill in decisions
    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb["Review Queue"]
    ws.cell(row=2, column=15, value="EXCLUDE")
    ws.cell(row=2, column=16, value="CV only, no robot")
    ws.cell(row=3, column=15, value="INCLUDE")
    ws.cell(row=4, column=15, value="EXCLUDE")
    wb.save(out)

    result = import_adjudication_decisions(flagged_db, out)
    assert result["stats"]["include"] == 1
    assert result["stats"]["exclude"] == 2
    assert result["stats"]["missing"] == 0


def test_import_updates_status(flagged_db, tmp_path):
    """Import should transition ABSTRACT_SCREEN_FLAGGED → ABSTRACT_SCREENED_IN/OUT."""
    out = tmp_path / "queue.xlsx"
    export_adjudication_queue(flagged_db, out)

    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb["Review Queue"]
    # Set all to EXCLUDE
    for row_num in range(2, 5):
        ws.cell(row=row_num, column=15, value="EXCLUDE")
    wb.save(out)

    import_adjudication_decisions(flagged_db, out)

    # All should now be ABSTRACT_SCREENED_OUT
    flagged = flagged_db.get_papers_by_status("ABSTRACT_SCREEN_FLAGGED")
    assert len(flagged) == 0
    screened_out = flagged_db.get_papers_by_status("ABSTRACT_SCREENED_OUT")
    assert len(screened_out) == 3


def test_import_rejects_all_blank(flagged_db, tmp_path):
    """All-blank decisions should reject the entire file with no DB changes."""
    out = tmp_path / "queue.xlsx"
    export_adjudication_queue(flagged_db, out)

    # Don't fill in any decisions — import should reject entirely
    result = import_adjudication_decisions(flagged_db, out)
    assert result["stats"]["missing"] == 3
    assert result["stats"]["include"] == 0
    assert result["stats"]["exclude"] == 0
    assert len(result["warnings"]) == 3

    # Verify no DB changes were made
    flagged = flagged_db.get_papers_by_status("ABSTRACT_SCREEN_FLAGGED")
    assert len(flagged) == 3


def test_import_rejects_partial_blank(flagged_db, tmp_path):
    """Even one blank decision cell rejects the entire file."""
    out = tmp_path / "queue.xlsx"
    export_adjudication_queue(flagged_db, out)

    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb["Review Queue"]
    ws.cell(row=2, column=15, value="INCLUDE")
    ws.cell(row=3, column=15, value="EXCLUDE")
    # Row 4 left blank
    wb.save(out)

    result = import_adjudication_decisions(flagged_db, out)
    assert result["stats"]["missing"] == 1
    assert result["stats"]["include"] == 0  # nothing applied

    # Verify no DB changes were made
    flagged = flagged_db.get_papers_by_status("ABSTRACT_SCREEN_FLAGGED")
    assert len(flagged) == 3


def test_import_rejects_invalid(flagged_db, tmp_path):
    """Invalid decision values reject the entire file with clear error."""
    out = tmp_path / "queue.xlsx"
    export_adjudication_queue(flagged_db, out)

    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb["Review Queue"]
    ws.cell(row=2, column=15, value="MAYBE")
    ws.cell(row=3, column=15, value="INCLUDE")
    ws.cell(row=4, column=15, value="EXCLUDE")
    wb.save(out)

    result = import_adjudication_decisions(flagged_db, out)
    assert result["stats"]["invalid"] == 1
    assert result["stats"]["include"] == 0  # nothing applied

    # Verify no DB changes were made
    flagged = flagged_db.get_papers_by_status("ABSTRACT_SCREEN_FLAGGED")
    assert len(flagged) == 3


# ── Gate Tests ──────────────────────────────────────────────────────


def test_adjudication_gate_with_flagged(flagged_db):
    assert check_adjudication_gate(flagged_db) == 3


def test_adjudication_gate_no_flagged(tmp_db):
    assert check_adjudication_gate(tmp_db) == 0
