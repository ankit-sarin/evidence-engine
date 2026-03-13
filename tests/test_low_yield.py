"""Tests for LOW_YIELD extraction threshold detection.

Covers: field counting, threshold flagging, configurable threshold,
audit queue integration, PRISMA reporting.
"""

import json
from pathlib import Path

import pytest

from engine.agents.auditor import check_low_yield, count_populated_fields
from engine.core.database import ReviewDatabase
from engine.core.review_spec import load_review_spec
from engine.exporters.prisma import generate_prisma_flow
from engine.search.models import Citation


SPEC_PATH = Path(__file__).resolve().parent.parent / "review_specs" / "surgical_autonomy_v1.yaml"


@pytest.fixture
def spec():
    return load_review_spec(SPEC_PATH)


@pytest.fixture
def tmp_db(tmp_path):
    db = ReviewDatabase("test_review", data_root=tmp_path)
    yield db
    db.close()


def _add_paper(db, title="Test Paper", pmid=None):
    cit = Citation(
        title=title, abstract="Abstract text",
        pmid=pmid, doi=None, source="pubmed",
        authors=["A"], journal="J Test", year=2024,
    )
    db.add_papers([cit])
    row = db._conn.execute(
        "SELECT id FROM papers WHERE title = ?", (title,)
    ).fetchone()
    return row["id"]


def _advance_to_ai_audit(db, pid, extracted_data, spec):
    """Move paper through to AI_AUDIT_COMPLETE with given extracted_data."""
    db.update_status(pid, "ABSTRACT_SCREENED_IN")
    db.update_status(pid, "PDF_ACQUIRED")
    db.update_status(pid, "PARSED")
    db.update_status(pid, "EXTRACTED")

    ext_id = db.add_extraction(
        pid, spec.extraction_hash(), extracted_data,
        "reasoning trace", "deepseek-r1:32b",
    )

    # Add evidence spans for populated fields
    for fname, value in extracted_data.items():
        if value and value not in ("NOT_FOUND", "NR", "Not discussed"):
            db.add_evidence_span(ext_id, fname, value, "Source text here.", 0.9)

    # Audit all spans as verified
    spans = db._conn.execute(
        "SELECT id FROM evidence_spans WHERE extraction_id = ?", (ext_id,)
    ).fetchall()
    for s in spans:
        db.update_audit(s["id"], "verified", "gemma3:27b", "OK")

    db.update_status(pid, "AI_AUDIT_COMPLETE")
    return ext_id


# ── count_populated_fields Tests ─────────────────────────────────


class TestCountPopulatedFields:

    def test_all_populated(self):
        data = {
            "study_type": "Original Research",
            "robot_platform": "STAR",
            "task_performed": "suturing",
            "sample_size": "20 trials",
            "country": "USA",
        }
        assert count_populated_fields(data) == 5

    def test_with_absence_values(self):
        data = {
            "study_type": "Original Research",
            "robot_platform": "STAR",
            "fda_status": "NR",
            "comparison_to_human": "No comparison reported",
            "key_limitation": "NOT_FOUND",
        }
        assert count_populated_fields(data) == 2  # only study_type and robot_platform

    def test_with_null_and_empty(self):
        data = {
            "study_type": "Original Research",
            "robot_platform": None,
            "task_performed": "",
            "sample_size": "   ",
        }
        assert count_populated_fields(data) == 1  # only study_type

    def test_empty_dict(self):
        assert count_populated_fields({}) == 0

    def test_all_absence(self):
        data = {
            "f1": "NR",
            "f2": "NOT_FOUND",
            "f3": "Not discussed",
            "f4": "Not assessable",
        }
        assert count_populated_fields(data) == 0


# ── check_low_yield Tests ────────────────────────────────────────


class TestCheckLowYield:

    def test_paper_below_threshold_flagged(self, tmp_db, spec):
        """Paper with 3/15 fields populated → flagged as LOW_YIELD."""
        pid = _add_paper(tmp_db, title="Sparse Paper", pmid="50001")
        sparse_data = {
            "study_type": "Original Research",
            "robot_platform": "STAR",
            "task_performed": "suturing",
            # Remaining 12 fields are absent
            "sample_size": "NR",
            "country": "NR",
            "autonomy_level": "NR",
            "validation_setting": "NR",
            "human_oversight_model": "NR",
            "fda_status": "NR",
            "study_design": "NR",
            "primary_outcome_metric": "NOT_FOUND",
            "primary_outcome_value": "NOT_FOUND",
            "comparison_to_human": "No comparison reported",
            "key_limitation": "NOT_FOUND",
            "clinical_readiness_assessment": "Not assessable",
        }
        ext_id = _advance_to_ai_audit(tmp_db, pid, sparse_data, spec)

        stats = check_low_yield(tmp_db, threshold=4)
        assert stats["low_yield"] == 1
        assert stats["ok"] == 0

        # Verify column set in DB
        row = tmp_db._conn.execute(
            "SELECT low_yield FROM extractions WHERE id = ?", (ext_id,)
        ).fetchone()
        assert row["low_yield"] == 1

    def test_paper_above_threshold_not_flagged(self, tmp_db, spec):
        """Paper with 5/15 fields populated → NOT flagged."""
        pid = _add_paper(tmp_db, title="Rich Paper", pmid="50002")
        rich_data = {
            "study_type": "Original Research",
            "robot_platform": "STAR",
            "task_performed": "suturing",
            "sample_size": "20 trials",
            "country": "USA",
            "autonomy_level": "Level 3",
            # Rest absent
            "validation_setting": "NR",
            "human_oversight_model": "NR",
            "fda_status": "NR",
            "study_design": "NR",
            "primary_outcome_metric": "NOT_FOUND",
            "primary_outcome_value": "NOT_FOUND",
            "comparison_to_human": "No comparison reported",
            "key_limitation": "NOT_FOUND",
            "clinical_readiness_assessment": "Not assessable",
        }
        ext_id = _advance_to_ai_audit(tmp_db, pid, rich_data, spec)

        stats = check_low_yield(tmp_db, threshold=4)
        assert stats["low_yield"] == 0
        assert stats["ok"] == 1

        row = tmp_db._conn.execute(
            "SELECT low_yield FROM extractions WHERE id = ?", (ext_id,)
        ).fetchone()
        assert row["low_yield"] == 0

    def test_threshold_configurable(self, tmp_db, spec):
        """Same paper flagged at threshold=6, not flagged at threshold=3."""
        pid = _add_paper(tmp_db, title="Border Paper", pmid="50003")
        data = {
            "study_type": "Original Research",
            "robot_platform": "STAR",
            "task_performed": "suturing",
            "sample_size": "20",
            "country": "NR",
        }
        _advance_to_ai_audit(tmp_db, pid, data, spec)

        # With threshold=6, paper has 4 populated → flagged
        stats = check_low_yield(tmp_db, threshold=6)
        assert stats["low_yield"] == 1

        # With threshold=3, paper has 4 populated → OK
        stats = check_low_yield(tmp_db, threshold=3)
        assert stats["low_yield"] == 0

    def test_threshold_from_review_spec(self, spec):
        """Verify spec loads the threshold correctly."""
        assert spec.low_yield_threshold == 4


# ── Audit Queue Integration Tests ────────────────────────────────


class TestLowYieldInAuditQueue:

    def test_low_yield_papers_in_audit_export(self, tmp_db, tmp_path, spec):
        """LOW_YIELD papers should appear in the exported audit queue."""
        from engine.adjudication.audit_adjudicator import (
            _collect_papers_for_review,
            export_audit_review_queue,
        )

        pid = _add_paper(tmp_db, title="Sparse Export Paper", pmid="60001")
        sparse_data = {
            "study_type": "Original Research",
            "robot_platform": "STAR",
            "task_performed": "NR",
            "sample_size": "NR",
        }
        _advance_to_ai_audit(tmp_db, pid, sparse_data, spec)

        # Flag as low_yield
        check_low_yield(tmp_db, threshold=4)

        # Collect papers for review
        papers = _collect_papers_for_review(tmp_db, spot_check_pct=0)
        assert len(papers) == 1
        assert papers[0]["paper_id"] == pid
        assert papers[0]["review_reason"] == "low_yield"
        assert papers[0]["low_yield"] is True

    def test_export_includes_low_yield_column(self, tmp_db, tmp_path, spec):
        """Exported XLSX should have a low_yield column."""
        from engine.adjudication.audit_adjudicator import export_audit_review_queue

        pid = _add_paper(tmp_db, title="LY XLSX Paper", pmid="60002")
        sparse_data = {
            "study_type": "Original Research",
            "robot_platform": "STAR",
            "task_performed": "NR",
        }
        _advance_to_ai_audit(tmp_db, pid, sparse_data, spec)
        check_low_yield(tmp_db, threshold=4)

        out = tmp_path / "audit_queue.xlsx"
        result = export_audit_review_queue(tmp_db, out, spot_check_pct=0)
        assert result["low_yield"] == 1

        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb["Audit Review"]
        headers = [cell.value for cell in ws[1]]
        assert "low_yield" in headers

        # Find the low_yield column index and check value
        ly_col = headers.index("low_yield")
        ly_value = ws.cell(row=2, column=ly_col + 1).value
        assert ly_value == "TRUE"


# ── PRISMA Tests ─────────────────────────────────────────────────


class TestPrismaLowYield:

    def test_prisma_includes_low_yield_rejected(self, tmp_db, spec):
        """PRISMA flow should report LOW_YIELD rejections as a distinct category."""
        pid = _add_paper(tmp_db, title="Rejected LY Paper", pmid="70001")
        sparse_data = {
            "study_type": "Original Research",
            "robot_platform": "NR",
        }
        _advance_to_ai_audit(tmp_db, pid, sparse_data, spec)
        check_low_yield(tmp_db, threshold=4)

        # Reject the paper with low_yield reason
        tmp_db.reject_paper(pid, "low_yield_excluded: too few populated fields")

        flow = generate_prisma_flow(tmp_db)
        assert flow["papers_rejected"] == 1
        assert flow["low_yield_rejected"] == 1
        assert "low_yield_excluded" in str(flow["rejection_reasons"])

    def test_prisma_no_low_yield_when_none_rejected(self, tmp_db):
        """PRISMA low_yield_rejected should be 0 when no such rejections exist."""
        flow = generate_prisma_flow(tmp_db)
        assert flow["low_yield_rejected"] == 0


# ── Database Schema Tests ────────────────────────────────────────


class TestLowYieldSchema:

    def test_extractions_has_low_yield_column(self, tmp_db):
        """The extractions table should have a low_yield column."""
        row = tmp_db._conn.execute(
            "PRAGMA table_info(extractions)"
        ).fetchall()
        col_names = [r["name"] for r in row]
        assert "low_yield" in col_names

    def test_low_yield_defaults_to_zero(self, tmp_db, spec):
        """New extractions should have low_yield=0 by default."""
        pid = _add_paper(tmp_db, title="Default Test", pmid="80001")
        tmp_db.update_status(pid, "ABSTRACT_SCREENED_IN")
        tmp_db.update_status(pid, "PDF_ACQUIRED")
        tmp_db.update_status(pid, "PARSED")
        tmp_db.update_status(pid, "EXTRACTED")

        ext_id = tmp_db.add_extraction(
            pid, spec.extraction_hash(),
            {"study_type": "RCT"}, "trace", "model",
        )
        row = tmp_db._conn.execute(
            "SELECT low_yield FROM extractions WHERE id = ?", (ext_id,)
        ).fetchone()
        assert row["low_yield"] == 0
