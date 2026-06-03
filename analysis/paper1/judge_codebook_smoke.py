"""Judge codebook-awareness A/B smoke harness (Paper 1).

Go/no-go gate before committing to the production ``build_judge_prompt()``
rebuild. Validates, against the real PI-audit disagreement cases, whether
injecting codebook context into the Pass 2 judge prompt resolves the
categorical/semantic false positives WITHOUT losing fabrication
sensitivity.

Design
------
Pass 2 is a joint three-arm call per (paper_id, field_name). The 100 audit
arm-rows belong to a set of underlying triples. This harness:

  1. Reconstructs each unique (paper_id, field_name) triple underlying the
     audit rows from the extraction store — all three arm values + spans +
     full parsed source text — via the SAME loader Pass 2 used
     (``load_ai_triples_csv``), so the JudgeInput is byte-identical to the
     recorded run.
  2. Runs the judge once per triple per CONDITION, reusing ``run_pass2``
     unchanged. The ONLY variable swapped between conditions is the
     prompt-construction function:
        - baseline       — the current ``build_pass2_prompt`` (control).
        - codebook_aware — a prototype that injects the active field's
          codebook spec (type, valid values + definitions, decision
          criteria, per-type scoring instruction) just before the OUTPUT
          FORMAT block. Everything else (model, temp 0, seed, arm-slot
          order, windowing, Pass2 parsing/validation) is identical.
  3. Reads off the per-arm verdict for the audited arm-rows and scores
     against the PI gold standard.

Source text is the full parsed text for both conditions (the existing
windowing path applies identically to both, so source is held constant and
the prompt is isolated as the only variable). Windowing was already ruled
out as the disagreement driver by the clean-subset re-estimate.

Acceptance gates (pre-registered)
---------------------------------
  1. Rig validation — under baseline, reproduce the recorded gemma_verdict
     on the subset that was full-text in Pass 2
     (source_text_windowed_in_pass2 == FALSE). No reproduction → rig
     unfaithful → results void.
  2. Recovery (primary) — substantial flip of the 39 PI=SUPPORTED
     disagreement rows toward agreement under codebook-aware, concentrated
     in categorical fields.
  3. Sensitivity guardrail (hard red line) — PI=UNSUPPORTED rows must not
     flip to SUPPORTED. The 4 the judge already catches must hold; ideally
     the 3 it currently misses get caught.
  4. Net summary — weighted kappa on the 75 substantive rows, baseline
     (~0) vs codebook-aware.

GO  = recovery up AND guardrail intact AND kappa clearly off zero.
NO-GO / iterate = flat kappa (design wrong) OR guardrail breach (too lenient).

No DB writes. One report file (+ raw JSON checkpoint). Local inference.

Usage
-----
    PYTHONPATH=. python -m analysis.paper1.judge_codebook_smoke \\
        --review surgical_autonomy \\
        --pairs-csv data/surgical_autonomy/exports/disagreement_pairs_3arm.csv \\
        --codebook data/surgical_autonomy/extraction_codebook.yaml \\
        [--limit N] [--background]
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
import time
from contextlib import contextmanager
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import yaml
from openpyxl import load_workbook

import analysis.paper1.judge as judge_mod
from analysis.paper1.judge import JudgeError, de_randomize_verdicts, run_pass2
from analysis.paper1.judge_loader import (
    compute_codebook_sha256,
    load_ai_triples_csv,
    load_codebook,
)
from analysis.paper1.judge_prompts import (
    _PASS2_OUTPUT_FORMAT_BLOCK,
    build_judge_prompt,
    build_pass2_prompt,
)
from analysis.paper1.judge_schema import JudgeInput
from analysis.paper1.pi_audit_unblind import (
    ORDINAL,
    weighted_kappa_with_ci,
    wilson,
)
from engine.core.database import ReviewDatabase
from engine.utils.background import maybe_background
from engine.utils.ollama_preflight import require_preflight

logger = logging.getLogger(__name__)

DEFAULT_MODEL = "gemma3:27b"

# Recorded Pass 2 run whose verdicts the audit sampled. MUST match
# pi_audit_sampler.PI_AUDIT_CONFIG["run_id"] — the seed (and thus the
# arm-slot permutation) is derived from this run_id, so baseline
# reproduction of recorded verdicts depends on reusing it verbatim.
RECORDED_RUN_ID = "surgical_autonomy_pass2_full_20260421T174729Z"

# Fields scored as numeric/verbatim regardless of codebook `type`. The
# spec names these explicitly: verify the figure/entity is present or
# correctly derived. (robot_platform / task_performed are free_text in
# the codebook but are entity-verbatim for scoring purposes.)
VERBATIM_FIELDS = frozenset({"sample_size", "robot_platform", "task_performed"})

# v2 — ordinal/graded family. These four fields are categorical in the
# codebook but their values are an ORDERED scale; v1's nominal rubric let
# the judge credit a higher level than the evidence earned (all 4 of v1's
# new PI=UNSUPPORTED→SUPPORTED leaks were ordinal fields). The ordinal
# rubric scores the assigned level against the level the evidence actually
# demonstrates.
ORDINAL_FIELDS = frozenset({
    "autonomy_level", "system_maturity",
    "clinical_readiness_assessment", "validation_setting",
})

# Per-field dimension name spliced into the ordinal rubric's [DIMENSION].
ORDINAL_DIMENSION = {
    "autonomy_level": "autonomy",
    "system_maturity": "technology readiness",
    "clinical_readiness_assessment": "clinical readiness",
    "validation_setting": "validation fidelity",
}

PROTOTYPE_VERSION = "v2"

SUBSTANTIVE = frozenset({"SUPPORTED", "PARTIALLY_SUPPORTED", "UNSUPPORTED"})
GEMMA_VERDICTS = ("SUPPORTED", "PARTIALLY_SUPPORTED", "UNSUPPORTED")

# Rig-validation pass threshold: at temp 0 + fixed seed + byte-identical
# prompt, reproduction should be near-total; we allow a small jitter band.
RIG_REPRODUCTION_THRESHOLD = 0.90

CHECKPOINT_EVERY = 20


# ═════════════════════════════════════════════════════════════════════
# Audit join
# ═════════════════════════════════════════════════════════════════════


@dataclass
class AuditRow:
    row_id: int
    paper_id: str
    field_name: str
    arm_name: str
    field_type: str
    gemma_recorded: str
    pi: str
    windowed_in_pass2: bool
    source_window_strategy: str
    # filled after runs
    baseline_verdict: Optional[str] = None
    codebook_verdict: Optional[str] = None

    @property
    def triple(self) -> tuple[str, str]:
        return (self.paper_id, self.field_name)


def _read_xlsx(path: Path, sheet: str) -> tuple[list[str], list[dict]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    header = [str(h) if h is not None else h for h in next(it)]
    out = []
    for raw in it:
        if raw is None or all(c is None for c in raw):
            continue
        out.append({header[i]: (raw[i] if i < len(raw) else None)
                    for i in range(len(header))})
    wb.close()
    return header, out


def load_audit_rows(
    completed_path: Path, key_path: Path, field_types: dict[str, str],
) -> list[AuditRow]:
    """Join the completed adjudication workbook against the key on row_id."""
    _, comp = _read_xlsx(completed_path, "Adjudication")
    _, key = _read_xlsx(key_path, "Key")
    pi_by_id = {int(r["row_id"]): str(r["adjudication"]).strip().upper()
                for r in comp}
    rows: list[AuditRow] = []
    for r in key:
        rid = int(r["row_id"])
        fname = str(r["field_name"])
        rows.append(AuditRow(
            row_id=rid,
            paper_id=str(r["paper_id"]),
            field_name=fname,
            arm_name=str(r["arm_name"]),
            field_type=field_types.get(fname, "?"),
            gemma_recorded=str(r["gemma_verdict"]).strip().upper(),
            pi=pi_by_id[rid],
            windowed_in_pass2=str(r["source_text_windowed_in_pass2"]).strip().upper()
            == "TRUE",
            source_window_strategy=str(r["source_window_strategy"]),
        ))
    return rows


# ═════════════════════════════════════════════════════════════════════
# Codebook-aware prototype prompt
# ═════════════════════════════════════════════════════════════════════


def load_raw_codebook(path: Path) -> dict[str, dict]:
    doc = yaml.safe_load(Path(path).read_text())
    return {f["name"]: f for f in (doc.get("fields") or []) if "name" in f}


_RUBRIC_HEADER = "=== FIELD-SPECIFIC SCORING RUBRIC (codebook-aware) ==="

_GUARD_CATEGORICAL = (
    "GUARDRAIL: this classification latitude exists to reward a FAITHFUL "
    "classification of what the paper describes — it is NOT license to "
    "accept an incorrect code. A wrong classification is UNSUPPORTED even "
    "when the assigned label is itself a valid vocabulary value."
)
_GUARD_FREETEXT = (
    "GUARDRAIL: semantic latitude is for faithful paraphrase / "
    "entity-normalization only — it is NOT license to accept a value the "
    "source does not support. An incorrect or invented value is UNSUPPORTED."
)
_GUARD_VERBATIM = (
    "GUARDRAIL: equivalence covers correct figures/entities and correct "
    "derivations only — it is NOT license to accept a wrong number or a "
    "different entity. A materially wrong figure/entity is UNSUPPORTED."
)

_SCOPE_NOTE = (
    "Apply this field-specific rubric to slots tagged \"CLEAN PRE-CHECK\" "
    "or \"NEEDS FULL VERIFICATION\". For slots tagged \"ABSENCE CLAIM\", "
    "use the absence-claim rubric."
)


def _instruction_line(raw_entry: dict) -> str:
    instr = str(raw_entry.get("instruction", "") or "").strip()
    if not instr:
        return ""
    instr = " ".join(instr.split())
    return (f"Codebook extraction instruction (what a correct value looks "
            f"like): {instr}\n")


def _categorical_rubric(field_name: str, raw_entry: dict) -> str:
    vvs = raw_entry.get("valid_values") or []
    vv_lines = []
    for item in vvs:
        if isinstance(item, dict):
            val = item.get("value", "")
            defn = " ".join(str(item.get("definition", "")).split())
            vv_lines.append(f"  - {val}: {defn}")
        else:
            vv_lines.append(f"  - {item}")
    crit = str(raw_entry.get("decision_criteria", "") or "").strip()
    crit_block = f"Decision criteria:\n{crit}\n\n" if crit else ""
    return (
        f"{_RUBRIC_HEADER}\n"
        f"Field '{field_name}' is a CONTROLLED-VOCABULARY CLASSIFICATION. "
        "The arm assigns one code from a fixed value set; the code label "
        "will USUALLY NOT appear verbatim in the paper — the paper "
        "describes the underlying concept in prose and the arm classifies "
        "it.\n\n"
        "Valid values and their definitions:\n"
        + "\n".join(vv_lines) + "\n\n"
        + crit_block
        + _instruction_line(raw_entry)
        + "\nScoring:\n"
        "  SUPPORTED            — the content the paper describes maps to "
        "the assigned code under the definitions/criteria above. Judge "
        "whether the CODE IS THE CORRECT CLASSIFICATION of what the paper "
        "describes; do NOT require the label to appear verbatim.\n"
        "  PARTIALLY_SUPPORTED  — the code is defensible but the paper is "
        "ambiguous between this and another code, or under-determined.\n"
        "  UNSUPPORTED          — misclassification: the paper describes "
        "content that maps to a DIFFERENT code (e.g., the paper describes "
        "urology but the value is \"Cardiac/Thoracic\"), or no code is "
        "supportable from the source.\n\n"
        f"{_GUARD_CATEGORICAL}\n\n{_SCOPE_NOTE}"
    )


def _freetext_rubric(field_name: str, raw_entry: dict) -> str:
    return (
        f"{_RUBRIC_HEADER}\n"
        f"Field '{field_name}' is FREE TEXT. Score by SEMANTIC "
        "EQUIVALENCE, not string match.\n\n"
        + _instruction_line(raw_entry)
        + "\nScoring:\n"
        "  SUPPORTED            — the value is a faithful representation of "
        "what the source states: exact match OR faithful paraphrase / "
        "entity-form normalization.\n"
        "  PARTIALLY_SUPPORTED  — overlapping but less complete or less "
        "specific than the source, or a defensible-but-partial reading.\n"
        "  UNSUPPORTED          — the value is not grounded in the source: "
        "it contradicts, materially distorts, or invents content.\n\n"
        f"{_GUARD_FREETEXT}\n\n{_SCOPE_NOTE}"
    )


def _verbatim_rubric(field_name: str, raw_entry: dict) -> str:
    return (
        f"{_RUBRIC_HEADER}\n"
        f"Field '{field_name}' records a SPECIFIC FIGURE OR NAMED ENTITY. "
        "Verify the figure/entity is present in the source OR correctly "
        "DERIVED from it (e.g., a summed sample size).\n\n"
        + _instruction_line(raw_entry)
        + "\nScoring:\n"
        "  SUPPORTED            — the figure/entity appears in the source "
        "or is a correct derivation (e.g., 4 pigs + 5 phantoms = 9). "
        "Formatting/unit differences that are numerically or semantically "
        "equal are SUPPORTED.\n"
        "  PARTIALLY_SUPPORTED  — the figure/entity is in the right area "
        "but differs in detail, or the derivation is defensible yet "
        "uncertain.\n"
        "  UNSUPPORTED          — the figure/entity is absent from the "
        "source or materially wrong.\n\n"
        f"{_GUARD_VERBATIM}\n\n{_SCOPE_NOTE}"
    )


def _ordinal_rubric(field_name: str, raw_entry: dict) -> str:
    """v2 ordinal/graded rubric — scores the assigned level against the
    level the paper's described evidence actually demonstrates."""
    vvs = raw_entry.get("valid_values") or []
    names: list[str] = []
    vv_lines: list[str] = []
    for item in vvs:
        if isinstance(item, dict):
            val = item.get("value", "")
            names.append(str(val))
            defn = " ".join(str(item.get("definition", "")).split())
            vv_lines.append(f"  - {val}: {defn}")
        else:
            names.append(str(item))
            vv_lines.append(f"  - {item}")
    ordered = " → ".join(names)
    crit = str(raw_entry.get("decision_criteria", "") or "").strip()
    dim = ORDINAL_DIMENSION.get(field_name, field_name)
    return (
        f"{_RUBRIC_HEADER}\n"
        f"Field '{field_name}' is an ORDERED SCALE; its values represent "
        f"increasing levels of {dim}, in order: {ordered}.\n\n"
        "Level definitions:\n" + "\n".join(vv_lines) + "\n\n"
        f"Decision rules: {crit}\n\n"
        + _instruction_line(raw_entry)
        + "\nScore the assigned level against the level the paper's "
        "described evidence ACTUALLY demonstrates.\n"
        "  SUPPORTED            — only if the evidence EARNS the assigned "
        "level: it must be demonstrated by described capability or results, "
        "not merely implied.\n"
        "  UNSUPPORTED          — if the assigned level is HIGHER than the "
        f"evidence supports. Crediting a higher level because the system "
        f"shows some {dim} is the PRIMARY FABRICATION MODE for this field "
        "and must be flagged.\n"
        "  PARTIALLY_SUPPORTED  — if the evidence is consistent with the "
        "level but does not pin it down, or supports an adjacent level — "
        "withhold full support rather than credit an unproven level.\n\n"
        f"{_SCOPE_NOTE}"
    )


def field_rubric_block(input: JudgeInput, raw_codebook: dict[str, dict]) -> str:
    raw_entry = raw_codebook.get(input.field_name, {})
    if input.field_name in VERBATIM_FIELDS:
        return _verbatim_rubric(input.field_name, raw_entry)
    if input.field_name in ORDINAL_FIELDS:
        return _ordinal_rubric(input.field_name, raw_entry)
    if input.field_type == "categorical":
        return _categorical_rubric(input.field_name, raw_entry)
    if input.field_type == "numeric":
        return _verbatim_rubric(input.field_name, raw_entry)
    return _freetext_rubric(input.field_name, raw_entry)


def make_codebook_aware_builder(raw_codebook: dict[str, dict]):
    """Return a drop-in replacement for build_pass2_prompt that injects the
    field-specific rubric immediately before the OUTPUT FORMAT block.

    By splicing into the real baseline prompt, EVERYTHING except the
    inserted rubric is byte-identical to baseline — system role, slots,
    source, task block, and the absence branch are untouched.
    """
    marker = "\n\n" + _PASS2_OUTPUT_FORMAT_BLOCK

    def builder(input, shuffled_arms, source_text, source_text_windowed):
        base = build_pass2_prompt(
            input, shuffled_arms, source_text, source_text_windowed
        )
        rubric = field_rubric_block(input, raw_codebook)
        if marker not in base:
            # Defensive: fall back to appending — should never happen.
            return base + "\n\n" + rubric
        return base.replace(marker, "\n\n" + rubric + marker, 1)

    return builder


def make_production_builder(raw_codebook: dict[str, dict]):
    """Return a drop-in replacement for build_pass2_prompt that calls the
    REAL production build_judge_prompt() (engine Pass 2 prompt construction,
    dispatching on the codebook's judge_rubric_family) — not the smoke
    splice. Used by the Phase 1 gate to validate the production builder.
    """
    def builder(input, shuffled_arms, source_text, source_text_windowed):
        return build_judge_prompt(
            input, shuffled_arms, source_text, source_text_windowed,
            raw_codebook[input.field_name],
        )

    return builder


def _builder_for(mode: str, raw_codebook: dict[str, dict]):
    if mode == "production":
        return make_production_builder(raw_codebook)
    if mode == "splice":
        return make_codebook_aware_builder(raw_codebook)
    raise ValueError(f"unknown builder mode {mode!r}")


@contextmanager
def _swap_prompt_builder(builder):
    """Temporarily replace judge.build_pass2_prompt (the name run_pass2
    calls). Restores on exit."""
    original = judge_mod.build_pass2_prompt
    judge_mod.build_pass2_prompt = builder
    try:
        yield
    finally:
        judge_mod.build_pass2_prompt = original


# ═════════════════════════════════════════════════════════════════════
# Execution
# ═════════════════════════════════════════════════════════════════════


@dataclass
class TripleRun:
    paper_id: str
    field_name: str
    baseline_by_arm: dict[str, str] = field(default_factory=dict)
    codebook_by_arm: dict[str, str] = field(default_factory=dict)
    baseline_error: Optional[str] = None
    codebook_error: Optional[str] = None
    baseline_windowed: bool = False
    baseline_sec: float = 0.0
    codebook_sec: float = 0.0


def _paper_text(review_dir: Path, paper_id: str) -> Optional[str]:
    parsed_dir = review_dir / "parsed_text"
    md = sorted(parsed_dir.glob(f"{paper_id}_v*.md"), reverse=True)
    if not md:
        return None
    try:
        return md[0].read_text()
    except OSError:
        return None


def run_conditions(
    db: ReviewDatabase,
    triples: list[tuple[str, str]],
    judge_inputs: dict[tuple[str, str], JudgeInput],
    raw_codebook: dict[str, dict],
    *,
    model: str,
    run_id: str,
    checkpoint_path: Path,
    builder_mode: str = "production",
) -> list[TripleRun]:
    review_dir = db.db_path.parent
    cb_builder = _builder_for(builder_mode, raw_codebook)
    runs: list[TripleRun] = []
    total = len(triples)

    for i, (pid, fname) in enumerate(triples, 1):
        tr = TripleRun(paper_id=pid, field_name=fname)
        inp = judge_inputs.get((pid, fname))
        source = _paper_text(review_dir, pid)

        if inp is None:
            tr.baseline_error = tr.codebook_error = "JudgeInput miss (not in pairs CSV)"
            logger.error("MISS [%d/%d] %s/%s", i, total, pid, fname)
            runs.append(tr)
            continue
        if not source:
            tr.baseline_error = tr.codebook_error = "no parsed text"
            logger.error("NO_TEXT [%d/%d] %s/%s", i, total, pid, fname)
            runs.append(tr)
            continue

        # --- baseline (control: current prompt) ---
        t0 = time.time()
        try:
            res = run_pass2(inp, run_id=run_id, source_text=source, model=model)
            tr.baseline_by_arm = {a: v.verdict
                                  for a, v in de_randomize_verdicts(res).items()}
            tr.baseline_windowed = res.source_text_windowed
        except JudgeError as exc:
            tr.baseline_error = f"{type(exc).__name__}: {exc}"
            logger.error("BASE_FAIL [%d/%d] %s/%s: %s", i, total, pid, fname, exc)
        tr.baseline_sec = time.time() - t0

        # --- codebook-aware (only the prompt builder changes) ---
        t0 = time.time()
        try:
            with _swap_prompt_builder(cb_builder):
                res = run_pass2(inp, run_id=run_id, source_text=source, model=model)
            tr.codebook_by_arm = {a: v.verdict
                                  for a, v in de_randomize_verdicts(res).items()}
        except JudgeError as exc:
            tr.codebook_error = f"{type(exc).__name__}: {exc}"
            logger.error("CB_FAIL [%d/%d] %s/%s: %s", i, total, pid, fname, exc)
        tr.codebook_sec = time.time() - t0

        runs.append(tr)
        logger.info(
            "OK [%d/%d] %s/%s base=%.1fs cb=%.1fs windowed=%s",
            i, total, pid, fname, tr.baseline_sec, tr.codebook_sec,
            tr.baseline_windowed,
        )

        if i % CHECKPOINT_EVERY == 0:
            _write_checkpoint(checkpoint_path, runs)
            logger.info("CHECKPOINT %d/%d -> %s", i, total, checkpoint_path)

    _write_checkpoint(checkpoint_path, runs)
    return runs


def _write_checkpoint(path: Path, runs: list[TripleRun]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps([asdict(r) for r in runs], indent=2))


# ═════════════════════════════════════════════════════════════════════
# Scoring & gates
# ═════════════════════════════════════════════════════════════════════


def _row_class(r: AuditRow) -> str:
    """v2 family bucket for a row: ordinal / nominal_categorical / free_text /
    numeric. Ordinal takes precedence over the codebook categorical type."""
    if r.field_name in ORDINAL_FIELDS:
        return "ordinal"
    if r.field_type == "categorical":
        return "nominal_categorical"
    return r.field_type


def attach_verdicts(rows: list[AuditRow], runs: list[TripleRun]) -> None:
    base = {(r.paper_id, r.field_name): r.baseline_by_arm for r in runs}
    cb = {(r.paper_id, r.field_name): r.codebook_by_arm for r in runs}
    for row in rows:
        row.baseline_verdict = base.get(row.triple, {}).get(row.arm_name)
        row.codebook_verdict = cb.get(row.triple, {}).get(row.arm_name)


def _kappa_block(rows: list[AuditRow], which: str) -> dict:
    """Weighted kappa (judge-vs-PI) on substantive rows for one condition."""
    pairs = []
    for r in rows:
        if r.pi not in SUBSTANTIVE:
            continue
        v = getattr(r, f"{which}_verdict")
        if v not in ORDINAL:
            continue
        pairs.append((ORDINAL[r.pi], ORDINAL[v]))
    return weighted_kappa_with_ci(pairs)


def compute_gates(rows: list[AuditRow]) -> dict:
    # ── Gate 1: rig validation ──
    rig_rows = [r for r in rows if not r.windowed_in_pass2]
    rig_scored = [r for r in rig_rows if r.baseline_verdict is not None]
    rig_match = sum(1 for r in rig_scored
                    if r.baseline_verdict == r.gemma_recorded)
    rig_rate = rig_match / len(rig_scored) if rig_scored else 0.0
    rig_mismatches = [
        {"row_id": r.row_id, "paper_id": r.paper_id, "field_name": r.field_name,
         "arm": r.arm_name, "recorded": r.gemma_recorded,
         "baseline": r.baseline_verdict}
        for r in rig_scored if r.baseline_verdict != r.gemma_recorded
    ]
    gate1 = {
        "n_full_text_rows": len(rig_rows),
        "n_scored": len(rig_scored),
        "n_match": rig_match,
        "reproduction_rate": rig_rate,
        "threshold": RIG_REPRODUCTION_THRESHOLD,
        "pass": rig_rate >= RIG_REPRODUCTION_THRESHOLD,
        "mismatches": rig_mismatches,
    }

    # ── Gate 2: recovery on PI=SUPPORTED disagreement rows ──
    recovery = [r for r in rows
                if r.pi == "SUPPORTED" and r.gemma_recorded != "SUPPORTED"]

    def _recovered(r: AuditRow, which: str) -> bool:
        return getattr(r, f"{which}_verdict") == "SUPPORTED"

    by_ftype: dict[str, dict] = {}
    for ft in ("categorical", "free_text", "numeric"):
        grp = [r for r in recovery if r.field_type == ft]
        scored = [r for r in grp if r.codebook_verdict is not None]
        by_ftype[ft] = {
            "n": len(grp),
            "baseline_supported": sum(1 for r in grp if _recovered(r, "baseline")),
            "codebook_supported": sum(1 for r in scored if _recovered(r, "codebook")),
            "n_scored": len(scored),
        }
    rec_scored = [r for r in recovery if r.codebook_verdict is not None]
    gate2 = {
        "n_recovery_rows": len(recovery),
        "baseline_agreement": sum(1 for r in recovery if _recovered(r, "baseline")),
        "codebook_agreement": sum(1 for r in rec_scored if _recovered(r, "codebook")),
        "n_scored": len(rec_scored),
        "by_field_type": by_ftype,
    }

    # ── Gate 3: sensitivity guardrail ──
    pi_unsup = [r for r in rows if r.pi == "UNSUPPORTED"]
    pi_unsup_scored = [r for r in pi_unsup if r.codebook_verdict is not None]
    leak_unsup = [r for r in pi_unsup_scored if r.codebook_verdict == "SUPPORTED"]
    base_unsup_supported = sum(1 for r in pi_unsup
                               if r.baseline_verdict == "SUPPORTED")
    # The 4 the judge catches (recorded gemma == UNSUP) — must hold.
    caught4 = [r for r in pi_unsup if r.gemma_recorded == "UNSUPPORTED"]
    caught4_hold = sum(1 for r in caught4
                       if r.codebook_verdict == "UNSUPPORTED")
    # The 3 it currently misses (recorded gemma == SUPPORTED) — ideally caught.
    missed3 = [r for r in pi_unsup if r.gemma_recorded == "SUPPORTED"]
    missed3_now_caught = sum(1 for r in missed3
                             if r.codebook_verdict in ("UNSUPPORTED",
                                                       "PARTIALLY_SUPPORTED"))
    # Broader leakage: PI not-supported (PARTIAL or UNSUP) → SUPPORTED.
    pi_notsup = [r for r in rows if r.pi in ("PARTIALLY_SUPPORTED", "UNSUPPORTED")]
    pi_notsup_scored = [r for r in pi_notsup if r.codebook_verdict is not None]
    base_notsup_leak = sum(1 for r in pi_notsup
                           if r.baseline_verdict == "SUPPORTED")
    cb_notsup_leak = sum(1 for r in pi_notsup_scored
                         if r.codebook_verdict == "SUPPORTED")
    gate3 = {
        "n_pi_unsupported": len(pi_unsup),
        "baseline_unsup_to_supported": base_unsup_supported,
        "codebook_unsup_to_supported": len(leak_unsup),
        "hard_red_line_pass": len(leak_unsup) <= base_unsup_supported,
        "leak_rows": [
            {"row_id": r.row_id, "paper_id": r.paper_id,
             "field_name": r.field_name, "arm": r.arm_name,
             "recorded": r.gemma_recorded, "baseline": r.baseline_verdict,
             "codebook": r.codebook_verdict}
            for r in leak_unsup
        ],
        "caught4_total": len(caught4),
        "caught4_hold": caught4_hold,
        "missed3_total": len(missed3),
        "missed3_now_caught": missed3_now_caught,
        "pi_notsupported_n": len(pi_notsup),
        "baseline_notsup_leak": base_notsup_leak,
        "codebook_notsup_leak": cb_notsup_leak,
    }

    # ── Gate 4: weighted kappa baseline vs codebook ──
    gate4 = {
        "baseline": _kappa_block(rows, "baseline"),
        "codebook": _kappa_block(rows, "codebook"),
    }

    # ── v2 diagnostic: ordinal vs nominal split of both cohorts ──
    classes = ("ordinal", "nominal_categorical", "free_text", "numeric")

    def _split(cohort: list[AuditRow], target: str) -> dict:
        """target='codebook'|'baseline' verdict==SUPPORTED counts by class."""
        out: dict[str, dict] = {}
        for cls in classes:
            grp = [r for r in cohort if _row_class(r) == cls]
            scored = [r for r in grp if r.codebook_verdict is not None]
            out[cls] = {
                "n": len(grp),
                "baseline_to_supported": sum(
                    1 for r in grp if r.baseline_verdict == "SUPPORTED"),
                "codebook_to_supported": sum(
                    1 for r in scored if r.codebook_verdict == "SUPPORTED"),
                "n_scored": len(scored),
            }
        return out

    ordinal_split = {
        "guardrail_cohort": _split(pi_unsup, "codebook"),   # 12 PI=UNSUPPORTED
        "recovery_cohort": _split(recovery, "codebook"),    # 39 PI=SUPPORTED
    }

    return {"gate1_rig": gate1, "gate2_recovery": gate2,
            "gate3_guardrail": gate3, "gate4_kappa": gate4,
            "ordinal_split": ordinal_split}


def overall_verdict(gates: dict) -> tuple[str, str]:
    g1 = gates["gate1_rig"]
    g2 = gates["gate2_recovery"]
    g3 = gates["gate3_guardrail"]
    g4 = gates["gate4_kappa"]

    if not g1["pass"]:
        return ("VOID", "Rig validation failed — baseline did not reproduce "
                "recorded verdicts; results are not interpretable.")

    k_base = g4["baseline"]["kappa_w"]
    k_cb = g4["codebook"]["kappa_w"]
    kappa_moved = (k_cb is not None and k_base is not None
                   and (k_cb - k_base) >= 0.10 and k_cb >= 0.15)
    recovered = g2["codebook_agreement"] - g2["baseline_agreement"]
    recovery_up = recovered >= 0.30 * max(1, g2["n_recovery_rows"])
    guardrail_ok = g3["hard_red_line_pass"] and (
        g3["codebook_notsup_leak"] <= g3["baseline_notsup_leak"] + 2
    )

    if not guardrail_ok:
        return ("NO-GO", "Guardrail breach — true non-support leaked into "
                "SUPPORTED; the recovery is a rubber-stamp artifact. Iterate "
                "the rubric on these 75 rows.")
    if recovery_up and kappa_moved:
        return ("GO", "Recovery up, guardrail intact, kappa clearly off zero "
                "— codebook-aware design validated. Proceed to the full "
                "build_judge_prompt() rebuild.")
    if not kappa_moved and not recovery_up:
        return ("NO-GO", "Flat kappa and no material recovery — the rubric "
                "design is wrong. Iterate before spending on the rebuild.")
    return ("MARGINAL", "Mixed signal — recovery or kappa moved but not both "
            "decisively, with guardrail intact. Inspect per-field-type "
            "breakdown and consider a rubric iteration before the rebuild.")


# ═════════════════════════════════════════════════════════════════════
# Report
# ═════════════════════════════════════════════════════════════════════


def _fmt_k(d: dict) -> str:
    if d["kappa_w"] is None:
        return f"undefined (n={d['n']})"
    lo = d.get("ci_lo")
    hi = d.get("ci_hi")
    ci = f" [{lo:.3f}, {hi:.3f}]" if lo is not None else ""
    return f"{d['kappa_w']:.3f}{ci} (n={d['n']})"


def _example_prompts(judge_inputs, raw_codebook, rows) -> list[tuple[str, str]]:
    """Render the rubric block for one field of each scoring family."""
    seen: dict[str, str] = {}
    examples: list[tuple[str, str]] = []
    for r in rows:
        inp = judge_inputs.get(r.triple)
        if inp is None:
            continue
        if r.field_name in VERBATIM_FIELDS:
            fam = "numeric/verbatim"
        elif r.field_name in ORDINAL_FIELDS:
            fam = "ordinal/graded"
        elif inp.field_type == "categorical":
            fam = "nominal categorical"
        else:
            fam = "free_text"
        if fam in seen:
            continue
        seen[fam] = r.field_name
        examples.append((f"{fam} (field: {r.field_name})",
                         field_rubric_block(inp, raw_codebook)))
        if len(seen) == 4:
            break
    return examples


def build_report(
    rows: list[AuditRow], runs: list[TripleRun], gates: dict, meta: dict,
    judge_inputs, raw_codebook, v2_gate: Optional[dict] = None,
) -> str:
    verdict, rationale = overall_verdict(gates)
    L: list[str] = []
    L.append(f"# Judge codebook-awareness smoke — A/B gate report "
             f"({PROTOTYPE_VERSION})")
    L.append("")
    L.append(f"**Prototype rubric version:** {PROTOTYPE_VERSION} — four "
             "families: ordinal/graded, nominal categorical, free_text, "
             "numeric/verbatim. Ordinal fields: "
             f"{', '.join(sorted(ORDINAL_FIELDS))}.")
    L.append(f"**Run timestamp (UTC):** {meta['run_timestamp']}")
    L.append(f"**Model:** `{meta['model']}` · digest `{meta['model_digest']}`")
    L.append(f"**Seed run_id (reused for reproduction):** `{meta['run_id']}`")
    L.append(f"**Codebook sha256:** `{meta['codebook_sha']}`")
    L.append(f"**Triples run:** {meta['n_triples']} · "
             f"**audit rows scored:** {meta['n_rows']} · "
             f"wall {meta['wall_h']:.2f}h")
    L.append("")
    L.append(f"## VERDICT: **{verdict}**")
    L.append("")
    L.append(f"> {rationale}")
    L.append("")
    L.append("_Decisive pair: recovery-up AND guardrail-intact (with kappa "
             "off zero). This is a recommendation; the PI owns the go/no-go._")
    L.append("")

    # ── Four gates summary ──
    g1, g2 = gates["gate1_rig"], gates["gate2_recovery"]
    g3, g4 = gates["gate3_guardrail"], gates["gate4_kappa"]

    L.append("## The four gates")
    L.append("")
    L.append("| # | Gate | Result | Pass? |")
    L.append("|---|---|---|---|")
    L.append(f"| 1 | Rig validation (baseline reproduces recorded gemma, "
             f"full-text subset) | {g1['n_match']}/{g1['n_scored']} = "
             f"{g1['reproduction_rate']*100:.1f}% (≥{g1['threshold']*100:.0f}%) "
             f"| {'✅' if g1['pass'] else '❌'} |")
    rec_delta = g2['codebook_agreement'] - g2['baseline_agreement']
    L.append(f"| 2 | Recovery on {g2['n_recovery_rows']} PI=SUPPORTED "
             f"disagreement rows | baseline {g2['baseline_agreement']} → "
             f"codebook {g2['codebook_agreement']} SUPPORTED "
             f"(Δ +{rec_delta}) | {'✅' if rec_delta > 0 else '—'} |")
    L.append(f"| 3 | Guardrail — PI=UNSUPPORTED→SUPPORTED leakage "
             f"(hard red line) | baseline {g3['baseline_unsup_to_supported']} "
             f"→ codebook {g3['codebook_unsup_to_supported']} of "
             f"{g3['n_pi_unsupported']} | {'✅' if g3['hard_red_line_pass'] else '❌'} |")
    L.append(f"| 4 | Weighted κ on 75 substantive rows | baseline "
             f"{_fmt_k(g4['baseline'])} → codebook {_fmt_k(g4['codebook'])} | "
             f"{'see verdict'} |")
    L.append("")

    # ── Phase 1 production-builder gate ──
    if v2_gate is not None:
        h = v2_gate["headline"]
        L.append("## Phase 1 gate — production build_judge_prompt() vs stored v2")
        L.append("")
        L.append(f"- Builder: **{meta.get('builder', '?')}** (real "
                 "build_judge_prompt, dispatch on judge_rubric_family).")
        L.append(f"- Verdict-identity vs v2: "
                 f"**{v2_gate['n_rows_compared'] - v2_gate['n_divergent']}/"
                 f"{v2_gate['n_rows_compared']}** rows identical — "
                 f"{'PASS ✅' if v2_gate['verdict_identity'] else 'DIVERGENT ❌'}.")
        L.append(f"- Recovery Δ: v2 +{h['recovery_delta']['v2']} vs now "
                 f"+{h['recovery_delta']['now']} "
                 f"{'✅' if h['recovery_delta']['match'] else '❌'}")
        L.append(f"- Guardrail PI=UNSUP→SUPPORTED: v2 "
                 f"{h['guardrail_codebook_unsup_to_supported']['v2']} vs now "
                 f"{h['guardrail_codebook_unsup_to_supported']['now']} "
                 f"{'✅' if h['guardrail_codebook_unsup_to_supported']['match'] else '❌'}")
        L.append(f"- κ codebook: v2 {h['kappa_codebook']['v2']} vs now "
                 f"{h['kappa_codebook']['now']} "
                 f"{'✅' if h['kappa_codebook']['match'] else '❌'}")
        if v2_gate["divergent_rows"]:
            L.append("")
            L.append("Divergent rows:")
            L.append("")
            L.append("| row_id | paper | field | arm | v2_codebook | now_codebook |")
            L.append("|---|---|---|---|---|---|")
            for d in v2_gate["divergent_rows"]:
                L.append(f"| {d['row_id']} | {d['paper_id']} | {d['field_name']} "
                         f"| {d['arm']} | {d['v2_codebook']} | "
                         f"{d['now_codebook']} |")
        L.append("")

    # ── Gate 1 detail ──
    L.append("## Gate 1 — Rig validation")
    L.append("")
    L.append(f"- Full-text-in-Pass-2 rows (source_text_windowed_in_pass2 == "
             f"FALSE): **{g1['n_full_text_rows']}**, scored "
             f"{g1['n_scored']}.")
    L.append(f"- Baseline reproduced recorded gemma_verdict on "
             f"**{g1['n_match']}/{g1['n_scored']}** = "
             f"**{g1['reproduction_rate']*100:.1f}%**.")
    if g1["mismatches"]:
        L.append(f"- Mismatches ({len(g1['mismatches'])}):")
        L.append("")
        L.append("| row_id | paper | field | arm | recorded | baseline |")
        L.append("|---|---|---|---|---|---|")
        for m in g1["mismatches"][:40]:
            L.append(f"| {m['row_id']} | {m['paper_id']} | {m['field_name']} "
                     f"| {m['arm']} | {m['recorded']} | {m['baseline']} |")
    L.append("")
    if not g1["pass"]:
        L.append("> ⚠️ Rig unfaithful — downstream gates are not "
                 "interpretable. Stop and diagnose before reading recovery/κ.")
        L.append("")

    # ── Gate 2 detail ──
    L.append("## Gate 2 — Recovery (primary), by field type")
    L.append("")
    L.append("Recovery = a PI=SUPPORTED row that Gemma had flagged "
             "(PARTIALLY_SUPPORTED/UNSUPPORTED) flipping to SUPPORTED under "
             "the condition (agreement with PI).")
    L.append("")
    L.append("| field type | rows | baseline→SUPPORTED | codebook→SUPPORTED | "
             "Δ |")
    L.append("|---|---:|---:|---:|---:|")
    for ft in ("categorical", "free_text", "numeric"):
        b = g2["by_field_type"][ft]
        d = b["codebook_supported"] - b["baseline_supported"]
        L.append(f"| {ft} | {b['n']} | {b['baseline_supported']} | "
                 f"{b['codebook_supported']} | +{d} |")
    L.append(f"| **all** | **{g2['n_recovery_rows']}** | "
             f"**{g2['baseline_agreement']}** | "
             f"**{g2['codebook_agreement']}** | "
             f"**+{rec_delta}** |")
    L.append("")

    # ── Gate 3 detail ──
    L.append("## Gate 3 — Sensitivity guardrail")
    L.append("")
    L.append(f"- **Hard red line:** PI=UNSUPPORTED rows flipping to "
             f"SUPPORTED. Baseline {g3['baseline_unsup_to_supported']} → "
             f"codebook {g3['codebook_unsup_to_supported']} (of "
             f"{g3['n_pi_unsupported']}). "
             f"{'PASS' if g3['hard_red_line_pass'] else 'BREACH'}.")
    L.append(f"- The {g3['caught4_total']} Gemma already catches "
             f"(recorded UNSUPPORTED, PI UNSUPPORTED): "
             f"**{g3['caught4_hold']}/{g3['caught4_total']}** still held as "
             f"UNSUPPORTED under codebook-aware.")
    L.append(f"- The {g3['missed3_total']} Gemma currently misses "
             f"(recorded SUPPORTED, PI UNSUPPORTED): "
             f"**{g3['missed3_now_caught']}/{g3['missed3_total']}** now "
             f"caught (moved off SUPPORTED) under codebook-aware.")
    L.append(f"- Broader leakage — PI∈{{PARTIALLY,UNSUPPORTED}} "
             f"({g3['pi_notsupported_n']} rows) flipping to SUPPORTED: "
             f"baseline {g3['baseline_notsup_leak']} → codebook "
             f"{g3['codebook_notsup_leak']}.")
    if g3["leak_rows"]:
        L.append("")
        L.append("PI=UNSUPPORTED → codebook SUPPORTED leak rows:")
        L.append("")
        L.append("| row_id | paper | field | arm | recorded | baseline | codebook |")
        L.append("|---|---|---|---|---|---|---|")
        for m in g3["leak_rows"]:
            L.append(f"| {m['row_id']} | {m['paper_id']} | {m['field_name']} | "
                     f"{m['arm']} | {m['recorded']} | {m['baseline']} | "
                     f"{m['codebook']} |")
    L.append("")

    # ── Gate 4 detail ──
    L.append("## Gate 4 — Weighted Cohen's κ (judge vs PI, 75 substantive rows)")
    L.append("")
    L.append("| condition | κ_w (linear) |")
    L.append("|---|---|")
    L.append(f"| baseline | {_fmt_k(g4['baseline'])} |")
    L.append(f"| codebook-aware | {_fmt_k(g4['codebook'])} |")
    L.append("")

    # ── v2 ordinal-vs-nominal split ──
    osp = gates["ordinal_split"]
    L.append("## v2 diagnostic — ordinal vs nominal split of both cohorts")
    L.append("")
    L.append("The four ordinal/graded fields (autonomy_level, "
             "system_maturity, clinical_readiness_assessment, "
             "validation_setting) now route to the ordinal rubric. The check: "
             "ordinal PI=UNSUPPORTED→SUPPORTED leakage should drop toward "
             "zero while the nominal/free-text recovery holds.")
    L.append("")
    L.append("**Guardrail cohort — 12 PI=UNSUPPORTED rows** "
             "(want codebook→SUPPORTED ≈ 0, especially ordinal):")
    L.append("")
    L.append("| class | rows | baseline→SUPPORTED | codebook→SUPPORTED |")
    L.append("|---|---:|---:|---:|")
    for cls in ("ordinal", "nominal_categorical", "free_text", "numeric"):
        b = osp["guardrail_cohort"][cls]
        if b["n"] == 0:
            continue
        L.append(f"| {cls} | {b['n']} | {b['baseline_to_supported']} | "
                 f"{b['codebook_to_supported']} |")
    L.append("")
    L.append("**Recovery cohort — 39 PI=SUPPORTED disagreement rows** "
             "(want nominal/free-text codebook→SUPPORTED to hold):")
    L.append("")
    L.append("| class | rows | baseline→SUPPORTED | codebook→SUPPORTED |")
    L.append("|---|---:|---:|---:|")
    for cls in ("ordinal", "nominal_categorical", "free_text", "numeric"):
        b = osp["recovery_cohort"][cls]
        if b["n"] == 0:
            continue
        L.append(f"| {cls} | {b['n']} | {b['baseline_to_supported']} | "
                 f"{b['codebook_to_supported']} |")
    L.append("")

    # ── Per-row table ──
    L.append("## Per-row: baseline → codebook-aware vs PI")
    L.append("")
    L.append("| row_id | paper | field | type | arm | recorded | baseline | "
             "codebook | PI | Δ |")
    L.append("|---|---|---|---|---|---|---|---|---|---|")
    for r in sorted(rows, key=lambda x: x.row_id):
        changed = "→" if r.baseline_verdict != r.codebook_verdict else ""
        L.append(f"| {r.row_id} | {r.paper_id} | {r.field_name} | "
                 f"{r.field_type} | {r.arm_name} | {r.gemma_recorded} | "
                 f"{r.baseline_verdict} | {r.codebook_verdict} | {r.pi} | "
                 f"{changed} |")
    L.append("")

    # ── Failures ──
    fails = [r for r in runs if r.baseline_error or r.codebook_error]
    if fails:
        L.append("## Triple failures")
        L.append("")
        L.append("| paper | field | baseline_error | codebook_error |")
        L.append("|---|---|---|---|")
        for r in fails:
            be = (r.baseline_error or "").replace("|", "\\|")[:80]
            ce = (r.codebook_error or "").replace("|", "\\|")[:80]
            L.append(f"| {r.paper_id} | {r.field_name} | {be} | {ce} |")
        L.append("")

    # ── Prototype prompt text ──
    L.append("## Prototype codebook-aware prompt (the rubric injected)")
    L.append("")
    L.append("The codebook-aware condition is the baseline Pass 2 prompt with "
             "the field-specific rubric below spliced in immediately before "
             "the OUTPUT FORMAT block. Everything else (system role, slots, "
             "source text, base task block, absence branch) is byte-identical "
             "to baseline.")
    L.append("")
    for title, block in _example_prompts(judge_inputs, raw_codebook, rows):
        L.append(f"### {title}")
        L.append("")
        L.append("```")
        L.append(block)
        L.append("```")
        L.append("")

    return "\n".join(L)


# ═════════════════════════════════════════════════════════════════════
# CLI / driver
# ═════════════════════════════════════════════════════════════════════


def _discover_audit(audit_dir: Path) -> tuple[Optional[Path], Optional[Path]]:
    comp = sorted(audit_dir.glob("*COMPLETED*.xlsx"))
    keys = sorted(audit_dir.glob("pi_audit_key_*.xlsx"))
    return (comp[-1] if comp else None, keys[-1] if keys else None)


def _setup_logging(log_dir: Path) -> Path:
    log_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    log_file = log_dir / f"judge_codebook_smoke_{ts}.log"
    fmt = logging.Formatter("%(asctime)s %(levelname)s: %(message)s",
                            datefmt="%Y-%m-%dT%H:%M:%S%z")
    root = logging.getLogger()
    if not root.handlers:
        root.setLevel(logging.INFO)
        sh = logging.StreamHandler(sys.stderr)
        sh.setFormatter(fmt)
        root.addHandler(sh)
        fh = logging.FileHandler(log_file)
        fh.setFormatter(fmt)
        root.addHandler(fh)
    return log_file


def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="analysis.paper1.judge_codebook_smoke")
    p.add_argument("--review", required=True)
    p.add_argument("--pairs-csv", required=True, type=Path)
    p.add_argument("--codebook", required=True, type=Path)
    p.add_argument("--audit-dir", type=Path,
                   default=Path("artifacts/paper1/pi_audit"))
    p.add_argument("--completed", type=Path, default=None)
    p.add_argument("--key", type=Path, default=None)
    p.add_argument("--out-dir", type=Path, default=Path("analysis/paper1/reports"))
    p.add_argument("--log-dir", type=Path, default=Path("analysis/paper1/logs"))
    p.add_argument("--model", default=DEFAULT_MODEL)
    p.add_argument("--run-id", default=RECORDED_RUN_ID,
                   help="run_id used to derive the per-triple seed; defaults "
                        "to the recorded Pass 2 run so baseline reproduces it.")
    p.add_argument("--limit", type=int, default=None,
                   help="Cap unique triples (smoke-of-the-smoke).")
    p.add_argument("--data-root", type=Path, default=None)
    p.add_argument("--builder", choices=("production", "splice"),
                   default="production",
                   help="codebook-aware condition: 'production' uses the real "
                        "build_judge_prompt(); 'splice' uses the v2 prototype "
                        "splice. Default production.")
    p.add_argument("--assert-v2", type=Path, default=None,
                   help="Path to a stored v2 *_results.json; the run asserts "
                        "verdict-identity per row + the headline gate numbers.")
    p.add_argument("--background", action="store_true",
                   help="Re-launch in a detached tmux session.")
    return p


def assert_reproduces_v2(
    scored_rows: list[AuditRow], gates: dict, v2_json_path: Path,
) -> dict:
    """Compare this run's codebook verdicts + headline gates against a stored
    v2 results JSON. Returns a result dict (divergent rows, headline checks)."""
    v2 = json.loads(Path(v2_json_path).read_text())
    v2_cb = {int(r["row_id"]): r.get("codebook_verdict") for r in v2["rows"]}
    divergent = []
    for r in scored_rows:
        prev = v2_cb.get(r.row_id, "__absent__")
        if prev != r.codebook_verdict:
            divergent.append({
                "row_id": r.row_id, "paper_id": r.paper_id,
                "field_name": r.field_name, "arm": r.arm_name,
                "v2_codebook": prev, "now_codebook": r.codebook_verdict,
            })

    g2v2 = v2["gates"]["gate2_recovery"]
    g3v2 = v2["gates"]["gate3_guardrail"]
    g4v2 = v2["gates"]["gate4_kappa"]
    g2, g3, g4 = (gates["gate2_recovery"], gates["gate3_guardrail"],
                  gates["gate4_kappa"])

    def _rec_delta(g):
        return g["codebook_agreement"] - g["baseline_agreement"]

    headline = {
        "recovery_delta": {"v2": _rec_delta(g2v2), "now": _rec_delta(g2),
                           "match": _rec_delta(g2v2) == _rec_delta(g2)},
        "guardrail_codebook_unsup_to_supported": {
            "v2": g3v2["codebook_unsup_to_supported"],
            "now": g3["codebook_unsup_to_supported"],
            "match": (g3v2["codebook_unsup_to_supported"]
                      == g3["codebook_unsup_to_supported"])},
        "kappa_codebook": {
            "v2": round(g4v2["codebook"]["kappa_w"], 3)
            if g4v2["codebook"]["kappa_w"] is not None else None,
            "now": round(g4["codebook"]["kappa_w"], 3)
            if g4["codebook"]["kappa_w"] is not None else None,
            "match": (g4v2["codebook"]["kappa_w"] is not None
                      and g4["codebook"]["kappa_w"] is not None
                      and round(g4v2["codebook"]["kappa_w"], 3)
                      == round(g4["codebook"]["kappa_w"], 3))},
    }
    return {
        "v2_results": str(v2_json_path),
        "n_rows_compared": len(scored_rows),
        "n_divergent": len(divergent),
        "divergent_rows": divergent,
        "headline": headline,
        "verdict_identity": len(divergent) == 0,
        "headline_match": all(h["match"] for h in headline.values()),
    }


def run(argv: Optional[list[str]] = None) -> int:
    maybe_background("judge_codebook_smoke", review_name="surgical_autonomy")
    args = _build_parser().parse_args(argv)
    log_file = _setup_logging(args.log_dir)
    logger.info("judge codebook smoke starting; log=%s", log_file)

    require_preflight([args.model], runner_name="judge_codebook_smoke")

    completed = args.completed
    key = args.key
    if completed is None or key is None:
        dc, dk = _discover_audit(args.audit_dir)
        completed = completed or dc
        key = key or dk
    if completed is None or key is None:
        raise SystemExit(f"ABORT: could not locate completed/key xlsx in "
                         f"{args.audit_dir}")
    logger.info("audit completed: %s", completed)
    logger.info("audit key:       %s", key)

    db = (ReviewDatabase(args.review, data_root=args.data_root)
          if args.data_root else ReviewDatabase(args.review))
    try:
        codebook = load_codebook(args.codebook)
        codebook_sha = compute_codebook_sha256(args.codebook)
        raw_codebook = load_raw_codebook(args.codebook)
        field_types = {name: e.field_type for name, e in codebook.items()}

        rows = load_audit_rows(completed, key, field_types)
        logger.info("loaded %d audit rows", len(rows))

        inputs_list = load_ai_triples_csv(args.pairs_csv, db, codebook, limit=None)
        judge_inputs = {(i.paper_id, i.field_name): i for i in inputs_list}

        # Unique triples underlying the audit rows, deterministically ordered.
        triples = sorted(
            {r.triple for r in rows},
            key=lambda t: (int(t[0]) if t[0].isdigit() else t[0], t[1]),
        )
        if args.limit is not None:
            triples = triples[: args.limit]
            logger.info("limited to first %d triples", args.limit)
        logger.info("running %d unique triples × 2 conditions = %d calls",
                    len(triples), len(triples) * 2)

        from engine.utils.ollama_client import fetch_model_digest
        model_digest = fetch_model_digest(args.model)

        ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        args.out_dir.mkdir(parents=True, exist_ok=True)
        checkpoint = args.out_dir / f"judge_codebook_smoke_{ts}_checkpoint.json"

        logger.info("codebook-aware condition builder: %s", args.builder)
        t0 = time.time()
        runs = run_conditions(
            db, triples, judge_inputs, raw_codebook,
            model=args.model, run_id=args.run_id, checkpoint_path=checkpoint,
            builder_mode=args.builder,
        )
        wall_h = (time.time() - t0) / 3600.0

        # Restrict scoring to rows whose triple was actually run.
        run_triples = {(r.paper_id, r.field_name) for r in runs}
        scored_rows = [r for r in rows if r.triple in run_triples]
        attach_verdicts(scored_rows, runs)
        gates = compute_gates(scored_rows)

        v2_gate = None
        if args.assert_v2 is not None:
            v2_gate = assert_reproduces_v2(scored_rows, gates, args.assert_v2)

        meta = {
            "run_timestamp": datetime.now(timezone.utc).isoformat(),
            "model": args.model,
            "model_digest": model_digest,
            "run_id": args.run_id,
            "codebook_sha": codebook_sha,
            "builder": args.builder,
            "n_triples": len(runs),
            "n_rows": len(scored_rows),
            "wall_h": wall_h,
        }

        report = build_report(scored_rows, runs, gates, meta,
                              judge_inputs, raw_codebook, v2_gate=v2_gate)
        report_path = args.out_dir / f"judge_codebook_smoke_{ts}.md"
        report_path.write_text(report)
        json_path = args.out_dir / f"judge_codebook_smoke_{ts}_results.json"
        json_path.write_text(json.dumps({
            "meta": meta, "gates": gates, "v2_gate": v2_gate,
            "rows": [asdict(r) for r in scored_rows],
        }, indent=2, default=str))
    finally:
        db.close()

    verdict, rationale = overall_verdict(gates)
    g1, g2 = gates["gate1_rig"], gates["gate2_recovery"]
    g3, g4 = gates["gate3_guardrail"], gates["gate4_kappa"]
    print("\n" + "=" * 64)
    print(f"VERDICT: {verdict} — {rationale}")
    print("=" * 64)
    print(f"GATE 1 rig:       {g1['n_match']}/{g1['n_scored']} = "
          f"{g1['reproduction_rate']*100:.1f}%  pass={g1['pass']}")
    print(f"GATE 2 recovery:  baseline {g2['baseline_agreement']} → codebook "
          f"{g2['codebook_agreement']} of {g2['n_recovery_rows']} "
          f"(cat/free/num breakdown in report)")
    print(f"GATE 3 guardrail: PI=UNSUP→SUPPORTED baseline "
          f"{g3['baseline_unsup_to_supported']} → codebook "
          f"{g3['codebook_unsup_to_supported']} of {g3['n_pi_unsupported']}  "
          f"red_line_pass={g3['hard_red_line_pass']}")
    print(f"GATE 4 kappa:     baseline {_fmt_k(g4['baseline'])} → codebook "
          f"{_fmt_k(g4['codebook'])}")
    if v2_gate is not None:
        h = v2_gate["headline"]
        print("=" * 64)
        print(f"PHASE 1 GATE (builder={meta['builder']}): "
              f"verdict-identity "
              f"{v2_gate['n_rows_compared'] - v2_gate['n_divergent']}/"
              f"{v2_gate['n_rows_compared']} "
              f"{'PASS' if v2_gate['verdict_identity'] else 'DIVERGENT'}")
        print(f"  recovery Δ: v2 +{h['recovery_delta']['v2']} / now "
              f"+{h['recovery_delta']['now']} "
              f"({'ok' if h['recovery_delta']['match'] else 'MISMATCH'})")
        print(f"  guardrail:  v2 {h['guardrail_codebook_unsup_to_supported']['v2']}"
              f" / now {h['guardrail_codebook_unsup_to_supported']['now']} "
              f"({'ok' if h['guardrail_codebook_unsup_to_supported']['match'] else 'MISMATCH'})")
        print(f"  kappa:      v2 {h['kappa_codebook']['v2']} / now "
              f"{h['kappa_codebook']['now']} "
              f"({'ok' if h['kappa_codebook']['match'] else 'MISMATCH'})")
        if v2_gate["divergent_rows"]:
            print(f"  divergent rows: "
                  f"{[d['row_id'] for d in v2_gate['divergent_rows']]}")
    print("=" * 64)
    print(f"report: {report_path}")
    print(f"json:   {json_path}")
    return 0


if __name__ == "__main__":
    sys.exit(run())
