"""Import human extractor workbooks into the human_extractions table.

Paper 1 analysis infrastructure — reads v2 extraction workbooks (.xlsx),
validates against the codebook, and stores field-level rows for concordance.
"""

import argparse
import logging
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import yaml

logger = logging.getLogger(__name__)

# Column layout: A–D identifiers, E–X extraction fields, Y–Z source quotes, AA notes
_HEADER_ROW = 2
_DATA_START_ROW = 3

_EXTRACTION_FIELDS = [
    "study_type", "robot_platform", "task_performed", "sample_size",
    "surgical_domain", "autonomy_level", "validation_setting",
    "task_monitor", "task_generate", "task_select", "task_execute",
    "system_maturity", "study_design", "country",
    "primary_outcome_metric", "primary_outcome_value",
    "comparison_to_human", "secondary_outcomes",
    "key_limitation", "clinical_readiness_assessment",
]

_SQ_FIELDS = {
    "sq_key_limitation": "key_limitation",
    "sq_clinical_readiness": "clinical_readiness_assessment",
}

_PAPER_ID_RE = re.compile(r"^EE-\d{3}$")


# ── Workbook Parsing ─────────────────────────────────────────────────


def _extract_extractor_id(filepath: Path) -> str:
    """Derive extractor ID from filename (e.g. *_A.xlsx -> A)."""
    stem = filepath.stem
    # Last segment after underscore
    parts = stem.rsplit("_", 1)
    if len(parts) == 2 and len(parts[1]) == 1 and parts[1].isalpha():
        return parts[1].upper()
    raise ValueError(
        f"Cannot derive extractor ID from filename '{filepath.name}'. "
        "Expected format: *_A.xlsx, *_B.xlsx, etc."
    )


def _normalize_value(val) -> str | None:
    """Normalize a cell value: blank and 'NR' -> None, else stripped string."""
    if val is None:
        return None
    s = str(val).strip()
    if s == "" or s.upper() == "NR":
        return None
    return s


def parse_workbook(filepath: Path) -> list[dict]:
    """Read the Extraction Form sheet and return one dict per paper row."""
    filepath = Path(filepath)
    extractor_id = _extract_extractor_id(filepath)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = wb["Extraction Form"]
    except KeyError:
        raise ValueError(f"Workbook missing 'Extraction Form' sheet: {filepath}")

    # Read header row and build column index
    headers = [c.value for c in ws[_HEADER_ROW]]
    col_idx: dict[str, int] = {}
    for i, h in enumerate(headers):
        if h is not None:
            col_idx[h.strip()] = i

    # Validate expected headers present
    missing = [f for f in _EXTRACTION_FIELDS if f not in col_idx]
    if missing:
        raise ValueError(f"Missing extraction columns in header: {missing}")

    # Map source quote and notes columns
    sq_limit_col = col_idx.get("SQ: key_limitation [REQ]")
    sq_cra_col = col_idx.get("SQ: clinical_readiness [REQ]")
    notes_col = col_idx.get("Extractor Notes")

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=_DATA_START_ROW, values_only=True):
        paper_id_raw = row[0] if row else None
        if paper_id_raw is None:
            continue  # skip blank rows

        paper_id = str(paper_id_raw).strip()
        record: dict = {
            "paper_id": paper_id,
            "extractor_id": extractor_id,
        }

        # Extraction fields
        for field in _EXTRACTION_FIELDS:
            idx = col_idx[field]
            record[field] = _normalize_value(row[idx] if idx < len(row) else None)

        # Source quotes
        record["sq_key_limitation"] = (
            _normalize_value(row[sq_limit_col]) if sq_limit_col is not None and sq_limit_col < len(row) else None
        )
        record["sq_clinical_readiness"] = (
            _normalize_value(row[sq_cra_col]) if sq_cra_col is not None and sq_cra_col < len(row) else None
        )

        # Notes
        record["notes"] = (
            _normalize_value(row[notes_col]) if notes_col is not None and notes_col < len(row) else None
        )

        rows.append(record)

    wb.close()
    return rows


# ── Validation ───────────────────────────────────────────────────────


def _load_codebook_valid_values(codebook_path: Path) -> dict[str, list[str]]:
    """Load categorical field -> list of valid values (lowercased) from codebook."""
    with open(codebook_path) as f:
        cb = yaml.safe_load(f)

    valid: dict[str, list[str]] = {}
    for field_def in cb.get("fields", []):
        if field_def.get("type") == "categorical" and "valid_values" in field_def:
            name = field_def["name"]
            valid[name] = [v["value"].lower() for v in field_def["valid_values"]]
    return valid


def validate_workbook(
    rows: list[dict],
    codebook_path: Path,
    db_path: Path | None = None,
) -> list[str]:
    """Validate parsed rows against codebook and DB. Returns list of error strings."""
    errors: list[str] = []

    if not rows:
        errors.append("No data rows found in workbook")
        return errors

    # (a) Check all 20 field names present (already enforced in parse, but belt-and-suspenders)
    sample = rows[0]
    missing_fields = [f for f in _EXTRACTION_FIELDS if f not in sample]
    if missing_fields:
        errors.append(f"Missing extraction fields: {missing_fields}")

    # Load codebook valid values
    valid_values = _load_codebook_valid_values(codebook_path)

    # Build set of DB paper IDs if db_path provided
    db_paper_ids: set[str] | None = None
    if db_path is not None:
        conn = sqlite3.connect(str(db_path))
        try:
            db_rows = conn.execute(
                "SELECT printf('EE-%03d', id) FROM papers"
            ).fetchall()
            db_paper_ids = {r[0] for r in db_rows}
        finally:
            conn.close()

    for i, row in enumerate(rows):
        row_label = f"Row {i + _DATA_START_ROW} ({row['paper_id']})"

        # (b) Paper ID format
        if not _PAPER_ID_RE.match(row["paper_id"]):
            errors.append(f"{row_label}: Paper ID '{row['paper_id']}' does not match EE-NNN format")

        # (c) Paper ID exists in DB
        if db_paper_ids is not None and row["paper_id"] not in db_paper_ids:
            errors.append(f"{row_label}: Paper ID '{row['paper_id']}' not found in database")

        # (d) Categorical field validation
        for field_name, allowed in valid_values.items():
            val = row.get(field_name)
            if val is None:
                continue  # blank/NR is OK
            # Handle semicolon-separated multi-values (e.g. validation_setting)
            parts = [v.strip() for v in val.split(";")]
            for part in parts:
                if part.lower() not in allowed:
                    errors.append(
                        f"{row_label}: {field_name}='{part}' not in codebook valid values"
                    )

        # (e) Mandatory source quotes
        if row.get("key_limitation") is not None and row.get("sq_key_limitation") is None:
            errors.append(f"{row_label}: key_limitation has a value but SQ: key_limitation is empty")
        if row.get("clinical_readiness_assessment") is not None and row.get("sq_clinical_readiness") is None:
            errors.append(f"{row_label}: clinical_readiness_assessment has a value but SQ: clinical_readiness is empty")

    return errors


# ── Storage ──────────────────────────────────────────────────────────


_CREATE_TABLE = """\
CREATE TABLE IF NOT EXISTS human_extractions (
    id INTEGER PRIMARY KEY,
    paper_id TEXT NOT NULL,
    extractor_id TEXT NOT NULL,
    field_name TEXT NOT NULL,
    value TEXT,
    source_quote TEXT,
    notes TEXT,
    imported_at TEXT NOT NULL,
    UNIQUE(paper_id, extractor_id, field_name)
)"""


def _source_quote_for_field(row: dict, field_name: str) -> str | None:
    """Return the source quote for fields that have one."""
    if field_name == "key_limitation":
        return row.get("sq_key_limitation")
    if field_name == "clinical_readiness_assessment":
        return row.get("sq_clinical_readiness")
    return None


def store_human_extractions(
    rows: list[dict],
    extractor_id: str,
    db_path: Path,
) -> int:
    """Store parsed rows into human_extractions table (long format).

    Returns the number of field rows inserted.
    """
    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute(_CREATE_TABLE)

    now = datetime.now(timezone.utc).isoformat()
    inserted = 0

    try:
        for row in rows:
            notes = row.get("notes")
            for field_name in _EXTRACTION_FIELDS:
                val = row.get(field_name)
                sq = _source_quote_for_field(row, field_name)
                conn.execute(
                    "INSERT INTO human_extractions "
                    "(paper_id, extractor_id, field_name, value, source_quote, notes, imported_at) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (row["paper_id"], extractor_id, field_name, val, sq, notes, now),
                )
                inserted += 1
        conn.commit()
    except sqlite3.IntegrityError as exc:
        conn.rollback()
        raise RuntimeError(
            f"Duplicate import detected for extractor '{extractor_id}'. "
            f"Clear existing rows before re-importing. Detail: {exc}"
        ) from exc
    finally:
        conn.close()

    return inserted


# ── CLI ──────────────────────────────────────────────────────────────


def main():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%H:%M:%S",
    )

    parser = argparse.ArgumentParser(
        description="Import human extractor workbook into human_extractions table"
    )
    parser.add_argument("--workbook", type=Path, required=True, help="Path to .xlsx workbook")
    parser.add_argument("--review", type=str, required=True, help="Review name (e.g. surgical_autonomy)")
    parser.add_argument("--codebook", type=Path, default=None, help="Path to extraction_codebook.yaml (auto-detected if omitted)")
    parser.add_argument("--dry-run", action="store_true", help="Parse and validate only, do not store")
    args = parser.parse_args()

    # Resolve paths via ReviewDatabase to match engine conventions
    from engine.core.database import ReviewDatabase
    db = ReviewDatabase(args.review)
    db_path = Path(db.db_path)
    data_dir = db_path.parent
    db.close()

    codebook_path = args.codebook or (data_dir / "extraction_codebook.yaml")
    if not codebook_path.exists():
        logger.error("Codebook not found: %s", codebook_path)
        raise SystemExit(1)

    # Parse
    logger.info("Parsing workbook: %s", args.workbook)
    rows = parse_workbook(args.workbook)
    logger.info("Parsed %d papers", len(rows))

    # Validate
    errors = validate_workbook(rows, codebook_path, db_path=db_path)
    if errors:
        logger.error("Validation failed with %d errors:", len(errors))
        for e in errors:
            logger.error("  %s", e)
        raise SystemExit(1)
    logger.info("Validation passed — 0 errors")

    if args.dry_run:
        # Report field coverage
        filled = {f: 0 for f in _EXTRACTION_FIELDS}
        for row in rows:
            for f in _EXTRACTION_FIELDS:
                if row.get(f) is not None:
                    filled[f] += 1
        print(f"\n{'Field':<35s} {'Filled':>6s} / {len(rows)}")
        print("-" * 50)
        for f in _EXTRACTION_FIELDS:
            print(f"{f:<35s} {filled[f]:>6d} / {len(rows)}")
        print(f"\nDry run complete — {len(rows)} papers, no data stored.")
        return

    # Store
    extractor_id = _extract_extractor_id(args.workbook)
    inserted = store_human_extractions(rows, extractor_id, db_path)
    logger.info(
        "Stored %d field rows for extractor %s (%d papers x %d fields)",
        inserted, extractor_id, len(rows), len(_EXTRACTION_FIELDS),
    )


if __name__ == "__main__":
    main()
