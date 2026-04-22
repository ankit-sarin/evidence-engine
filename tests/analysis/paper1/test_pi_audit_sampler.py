"""Tests for analysis/paper1/pi_audit_sampler.py.

Uses a fixture SQLite DB pre-populated with enough Pass 2 rows per
(verdict, arm) cell to satisfy a scaled-down allocation — the
production-DB run is a one-shot generation, not a unit test.
"""

from __future__ import annotations

import copy
import importlib
import sqlite3
from pathlib import Path

import pytest

from openpyxl import load_workbook

from analysis.paper1 import pi_audit_sampler as sampler

# Migrations we need for the fixture:
# - 007: judge_runs, judge_ratings, judge_pair_ratings
# - 008: fabrication_verifications
# - 009: judge_run_audit (not used in tests but keeps schema consistent)
m007 = importlib.import_module("engine.migrations.007_add_judge_tables")
m008 = importlib.import_module("engine.migrations.008_add_fabrication_verifications")


# ── Fixture construction ────────────────────────────────────────────


SMALL_ALLOCATION = {
    "UNSUPPORTED":         {"local": 3, "openai_o4_mini_high": 3, "anthropic_sonnet_4_6": 4},
    "PARTIALLY_SUPPORTED": {"local": 3, "openai_o4_mini_high": 3, "anthropic_sonnet_4_6": 4},
    "SUPPORTED":           {"local": 2, "openai_o4_mini_high": 2, "anthropic_sonnet_4_6": 1},
}
SMALL_TOTAL = sum(n for by_arm in SMALL_ALLOCATION.values()
                  for n in by_arm.values())  # = 25
assert SMALL_TOTAL < 100

SMALL_CONFIG = {
    "run_id": "fixture_run_1",
    "master_seed": 42,
    "allocation": SMALL_ALLOCATION,
}


def _build_fixture_db(tmp_path: Path) -> tuple[Path, Path]:
    """Create a fixture DB with enough rows to satisfy SMALL_ALLOCATION,
    plus the parsed_text files the sampler reads. Returns (db_path, review_dir).
    """
    review_dir = tmp_path / "review_fixture"
    review_dir.mkdir()
    db_path = review_dir / "review.db"

    # Apply migrations.
    m007.run_migration(str(db_path))
    m008.run_migration(str(db_path))

    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        # papers table — create minimal schema the sampler queries on.
        conn.execute("""
            CREATE TABLE IF NOT EXISTS papers (
                id INTEGER PRIMARY KEY,
                pmid TEXT,
                doi TEXT,
                title TEXT,
                ee_identifier TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS extractions (
                id INTEGER PRIMARY KEY,
                paper_id INTEGER
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS evidence_spans (
                id INTEGER PRIMARY KEY,
                extraction_id INTEGER,
                field_name TEXT,
                value TEXT,
                source_snippet TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS cloud_extractions (
                id INTEGER PRIMARY KEY,
                paper_id INTEGER,
                arm TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS cloud_evidence_spans (
                id INTEGER PRIMARY KEY,
                cloud_extraction_id INTEGER,
                field_name TEXT,
                value TEXT,
                source_snippet TEXT
            )
        """)

        # Seed judge_runs row.
        conn.execute(
            """INSERT INTO judge_runs
               (run_id, judge_model_name, judge_model_digest,
                codebook_sha256, pass_number, input_scope,
                started_at, completed_at, n_triples_attempted,
                n_triples_succeeded, n_triples_failed,
                run_config_json, notes)
               VALUES ('fixture_run_1', 'gemma3:27b',
                       'a' * 64, 'b' * 64, 2, 'AI_TRIPLES',
                       '2026-04-01T00:00:00Z', '2026-04-02T00:00:00Z',
                       10, 10, 0, '{}', NULL)""",
        )

        # Populate papers + extractions + verifications.
        # We need (verdict, arm) cells with >= N rows each — give each
        # cell 12 rows so sampling without replacement is safe.
        verdicts = ("UNSUPPORTED", "PARTIALLY_SUPPORTED", "SUPPORTED")
        arms = ("local", "openai_o4_mini_high", "anthropic_sonnet_4_6")
        per_cell = 12

        # Distinct papers per (verdict, arm) slot to keep field_name mix.
        paper_id_counter = 0
        extraction_id_counter = 0
        cloud_extraction_id_counter = 0
        span_id_counter = 0
        cloud_span_id_counter = 0

        fields_pool = [
            "study_design", "sample_size", "primary_outcome_metric",
            "secondary_outcomes", "autonomy_level", "task_monitor",
        ]

        paper_id_seen: set[int] = set()

        for verdict in verdicts:
            for arm in arms:
                for k in range(per_cell):
                    paper_id_counter += 1
                    pid = paper_id_counter
                    paper_id_seen.add(pid)
                    field = fields_pool[k % len(fields_pool)]
                    value = f"value_{verdict[:3]}_{arm[:3]}_{k}"

                    # paper
                    conn.execute(
                        "INSERT OR IGNORE INTO papers (id, title, ee_identifier) VALUES (?, ?, ?)",
                        (pid, f"Fixture paper {pid}", f"EE-{pid:03d}"),
                    )

                    # write parsed_text file (content seeded with value
                    # so window_source_text has something to anchor on).
                    parsed_dir = review_dir / "parsed_text"
                    parsed_dir.mkdir(exist_ok=True)
                    (parsed_dir / f"{pid}_v1.md").write_text(
                        "Fixture paper body. " * 50 +
                        f"\n\nEvidence span: {value} in context.\n\n" +
                        "More content. " * 50
                    )

                    # Evidence-span snippet text placed in the parsed
                    # paper body below so the sampler can locate it via
                    # substring match (matching the real pipeline).
                    snippet = f"Evidence span: {value} in context."
                    # arm value storage per arm (for _fetch_arm_value_and_snippet)
                    if arm == "local":
                        extraction_id_counter += 1
                        eid = extraction_id_counter
                        conn.execute(
                            "INSERT INTO extractions (id, paper_id) VALUES (?, ?)",
                            (eid, pid),
                        )
                        span_id_counter += 1
                        conn.execute(
                            """INSERT INTO evidence_spans
                               (id, extraction_id, field_name, value, source_snippet)
                               VALUES (?, ?, ?, ?, ?)""",
                            (span_id_counter, eid, field, value, snippet),
                        )
                    else:
                        cloud_extraction_id_counter += 1
                        ceid = cloud_extraction_id_counter
                        conn.execute(
                            "INSERT INTO cloud_extractions (id, paper_id, arm) VALUES (?, ?, ?)",
                            (ceid, pid, arm),
                        )
                        cloud_span_id_counter += 1
                        conn.execute(
                            """INSERT INTO cloud_evidence_spans
                               (id, cloud_extraction_id, field_name, value, source_snippet)
                               VALUES (?, ?, ?, ?, ?)""",
                            (cloud_span_id_counter, ceid, field, value, snippet),
                        )

                    # fabrication_verifications row
                    reasoning = (
                        "source span does not ground the arm_value"
                        if verdict != "SUPPORTED" else None
                    )
                    fab_hyp = (
                        "plausible-sounding default"
                        if verdict == "UNSUPPORTED" else None
                    )
                    conn.execute(
                        """INSERT INTO fabrication_verifications
                           (judge_run_id, paper_id, field_name, arm_name,
                            pre_check_short_circuit, verdict,
                            verification_span, reasoning,
                            fabrication_hypothesis, verified_at)
                           VALUES ('fixture_run_1', ?, ?, ?, ?, ?, ?, ?, ?,
                                   '2026-04-01T00:00:00Z')""",
                        (
                            str(pid), field, arm,
                            1 if (k % 5 == 0) else 0,
                            verdict,
                            f"Evidence span: {value} in context.",
                            reasoning, fab_hyp,
                        ),
                    )

        conn.commit()
    finally:
        conn.close()

    return db_path, review_dir


@pytest.fixture()
def fixture_db(tmp_path):
    db_path, review_dir = _build_fixture_db(tmp_path)
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    yield conn, review_dir
    conn.close()


# ── Determinism ─────────────────────────────────────────────────────


def test_deterministic_sampling_byte_identical(fixture_db, tmp_path):
    conn, review_dir = fixture_db
    out1 = tmp_path / "out1"
    out2 = tmp_path / "out2"

    blinded1, key1, summary1 = sampler.generate(
        conn, review_dir, out1,
        config=SMALL_CONFIG, timestamp="T1",
    )
    blinded2, key2, summary2 = sampler.generate(
        conn, review_dir, out2,
        config=SMALL_CONFIG, timestamp="T2",
    )

    # Selected verification_ids (by row_id) identical.
    def _pairs(wb_path):
        wb = load_workbook(wb_path)
        ws = wb["Key"]
        out = []
        for r in ws.iter_rows(min_row=2, max_row=SMALL_TOTAL + 1,
                              values_only=True):
            out.append((r[0], r[1]))  # (row_id, verification_id)
        wb.close()
        return sorted(out)

    assert _pairs(key1) == _pairs(key2)

    # Row_id → verification_id mapping identical (randomization determinism).
    def _map(wb_path):
        wb = load_workbook(wb_path)
        ws = wb["Key"]
        m = {}
        for r in ws.iter_rows(min_row=2, max_row=SMALL_TOTAL + 1,
                              values_only=True):
            m[r[0]] = r[1]
        wb.close()
        return m

    assert _map(key1) == _map(key2)


# ── Allocation exactness ───────────────────────────────────────────


def test_allocation_exactness(fixture_db, tmp_path):
    conn, review_dir = fixture_db
    _, key_path, _ = sampler.generate(
        conn, review_dir, tmp_path / "out",
        config=SMALL_CONFIG, timestamp="TX",
    )
    wb = load_workbook(key_path)
    ws = wb["Key"]
    # Column positions: row_id=1, verification_id=2, paper_id=3,
    # ee_identifier=4, arm_name=5, field_name=6, arm_value=7,
    # sampling_stratum=8, gemma_verdict=9, ...
    counts: dict[tuple[str, str], int] = {}
    for r in ws.iter_rows(min_row=2, max_row=SMALL_TOTAL + 1,
                          values_only=True):
        stratum = r[7]
        arm = r[4]
        counts[(stratum, arm)] = counts.get((stratum, arm), 0) + 1
    wb.close()
    for verdict, by_arm in SMALL_ALLOCATION.items():
        for arm, want in by_arm.items():
            assert counts.get((verdict, arm), 0) == want, (
                f"cell ({verdict},{arm}) = {counts.get((verdict, arm), 0)}, "
                f"expected {want}"
            )


# ── Blinding ───────────────────────────────────────────────────────


def test_blinded_sheet_contains_no_forbidden_strings(fixture_db, tmp_path):
    conn, review_dir = fixture_db
    blinded_path, _, _ = sampler.generate(
        conn, review_dir, tmp_path / "out",
        config=SMALL_CONFIG, timestamp="TB",
    )
    hits = sampler._forbidden_strings_in_adjudication(blinded_path)
    assert hits == [], f"forbidden strings found: {hits}"


# ── Row randomization (anti-sort) ───────────────────────────────────


def test_row_ids_not_sorted_by_any_leakage_key(fixture_db, tmp_path):
    conn, review_dir = fixture_db
    _, key_path, _ = sampler.generate(
        conn, review_dir, tmp_path / "out",
        config=SMALL_CONFIG, timestamp="TR",
    )
    wb = load_workbook(key_path)
    ws = wb["Key"]
    ordered = []
    for r in ws.iter_rows(min_row=2, max_row=SMALL_TOTAL + 1,
                          values_only=True):
        ordered.append({
            "row_id": r[0], "verification_id": r[1], "paper_id": r[2],
            "arm_name": r[4], "field_name": r[5],
            "sampling_stratum": r[7], "gemma_verdict": r[8],
        })
    wb.close()
    ordered.sort(key=lambda d: d["row_id"])
    for key in ("paper_id", "arm_name", "field_name",
                "sampling_stratum", "gemma_verdict"):
        vals = [d[key] for d in ordered]
        asc = all(a <= b for a, b in zip(vals, vals[1:]))
        desc = all(a >= b for a, b in zip(vals, vals[1:]))
        assert not (asc or desc), f"row_id is monotonic in {key}"


# ── Key ↔ blinded pairing ──────────────────────────────────────────


def test_row_id_pairing_between_blinded_and_key(fixture_db, tmp_path):
    conn, review_dir = fixture_db
    blinded_path, key_path, _ = sampler.generate(
        conn, review_dir, tmp_path / "out",
        config=SMALL_CONFIG, timestamp="TP",
    )
    wb_b = load_workbook(blinded_path)
    wb_k = load_workbook(key_path)
    blinded_ids = {
        r[0] for r in wb_b["Adjudication"].iter_rows(
            min_row=2, max_row=SMALL_TOTAL + 1, max_col=1, values_only=True
        )
    }
    key_ids = {
        r[0] for r in wb_k["Key"].iter_rows(
            min_row=2, max_row=SMALL_TOTAL + 1, max_col=1, values_only=True
        )
    }
    wb_b.close()
    wb_k.close()
    assert blinded_ids == key_ids == set(range(1, SMALL_TOTAL + 1))


def test_stub_raises_not_implemented():
    from analysis.paper1 import pi_audit_unblind
    with pytest.raises(NotImplementedError):
        pi_audit_unblind.unblind_and_compute_precision("a", "b", "c")


# ── v2 windowing regression ─────────────────────────────────────────


def test_strategy_column_present_in_both_workbooks(fixture_db, tmp_path):
    conn, review_dir = fixture_db
    blinded_path, key_path, _ = sampler.generate(
        conn, review_dir, tmp_path / "out",
        config=SMALL_CONFIG, timestamp="TS",
    )
    wb_b = load_workbook(blinded_path)
    wb_k = load_workbook(key_path)

    blinded_headers = [
        c.value for c in next(wb_b["Adjudication"].iter_rows(max_row=1))
    ]
    key_headers = [
        c.value for c in next(wb_k["Key"].iter_rows(max_row=1))
    ]

    assert "source_window_strategy" in blinded_headers, (
        f"Adjudication headers: {blinded_headers}"
    )
    assert "source_window_strategy" in key_headers, (
        f"Key headers: {key_headers}"
    )
    # Blinded column order — strategy immediately after truncated flag.
    assert blinded_headers[:6] == [
        "row_id", "field_name", "arm_value",
        "source_text_truncated_for_workbook",
        "source_window_strategy",
        "source_text",
    ]
    wb_b.close()
    wb_k.close()


def test_strategy_distribution_totals_to_sample_size(fixture_db, tmp_path):
    conn, review_dir = fixture_db
    _, key_path, _ = sampler.generate(
        conn, review_dir, tmp_path / "out",
        config=SMALL_CONFIG, timestamp="TD",
    )
    wb = load_workbook(key_path)
    ws = wb["Key"]
    # source_window_strategy is col N (index 13)
    strat_col_idx = [c.value for c in next(ws.iter_rows(max_row=1))].index(
        "source_window_strategy"
    ) + 1
    totals: dict[str, int] = {}
    for r in ws.iter_rows(min_row=2, max_row=SMALL_TOTAL + 1,
                          values_only=True):
        strat = r[strat_col_idx - 1]
        totals[strat] = totals.get(strat, 0) + 1
    wb.close()
    # Every row must have a recognized strategy.
    recognized = {
        sampler.WindowStrategy.FULL_TEXT,
        sampler.WindowStrategy.PASS2_WINDOW,
        sampler.WindowStrategy.ARM_SPAN_WINDOW,
        sampler.WindowStrategy.ABSENCE_FALLBACK_HEAD,
        sampler.WindowStrategy.MISSING_SPAN_FALLBACK_HEAD,
    }
    assert set(totals.keys()) <= recognized, (
        f"unknown strategies: {set(totals.keys()) - recognized}"
    )
    assert sum(totals.values()) == SMALL_TOTAL


def test_arm_span_window_rows_contain_span_text(fixture_db, tmp_path):
    """arm_span_window source_text must contain a locatable prefix of
    the arm's evidence span. Guards against the v1 regression where
    long papers were hard-truncated at 32,767 chars from the front,
    stripping out the actual span being adjudicated."""
    conn, review_dir = fixture_db

    # Replace all 12 UNSUPPORTED/local fixture rows with a single big
    # paper row — then allocate exactly 1 in that cell so selection is
    # forced. Other cells stay at their original size; we zero them
    # out in the mini-allocation to keep the run small.
    phrase = "ANCHOR_TAG_92173"
    filler = "Lorem ipsum dolor sit amet, consectetur adipiscing elit. " * 800
    body = (
        filler
        + f"\n\n## Results\n\nSpan: {phrase} supporting this.\n\n"
        + filler
    )
    # Must exceed the Excel cell cap, but stay under the 20K-token
    # Pass 2 budget so arm_span_window (not pass2_window) fires.
    assert len(body) > sampler.EXCEL_CELL_MAX_CHARS

    big_pid = 9999
    parsed_dir = review_dir / "parsed_text"
    parsed_dir.mkdir(exist_ok=True)
    (parsed_dir / f"{big_pid}_v1.md").write_text(body)

    conn.execute(
        "INSERT OR IGNORE INTO papers (id, title, ee_identifier) VALUES (?, ?, ?)",
        (big_pid, "Big Fixture Paper", "EE-9999"),
    )
    conn.execute(
        "INSERT INTO extractions (id, paper_id) VALUES (?, ?)",
        (1000, big_pid),
    )
    conn.execute(
        """INSERT INTO evidence_spans
           (id, extraction_id, field_name, value, source_snippet)
           VALUES (?, ?, ?, ?, ?)""",
        (1000, 1000, "primary_outcome_metric", "accuracy",
         f"Span: {phrase} supporting this."),
    )
    # Clear UNSUPPORTED/local and insert just the big-paper row.
    conn.execute(
        """DELETE FROM fabrication_verifications
           WHERE judge_run_id = 'fixture_run_1'
             AND verdict = 'UNSUPPORTED' AND arm_name = 'local'""",
    )
    conn.execute(
        """INSERT INTO fabrication_verifications
           (judge_run_id, paper_id, field_name, arm_name,
            pre_check_short_circuit, verdict, verification_span,
            reasoning, fabrication_hypothesis, verified_at)
           VALUES ('fixture_run_1', ?, ?, 'local', 0, 'UNSUPPORTED',
                   ?, 'span does not ground the value',
                   'plausible-sounding default',
                   '2026-04-01T00:00:00Z')""",
        (str(big_pid), "primary_outcome_metric",
         f"Span: {phrase} supporting this."),
    )
    conn.commit()

    # Minimal allocation: exactly the UNSUPPORTED/local slot (1 row).
    mini_config = {
        "run_id": "fixture_run_1",
        "master_seed": 42,
        "allocation": {
            "UNSUPPORTED": {
                "local": 1, "openai_o4_mini_high": 0,
                "anthropic_sonnet_4_6": 0,
            },
            "PARTIALLY_SUPPORTED": {
                "local": 0, "openai_o4_mini_high": 0,
                "anthropic_sonnet_4_6": 0,
            },
            "SUPPORTED": {
                "local": 0, "openai_o4_mini_high": 0,
                "anthropic_sonnet_4_6": 0,
            },
        },
    }

    blinded_path, key_path, summary = sampler.generate(
        conn, review_dir, tmp_path / "out",
        config=mini_config, timestamp="TA",
    )

    wb = load_workbook(blinded_path)
    ws = wb["Adjudication"]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    strat_idx = headers.index("source_window_strategy")
    text_idx = headers.index("source_text")

    data_rows = list(ws.iter_rows(
        min_row=2, max_row=2, values_only=True,
    ))
    wb.close()
    assert len(data_rows) == 1
    r = data_rows[0]
    assert r[strat_idx] == sampler.WindowStrategy.ARM_SPAN_WINDOW, (
        f"strategy={r[strat_idx]} (expected arm_span_window); "
        f"summary={summary['strategy_totals']}"
    )
    assert phrase in r[text_idx], (
        f"arm_span_window row does NOT contain anchor phrase {phrase!r}: "
        f"{r[text_idx][:200]!r}"
    )
    assert len(r[text_idx]) <= sampler.EXCEL_CELL_MAX_CHARS
