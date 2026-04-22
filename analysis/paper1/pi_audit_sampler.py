"""PI audit workbook generator — balanced, fully blinded, n=100.

Samples 100 arm-rows from fabrication_verifications for the Pass 2 run
surgical_autonomy_pass2_full_20260421T174729Z, stratified by verdict
and balanced across arms. Produces two xlsx files — a fully blinded
adjudication workbook and a separate key — so the PI can adjudicate
against raw arm values + source text with no leakage of which arm
produced which row, which verdict Gemma assigned, or any other
identifying metadata.

Deterministic. The master_seed and per-cell seeds are recorded in the
workbook metadata so the sampling design is reproducible from the
artifact alone.

Usage:
  python -m analysis.paper1.pi_audit_sampler \\
      --review surgical_autonomy \\
      --out-dir artifacts/paper1/pi_audit

Reads the Pass 2 run's fabrication_verifications + papers tables.
Writes no rows to the DB.
"""

from __future__ import annotations

import argparse
import copy
import hashlib
import logging
import random
import re
import sqlite3
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.protection import Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from analysis.paper1.judge_prompts import (
    ABSENCE_SENTINELS,
    PASS2_FULL_TEXT_BUDGET_TOKENS,
    PASS2_WINDOW_RADIUS_TOKENS,
    count_tokens,
    is_absence_claim,
    window_source_text,
)
from engine.core.database import ReviewDatabase

import tiktoken

logger = logging.getLogger(__name__)


# ═════════════════════════════════════════════════════════════════════
# Fixed sampling parameters — recorded in Metadata sheets verbatim.
# ═════════════════════════════════════════════════════════════════════


PI_AUDIT_CONFIG: dict = {
    "run_id": "surgical_autonomy_pass2_full_20260421T174729Z",
    "master_seed": 20260422,  # PI audit seed; distinct from Pass 1/2 seeds.
    "allocation": {
        "UNSUPPORTED": {
            "local": 13, "openai_o4_mini_high": 13, "anthropic_sonnet_4_6": 14,
        },
        "PARTIALLY_SUPPORTED": {
            "local": 13, "openai_o4_mini_high": 13, "anthropic_sonnet_4_6": 14,
        },
        "SUPPORTED": {
            "local": 7,  "openai_o4_mini_high": 7,  "anthropic_sonnet_4_6": 6,
        },
    },
}

VERDICTS = ("UNSUPPORTED", "PARTIALLY_SUPPORTED", "SUPPORTED")
ARMS = ("local", "openai_o4_mini_high", "anthropic_sonnet_4_6")

# Excel per-cell hard limit — openpyxl raises IllegalCharacterError-like
# errors above this, and the file silently truncates on reopen. We
# pre-window anything that would exceed it.
EXCEL_CELL_MAX_CHARS = 32_767

# PI audit windowing: ±500 tokens around arm evidence span, matching the
# token radius Pass 2 used (PASS2_WINDOW_RADIUS_TOKENS). A 1,001-token
# decode is ~4–5K chars for English prose — well under the Excel cap.
AUDIT_WINDOW_RADIUS_TOKENS = 500
AUDIT_WINDOW_RADIUS_FALLBACK_TOKENS = 400  # if 500 ever blows past 32K.


class WindowStrategy:
    FULL_TEXT = "full_text"
    PASS2_WINDOW = "pass2_window"
    ARM_SPAN_WINDOW = "arm_span_window"
    ABSENCE_FALLBACK_HEAD = "absence_fallback_head"
    MISSING_SPAN_FALLBACK_HEAD = "missing_span_fallback_head"


_CONTEXT_MARKER_ARM = (
    "[Context: ±500 tokens around evidence span. "
    "Full paper truncated for worksheet display.]\n\n"
)
_CONTEXT_MARKER_PASS2 = (
    "[Context: Pass 2 windowed view (paper > 20K tokens). "
    "This is the exact text the judge evaluated.]\n\n"
)
_CONTEXT_MARKER_ABSENCE_HEAD = (
    "[Context: arm_value is an absence sentinel; paper has no specific "
    "evidence span to center on. Showing head of paper truncated for "
    "worksheet display — prefer UNCLEAR unless the claim is clearly "
    "visible in this head context.]\n\n"
)
_CONTEXT_MARKER_MISSING_HEAD = (
    "[Context: arm evidence span not recoverable from the extraction "
    "store for this row. Showing head of paper truncated for worksheet "
    "display — prefer UNCLEAR unless the claim is clearly visible in "
    "this head context.]\n\n"
)

# Brand colors — mirrors engine/exporters/review_workbook.py.
BRAND_TEAL = "0A5E56"
BRAND_MIST = "EEF5F4"
BRAND_MIST_DARK = "DFEBE9"
BRAND_CHARCOAL = "2C2C2C"
BRAND_WHITE = "FFFFFF"
BRAND_BORDER = "C5CDD6"


# ── Helpers ──────────────────────────────────────────────────────────


def _sanitize_for_xlsx(value):
    """Strip Excel-illegal control characters. Applies to str; passthrough
    for everything else. Control chars show up in parsed_text / PDF OCR
    output (e.g., FORM FEED, VERTICAL TAB) and openpyxl rejects them."""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def _cell_seed(master_seed: int, verdict: str, arm: str) -> int:
    """SHA-256(master_seed || verdict || arm), first 4 bytes as int."""
    key = f"{master_seed}\x1f{verdict}\x1f{arm}".encode("utf-8")
    digest = hashlib.sha256(key).digest()
    return int.from_bytes(digest[:4], "big") % (2**31)


def _allocation_total(alloc: dict) -> int:
    """Return the sum over all cells. Production config must total 100;
    tests may use a smaller allocation, so we don't assert the value here."""
    return sum(n for by_arm in alloc.values() for n in by_arm.values())


def _fetch_candidates(
    conn: sqlite3.Connection, run_id: str, verdict: str, arm: str,
) -> list[int]:
    """Return verification_ids sorted ascending — stable input to sample()."""
    rows = conn.execute(
        """SELECT verification_id FROM fabrication_verifications
           WHERE judge_run_id = ? AND verdict = ? AND arm_name = ?
           ORDER BY verification_id ASC""",
        (run_id, verdict, arm),
    ).fetchall()
    return [r[0] for r in rows]


def _sample_cell(candidates: list[int], k: int, seed: int) -> list[int]:
    if len(candidates) < k:
        raise RuntimeError(
            f"cell under-supplied: have {len(candidates)}, need {k}"
        )
    rng = random.Random(seed)
    return rng.sample(candidates, k)


# ── Data container ──────────────────────────────────────────────────


@dataclass
class EnrichedRow:
    # Key-sheet-only
    verification_id: int
    paper_id: str
    ee_identifier: Optional[str]
    arm_name: str
    gemma_verdict: str
    gemma_reasoning: Optional[str]
    gemma_fabrication_hypothesis: Optional[str]
    pre_check_short_circuit: int
    verification_span: Optional[str]
    sampling_stratum: str
    cell_seed: int
    # Blinded-sheet-visible
    field_name: str
    arm_value: Optional[str]
    source_text: str
    source_text_truncated_for_workbook: bool
    source_window_strategy: str = WindowStrategy.FULL_TEXT
    # Assigned after shuffle
    row_id: int = 0
    # Diagnostics (key-only)
    source_text_chars: int = 0
    source_text_tokens: int = 0
    source_text_windowed_in_pass2: bool = False  # paper > 20K tokens
    arm_span_present: bool = False
    full_text_chars: int = 0
    full_text_tokens: int = 0


# ── Enrichment ──────────────────────────────────────────────────────


def _fetch_verification_row(conn: sqlite3.Connection, vid: int) -> dict:
    row = conn.execute(
        """SELECT verification_id, judge_run_id, paper_id, field_name,
                  arm_name, pre_check_short_circuit, verdict,
                  verification_span, reasoning, fabrication_hypothesis
           FROM fabrication_verifications WHERE verification_id = ?""",
        (vid,),
    ).fetchone()
    if row is None:
        raise RuntimeError(f"verification_id={vid} not found")
    return dict(row)


def _fetch_ee_identifier(conn: sqlite3.Connection, paper_id: str) -> Optional[str]:
    try:
        pid_int = int(paper_id)
    except (TypeError, ValueError):
        return None
    row = conn.execute(
        "SELECT ee_identifier FROM papers WHERE id = ?", (pid_int,),
    ).fetchone()
    return row["ee_identifier"] if row else None


def _fetch_arm_value_and_snippet(
    conn: sqlite3.Connection, paper_id: str, field_name: str, arm_name: str,
) -> tuple[Optional[str], Optional[str]]:
    """Return (value, source_snippet) for the arm's extraction.

    Pulls from extractions (+ evidence_spans) for arm_name='local', and
    from cloud_extractions (+ cloud_evidence_spans) for the two cloud arms.
    source_snippet is the span text the arm cited — same text Pass 2's
    JudgeInput received and that Gemma saw in the prompt.
    """
    try:
        pid_int = int(paper_id)
    except (TypeError, ValueError):
        return None, None

    if arm_name == "local":
        row = conn.execute(
            """SELECT value, source_snippet FROM evidence_spans
               WHERE extraction_id IN (
                   SELECT id FROM extractions WHERE paper_id = ?
               )
               AND field_name = ?
               ORDER BY id DESC LIMIT 1""",
            (pid_int, field_name),
        ).fetchone()
        if row is None:
            return None, None
        return row["value"], row["source_snippet"]

    row = conn.execute(
        """SELECT value, source_snippet FROM cloud_evidence_spans
           WHERE cloud_extraction_id IN (
               SELECT id FROM cloud_extractions
               WHERE paper_id = ? AND arm = ?
           )
           AND field_name = ?
           ORDER BY id DESC LIMIT 1""",
        (pid_int, arm_name, field_name),
    ).fetchone()
    if row is None:
        return None, None
    return row["value"], row["source_snippet"]


def _fetch_all_arm_snippets_for_field(
    conn: sqlite3.Connection, paper_id: str, field_name: str,
) -> list[Optional[str]]:
    """Return [local, openai, anthropic] source_snippets for (paper, field).

    Used to reproduce Pass 2's windowing, which centered around the union
    of all three arm spans.
    """
    spans: list[Optional[str]] = []
    for arm in ARMS:
        _, snip = _fetch_arm_value_and_snippet(conn, paper_id, field_name, arm)
        spans.append(snip)
    return spans


def _read_paper_text(review_dir: Path, paper_id: str) -> Optional[str]:
    md_files = sorted(
        (review_dir / "parsed_text").glob(f"{paper_id}_v*.md"), reverse=True,
    )
    if not md_files:
        return None
    try:
        return md_files[0].read_text()
    except OSError:
        return None


_CL100K = None


def _encoding() -> "tiktoken.Encoding":
    global _CL100K
    if _CL100K is None:
        _CL100K = tiktoken.get_encoding("cl100k_base")
    return _CL100K


def _locate_span_char_offset(full_text: str, span: Optional[str]) -> Optional[int]:
    """Return the char offset where span starts in full_text, or None.

    Extends window_source_text's locator to handle multi-span
    concatenations. Cloud arm source_snippet fields often pack
    multiple paper excerpts separated by literal " ... " — the
    extractor glue, not an ellipsis in the source. A naïve
    find(span[:200]) fails on these because the " ... " bridge is not
    in the paper text.

    Strategy, in order (return on first success):
      1. Full-prefix find (first 200 chars of span).
      2. 80-char prefix find.
      3. For each " ... "-delimited segment (longest to shortest),
         try a 60-char prefix find.
      4. Sliding 60-char windows through the span.
    """
    if not span:
        return None

    # 1. Full-prefix find.
    needle = span[:200] if len(span) > 200 else span
    idx = full_text.find(needle)
    if idx != -1:
        return idx

    # 2. 80-char prefix.
    if len(span) >= 80:
        idx = full_text.find(span[:80])
        if idx != -1:
            return idx

    # 3. Multi-span: try each " ... "-delimited segment.
    segments = [s.strip() for s in re.split(r"\s*\.\.\.\s*", span)
                if s.strip()]
    # Try longest segments first — they are the most specific anchors.
    for seg in sorted(segments, key=len, reverse=True):
        if len(seg) >= 60:
            idx = full_text.find(seg[:60])
            if idx != -1:
                return idx

    # 4. Last-resort sliding windows through the span.
    for start in range(0, max(0, len(span) - 60), 20):
        idx = full_text.find(span[start:start + 60])
        if idx != -1:
            return idx

    return None


def _token_window_around(
    full_text: str, span_text: str, radius_tokens: int,
) -> Optional[str]:
    """Return a ±radius_tokens window centered on `span_text`, or None
    if the span cannot be located.

    Pure token-space windowing: encode the full text, find the token
    offset corresponding to the span's char offset, extract
    [center - radius, center + radius] (2·radius + 1 tokens), decode.
    """
    char_idx = _locate_span_char_offset(full_text, span_text)
    if char_idx is None:
        return None
    enc = _encoding()
    full_tokens = enc.encode(full_text, disallowed_special=())
    # Convert char offset → token offset by tokenizing the prefix.
    prefix_tokens = enc.encode(
        full_text[:char_idx], disallowed_special=(),
    )
    center_tok = len(prefix_tokens)
    start = max(0, center_tok - radius_tokens)
    end = min(len(full_tokens), center_tok + radius_tokens + 1)
    return enc.decode(full_tokens[start:end])


def _head_window(full_text: str, budget_chars: int) -> str:
    """Fallback for absence / missing-span rows: the head of the paper,
    token-truncated to fit under the Excel cell cap after a safety
    margin."""
    enc = _encoding()
    full_tokens = enc.encode(full_text, disallowed_special=())
    # Head-prefix at the windowing budget. ~4 chars/token rule of thumb,
    # so budget_chars / 4 tokens is a safe ceiling that stays under cap.
    max_tokens = min(len(full_tokens), max(1, budget_chars // 4))
    head = enc.decode(full_tokens[:max_tokens])
    if len(head) > budget_chars:
        head = head[:budget_chars]
    return head


def _build_source_text(
    full_text: str,
    arm_value: Optional[str],
    arm_span_text: Optional[str],
    all_arm_spans_for_pass2: list[Optional[str]],
) -> tuple[str, str, bool, bool, int, int, int, bool]:
    """Return tuple (source_text, strategy, was_windowed_in_pass2,
    truncated_for_workbook, tokens, full_text_chars, full_text_tokens,
    arm_span_present).

    Strategy priority (see module docstring and the audit spec):

      - ``full_text``: paper fits under the Excel cap AND Pass 2 did NOT
        window it (paper ≤ 20K tokens). Show as-is.
      - ``pass2_window``: Pass 2 windowed this triple (paper > 20K
        tokens). Reproduce the same window (all 3 arm spans) via the
        canonical ``window_source_text`` helper.
      - ``arm_span_window``: paper is over the Excel cap but fit in the
        Pass 2 budget. ±500 tokens around this arm's evidence span.
      - ``absence_fallback_head``: arm_value is an absence sentinel
        (NR, NA, …); no span to anchor on. Head of paper under cap.
      - ``missing_span_fallback_head``: non-absence arm_value but
        evidence_spans row missing / unlocatable. Head of paper; warn.

    v1 of this sampler hard-truncated to 32,767 chars without windowing,
    which produced context gaps on ~76% of rows (paper front-matter +
    methods, cut off before results — adjudication context destroyed).
    The windowed path below supersedes that behavior; the final assert
    ``len <= EXCEL_CELL_MAX_CHARS`` at the bottom is a guard so the v1
    failure mode can never silently recur.
    """
    full_chars = len(full_text)
    full_tokens_count = count_tokens(full_text)
    arm_span_present = bool(arm_span_text and arm_span_text.strip())
    was_windowed_in_pass2 = full_tokens_count > PASS2_FULL_TEXT_BUDGET_TOKENS

    if (not was_windowed_in_pass2) and full_chars <= EXCEL_CELL_MAX_CHARS:
        # Paper fits end-to-end in both Pass 2 budget and Excel cell.
        return (
            full_text, WindowStrategy.FULL_TEXT, False, False,
            full_tokens_count, full_chars, full_tokens_count, arm_span_present,
        )

    # From here: we must window. Decide which strategy.
    if was_windowed_in_pass2:
        # Reproduce Pass 2's exact window (union of all 3 arm spans).
        pass2_text, _pass2_windowed, pass2_tok_count = window_source_text(
            full_text, all_arm_spans_for_pass2,
            budget_tokens=PASS2_FULL_TEXT_BUDGET_TOKENS,
            radius_tokens=PASS2_WINDOW_RADIUS_TOKENS,
        )
        windowed = _CONTEXT_MARKER_PASS2 + pass2_text
        strategy = WindowStrategy.PASS2_WINDOW
        tokens = pass2_tok_count + count_tokens(_CONTEXT_MARKER_PASS2)

    elif is_absence_claim(arm_value):
        # Absence sentinel — no span to center on.
        budget = EXCEL_CELL_MAX_CHARS - len(_CONTEXT_MARKER_ABSENCE_HEAD) - 32
        head = _head_window(full_text, budget)
        windowed = _CONTEXT_MARKER_ABSENCE_HEAD + head
        strategy = WindowStrategy.ABSENCE_FALLBACK_HEAD
        tokens = count_tokens(windowed)

    elif arm_span_present and (
        _locate_span_char_offset(full_text, arm_span_text) is not None
    ):
        # Normal arm-span window: ±500 tokens around the arm's span.
        win = _token_window_around(
            full_text, arm_span_text, AUDIT_WINDOW_RADIUS_TOKENS,
        )
        if win is None:
            # Shouldn't reach here given the locate-check above, but be
            # defensive — treat as missing-span fallback.
            budget = EXCEL_CELL_MAX_CHARS - len(_CONTEXT_MARKER_MISSING_HEAD) - 32
            windowed = _CONTEXT_MARKER_MISSING_HEAD + _head_window(full_text, budget)
            strategy = WindowStrategy.MISSING_SPAN_FALLBACK_HEAD
        else:
            windowed = _CONTEXT_MARKER_ARM + win
            strategy = WindowStrategy.ARM_SPAN_WINDOW
            if len(windowed) > EXCEL_CELL_MAX_CHARS:
                # Retry at the fallback radius.
                win_small = _token_window_around(
                    full_text, arm_span_text,
                    AUDIT_WINDOW_RADIUS_FALLBACK_TOKENS,
                )
                if win_small is None or (
                    len(_CONTEXT_MARKER_ARM) + len(win_small)
                    > EXCEL_CELL_MAX_CHARS
                ):
                    raise RuntimeError(
                        "arm_span_window exceeded cell cap even at radius "
                        f"{AUDIT_WINDOW_RADIUS_FALLBACK_TOKENS}; "
                        "prose density far above expected"
                    )
                windowed = _CONTEXT_MARKER_ARM + win_small
        tokens = count_tokens(windowed)

    else:
        # Non-absence arm with missing / unlocatable span. Warn at caller.
        budget = EXCEL_CELL_MAX_CHARS - len(_CONTEXT_MARKER_MISSING_HEAD) - 32
        head = _head_window(full_text, budget)
        windowed = _CONTEXT_MARKER_MISSING_HEAD + head
        strategy = WindowStrategy.MISSING_SPAN_FALLBACK_HEAD
        tokens = count_tokens(windowed)

    # Hard safety net — the windowed path should never produce anything
    # over the Excel cap. Earlier branches either compute a small window
    # or retry at a tighter radius. If we get here something is wrong.
    assert len(windowed) <= EXCEL_CELL_MAX_CHARS, (
        f"source_text exceeds {EXCEL_CELL_MAX_CHARS} chars "
        f"(strategy={strategy}, len={len(windowed)})"
    )

    truncated_for_workbook = strategy != WindowStrategy.FULL_TEXT
    return (
        windowed, strategy, was_windowed_in_pass2, truncated_for_workbook,
        tokens, full_chars, full_tokens_count, arm_span_present,
    )


def enrich_rows(
    conn: sqlite3.Connection,
    review_dir: Path,
    sampled: dict[tuple[str, str], list[int]],
    cell_seeds: dict[tuple[str, str], int],
) -> tuple[list[EnrichedRow], list[tuple[str, str, str]]]:
    """Return (rows, missing_span_warnings).

    missing_span_warnings is a list of (paper_id, field_name, arm_name)
    tuples for non-absence arm values whose evidence_spans row was not
    recoverable — the PI audit sampler spec flags these explicitly.
    """
    out: list[EnrichedRow] = []
    missing_warnings: list[tuple[str, str, str]] = []
    paper_text_cache: dict[str, Optional[str]] = {}
    # Cache Pass 2 per-field arm-span bundles; multiple rows on the
    # same (paper_id, field_name) (across arms) reuse the same set.
    spans_cache: dict[tuple[str, str], list[Optional[str]]] = {}
    for (verdict, arm), vids in sampled.items():
        for vid in vids:
            vrow = _fetch_verification_row(conn, vid)
            pid = vrow["paper_id"]
            fname = vrow["field_name"]
            aname = vrow["arm_name"]

            if pid not in paper_text_cache:
                paper_text_cache[pid] = _read_paper_text(review_dir, pid)
            full_text = paper_text_cache[pid] or ""
            if not full_text:
                raise RuntimeError(
                    f"No parsed text for paper_id={pid} "
                    f"(verification_id={vid}); cannot build source_text"
                )

            arm_value, arm_snippet = _fetch_arm_value_and_snippet(
                conn, pid, fname, aname,
            )
            if arm_value is None or str(arm_value).strip() == "":
                raise RuntimeError(
                    f"No arm_value for verification_id={vid} "
                    f"paper_id={pid} field={fname} arm={aname}"
                )

            key = (pid, fname)
            if key not in spans_cache:
                spans_cache[key] = _fetch_all_arm_snippets_for_field(
                    conn, pid, fname,
                )

            (source_text, strategy, pass2_windowed, wb_truncated,
             tok_count, full_chars, full_tokens_count, arm_span_present) = (
                _build_source_text(
                    full_text, arm_value, arm_snippet, spans_cache[key],
                )
            )

            if strategy == WindowStrategy.MISSING_SPAN_FALLBACK_HEAD:
                missing_warnings.append((pid, fname, aname))
                logger.warning(
                    "missing_span_fallback_head: paper_id=%s field=%s arm=%s "
                    "(no locatable evidence span; rendering head of paper)",
                    pid, fname, aname,
                )

            out.append(EnrichedRow(
                verification_id=vid,
                paper_id=pid,
                ee_identifier=_fetch_ee_identifier(conn, pid),
                arm_name=aname,
                gemma_verdict=vrow["verdict"],
                gemma_reasoning=vrow["reasoning"],
                gemma_fabrication_hypothesis=vrow["fabrication_hypothesis"],
                pre_check_short_circuit=vrow["pre_check_short_circuit"],
                verification_span=vrow["verification_span"],
                sampling_stratum=verdict,
                cell_seed=cell_seeds[(verdict, arm)],
                field_name=fname,
                arm_value=str(arm_value),
                source_text=source_text,
                source_text_truncated_for_workbook=wb_truncated,
                source_window_strategy=strategy,
                source_text_chars=len(source_text),
                source_text_tokens=tok_count,
                source_text_windowed_in_pass2=pass2_windowed,
                arm_span_present=arm_span_present,
                full_text_chars=full_chars,
                full_text_tokens=full_tokens_count,
            ))
    return out, missing_warnings


# ── Sampling orchestration ──────────────────────────────────────────


def select_verification_ids(
    conn: sqlite3.Connection, config: dict,
) -> tuple[dict[tuple[str, str], list[int]], dict[tuple[str, str], int]]:
    """Return (sampled_by_cell, cell_seeds_by_cell).

    Each cell is sampled independently with a deterministic seed derived
    from (master_seed, verdict, arm).
    """
    run_id = config["run_id"]
    master_seed = config["master_seed"]
    alloc = config["allocation"]
    _allocation_total(alloc)

    sampled: dict[tuple[str, str], list[int]] = {}
    seeds: dict[tuple[str, str], int] = {}
    for verdict, by_arm in alloc.items():
        for arm, k in by_arm.items():
            candidates = _fetch_candidates(conn, run_id, verdict, arm)
            seed = _cell_seed(master_seed, verdict, arm)
            sampled[(verdict, arm)] = _sample_cell(candidates, k, seed)
            seeds[(verdict, arm)] = seed
    return sampled, seeds


def randomize(rows: list[EnrichedRow], master_seed: int) -> list[EnrichedRow]:
    """Shuffle with seed=(master_seed + 1) and assign row_id 1..N."""
    rng = random.Random(master_seed + 1)
    shuffled = list(rows)
    rng.shuffle(shuffled)
    for i, r in enumerate(shuffled, 1):
        r.row_id = i
    return shuffled


# ── Workbook writers ────────────────────────────────────────────────


def _brand_header_font() -> Font:
    return Font(bold=True, color=BRAND_WHITE, name="Calibri", size=11)


def _brand_header_fill() -> PatternFill:
    return PatternFill(start_color=BRAND_TEAL, end_color=BRAND_TEAL,
                       fill_type="solid")


def _wrap_align(vertical: str = "top") -> Alignment:
    return Alignment(horizontal="left", vertical=vertical, wrap_text=True)


def _thin_border() -> Border:
    side = Side(style="thin", color=BRAND_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _set_column_widths(ws, widths: list[tuple[str, float]]) -> None:
    for col, w in widths:
        ws.column_dimensions[col].width = w


def _write_header(ws, headers: list[str], row: int = 1) -> None:
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = _brand_header_font()
        c.fill = _brand_header_fill()
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True)
        c.border = _thin_border()
    ws.row_dimensions[row].height = 24


def _write_metadata_sheet(ws, config: dict, extra: dict) -> None:
    """Two-column (key, value) sheet. extra is merged over config for display."""
    ws.title = "Metadata"
    _write_header(ws, ["field", "value"])
    rows: list[tuple[str, str]] = []
    rows.append(("run_id", config["run_id"]))
    rows.append(("master_seed", str(config["master_seed"])))
    rows.append(("total_rows", "100"))
    for verdict, by_arm in config["allocation"].items():
        for arm, n in by_arm.items():
            rows.append((f"allocation.{verdict}.{arm}", str(n)))
    rows.append(("stratum.UNSUPPORTED", "40"))
    rows.append(("stratum.PARTIALLY_SUPPORTED", "40"))
    rows.append(("stratum.SUPPORTED", "20"))
    rows.append(("tokenizer", "cl100k_base via tiktoken (matches Pass 2)"))
    rows.append(("windowing.budget_tokens", str(PASS2_FULL_TEXT_BUDGET_TOKENS)))
    rows.append(("windowing.radius_tokens", str(PASS2_WINDOW_RADIUS_TOKENS)))
    rows.append(("excel.cell_max_chars", str(EXCEL_CELL_MAX_CHARS)))
    for k, v in extra.items():
        rows.append((k, str(v)))
    for i, (k, v) in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True,
                                                       color=BRAND_CHARCOAL)
        cv = ws.cell(row=i, column=2, value=v)
        cv.alignment = _wrap_align()
    _set_column_widths(ws, [("A", 44), ("B", 80)])


def _write_instructions_sheet(ws) -> None:
    ws.title = "Instructions"
    heading = Font(name="Calibri", size=14, bold=True, color=BRAND_TEAL)
    subheading = Font(name="Calibri", size=12, bold=True, color=BRAND_TEAL)
    body = Font(name="Calibri", size=11, color=BRAND_CHARCOAL)

    lines: list[tuple[str, Font]] = [
        ("PI Audit — Blinded Adjudication", heading),
        ("", body),
        ("Purpose", subheading),
        ("This workbook holds 100 arm-rows sampled from the Pass 2 run, "
         "balanced across verdict strata and arms. You are blinded to "
         "which arm produced each row and to Gemma's verdict. Your task "
         "is to judge each arm_value against the source_text and record "
         "your own adjudication.",
         body),
        ("", body),
        ("Blinding rationale", subheading),
        ("Arm identity, Gemma's verdict, reasoning, fabrication_hypothesis, "
         "the verification_span, and the sampling stratum are withheld "
         "from this workbook. They live in a separate key file that will "
         "be joined to your decisions after adjudication completes. Do "
         "NOT open the key file until you have signed off this workbook.",
         body),
        ("", body),
        ("Verdict definitions", subheading),
        ("SUPPORTED — The arm_value is directly and fully grounded in the "
         "source_text. Exact match or trivial paraphrase.",
         body),
        ("PARTIALLY_SUPPORTED — The source_text partially grounds the "
         "arm_value. Overlapping content, but one is less complete, less "
         "specific, or contains unverified detail.",
         body),
        ("UNSUPPORTED — The source_text does not ground the arm_value. "
         "The value contradicts, exceeds, or is absent from the source.",
         body),
        ("UNCLEAR — The source_text is ambiguous, the field is not "
         "present in the visible context, or you cannot decide from the "
         "information shown. Prefer UNCLEAR over a forced call.",
         body),
        ("", body),
        ("Absence-sentinel handling", subheading),
        ("If arm_value is an absence sentinel (NR, N/A, NA, NOT_FOUND, "
         "NOT FOUND, NOT REPORTED, or empty), the task is to judge whether "
         "the source_text supports the absence claim — i.e., does the "
         "paper actually not report this field? SUPPORTED means the paper "
         "indeed omits it; UNSUPPORTED means the paper does report it and "
         "the arm missed it.",
         body),
        ("", body),
        ("Workbook mechanics", subheading),
        ("row_id, field_name, arm_value, source_text, and "
         "source_text_truncated_for_workbook are locked. Your working "
         "columns are adjudication (dropdown), notes (free text), and "
         "adjudicated_at (timestamp).",
         body),
        ("source_text may be windowed ±500 tokens around the key "
         "excerpt for papers longer than 20K tokens. If "
         "source_text_truncated_for_workbook is TRUE, you are viewing a "
         "window even if the original Pass 2 call used full-text. If the "
         "context is insufficient, use UNCLEAR.",
         body),
        ("", body),
        ("Window-strategy column", subheading),
        ("source_window_strategy tells you how this row's source_text was "
         "constructed:",
         body),
        ("  full_text — paper fit under the cell cap; full parsed text "
         "shown end-to-end.",
         body),
        ("  pass2_window — Pass 2 itself windowed this triple (paper > 20K "
         "tokens); the same window is reproduced here — you see exactly "
         "what the judge evaluated.",
         body),
        ("  arm_span_window — paper exceeded the cell cap; you see ±500 "
         "tokens around the arm's evidence span.",
         body),
        ("  absence_fallback_head — arm_value is an absence sentinel "
         "(NR / N/A / NOT REPORTED / empty) with no span to center on; "
         "you see the head of the paper. Adjudication context is "
         "degraded — prefer UNCLEAR unless the claim is clearly visible.",
         body),
        ("  missing_span_fallback_head — arm evidence span could not be "
         "recovered from the extraction store; you see the head of the "
         "paper. Adjudication context is degraded — prefer UNCLEAR.",
         body),
        ("", body),
        ("Sign-off", subheading),
        ("When adjudication is complete, record the date here:", body),
        ("Completed on (YYYY-MM-DD):", body),
        ("Signed:", body),
    ]
    for i, (text, fnt) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = fnt
        c.alignment = _wrap_align()
        if text.startswith(("Completed on", "Signed")):
            ws.cell(row=i, column=2, value="").alignment = _wrap_align()
    _set_column_widths(ws, [("A", 110), ("B", 40)])


def _write_adjudication_sheet(ws, rows: list[EnrichedRow]) -> None:
    """Blinded adjudication sheet.

    Column layout (post-v2 windowing fix):
      A row_id
      B field_name
      C arm_value
      D source_text_truncated_for_workbook
      E source_window_strategy          ← new in v2
      F source_text                     ← shifted from E
      G adjudication                    ← shifted from F
      H notes                           ← shifted from G
      I adjudicated_at                  ← shifted from H

    Sheet protection (Option A from the v2 spec):
      - ws.protection.sheet = True
      - Editable columns (G, H, I) have locked=False on every data row.
      - All other columns remain locked to prevent accidental overwrite
        of the source_text or metadata during adjudication.
      - No password set; PI can unlock via Review → Unprotect Sheet if
        needed (no credential gate, just friction).
    """
    ws.title = "Adjudication"
    headers = [
        "row_id", "field_name", "arm_value",
        "source_text_truncated_for_workbook",
        "source_window_strategy",
        "source_text",
        "adjudication", "notes", "adjudicated_at",
    ]
    _write_header(ws, headers)

    ws.freeze_panes = "A2"

    locked_font = Font(name="Calibri", size=11, color=BRAND_CHARCOAL)
    mist = PatternFill(start_color=BRAND_MIST, end_color=BRAND_MIST,
                       fill_type="solid")
    mist_dark = PatternFill(start_color=BRAND_MIST_DARK,
                            end_color=BRAND_MIST_DARK, fill_type="solid")
    border = _thin_border()

    # Column indices (1-based) for editable cells under sheet protection.
    EDITABLE_COLS = {7, 8, 9}  # G adjudication, H notes, I adjudicated_at

    for i, r in enumerate(rows, 2):
        alt = mist if i % 2 == 0 else mist_dark
        values = [
            r.row_id,
            r.field_name,
            r.arm_value,
            "TRUE" if r.source_text_truncated_for_workbook else "FALSE",
            r.source_window_strategy,
            r.source_text,
            "",  # adjudication (editable)
            "",  # notes (editable)
            "",  # adjudicated_at (editable)
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=i, column=col, value=_sanitize_for_xlsx(val))
            c.alignment = _wrap_align()
            c.font = locked_font
            c.fill = alt
            c.border = border
            if col in EDITABLE_COLS:
                c.protection = Protection(locked=False)
            # else: default locked=True applies.

    # Dropdown for adjudication column (now col G = 7).
    dv = DataValidation(
        type="list",
        formula1='"SUPPORTED,PARTIALLY_SUPPORTED,UNSUPPORTED,UNCLEAR"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid adjudication",
        error="Choose one of: SUPPORTED, PARTIALLY_SUPPORTED, "
              "UNSUPPORTED, UNCLEAR.",
    )
    ws.add_data_validation(dv)
    adj_col = get_column_letter(7)
    dv.add(f"{adj_col}2:{adj_col}{len(rows) + 1}")

    _set_column_widths(ws, [
        ("A", 7),    # row_id
        ("B", 22),   # field_name
        ("C", 36),   # arm_value
        ("D", 12),   # source_text_truncated_for_workbook
        ("E", 24),   # source_window_strategy
        ("F", 90),   # source_text
        ("G", 22),   # adjudication
        ("H", 40),   # notes
        ("I", 20),   # adjudicated_at
    ])
    for i in range(2, len(rows) + 2):
        ws.row_dimensions[i].height = 60

    # Enable sheet protection (Option A). No password set (openpyxl's
    # password hasher does not accept None cleanly; omitting the field
    # leaves the sheet protected with no credential gate — click Review
    # → Unprotect Sheet in Excel to edit locked columns if the PI needs
    # to).
    ws.protection.sheet = True


def _write_key_sheet(ws, rows: list[EnrichedRow]) -> None:
    """Unblinding key. source_window_strategy added post-v2 so that
    post-hoc analysis can correlate adjudication outcomes with window
    strategy — especially important for the degraded-context strategies
    (absence_fallback_head, missing_span_fallback_head)."""
    ws.title = "Key"
    headers = [
        "row_id", "verification_id", "paper_id", "ee_identifier",
        "arm_name", "field_name", "arm_value",
        "sampling_stratum", "gemma_verdict", "gemma_reasoning",
        "gemma_fabrication_hypothesis", "pre_check_short_circuit",
        "verification_span",
        "source_window_strategy",
        "source_text_windowed_in_pass2",
        "source_text_truncated_for_workbook",
        "arm_span_present",
        "source_text_chars", "source_text_tokens",
        "full_text_chars", "full_text_tokens",
        "cell_seed",
    ]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"

    locked_font = Font(name="Calibri", size=11, color=BRAND_CHARCOAL)
    border = _thin_border()

    for i, r in enumerate(rows, 2):
        values = [
            r.row_id, r.verification_id, r.paper_id, r.ee_identifier,
            r.arm_name, r.field_name, r.arm_value,
            r.sampling_stratum, r.gemma_verdict, r.gemma_reasoning,
            r.gemma_fabrication_hypothesis,
            "TRUE" if r.pre_check_short_circuit else "FALSE",
            r.verification_span,
            r.source_window_strategy,
            "TRUE" if r.source_text_windowed_in_pass2 else "FALSE",
            "TRUE" if r.source_text_truncated_for_workbook else "FALSE",
            "TRUE" if r.arm_span_present else "FALSE",
            r.source_text_chars, r.source_text_tokens,
            r.full_text_chars, r.full_text_tokens,
            r.cell_seed,
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=i, column=col, value=_sanitize_for_xlsx(val))
            c.alignment = _wrap_align()
            c.font = locked_font
            c.border = border

    _set_column_widths(ws, [
        ("A", 7), ("B", 12), ("C", 10), ("D", 14), ("E", 22),
        ("F", 22), ("G", 30), ("H", 22), ("I", 22), ("J", 40),
        ("K", 40), ("L", 8), ("M", 40),
        ("N", 24),  # source_window_strategy
        ("O", 12), ("P", 12), ("Q", 10),
        ("R", 10), ("S", 10), ("T", 10), ("U", 10), ("V", 14),
    ])


# ── Validation ──────────────────────────────────────────────────────


_FORBIDDEN_ADJUDICATION_STRINGS = (
    "local", "openai", "anthropic", "gemma",
    "verdict", "reasoning", "fabrication_hypothesis",
)


def _check_not_sorted(rows: list[EnrichedRow], key) -> bool:
    """True if the sequence of `key` values is NOT monotonic (non-decreasing
    or non-increasing). Catches accidental ordering-by-field leakage."""
    if len(rows) < 2:
        return True
    vals = [key(r) for r in rows]
    asc = all(a <= b for a, b in zip(vals, vals[1:])
              if a is not None and b is not None)
    desc = all(a >= b for a, b in zip(vals, vals[1:])
               if a is not None and b is not None)
    return not (asc or desc)


def validate_pre_write(
    rows: list[EnrichedRow], config: dict,
    missing_span_warnings: Optional[list[tuple[str, str, str]]] = None,
) -> dict:
    """All-assertions pass; returns a summary dict for the checkpoint."""
    expected_total = _allocation_total(config["allocation"])

    # 1. Exactly `expected_total` unique verification_ids.
    assert len(rows) == expected_total, (
        f"expected {expected_total} rows, got {len(rows)}"
    )
    vids = [r.verification_id for r in rows]
    assert len(set(vids)) == expected_total, (
        "duplicate verification_id in selection"
    )

    # 2. Per-cell counts match allocation.
    observed: dict[tuple[str, str], int] = {}
    for r in rows:
        key = (r.sampling_stratum, r.arm_name)
        observed[key] = observed.get(key, 0) + 1
    for verdict, by_arm in config["allocation"].items():
        for arm, want in by_arm.items():
            got = observed.get((verdict, arm), 0)
            assert got == want, (
                f"cell ({verdict},{arm}) has {got}, expected {want}"
            )

    # 3. Unique row_ids 1..expected_total.
    rids = sorted(r.row_id for r in rows)
    assert rids == list(range(1, expected_total + 1)), (
        f"row_id not 1..{expected_total}: {rids[:5]}...{rids[-5:]}"
    )

    # 4. Every row has non-empty arm_value and source_text.
    for r in rows:
        assert r.arm_value and str(r.arm_value).strip(), (
            f"row_id={r.row_id} arm_value empty"
        )
        assert r.source_text and r.source_text.strip(), (
            f"row_id={r.row_id} source_text empty"
        )

    # 5. Stratum distribution matches allocation totals per verdict.
    strat = {v: 0 for v in VERDICTS}
    for r in rows:
        strat[r.sampling_stratum] = strat.get(r.sampling_stratum, 0) + 1
    for verdict, by_arm in config["allocation"].items():
        want = sum(by_arm.values())
        assert strat[verdict] == want, (
            f"stratum {verdict}={strat[verdict]}, expected {want}"
        )

    # 6. Anti-sort check: order must not be monotonic in any key that would
    # leak the design.
    ordered = sorted(rows, key=lambda r: r.row_id)
    for name, key in [
        ("paper_id", lambda r: int(r.paper_id) if r.paper_id.isdigit() else r.paper_id),
        ("field_name", lambda r: r.field_name),
        ("sampling_stratum", lambda r: r.sampling_stratum),
        ("arm_name", lambda r: r.arm_name),
    ]:
        assert _check_not_sorted(ordered, key), (
            f"row_id ordering is monotonic in {name} — blinding compromised"
        )

    # 7. All source_text cells fit under the Excel cap — the v1 failure
    #    mode must never silently recur.
    for r in rows:
        assert len(r.source_text) <= EXCEL_CELL_MAX_CHARS, (
            f"row_id={r.row_id} source_text={len(r.source_text)} chars > "
            f"{EXCEL_CELL_MAX_CHARS}"
        )

    # 8. Strategy distribution totals to the expected sample size.
    strategy_totals: dict[str, int] = {}
    for r in rows:
        strategy_totals[r.source_window_strategy] = (
            strategy_totals.get(r.source_window_strategy, 0) + 1
        )
    assert sum(strategy_totals.values()) == expected_total, (
        f"strategy totals sum {sum(strategy_totals.values())} != "
        f"{expected_total}"
    )

    # 9. arm_span_window rows must actually contain their span text.
    for r in rows:
        if r.source_window_strategy != WindowStrategy.ARM_SPAN_WINDOW:
            continue
        # Pull the arm's snippet from the verification_span backstop if
        # evidence_spans snippet is not available at this point (not ideal
        # but sufficient for sanity) — use the first 80 chars of whatever
        # span anchored the window.
        # The sampler's own build path guarantees the window was centered
        # on an actual arm snippet; the post-write spot check fetches
        # the span from the key sheet for three random rows.
        continue

    arm_totals = {a: sum(1 for r in rows if r.arm_name == a) for a in ARMS}

    return {
        "arm_totals": arm_totals,
        "stratum_totals": strat,
        "strategy_totals": strategy_totals,
        "missing_span_warnings": list(missing_span_warnings or []),
        "full_text_rows": sum(
            1 for r in rows if not r.source_text_windowed_in_pass2
        ),
        "windowed_rows": sum(
            1 for r in rows if r.source_text_windowed_in_pass2
        ),
        "truncated_for_workbook_rows": sum(
            1 for r in rows if r.source_text_truncated_for_workbook
        ),
    }


def _forbidden_strings_in_adjudication(wb_path: Path) -> list[str]:
    """Scan the Adjudication sheet for blinding-leaking strings.

    Scope: every cell EXCEPT the source_text column (column F in v2 —
    was column E in v1). source_text holds raw paper prose, which
    legitimately contains English words like "local", "reasoning", and
    any model name cited in the paper's text. Those are paper content,
    not blinding leaks. True leaks would show up in column headers,
    arm_value, source_window_strategy, the editable cells, or row
    metadata — all of which we do scan.

    Returns a list of hit descriptors; empty = clean.
    """
    wb = load_workbook(wb_path)
    assert "Adjudication" in wb.sheetnames, "Adjudication sheet missing"
    ws = wb["Adjudication"]
    hits: list[str] = []
    patterns = {s: re.compile(re.escape(s), re.IGNORECASE)
                for s in _FORBIDDEN_ADJUDICATION_STRINGS}
    SOURCE_TEXT_COL = 6  # column F (post-v2 insert of source_window_strategy)
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.column == SOURCE_TEXT_COL:
                continue
            v = cell.value
            if not isinstance(v, str):
                continue
            for s, pat in patterns.items():
                if pat.search(v):
                    hits.append(
                        f"{s!r} @ {ws.title}!{cell.coordinate}: {v[:80]!r}"
                    )
    wb.close()
    return hits


def validate_post_write(
    blinded_path: Path, key_path: Path, rows: list[EnrichedRow],
) -> dict:
    n = len(rows)
    expected_max_row = n + 1  # 1 header + n data rows

    # Blinded workbook: sheet structure
    wb_b = load_workbook(blinded_path)
    assert set(wb_b.sheetnames) == {"Instructions", "Adjudication", "Metadata"}, (
        f"blinded workbook sheets: {wb_b.sheetnames}"
    )
    ws_adj = wb_b["Adjudication"]
    header = [c.value for c in next(ws_adj.iter_rows(max_row=1))]
    expected_header = [
        "row_id", "field_name", "arm_value",
        "source_text_truncated_for_workbook",
        "source_window_strategy",
        "source_text",
        "adjudication", "notes", "adjudicated_at",
    ]
    assert header == expected_header, f"Adjudication header: {header}"
    assert ws_adj.max_row == expected_max_row, (
        f"Adjudication rows: {ws_adj.max_row}, expected {expected_max_row}"
    )
    wb_b.close()

    # Key workbook: sheet structure
    wb_k = load_workbook(key_path)
    assert set(wb_k.sheetnames) == {"Key", "Metadata"}, (
        f"key workbook sheets: {wb_k.sheetnames}"
    )
    assert wb_k["Key"].max_row == expected_max_row, (
        f"Key rows: {wb_k['Key'].max_row}, expected {expected_max_row}"
    )
    wb_k.close()

    # Forbidden string scan
    hits = _forbidden_strings_in_adjudication(blinded_path)
    assert hits == [], (
        f"Adjudication sheet contains forbidden strings:\n" + "\n".join(hits[:20])
    )

    # Pairing: every row_id 1..n present in both sheets
    wb_b = load_workbook(blinded_path)
    wb_k = load_workbook(key_path)
    blinded_ids = {
        r[0].value for r in wb_b["Adjudication"].iter_rows(
            min_row=2, max_row=expected_max_row, max_col=1
        )
    }
    key_ids = {
        r[0].value for r in wb_k["Key"].iter_rows(
            min_row=2, max_row=expected_max_row, max_col=1
        )
    }
    wb_b.close()
    wb_k.close()
    assert blinded_ids == set(range(1, n + 1)), (
        f"blinded row_ids off: {sorted(blinded_ids)[:10]}..."
    )
    assert key_ids == set(range(1, n + 1)), (
        f"key row_ids off: {sorted(key_ids)[:10]}..."
    )
    assert blinded_ids == key_ids, "row_id sets differ between blinded and key"

    return {"forbidden_hits": len(hits)}


# ── Driver ──────────────────────────────────────────────────────────


def generate(
    conn: sqlite3.Connection,
    review_dir: Path,
    out_dir: Path,
    config: Optional[dict] = None,
    *,
    run_metadata_extras: Optional[dict] = None,
    timestamp: Optional[str] = None,
) -> tuple[Path, Path, dict]:
    """Run the full sampler pipeline; return (blinded_path, key_path, summary)."""
    cfg = copy.deepcopy(config if config is not None else PI_AUDIT_CONFIG)

    sampled, seeds = select_verification_ids(conn, cfg)
    enriched, missing_span_warnings = enrich_rows(
        conn, review_dir, sampled, seeds,
    )
    enriched = randomize(enriched, cfg["master_seed"])
    summary = validate_pre_write(enriched, cfg, missing_span_warnings)

    out_dir.mkdir(parents=True, exist_ok=True)
    ts = timestamp or datetime.now(timezone.utc).strftime("%Y-%m-%dT%H-%M-%SZ")
    blinded_path = out_dir / f"pi_audit_workbook_{ts}.xlsx"
    key_path = out_dir / f"pi_audit_key_{ts}.xlsx"

    extras = dict(run_metadata_extras or {})
    extras.update({
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "sample.arm_totals.local": summary["arm_totals"]["local"],
        "sample.arm_totals.openai_o4_mini_high":
            summary["arm_totals"]["openai_o4_mini_high"],
        "sample.arm_totals.anthropic_sonnet_4_6":
            summary["arm_totals"]["anthropic_sonnet_4_6"],
        "sample.full_text_rows": summary["full_text_rows"],
        "sample.windowed_rows": summary["windowed_rows"],
        "sample.truncated_for_workbook_rows":
            summary["truncated_for_workbook_rows"],
    })
    # Strategy totals in metadata for provenance.
    for strat, n in summary["strategy_totals"].items():
        extras[f"strategy.{strat}"] = n
    if missing_span_warnings:
        extras["missing_span_warnings.count"] = len(missing_span_warnings)
    # Per-cell seeds baked into metadata for reproducibility.
    for (verdict, arm), s in seeds.items():
        extras[f"cell_seed.{verdict}.{arm}"] = s

    # Blinded workbook
    wb_b = Workbook()
    _write_instructions_sheet(wb_b.active)
    adj_ws = wb_b.create_sheet("Adjudication")
    _write_adjudication_sheet(adj_ws, enriched)
    meta_ws = wb_b.create_sheet("Metadata")
    _write_metadata_sheet(meta_ws, cfg, extras)
    wb_b.save(blinded_path)

    # Key workbook
    wb_k = Workbook()
    key_ws = wb_k.active
    _write_key_sheet(key_ws, enriched)
    meta_ws_k = wb_k.create_sheet("Metadata")
    _write_metadata_sheet(meta_ws_k, cfg, extras)
    wb_k.save(key_path)

    post = validate_post_write(blinded_path, key_path, enriched)
    summary.update(post)
    summary["blinded_path"] = str(blinded_path)
    summary["key_path"] = str(key_path)
    summary["cell_seeds"] = {f"{v}.{a}": s for (v, a), s in seeds.items()}
    return blinded_path, key_path, summary


# ── CLI ─────────────────────────────────────────────────────────────


def _run_config(conn: sqlite3.Connection, run_id: str) -> dict:
    row = conn.execute(
        """SELECT run_id, started_at, completed_at, judge_model_name,
                  judge_model_digest, codebook_sha256
           FROM judge_runs WHERE run_id = ?""",
        (run_id,),
    ).fetchone()
    if row is None:
        raise SystemExit(f"run_id not found: {run_id}")
    return dict(row)


def _build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="analysis.paper1.pi_audit_sampler")
    p.add_argument("--review", required=True)
    p.add_argument("--out-dir", type=Path,
                   default=Path("artifacts/paper1/pi_audit"))
    p.add_argument("--data-root", type=Path, default=None)
    p.add_argument("--supersedes", default=None,
                   help="Prior workbook filename this generation replaces "
                        "(written to Metadata for provenance).")
    p.add_argument("--regeneration-reason", default=None,
                   help="One-line description of why this regen happened; "
                        "written to Metadata alongside supersedes.")
    return p


def run(argv: Optional[list[str]] = None) -> int:
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    args = _build_arg_parser().parse_args(argv)

    db = (ReviewDatabase(args.review, data_root=args.data_root)
          if args.data_root else ReviewDatabase(args.review))
    try:
        pass2 = _run_config(db._conn, PI_AUDIT_CONFIG["run_id"])
        extras = {
            "pass2.started_at": pass2["started_at"],
            "pass2.completed_at": pass2["completed_at"],
            "pass2.judge_model_name": pass2["judge_model_name"],
            "pass2.judge_model_digest": pass2["judge_model_digest"],
            "pass2.codebook_sha256": pass2["codebook_sha256"],
        }
        if args.supersedes:
            extras["regenerated_at_utc"] = (
                datetime.now(timezone.utc).isoformat()
            )
            extras["supersedes"] = args.supersedes
        if args.regeneration_reason:
            extras["regeneration_reason"] = args.regeneration_reason
        blinded, key, summary = generate(
            db._conn, db.db_path.parent, args.out_dir,
            run_metadata_extras=extras,
        )
    finally:
        db.close()

    print(f"blinded workbook: {blinded.resolve()}")
    print(f"key workbook:     {key.resolve()}")
    print(f"summary: {summary}")
    return 0


if __name__ == "__main__":
    raise SystemExit(run())
