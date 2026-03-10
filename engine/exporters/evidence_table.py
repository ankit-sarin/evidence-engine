"""Evidence table exports: CSV and Excel."""

import csv
import json
import logging
from pathlib import Path

import openpyxl

from engine.core.database import ReviewDatabase
from engine.core.review_spec import ReviewSpec

logger = logging.getLogger(__name__)


# ── Helpers ──────────────────────────────────────────────────────────


def _build_evidence_rows(
    db: ReviewDatabase, spec: ReviewSpec,
    min_status: str = "AI_AUDIT_COMPLETE",
) -> tuple[list[str], list[list]]:
    """Build header and data rows for the evidence table.

    Args:
        min_status: Minimum paper status to include. Papers at or beyond this
            status are included. Default "AI_AUDIT_COMPLETE" for raw AI output.
            Use "HUMAN_AUDIT_COMPLETE" for human-verified production exports.

    Returns (headers, rows) where each row is one included paper.
    """
    from engine.core.database import _STATUS_ORDER

    field_names = [f.name for f in spec.extraction_schema.fields]

    # Base columns
    headers = ["paper_id", "pmid", "doi", "title", "authors", "year", "journal"]
    # Per extraction field: value, source_snippet, confidence, audit_status
    for fname in field_names:
        headers.extend([fname, f"{fname}_snippet", f"{fname}_confidence", f"{fname}_audit"])

    # Get papers that meet or exceed min_status
    min_level = _STATUS_ORDER.get(min_status, 0)
    qualifying_statuses = [s for s, level in _STATUS_ORDER.items() if level >= min_level]
    placeholders = ", ".join("?" for _ in qualifying_statuses)
    papers = db._conn.execute(
        f"SELECT * FROM papers WHERE status IN ({placeholders}) ORDER BY id",
        qualifying_statuses,
    ).fetchall()

    rows = []
    for paper in papers:
        pid = paper["id"]
        row = [
            pid,
            paper["pmid"],
            paper["doi"],
            paper["title"],
            paper["authors"],
            paper["year"],
            paper["journal"],
        ]

        # Get latest extraction's spans
        extraction = db._conn.execute(
            "SELECT id FROM extractions WHERE paper_id = ? ORDER BY id DESC LIMIT 1",
            (pid,),
        ).fetchone()

        span_map = {}
        if extraction:
            spans = db._conn.execute(
                "SELECT * FROM evidence_spans WHERE extraction_id = ?",
                (extraction["id"],),
            ).fetchall()
            for s in spans:
                span_map[s["field_name"]] = s

        for fname in field_names:
            span = span_map.get(fname)
            if span:
                row.extend([
                    span["value"],
                    span["source_snippet"] or "",
                    span["confidence"],
                    span["audit_status"],
                ])
            else:
                row.extend(["", "", "", ""])

        rows.append(row)

    return headers, rows


# ── CSV Export ───────────────────────────────────────────────────────


def export_evidence_csv(
    db: ReviewDatabase, spec: ReviewSpec, output_path: str,
    min_status: str = "AI_AUDIT_COMPLETE",
) -> None:
    """Export evidence table as CSV."""
    headers, rows = _build_evidence_rows(db, spec, min_status=min_status)

    with open(output_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)

    logger.info("Evidence CSV exported to %s (%d rows)", output_path, len(rows))


# ── Excel Export ─────────────────────────────────────────────────────


def export_evidence_excel(
    db: ReviewDatabase, spec: ReviewSpec, output_path: str,
    min_status: str = "AI_AUDIT_COMPLETE",
) -> None:
    """Export evidence table as Excel with 3 sheets."""
    wb = openpyxl.Workbook()

    # Sheet 1: Evidence Table
    ws1 = wb.active
    ws1.title = "Evidence Table"
    headers, rows = _build_evidence_rows(db, spec, min_status=min_status)
    ws1.append(headers)
    for row in rows:
        ws1.append(row)
    _style_header(ws1)

    # Sheet 2: Screening Log
    ws2 = wb.create_sheet("Screening Log")
    ws2.append(["paper_id", "title", "pass_number", "decision", "rationale", "model", "decided_at"])
    screen_rows = db._conn.execute(
        """SELECT p.id, p.title, sd.pass_number, sd.decision,
                  sd.rationale, sd.model, sd.decided_at
           FROM screening_decisions sd
           JOIN papers p ON p.id = sd.paper_id
           ORDER BY p.id, sd.pass_number"""
    ).fetchall()
    for r in screen_rows:
        ws2.append(list(dict(r).values()))
    _style_header(ws2)

    # Sheet 3: Audit Log
    ws3 = wb.create_sheet("Audit Log")
    ws3.append([
        "paper_id", "title", "field_name", "value",
        "source_snippet", "confidence", "audit_status",
        "auditor_model", "audit_rationale",
    ])
    audit_rows = db._conn.execute(
        """SELECT p.id, p.title, es.field_name, es.value,
                  es.source_snippet, es.confidence, es.audit_status,
                  es.auditor_model, es.audit_rationale
           FROM evidence_spans es
           JOIN extractions e ON e.id = es.extraction_id
           JOIN papers p ON p.id = e.paper_id
           ORDER BY p.id, es.id"""
    ).fetchall()
    for r in audit_rows:
        ws3.append(list(dict(r).values()))
    _style_header(ws3)

    wb.save(output_path)
    logger.info("Evidence Excel exported to %s", output_path)


def _style_header(ws) -> None:
    """Bold the header row."""
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)
