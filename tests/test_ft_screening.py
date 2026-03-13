"""Tests for full-text screening pipeline — ft_screener, ft_screening_adjudicator,
database FT tables, workflow stages, specialty scope in prompts, text truncation.
"""

import json
import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

from engine.adjudication.ft_screening_adjudicator import (
    _collect_ft_flagged,
    check_ft_adjudication_gate,
    export_ft_adjudication_queue,
    import_ft_adjudication_decisions,
)
from engine.adjudication.workflow import (
    complete_stage,
    is_stage_done,
)
from engine.agents.ft_screener import (
    FTScreeningDecision,
    FTVerificationDecision,
    build_ft_screening_prompt,
    build_ft_verification_prompt,
    truncate_paper_text,
)
from engine.core.constants import FT_MAX_TEXT_CHARS, FT_REASON_CODES
from engine.core.database import ReviewDatabase
from engine.core.review_spec import load_review_spec
from engine.search.models import Citation


# ── Fixtures ────────────────────────────────────────────────────────


@pytest.fixture
def tmp_db(tmp_path):
    """Create a ReviewDatabase in a temp directory."""
    db = ReviewDatabase("test_review", data_root=tmp_path)
    yield db
    db.close()


@pytest.fixture
def spec():
    """Load the surgical autonomy review spec."""
    return load_review_spec("review_specs/surgical_autonomy_v1.yaml")


def _add_paper(db, title="Test Paper", pmid=None, doi=None, abstract="Test abstract"):
    """Helper to add a paper and return its id."""
    cit = Citation(
        title=title, abstract=abstract,
        pmid=pmid, doi=doi, source="pubmed",
        authors=["A"], journal="J Test", year=2024,
    )
    db.add_papers([cit])
    row = db._conn.execute(
        "SELECT id FROM papers WHERE title = ?", (title,)
    ).fetchone()
    return row["id"]


def _advance_to_parsed(db, paper_id):
    """Move a paper through INGESTED → ABSTRACT_SCREENED_IN → PDF_ACQUIRED → PARSED."""
    db.update_status(paper_id, "ABSTRACT_SCREENED_IN")
    db.update_status(paper_id, "PDF_ACQUIRED")
    db.update_status(paper_id, "PARSED")


def _advance_to_ft_flagged(db, paper_id):
    """Move a paper through to FT_FLAGGED, adding FT decisions along the way."""
    _advance_to_parsed(db, paper_id)
    db.update_status(paper_id, "FT_ELIGIBLE")
    db.add_ft_screening_decision(
        paper_id, "qwen3.5:27b", "FT_ELIGIBLE", "eligible",
        "Paper describes autonomous suturing", 0.9,
    )
    db.add_ft_verification_decision(
        paper_id, "gemma3:27b", "FT_FLAGGED",
        "Verifier disagrees — no autonomous component", 0.7,
    )
    db.update_status(paper_id, "FT_FLAGGED")


# ── Text Truncation Tests ──────────────────────────────────────────


class TestTruncation:

    def test_short_text_no_truncation(self):
        text = "This is a short paper."
        result = truncate_paper_text(text, title="Title", abstract="Abstract")
        assert "Title: Title" in result
        assert "Abstract: Abstract" in result
        assert "This is a short paper." in result

    def test_truncation_at_references(self):
        body = "Introduction section content. " * 200  # ~6000 chars
        body += "\n## References\n\n[1] Foo et al. 2024\n[2] Bar et al. 2023\n"
        result = truncate_paper_text(body, max_chars=2000)
        assert "References" not in result
        assert "Introduction" in result

    def test_truncation_respects_max_chars(self):
        body = "x" * 50_000
        result = truncate_paper_text(body, max_chars=1000)
        assert len(result) <= 1000

    def test_title_and_abstract_always_included(self):
        body = "x" * 50_000
        result = truncate_paper_text(body, title="My Title", abstract="My Abstract", max_chars=500)
        assert "Title: My Title" in result
        assert "Abstract: My Abstract" in result

    def test_empty_text(self):
        result = truncate_paper_text("", title="T", abstract="A")
        assert "Title: T" in result
        assert "Abstract: A" in result

    def test_default_max_chars(self):
        assert FT_MAX_TEXT_CHARS == 32_000


# ── Pydantic Model Tests ──────────────────────────────────────────


class TestDecisionModels:

    def test_screening_decision_valid(self):
        d = FTScreeningDecision(
            decision="FT_ELIGIBLE", reason_code="eligible",
            rationale="Paper qualifies", confidence=0.95,
        )
        assert d.decision == "FT_ELIGIBLE"
        assert d.reason_code == "eligible"

    def test_screening_decision_exclude(self):
        d = FTScreeningDecision(
            decision="FT_EXCLUDE", reason_code="wrong_specialty",
            rationale="Dental surgery", confidence=0.8,
        )
        assert d.decision == "FT_EXCLUDE"

    def test_screening_decision_invalid_decision(self):
        with pytest.raises(Exception):
            FTScreeningDecision(
                decision="INCLUDE", reason_code="eligible",
                rationale="test", confidence=0.5,
            )

    def test_verification_decision_valid(self):
        d = FTVerificationDecision(
            decision="FT_ELIGIBLE", rationale="Confirmed eligible", confidence=0.9,
        )
        assert d.decision == "FT_ELIGIBLE"

    def test_verification_decision_flagged(self):
        d = FTVerificationDecision(
            decision="FT_FLAGGED", rationale="No autonomy found", confidence=0.85,
        )
        assert d.decision == "FT_FLAGGED"

    def test_confidence_bounds(self):
        with pytest.raises(Exception):
            FTScreeningDecision(
                decision="FT_ELIGIBLE", reason_code="eligible",
                rationale="test", confidence=1.5,
            )


# ── Constants Tests ───────────────────────────────────────────────


class TestConstants:

    def test_reason_codes_tuple(self):
        assert "eligible" in FT_REASON_CODES
        assert "wrong_specialty" in FT_REASON_CODES
        assert "no_autonomy_content" in FT_REASON_CODES
        assert len(FT_REASON_CODES) == 7


# ── Database FT Decision Tests ────────────────────────────────────


class TestFTDecisionDB:

    def test_ft_screening_decisions_table_exists(self, tmp_db):
        row = tmp_db._conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='ft_screening_decisions'"
        ).fetchone()
        assert row is not None

    def test_ft_verification_decisions_table_exists(self, tmp_db):
        row = tmp_db._conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='ft_verification_decisions'"
        ).fetchone()
        assert row is not None

    def test_add_ft_screening_decision(self, tmp_db):
        pid = _add_paper(tmp_db, title="FT Test", pmid="99001")
        dec_id = tmp_db.add_ft_screening_decision(
            pid, "qwen3.5:27b", "FT_ELIGIBLE", "eligible",
            "Paper qualifies", 0.95,
        )
        assert dec_id > 0

        row = tmp_db._conn.execute(
            "SELECT * FROM ft_screening_decisions WHERE id = ?", (dec_id,)
        ).fetchone()
        assert row["paper_id"] == pid
        assert row["decision"] == "FT_ELIGIBLE"
        assert row["reason_code"] == "eligible"
        assert row["confidence"] == 0.95

    def test_add_ft_verification_decision(self, tmp_db):
        pid = _add_paper(tmp_db, title="FT Verify", pmid="99002")
        dec_id = tmp_db.add_ft_verification_decision(
            pid, "gemma3:27b", "FT_FLAGGED",
            "No autonomous component found", 0.7,
        )
        assert dec_id > 0

        row = tmp_db._conn.execute(
            "SELECT * FROM ft_verification_decisions WHERE id = ?", (dec_id,)
        ).fetchone()
        assert row["decision"] == "FT_FLAGGED"
        assert row["rationale"] == "No autonomous component found"

    def test_ft_screening_decision_check_constraint(self, tmp_db):
        pid = _add_paper(tmp_db, title="FT Check", pmid="99003")
        with pytest.raises(Exception):
            tmp_db._conn.execute(
                """INSERT INTO ft_screening_decisions
                   (paper_id, model, decision, reason_code, rationale, confidence, decided_at)
                   VALUES (?, 'test', 'INVALID', 'eligible', 'test', 0.5, '2024-01-01')""",
                (pid,),
            )

    def test_ft_verification_decision_check_constraint(self, tmp_db):
        pid = _add_paper(tmp_db, title="FT Check V", pmid="99004")
        with pytest.raises(Exception):
            tmp_db._conn.execute(
                """INSERT INTO ft_verification_decisions
                   (paper_id, model, decision, rationale, confidence, decided_at)
                   VALUES (?, 'test', 'INVALID', 'test', 0.5, '2024-01-01')""",
                (pid,),
            )


# ── Status Transition Tests ──────────────────────────────────────


class TestFTTransitions:

    def test_parsed_to_ft_eligible(self, tmp_db):
        pid = _add_paper(tmp_db, title="Trans1", pmid="88001")
        _advance_to_parsed(tmp_db, pid)
        tmp_db.update_status(pid, "FT_ELIGIBLE")
        row = tmp_db._conn.execute(
            "SELECT status FROM papers WHERE id = ?", (pid,)
        ).fetchone()
        assert row["status"] == "FT_ELIGIBLE"

    def test_parsed_to_ft_screened_out(self, tmp_db):
        pid = _add_paper(tmp_db, title="Trans2", pmid="88002")
        _advance_to_parsed(tmp_db, pid)
        tmp_db.update_status(pid, "FT_SCREENED_OUT")
        row = tmp_db._conn.execute(
            "SELECT status FROM papers WHERE id = ?", (pid,)
        ).fetchone()
        assert row["status"] == "FT_SCREENED_OUT"

    def test_parsed_to_ft_flagged(self, tmp_db):
        pid = _add_paper(tmp_db, title="Trans3", pmid="88003")
        _advance_to_parsed(tmp_db, pid)
        tmp_db.update_status(pid, "FT_FLAGGED")
        row = tmp_db._conn.execute(
            "SELECT status FROM papers WHERE id = ?", (pid,)
        ).fetchone()
        assert row["status"] == "FT_FLAGGED"

    def test_ft_eligible_to_extracted(self, tmp_db):
        pid = _add_paper(tmp_db, title="Trans4", pmid="88004")
        _advance_to_parsed(tmp_db, pid)
        tmp_db.update_status(pid, "FT_ELIGIBLE")
        tmp_db.update_status(pid, "EXTRACTED")
        row = tmp_db._conn.execute(
            "SELECT status FROM papers WHERE id = ?", (pid,)
        ).fetchone()
        assert row["status"] == "EXTRACTED"

    def test_ft_flagged_to_ft_eligible(self, tmp_db):
        pid = _add_paper(tmp_db, title="Trans5", pmid="88005")
        _advance_to_parsed(tmp_db, pid)
        tmp_db.update_status(pid, "FT_FLAGGED")
        tmp_db.update_status(pid, "FT_ELIGIBLE")
        row = tmp_db._conn.execute(
            "SELECT status FROM papers WHERE id = ?", (pid,)
        ).fetchone()
        assert row["status"] == "FT_ELIGIBLE"

    def test_ft_flagged_to_ft_screened_out(self, tmp_db):
        pid = _add_paper(tmp_db, title="Trans6", pmid="88006")
        _advance_to_parsed(tmp_db, pid)
        tmp_db.update_status(pid, "FT_FLAGGED")
        tmp_db.update_status(pid, "FT_SCREENED_OUT")
        row = tmp_db._conn.execute(
            "SELECT status FROM papers WHERE id = ?", (pid,)
        ).fetchone()
        assert row["status"] == "FT_SCREENED_OUT"

    def test_ft_screened_out_is_terminal(self, tmp_db):
        pid = _add_paper(tmp_db, title="Trans7", pmid="88007")
        _advance_to_parsed(tmp_db, pid)
        tmp_db.update_status(pid, "FT_SCREENED_OUT")
        with pytest.raises(ValueError, match="Invalid transition"):
            tmp_db.update_status(pid, "FT_ELIGIBLE")

    def test_parsed_can_skip_ft_to_extracted(self, tmp_db):
        """PARSED can go directly to EXTRACTED (for reviews without FT screening)."""
        pid = _add_paper(tmp_db, title="Trans8", pmid="88008")
        _advance_to_parsed(tmp_db, pid)
        tmp_db.update_status(pid, "EXTRACTED")
        row = tmp_db._conn.execute(
            "SELECT status FROM papers WHERE id = ?", (pid,)
        ).fetchone()
        assert row["status"] == "EXTRACTED"


# ── Prompt Builder Tests ─────────────────────────────────────────


class TestPromptBuilders:

    def test_ft_screening_prompt_contains_pico(self, spec):
        prompt = build_ft_screening_prompt("Paper text here", spec)
        assert "Population:" in prompt
        assert "Intervention:" in prompt
        assert "autonomous" in prompt.lower()

    def test_ft_screening_prompt_contains_criteria(self, spec):
        prompt = build_ft_screening_prompt("Paper text here", spec)
        assert "INCLUSION CRITERIA:" in prompt
        assert "EXCLUSION CRITERIA:" in prompt

    def test_ft_screening_prompt_contains_specialty_scope(self, spec):
        prompt = build_ft_screening_prompt("Paper text here", spec)
        assert "SPECIALTY SCOPE" in prompt or "specialty" in prompt.lower()
        assert "abdominal surgery" in prompt.lower() or "Included" in prompt

    def test_ft_screening_prompt_contains_reason_codes(self, spec):
        prompt = build_ft_screening_prompt("Paper text here", spec)
        assert "eligible" in prompt
        assert "wrong_specialty" in prompt
        assert "no_autonomy_content" in prompt

    def test_ft_screening_prompt_contains_paper_text(self, spec):
        prompt = build_ft_screening_prompt("MY UNIQUE PAPER CONTENT", spec)
        assert "MY UNIQUE PAPER CONTENT" in prompt

    def test_ft_screening_prompt_has_no_think(self, spec):
        prompt = build_ft_screening_prompt("text", spec)
        assert prompt.startswith("/no_think")

    def test_ft_verification_prompt_strict(self, spec):
        prompt = build_ft_verification_prompt("Paper text here", spec)
        assert "VERIFICATION" in prompt
        assert "false positives" in prompt.lower() or "FP" in prompt
        assert "FT_FLAGGED" in prompt
        assert "FT_ELIGIBLE" in prompt

    def test_ft_verification_prompt_contains_5_tests(self, spec):
        prompt = build_ft_verification_prompt("Paper text here", spec)
        assert "1." in prompt
        assert "5." in prompt

    def test_ft_verification_prompt_has_no_think(self, spec):
        prompt = build_ft_verification_prompt("text", spec)
        assert prompt.startswith("/no_think")


# ── ReviewSpec FT Models Tests ───────────────────────────────────


class TestReviewSpecFTModels:

    def test_ft_screening_models_loaded(self, spec):
        assert spec.ft_screening_models is not None
        assert spec.ft_screening_models.primary == "qwen3.5:27b"
        assert spec.ft_screening_models.verifier == "gemma3:27b"

    def test_ft_screening_models_think_false(self, spec):
        assert spec.ft_screening_models.think is False

    def test_ft_screening_models_temperature_zero(self, spec):
        assert spec.ft_screening_models.temperature == 0.0


# ── FT Adjudication Export/Import Tests ──────────────────────────


class TestFTAdjudication:

    def test_collect_no_flagged(self, tmp_db):
        result = _collect_ft_flagged(tmp_db)
        assert result == []

    def test_collect_ft_flagged(self, tmp_db):
        pid = _add_paper(tmp_db, title="Flagged Paper", pmid="77001")
        _advance_to_ft_flagged(tmp_db, pid)

        result = _collect_ft_flagged(tmp_db)
        assert len(result) == 1
        assert result[0]["paper_id"] == pid
        assert result[0]["title"] == "Flagged Paper"
        assert result[0]["primary_decision"] == "FT_ELIGIBLE"
        assert result[0]["verifier_decision"] == "FT_FLAGGED"

    def test_export_empty_queue(self, tmp_db, tmp_path):
        out = tmp_path / "empty.xlsx"
        result = export_ft_adjudication_queue(tmp_db, out)
        assert result["total"] == 0

    def test_export_ft_queue(self, tmp_db, tmp_path):
        pid = _add_paper(tmp_db, title="Export Test", pmid="77002")
        _advance_to_ft_flagged(tmp_db, pid)

        out = tmp_path / "ft_queue.xlsx"
        result = export_ft_adjudication_queue(tmp_db, out)
        assert result["total"] == 1
        assert out.exists()

    def test_export_creates_xlsx_sheets(self, tmp_db, tmp_path):
        pid = _add_paper(tmp_db, title="Sheet Test", pmid="77003")
        _advance_to_ft_flagged(tmp_db, pid)

        out = tmp_path / "ft_sheets.xlsx"
        export_ft_adjudication_queue(tmp_db, out)

        from openpyxl import load_workbook
        wb = load_workbook(out)
        assert "Instructions" in wb.sheetnames
        assert "Review Queue" in wb.sheetnames
        assert "FT Screening Criteria" in wb.sheetnames

    def test_import_ft_decisions(self, tmp_db, tmp_path):
        pid1 = _add_paper(tmp_db, title="Import Eligible", pmid="77010")
        _advance_to_ft_flagged(tmp_db, pid1)

        pid2 = _add_paper(tmp_db, title="Import Excluded", pmid="77011")
        _advance_to_ft_flagged(tmp_db, pid2)

        # Export
        out = tmp_path / "ft_import.xlsx"
        export_ft_adjudication_queue(tmp_db, out)

        # Fill in decisions using header-based lookup
        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb["Review Queue"]
        headers = {cell.value: cell.column - 1 for cell in ws[1] if cell.value}
        pid_col = next(i for h, i in headers.items() if "Paper ID" in h)
        dec_col = next(i for h, i in headers.items() if "PI_decision" in h)
        for row in ws.iter_rows(min_row=2, values_only=False):
            paper_id_val = row[pid_col].value
            if paper_id_val == pid1:
                row[dec_col].value = "FT_ELIGIBLE"
            elif paper_id_val == pid2:
                row[dec_col].value = "FT_SCREENED_OUT"
        wb.save(out)

        # Import
        result = import_ft_adjudication_decisions(tmp_db, out)
        assert result["stats"]["ft_eligible"] == 1
        assert result["stats"]["ft_screened_out"] == 1
        assert result["stats"]["missing"] == 0

        # Check statuses updated
        r1 = tmp_db._conn.execute(
            "SELECT status FROM papers WHERE id = ?", (pid1,)
        ).fetchone()
        assert r1["status"] == "FT_ELIGIBLE"

        r2 = tmp_db._conn.execute(
            "SELECT status FROM papers WHERE id = ?", (pid2,)
        ).fetchone()
        assert r2["status"] == "FT_SCREENED_OUT"

    def test_import_rejects_blank_decisions(self, tmp_db, tmp_path):
        """Blank decision cells cause full import rejection with zero DB changes."""
        pid = _add_paper(tmp_db, title="Missing Decision", pmid="77020")
        _advance_to_ft_flagged(tmp_db, pid)

        out = tmp_path / "ft_missing.xlsx"
        export_ft_adjudication_queue(tmp_db, out)

        # Don't fill in any decisions — import as-is
        result = import_ft_adjudication_decisions(tmp_db, out)
        assert result["stats"]["missing"] == 1
        assert result["stats"]["ft_eligible"] == 0

        # Verify paper status unchanged
        row = tmp_db._conn.execute(
            "SELECT status FROM papers WHERE id = ?", (pid,)
        ).fetchone()
        assert row["status"] == "FT_FLAGGED"

    def test_import_rejects_invalid_decision(self, tmp_db, tmp_path):
        """Invalid decision values cause full import rejection with zero DB changes."""
        pid = _add_paper(tmp_db, title="Invalid Decision", pmid="77021")
        _advance_to_ft_flagged(tmp_db, pid)

        out = tmp_path / "ft_invalid.xlsx"
        export_ft_adjudication_queue(tmp_db, out)

        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb["Review Queue"]
        headers = {cell.value: cell.column - 1 for cell in ws[1] if cell.value}
        dec_col = next(i for h, i in headers.items() if "PI_decision" in h)
        for row in ws.iter_rows(min_row=2, values_only=False):
            row[dec_col].value = "INCLUDE"  # wrong value
        wb.save(out)

        result = import_ft_adjudication_decisions(tmp_db, out)
        assert result["stats"]["invalid"] == 1

        # Verify paper status unchanged
        row = tmp_db._conn.execute(
            "SELECT status FROM papers WHERE id = ?", (pid,)
        ).fetchone()
        assert row["status"] == "FT_FLAGGED"

    def test_import_advances_workflow(self, tmp_db, tmp_path):
        pid = _add_paper(tmp_db, title="Workflow Test", pmid="77030")
        _advance_to_ft_flagged(tmp_db, pid)

        out = tmp_path / "ft_workflow.xlsx"
        export_ft_adjudication_queue(tmp_db, out)

        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb["Review Queue"]
        headers = {cell.value: cell.column - 1 for cell in ws[1] if cell.value}
        pid_col = next(i for h, i in headers.items() if "Paper ID" in h)
        dec_col = next(i for h, i in headers.items() if "PI_decision" in h)
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[pid_col].value == pid:
                row[dec_col].value = "FT_ELIGIBLE"
        wb.save(out)

        import_ft_adjudication_decisions(tmp_db, out)
        assert is_stage_done(tmp_db._conn, "FULL_TEXT_ADJUDICATION_COMPLETE")

    def test_ft_adjudication_gate(self, tmp_db):
        assert check_ft_adjudication_gate(tmp_db) == 0

        pid = _add_paper(tmp_db, title="Gate Test", pmid="77040")
        _advance_to_ft_flagged(tmp_db, pid)
        assert check_ft_adjudication_gate(tmp_db) == 1

    def test_ft_adjudication_records_in_table(self, tmp_db, tmp_path):
        pid = _add_paper(tmp_db, title="Record Test", pmid="77050")
        _advance_to_ft_flagged(tmp_db, pid)

        out = tmp_path / "ft_record.xlsx"
        export_ft_adjudication_queue(tmp_db, out)

        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb["Review Queue"]
        headers = {cell.value: cell.column - 1 for cell in ws[1] if cell.value}
        pid_col = next(i for h, i in headers.items() if "Paper ID" in h)
        dec_col = next(i for h, i in headers.items() if "PI_decision" in h)
        notes_col = next(i for h, i in headers.items() if "PI_notes" in h)
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[pid_col].value == pid:
                row[dec_col].value = "FT_ELIGIBLE"
                row[notes_col].value = "Reviewer confirmed"
        wb.save(out)

        import_ft_adjudication_decisions(tmp_db, out)

        row = tmp_db._conn.execute(
            "SELECT * FROM ft_screening_adjudication WHERE paper_id = ?", (pid,)
        ).fetchone()
        assert row is not None
        assert row["adjudication_decision"] == "FT_ELIGIBLE"
        assert row["adjudication_reason"] == "Reviewer confirmed"


# ── Workflow FT Stage Tests ──────────────────────────────────────


class TestWorkflowFTStages:

    def test_ft_screening_complete_stage_exists(self, tmp_db):
        from engine.adjudication.workflow import WORKFLOW_STAGES
        assert "FULL_TEXT_SCREENING_COMPLETE" in WORKFLOW_STAGES

    def test_ft_adjudication_complete_stage_exists(self, tmp_db):
        from engine.adjudication.workflow import WORKFLOW_STAGES
        assert "FULL_TEXT_ADJUDICATION_COMPLETE" in WORKFLOW_STAGES

    def test_ft_stages_after_acquisition(self, tmp_db):
        from engine.adjudication.workflow import WORKFLOW_STAGES
        acq_idx = WORKFLOW_STAGES.index("PDF_ACQUISITION")
        ft_screen_idx = WORKFLOW_STAGES.index("FULL_TEXT_SCREENING_COMPLETE")
        ft_adj_idx = WORKFLOW_STAGES.index("FULL_TEXT_ADJUDICATION_COMPLETE")
        assert ft_screen_idx > acq_idx
        assert ft_adj_idx > ft_screen_idx

    def test_ft_stages_before_extraction(self, tmp_db):
        from engine.adjudication.workflow import WORKFLOW_STAGES
        ft_adj_idx = WORKFLOW_STAGES.index("FULL_TEXT_ADJUDICATION_COMPLETE")
        ext_idx = WORKFLOW_STAGES.index("EXTRACTION_COMPLETE")
        assert ext_idx > ft_adj_idx

    def test_complete_ft_screening_stage(self, tmp_db):
        # Complete all prerequisite stages first
        for stage in [
            "ABSTRACT_SCREENING_COMPLETE",
            "ABSTRACT_DIAGNOSTIC_COMPLETE",
            "ABSTRACT_CATEGORIES_CONFIGURED",
            "ABSTRACT_QUEUE_EXPORTED",
            "ABSTRACT_ADJUDICATION_COMPLETE",
            "PDF_ACQUISITION",
        ]:
            complete_stage(tmp_db._conn, stage, metadata="test")

        complete_stage(tmp_db._conn, "FULL_TEXT_SCREENING_COMPLETE", metadata="test")
        assert is_stage_done(tmp_db._conn, "FULL_TEXT_SCREENING_COMPLETE")

    def test_ft_guidance_contains_ft_screener(self, tmp_db):
        from engine.adjudication.workflow import _NEXT_STEP_GUIDANCE
        guidance = _NEXT_STEP_GUIDANCE.get("FULL_TEXT_SCREENING_COMPLETE", "")
        assert "ft_screener" in guidance

    def test_ft_adjudication_guidance(self, tmp_db):
        from engine.adjudication.workflow import _NEXT_STEP_GUIDANCE
        guidance = _NEXT_STEP_GUIDANCE.get("FULL_TEXT_ADJUDICATION_COMPLETE", "")
        assert "ft_screening_adjudicator" in guidance


# ── Specialty Scope in Prompt Tests ──────────────────────────────


class TestSpecialtyScopeInPrompt:

    def test_specialty_scope_loaded(self, spec):
        assert spec.specialty_scope is not None
        assert "abdominal surgery" in spec.specialty_scope.included
        assert "dental surgery" in spec.specialty_scope.excluded

    def test_specialty_format_for_prompt(self, spec):
        formatted = spec.specialty_scope.format_for_prompt()
        assert "Included" in formatted or "included" in formatted
        assert "Excluded" in formatted or "excluded" in formatted
        assert "abdominal surgery" in formatted.lower()
        assert "dental surgery" in formatted.lower()

    def test_screening_prompt_includes_specialty_scope(self, spec):
        prompt = build_ft_screening_prompt("Paper text", spec)
        # Should contain specialty info from spec
        assert "dental" in prompt.lower() or "ophthalmic" in prompt.lower()

    def test_verification_prompt_includes_specialty_scope(self, spec):
        prompt = build_ft_verification_prompt("Paper text", spec)
        assert "dental" in prompt.lower() or "ophthalmic" in prompt.lower()
