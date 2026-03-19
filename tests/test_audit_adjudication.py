"""Tests for audit adjudication: export/import round-trip, spot-check, reject, min_status.

Tests cover the per-span export format with PI_decision (ACCEPT/REJECT/CORRECT)
and the hardened two-pass validation importer.
"""

import json
from pathlib import Path

import pytest

from engine.adjudication.audit_adjudicator import (
    _collect_papers_for_review,
    _flatten_to_span_rows,
    check_audit_review_gate,
    export_audit_review_queue,
    import_audit_review_decisions,
)
from engine.adjudication.workflow import (
    complete_stage,
    is_stage_done,
)
from engine.core.database import ReviewDatabase, _STATUS_ORDER
from engine.search.models import Citation


# ── Fixtures ──────────────────────────────────────────────────────────


def _make_citation(pmid: str, title: str = "Test Paper") -> Citation:
    return Citation(
        title=title, abstract="test abstract", pmid=pmid,
        source="pubmed", authors=["Author A"], journal="J Test", year=2024,
    )


def _add_paper_with_extraction(db, pmid, *, spans, status="AI_AUDIT_COMPLETE"):
    """Add a paper, extraction, and evidence spans. Returns paper_id."""
    db.add_papers([_make_citation(pmid, title=f"Paper {pmid}")])
    paper = db._conn.execute(
        "SELECT id FROM papers WHERE pmid = ?", (pmid,)
    ).fetchone()
    pid = paper["id"]

    # Walk through status transitions to reach AI_AUDIT_COMPLETE
    _transition_to(db, pid, status)

    # Insert extraction
    db._conn.execute(
        "INSERT INTO extractions (paper_id, extraction_schema_hash, extracted_data, model, extracted_at) "
        "VALUES (?, 'testhash', '{}', 'test', datetime('now'))",
        (pid,),
    )
    ext = db._conn.execute(
        "SELECT id FROM extractions WHERE paper_id = ? ORDER BY id DESC LIMIT 1",
        (pid,),
    ).fetchone()
    ext_id = ext["id"]

    # Insert spans
    for s in spans:
        db._conn.execute(
            """INSERT INTO evidence_spans
               (extraction_id, field_name, value, source_snippet, confidence,
                audit_status, auditor_model, audit_rationale, audited_at)
               VALUES (?, ?, ?, ?, ?, ?, 'test_model', ?, datetime('now'))""",
            (ext_id, s["field_name"], s["value"], s.get("snippet", "some text"),
             s.get("confidence", 0.8), s["audit_status"],
             s.get("rationale", "")),
        )
    db._conn.commit()
    return pid


def _transition_to(db, pid, target):
    """Walk paper through valid transitions to reach target status."""
    transitions = {
        "INGESTED": [],
        "ABSTRACT_SCREENED_IN": ["ABSTRACT_SCREENED_IN"],
        "PDF_ACQUIRED": ["ABSTRACT_SCREENED_IN", "PDF_ACQUIRED"],
        "PARSED": ["ABSTRACT_SCREENED_IN", "PDF_ACQUIRED", "PARSED"],
        "EXTRACTED": ["ABSTRACT_SCREENED_IN", "PDF_ACQUIRED", "PARSED", "EXTRACTED"],
        "AI_AUDIT_COMPLETE": ["ABSTRACT_SCREENED_IN", "PDF_ACQUIRED", "PARSED", "EXTRACTED", "AI_AUDIT_COMPLETE"],
        "HUMAN_AUDIT_COMPLETE": ["ABSTRACT_SCREENED_IN", "PDF_ACQUIRED", "PARSED", "EXTRACTED",
                                 "AI_AUDIT_COMPLETE", "HUMAN_AUDIT_COMPLETE"],
    }
    for step in transitions.get(target, []):
        try:
            db.update_status(pid, step)
        except ValueError:
            pass  # already at or past this status


def _complete_prereq_stages(db):
    """Complete all workflow stages up to AUDIT_QUEUE_EXPORTED prerequisite."""
    for stage in ("ABSTRACT_SCREENING_COMPLETE", "ABSTRACT_DIAGNOSTIC_COMPLETE",
                   "ABSTRACT_CATEGORIES_CONFIGURED", "ABSTRACT_QUEUE_EXPORTED",
                   "ABSTRACT_ADJUDICATION_COMPLETE",
                   "FULL_TEXT_SCREENING_COMPLETE", "FULL_TEXT_ADJUDICATION_COMPLETE",
                   "EXTRACTION_COMPLETE",
                   "AI_AUDIT_COMPLETE_STAGE"):
        complete_stage(db._conn, stage)


def _find_header_col(ws, name):
    """Find 0-indexed column by partial header name match."""
    for cell in ws[1]:
        if cell.value and name.lower() in str(cell.value).lower():
            return cell.column - 1
    raise ValueError(f"Column '{name}' not found in headers")


@pytest.fixture
def db(tmp_path):
    d = ReviewDatabase("test_audit_adj", data_root=tmp_path)
    yield d
    d.close()


# ── Collection Tests ──────────────────────────────────────────────────


def test_collect_flagged_paper(db):
    """Papers with flagged spans should be collected for review."""
    _add_paper_with_extraction(db, "10001", spans=[
        {"field_name": "study_design", "value": "RCT", "audit_status": "verified"},
        {"field_name": "sample_size", "value": "50", "audit_status": "flagged"},
    ])

    papers = _collect_papers_for_review(db, spot_check_pct=0)
    assert len(papers) == 1
    assert papers[0]["worst_state"] == "flagged"
    assert papers[0]["review_reason"] == "audit_issues"
    assert len(papers[0]["problem_spans"]) == 1


def test_collect_contested_paper(db):
    """Papers with contested spans should be collected."""
    _add_paper_with_extraction(db, "10002", spans=[
        {"field_name": "study_design", "value": "RCT", "audit_status": "verified"},
        {"field_name": "sample_size", "value": "50", "audit_status": "contested"},
    ])

    papers = _collect_papers_for_review(db, spot_check_pct=0)
    assert len(papers) == 1
    assert papers[0]["worst_state"] == "contested"


def test_collect_spot_check(db):
    """All-verified papers should be spot-checked at configured rate."""
    for i in range(10):
        _add_paper_with_extraction(db, str(20000 + i), spans=[
            {"field_name": "study_design", "value": "RCT", "audit_status": "verified"},
        ])

    papers = _collect_papers_for_review(db, spot_check_pct=0.50)
    spot = [p for p in papers if p["review_reason"] == "spot_check"]
    assert len(spot) == 5  # 50% of 10


def test_collect_minimum_one_spot_check(db):
    """Even at spot_check_pct=0, at least 1 all-verified paper is sampled."""
    for i in range(5):
        _add_paper_with_extraction(db, str(30000 + i), spans=[
            {"field_name": "study_design", "value": "RCT", "audit_status": "verified"},
        ])

    papers = _collect_papers_for_review(db, spot_check_pct=0)
    assert len(papers) == 1
    assert papers[0]["review_reason"] == "spot_check"


# ── Flatten Tests ─────────────────────────────────────────────────────


def test_flatten_exports_problem_spans_only(db):
    """For audit_issues papers, only problem spans are exported."""
    _add_paper_with_extraction(db, "11001", spans=[
        {"field_name": "study_design", "value": "RCT", "audit_status": "verified"},
        {"field_name": "sample_size", "value": "50", "audit_status": "flagged"},
    ])

    papers = _collect_papers_for_review(db, spot_check_pct=0)
    rows = _flatten_to_span_rows(papers)
    assert len(rows) == 1
    assert rows[0]["field_name"] == "sample_size"
    assert rows[0]["audit_state"] == "flagged"


def test_flatten_spot_check_exports_all_spans(db):
    """For spot-check papers, all spans are exported."""
    _add_paper_with_extraction(db, "11002", spans=[
        {"field_name": "study_design", "value": "RCT", "audit_status": "verified"},
        {"field_name": "sample_size", "value": "50", "audit_status": "verified"},
    ])

    papers = _collect_papers_for_review(db, spot_check_pct=1.0)
    rows = _flatten_to_span_rows(papers)
    assert len(rows) == 2


# ── Export Tests ──────────────────────────────────────────────────────


def test_export_creates_xlsx(db, tmp_path):
    """Export should create an Excel file with per-span rows and expected sheets."""
    _add_paper_with_extraction(db, "40001", spans=[
        {"field_name": "study_design", "value": "RCT", "audit_status": "flagged"},
        {"field_name": "sample_size", "value": "100", "audit_status": "verified"},
    ])

    _complete_prereq_stages(db)

    out = tmp_path / "audit_queue.xlsx"
    stats = export_audit_review_queue(db, out, spot_check_pct=0)

    assert out.exists()
    assert stats["total"] == 1
    assert stats["flagged"] == 1

    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert "Instructions" in wb.sheetnames
    assert "Review Queue" in wb.sheetnames
    assert "Audit Reference" in wb.sheetnames

    ws = wb["Review Queue"]
    headers = [cell.value for cell in ws[1]]
    assert "paper_id" in headers
    assert "Field Name" in headers
    assert "Extracted Value" in headers
    assert "Audit State" in headers
    # Decision column with valid values in header
    assert any("PI_decision" in str(h) for h in headers if h)
    # Free text columns
    assert any("corrected_value" in str(h) for h in headers if h)
    assert any("PI_notes" in str(h) for h in headers if h)

    # Only the flagged span should appear (not verified)
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    non_empty = [r for r in data_rows if r[0] is not None]
    assert len(non_empty) == 1


def test_export_sets_workflow_stage(db, tmp_path):
    """Export should auto-advance AUDIT_QUEUE_EXPORTED."""
    _add_paper_with_extraction(db, "40002", spans=[
        {"field_name": "study_design", "value": "RCT", "audit_status": "contested"},
    ])

    _complete_prereq_stages(db)

    out = tmp_path / "audit_queue.xlsx"
    export_audit_review_queue(db, out, spot_check_pct=0)

    assert is_stage_done(db._conn, "AUDIT_QUEUE_EXPORTED")


# ── Import / Round-Trip Tests ─────────────────────────────────────────


def test_import_accept_spans(db, tmp_path):
    """ACCEPT should mark contested/flagged spans as verified."""
    pid = _add_paper_with_extraction(db, "50001", spans=[
        {"field_name": "study_design", "value": "RCT", "audit_status": "flagged"},
        {"field_name": "sample_size", "value": "100", "audit_status": "contested"},
    ])

    _complete_prereq_stages(db)

    out = tmp_path / "queue.xlsx"
    export_audit_review_queue(db, out, spot_check_pct=0)

    # Fill in ACCEPT for all span rows
    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb["Review Queue"]
    dec_col = _find_header_col(ws, "PI_decision")
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value is not None:
            row[dec_col].value = "ACCEPT"
    wb.save(out)

    result = import_audit_review_decisions(db, out)
    assert result["stats"]["accepted"] == 2

    # Verify spans are now "verified"
    ext = db._conn.execute(
        "SELECT id FROM extractions WHERE paper_id = ? ORDER BY id DESC LIMIT 1",
        (pid,),
    ).fetchone()
    spans = db._conn.execute(
        "SELECT audit_status FROM evidence_spans WHERE extraction_id = ?",
        (ext["id"],),
    ).fetchall()
    assert all(s["audit_status"] == "verified" for s in spans)


def test_import_correct_records_original(db, tmp_path):
    """CORRECT should update span value and record original in audit_adjudication."""
    pid = _add_paper_with_extraction(db, "50002", spans=[
        {"field_name": "study_design", "value": "RCT", "audit_status": "flagged"},
    ])

    _complete_prereq_stages(db)

    out = tmp_path / "queue.xlsx"
    export_audit_review_queue(db, out, spot_check_pct=0)

    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb["Review Queue"]
    dec_col = _find_header_col(ws, "PI_decision")
    corrected_col = _find_header_col(ws, "corrected_value")
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value is not None:
            row[dec_col].value = "CORRECT"
            row[corrected_col].value = "Prospective cohort"
    wb.save(out)

    result = import_audit_review_decisions(db, out)
    assert result["stats"]["corrected_fields"] == 1

    # Check span value was updated
    ext = db._conn.execute(
        "SELECT id FROM extractions WHERE paper_id = ? ORDER BY id DESC LIMIT 1",
        (pid,),
    ).fetchone()
    span = db._conn.execute(
        "SELECT value FROM evidence_spans WHERE extraction_id = ? AND field_name = 'study_design'",
        (ext["id"],),
    ).fetchone()
    assert span["value"] == "Prospective cohort"

    # Check audit_adjudication table recorded the original
    adj = db._conn.execute(
        "SELECT * FROM audit_adjudication WHERE paper_id = ? AND field_name = 'study_design'",
        (pid,),
    ).fetchone()
    assert adj["original_value"] == "RCT"
    assert adj["override_value"] == "Prospective cohort"
    assert adj["human_decision"] == "override"


def test_import_transitions_to_human_audit_complete(db, tmp_path):
    """Paper should transition to HUMAN_AUDIT_COMPLETE when all spans resolved."""
    pid = _add_paper_with_extraction(db, "50003", spans=[
        {"field_name": "study_design", "value": "RCT", "audit_status": "flagged"},
    ])

    _complete_prereq_stages(db)

    out = tmp_path / "queue.xlsx"
    export_audit_review_queue(db, out, spot_check_pct=0)

    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb["Review Queue"]
    dec_col = _find_header_col(ws, "PI_decision")
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value is not None:
            row[dec_col].value = "ACCEPT"
    wb.save(out)

    import_audit_review_decisions(db, out)

    paper = db._conn.execute("SELECT status FROM papers WHERE id = ?", (pid,)).fetchone()
    assert paper["status"] == "HUMAN_AUDIT_COMPLETE"


def test_import_sets_audit_review_complete(db, tmp_path):
    """When all papers resolved, AUDIT_REVIEW_COMPLETE should be set."""
    _add_paper_with_extraction(db, "50004", spans=[
        {"field_name": "study_design", "value": "RCT", "audit_status": "contested"},
    ])

    _complete_prereq_stages(db)

    out = tmp_path / "queue.xlsx"
    export_audit_review_queue(db, out, spot_check_pct=0)

    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb["Review Queue"]
    dec_col = _find_header_col(ws, "PI_decision")
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value is not None:
            row[dec_col].value = "ACCEPT"
    wb.save(out)

    import_audit_review_decisions(db, out)

    assert is_stage_done(db._conn, "AUDIT_REVIEW_COMPLETE")


def test_import_rejects_blank_decisions(db, tmp_path):
    """Blank decision cells cause full import rejection with zero DB changes."""
    pid = _add_paper_with_extraction(db, "50005", spans=[
        {"field_name": "study_design", "value": "RCT", "audit_status": "flagged"},
    ])

    _complete_prereq_stages(db)

    out = tmp_path / "queue.xlsx"
    export_audit_review_queue(db, out, spot_check_pct=0)

    # Don't fill in any decisions
    result = import_audit_review_decisions(db, out)

    assert result["stats"]["missing"] >= 1
    assert not is_stage_done(db._conn, "AUDIT_REVIEW_COMPLETE")

    # Paper status unchanged
    paper = db._conn.execute("SELECT status FROM papers WHERE id = ?", (pid,)).fetchone()
    assert paper["status"] == "AI_AUDIT_COMPLETE"


def test_import_rejects_invalid_decision(db, tmp_path):
    """Invalid decision values cause full import rejection with zero DB changes."""
    pid = _add_paper_with_extraction(db, "50006", spans=[
        {"field_name": "study_design", "value": "RCT", "audit_status": "flagged"},
    ])

    _complete_prereq_stages(db)

    out = tmp_path / "queue.xlsx"
    export_audit_review_queue(db, out, spot_check_pct=0)

    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb["Review Queue"]
    dec_col = _find_header_col(ws, "PI_decision")
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value is not None:
            row[dec_col].value = "MAYBE"  # invalid
    wb.save(out)

    result = import_audit_review_decisions(db, out)
    assert result["stats"]["invalid"] == 1

    # Paper status unchanged
    paper = db._conn.execute("SELECT status FROM papers WHERE id = ?", (pid,)).fetchone()
    assert paper["status"] == "AI_AUDIT_COMPLETE"


def test_import_rejects_correct_without_value(db, tmp_path):
    """CORRECT without corrected_value causes full import rejection."""
    pid = _add_paper_with_extraction(db, "50007", spans=[
        {"field_name": "study_design", "value": "RCT", "audit_status": "flagged"},
    ])

    _complete_prereq_stages(db)

    out = tmp_path / "queue.xlsx"
    export_audit_review_queue(db, out, spot_check_pct=0)

    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb["Review Queue"]
    dec_col = _find_header_col(ws, "PI_decision")
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value is not None:
            row[dec_col].value = "CORRECT"
            # Don't fill corrected_value
    wb.save(out)

    result = import_audit_review_decisions(db, out)
    assert result["stats"]["missing"] >= 1  # CORRECT-without-value counted as missing

    # Paper status unchanged
    paper = db._conn.execute("SELECT status FROM papers WHERE id = ?", (pid,)).fetchone()
    assert paper["status"] == "AI_AUDIT_COMPLETE"


# ── Reject Span ──────────────────────────────────────────────────────


def test_reject_span(db, tmp_path):
    """REJECT should mark span as rejected and transition paper to HUMAN_AUDIT_COMPLETE."""
    pid = _add_paper_with_extraction(db, "60001", spans=[
        {"field_name": "study_design", "value": "RCT", "audit_status": "flagged"},
    ])

    _complete_prereq_stages(db)

    out = tmp_path / "queue.xlsx"
    export_audit_review_queue(db, out, spot_check_pct=0)

    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb["Review Queue"]
    dec_col = _find_header_col(ws, "PI_decision")
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value is not None:
            row[dec_col].value = "REJECT"
    wb.save(out)

    result = import_audit_review_decisions(db, out)
    assert result["stats"]["rejected"] == 1

    # Paper transitions to HUMAN_AUDIT_COMPLETE (all spans resolved)
    paper = db._conn.execute("SELECT status FROM papers WHERE id = ?", (pid,)).fetchone()
    assert paper["status"] == "HUMAN_AUDIT_COMPLETE"

    # Audit adjudication record created
    adj = db._conn.execute(
        "SELECT * FROM audit_adjudication WHERE paper_id = ?", (pid,)
    ).fetchone()
    assert adj is not None
    assert adj["human_decision"] == "reject_paper"


# ── min_status Filtering ──────────────────────────────────────────────


def test_min_status_filtering(db):
    """_STATUS_ORDER filtering: AI_AUDIT includes both, HUMAN_AUDIT only human-verified."""
    from engine.core.database import _STATUS_ORDER

    # Add an AI_AUDIT_COMPLETE paper
    _add_paper_with_extraction(db, "70001", spans=[
        {"field_name": "study_design", "value": "RCT", "audit_status": "verified"},
    ], status="AI_AUDIT_COMPLETE")

    # Add a HUMAN_AUDIT_COMPLETE paper
    _add_paper_with_extraction(db, "70002", spans=[
        {"field_name": "study_design", "value": "cohort", "audit_status": "verified"},
    ], status="HUMAN_AUDIT_COMPLETE")

    # AI_AUDIT_COMPLETE level should include both
    min_level = _STATUS_ORDER["AI_AUDIT_COMPLETE"]
    qualifying = [s for s, lvl in _STATUS_ORDER.items() if lvl >= min_level]
    placeholders = ", ".join("?" for _ in qualifying)
    rows = db._conn.execute(
        f"SELECT id FROM papers WHERE status IN ({placeholders})",
        qualifying,
    ).fetchall()
    assert len(rows) == 2

    # HUMAN_AUDIT_COMPLETE level should include only the human-verified paper
    min_level = _STATUS_ORDER["HUMAN_AUDIT_COMPLETE"]
    qualifying = [s for s, lvl in _STATUS_ORDER.items() if lvl >= min_level]
    placeholders = ", ".join("?" for _ in qualifying)
    rows = db._conn.execute(
        f"SELECT id FROM papers WHERE status IN ({placeholders})",
        qualifying,
    ).fetchall()
    assert len(rows) == 1


# ── Gate Check ────────────────────────────────────────────────────────


def test_check_audit_review_gate(db):
    """Gate should count papers with unresolved spans."""
    _add_paper_with_extraction(db, "80001", spans=[
        {"field_name": "study_design", "value": "RCT", "audit_status": "flagged"},
    ])
    _add_paper_with_extraction(db, "80002", spans=[
        {"field_name": "study_design", "value": "RCT", "audit_status": "verified"},
    ])

    count = check_audit_review_gate(db)
    assert count == 1


def test_check_audit_review_gate_zero_when_clean(db):
    """Gate should return 0 when no papers have issues."""
    _add_paper_with_extraction(db, "80003", spans=[
        {"field_name": "study_design", "value": "RCT", "audit_status": "verified"},
    ])

    count = check_audit_review_gate(db)
    assert count == 0


# ── H9: Missing span stats accuracy ─────────────────────────────────


def test_missing_span_not_counted_as_success(db, tmp_path):
    """Adjudication for a nonexistent span is rejected — not counted as success."""
    pid = _add_paper_with_extraction(db, "90001", spans=[
        {"field_name": "study_design", "value": "RCT", "audit_status": "contested"},
    ])

    # Create a decision referencing a span that doesn't exist
    decisions = [{
        "span_id": 999999,  # nonexistent
        "paper_id": pid,
        "field_name": "nonexistent_field",
        "decision": "ACCEPT",
    }]

    json_path = tmp_path / "missing_span_decisions.json"
    json_path.write_text(json.dumps(decisions))

    # JSON import path validates first: nonexistent span → validation error → no writes
    result = import_audit_review_decisions(db, str(json_path))

    # Errors reported, nothing applied
    assert len(result["errors"]) >= 1
    assert result["applied"] == 0

    # Original span unchanged
    span = db._conn.execute(
        "SELECT audit_status FROM evidence_spans WHERE field_name = 'study_design'"
    ).fetchone()
    assert span["audit_status"] == "contested"
