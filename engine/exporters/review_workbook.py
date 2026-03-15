"""Shared xlsx workbook generator for reference exports and backward compatibility.

For interactive human review, use the HTML/JSON round-trip tools instead:
  - Abstract adjudication: engine/adjudication/abstract_adjudication_html.py
  - FT adjudication: engine/adjudication/ft_adjudication_html.py
  - Extraction audit: engine/review/extraction_audit_html.py

xlsx generation is retained for archival/reference exports only.

Produces multi-sheet xlsx workbooks with:
  - Instructions sheet (opens by default) with review context, decision criteria,
    edge case guidance, and complete import procedure
  - Review Queue sheet with data validation dropdowns, conditional formatting
    for blank decision cells, frozen header, auto-filter
  - Optional Reference sheet with verbatim review spec criteria

Usage:
    from engine.exporters.review_workbook import create_review_workbook

    create_review_workbook(
        output_path="queue.xlsx",
        rows=[{"paper_id": 1, "title": "...", ...}],
        columns=[
            {"key": "paper_id", "header": "Paper ID", "width": 10},
            {"key": "title", "header": "Title", "width": 60},
            ...
        ],
        decision_columns=[
            {"key": "PI_decision", "header": "PI_decision (INCLUDE/EXCLUDE)",
             "valid_values": ["INCLUDE", "EXCLUDE"], "width": 25},
        ],
        instructions=InstructionsConfig(...),
    )
"""

import logging
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger(__name__)


@dataclass
class ColumnDef:
    """Definition for a data column in the Review Queue sheet."""
    key: str
    header: str
    width: int = 20
    wrap: bool = False
    truncate: int | None = None  # max chars to write (None = no truncation)
    fill: PatternFill | None = None  # optional background fill


@dataclass
class DecisionColumnDef:
    """Definition for a decision column with dropdown validation."""
    key: str
    header: str  # should include valid values reminder, e.g. "PI_decision (INCLUDE/EXCLUDE)"
    valid_values: list[str]
    width: int = 25


@dataclass
class FreeTextColumnDef:
    """Definition for an editable free-text column (no validation)."""
    key: str
    header: str
    width: int = 30


@dataclass
class InstructionsConfig:
    """All content for the Instructions sheet."""
    review_name: str
    review_spec_id: str = ""
    db_path: str = ""
    export_trigger: str = ""  # e.g. "86 papers flagged during specialty re-screen"
    row_count: int = 0
    decision_column_name: str = ""
    valid_values: list[str] = field(default_factory=list)
    decision_criteria: list[str] = field(default_factory=list)
    edge_case_guidance: str = ""
    import_command: str = ""
    columns_importer_reads: list[str] = field(default_factory=list)
    columns_importer_ignores: str = "All other columns are read-only context and are ignored by the importer."
    notes_on_rejection: str = (
        "If any decision cell is blank, the entire import is rejected with a list of blank rows. "
        "If any decision value is not in the valid set, the entire import is rejected with a list "
        "of invalid rows and values. Fix the workbook and re-run the import command."
    )
    pre_import_validation: str = (
        "The importer validates every row before making any database changes. "
        "If validation fails, no changes are made."
    )


def create_review_workbook(
    output_path: str | Path,
    rows: list[dict],
    columns: list[ColumnDef],
    decision_columns: list[DecisionColumnDef],
    instructions: InstructionsConfig,
    free_text_columns: list[FreeTextColumnDef] | None = None,
    reference_content: str | None = None,
    reference_sheet_title: str = "Reference",
) -> Path:
    """Create a self-documenting review workbook.

    Returns the output path.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ── Sheet 1: Instructions (default sheet, opens first) ─────────
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    _build_instructions_sheet(ws_instr, instructions)

    # ── Sheet 2: Review Queue ──────────────────────────────────────
    ws_queue = wb.create_sheet("Review Queue")
    _build_review_queue_sheet(
        ws_queue, rows, columns, decision_columns,
        free_text_columns=free_text_columns or [],
    )

    # ── Sheet 3: Reference (optional) ──────────────────────────────
    if reference_content:
        ws_ref = wb.create_sheet(reference_sheet_title)
        _build_reference_sheet(ws_ref, reference_content)

    wb.save(output_path)

    # Terminal summary
    all_decision_cols = ", ".join(dc.header for dc in decision_columns)
    all_valid_vals = ", ".join(
        f"{dc.key}: [{', '.join(dc.valid_values)}]" for dc in decision_columns
    )
    print(f"\n{'='*60}")
    print("REVIEW WORKBOOK EXPORTED")
    print(f"{'='*60}")
    print(f"  File:            {output_path}")
    print(f"  Rows:            {len(rows)}")
    print(f"  Decision column: {all_decision_cols}")
    print(f"  Valid values:    {all_valid_vals}")
    if instructions.import_command:
        print(f"  Import command:  {instructions.import_command}")
    print(f"{'='*60}\n")

    return output_path


# ── Instructions Sheet ─────────────────────────────────────────────


def _build_instructions_sheet(ws, config: InstructionsConfig) -> None:
    """Build structured-form instructions sheet."""
    bold = Font(bold=True, size=11)
    heading = Font(bold=True, size=14, color="0A5E56")
    subheading = Font(bold=True, size=12, color="0A5E56")
    normal = Font(size=11)
    wrap = Alignment(wrap_text=True, vertical="top")
    label_fill = PatternFill(start_color="EEF5F4", end_color="EEF5F4", fill_type="solid")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90

    row = 1

    def add_heading(text):
        nonlocal row
        if row > 1:
            row += 1  # blank spacer
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = heading
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1

    def add_subheading(text):
        nonlocal row
        row += 1  # blank spacer
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = subheading
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1

    def add_field(label, value):
        nonlocal row
        lbl = ws.cell(row=row, column=1, value=label)
        lbl.font = bold
        lbl.fill = label_fill
        lbl.alignment = Alignment(vertical="top")
        val = ws.cell(row=row, column=2, value=value)
        val.font = normal
        val.alignment = wrap
        row += 1

    export_date = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    add_heading("REVIEW WORKBOOK — INSTRUCTIONS")

    add_subheading("Export Context")
    add_field("Review name", config.review_name)
    if config.review_spec_id:
        add_field("Review spec", config.review_spec_id)
    add_field("Export date", export_date)
    if config.db_path:
        add_field("Database", config.db_path)
    if config.export_trigger:
        add_field("What triggered this export", config.export_trigger)
    add_field("Total rows in Review Queue", str(config.row_count))

    add_subheading("Decision Column")
    add_field("Column name", config.decision_column_name)
    add_field("Valid values (dropdown)", ", ".join(config.valid_values))

    if config.decision_criteria:
        add_subheading("Decision Criteria")
        for criterion in config.decision_criteria:
            ws.cell(row=row, column=1, value="•")
            c = ws.cell(row=row, column=2, value=criterion)
            c.font = normal
            c.alignment = wrap
            row += 1

    if config.edge_case_guidance:
        add_subheading("Edge Cases")
        c = ws.cell(row=row, column=2, value=config.edge_case_guidance)
        c.font = normal
        c.alignment = wrap
        row += 1

    add_subheading("Import Procedure")
    if config.import_command:
        add_field("Import command", config.import_command)
    if config.columns_importer_reads:
        add_field("Columns read by importer", ", ".join(config.columns_importer_reads))
    if config.columns_importer_ignores:
        add_field("Columns ignored", config.columns_importer_ignores)
    add_field("Blank decision cells", config.notes_on_rejection)
    add_field("Pre-import validation", config.pre_import_validation)


# ── Review Queue Sheet ─────────────────────────────────────────────


def _build_review_queue_sheet(
    ws,
    rows: list[dict],
    columns: list[ColumnDef],
    decision_columns: list[DecisionColumnDef],
    free_text_columns: list[FreeTextColumnDef] | None = None,
) -> None:
    """Build the working data sheet with dropdowns and conditional formatting."""
    free_text_columns = free_text_columns or []
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0A5E56", end_color="0A5E56", fill_type="solid")
    header_align = Alignment(horizontal="center", wrap_text=True)
    wrap_align = Alignment(wrap_text=True, vertical="top")

    all_cols = (
        columns
        + [ColumnDef(key=dc.key, header=dc.header, width=dc.width) for dc in decision_columns]
        + [ColumnDef(key=ft.key, header=ft.header, width=ft.width) for ft in free_text_columns]
    )
    total_cols = len(all_cols)

    # Write headers
    for col_idx, col_def in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_def.header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Column widths
    for col_idx, col_def in enumerate(all_cols, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_def.width

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_def in enumerate(columns, 1):
            value = row_data.get(col_def.key, "")
            if value is None:
                value = ""
            if col_def.truncate and isinstance(value, str) and len(value) > col_def.truncate:
                value = value[:col_def.truncate]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_def.wrap:
                cell.alignment = wrap_align
            if col_def.fill:
                cell.fill = col_def.fill

        # Decision columns — leave blank for human input
        for dc_idx, dc in enumerate(decision_columns):
            col_num = len(columns) + dc_idx + 1
            ws.cell(row=row_idx, column=col_num, value="")

    # Data validation dropdowns on decision columns
    last_data_row = len(rows) + 1
    for dc_idx, dc in enumerate(decision_columns):
        col_num = len(columns) + dc_idx + 1
        col_letter = get_column_letter(col_num)

        formula_list = ",".join(dc.valid_values)
        dv = DataValidation(
            type="list",
            formula1=f'"{formula_list}"',
            allow_blank=True,
            showDropDown=False,  # openpyxl: False = show dropdown arrow
        )
        dv.error = f"Value must be one of: {formula_list}"
        dv.errorTitle = "Invalid Decision"
        dv.prompt = f"Select: {formula_list}"
        dv.promptTitle = dc.key
        ws.add_data_validation(dv)
        if last_data_row >= 2:
            dv.add(f"{col_letter}2:{col_letter}{last_data_row}")

    # Conditional formatting: red fill for blank decision cells
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for dc_idx, dc in enumerate(decision_columns):
        col_num = len(columns) + dc_idx + 1
        col_letter = get_column_letter(col_num)
        cell_range = f"{col_letter}2:{col_letter}{last_data_row}"
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="equal",
                formula=['""'],
                fill=red_fill,
            ),
        )

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter on all columns
    if total_cols > 0 and len(rows) > 0:
        last_col_letter = get_column_letter(total_cols)
        ws.auto_filter.ref = f"A1:{last_col_letter}{last_data_row}"


# ── Reference Sheet ────────────────────────────────────────────────


def _build_reference_sheet(ws, content: str) -> None:
    """Build reference sheet with verbatim review spec criteria."""
    heading = Font(bold=True, size=12, color="0A5E56")
    normal = Font(size=11)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 120

    ws.cell(row=1, column=1, value="Review Specification — Reference Criteria").font = heading

    for i, line in enumerate(content.split("\n"), 2):
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = normal
        cell.alignment = wrap
