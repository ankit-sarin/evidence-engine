"""Export 3-arm AI disagreement pairs for PLUM Lab review.

Produces CSV, XLSX (4-sheet workbook), and HTML exports of all paper × field
combinations where at least one arm-pair disagrees (MISMATCH or AMBIGUOUS).

Usage:
    PYTHONPATH=. python analysis/paper1/export_disagreement_pairs.py
    PYTHONPATH=. python analysis/paper1/export_disagreement_pairs.py --review surgical_autonomy
"""

import argparse
import csv
import html as html_mod
import json
import math
import sqlite3
from collections import defaultdict
from datetime import datetime, timezone
from itertools import combinations
from pathlib import Path

from engine.analysis.concordance import load_arm
from engine.analysis.metrics import FieldSummary, field_summary
from engine.analysis.scoring import FieldScore, score_pair
from engine.core.database import DATA_ROOT
from engine.core.review_spec import load_review_spec

# ── Constants ────────────────────────────────────────────────────────

ARMS = ["local", "openai_o4_mini_high", "anthropic_sonnet_4_6"]
ARM_LABELS = {
    "local": "Local (DeepSeek-R1:32b)",
    "openai_o4_mini_high": "o4-mini",
    "anthropic_sonnet_4_6": "Sonnet 4.6",
}
ARM_PAIRS = list(combinations(ARMS, 2))

FREE_TEXT_FIELDS = {
    "robot_platform", "task_performed", "primary_outcome_metric",
    "primary_outcome_value", "comparison_to_human", "secondary_outcomes",
    "key_limitation",
}


def _field_type(field_name: str, spec) -> str:
    """Return 'categorical', 'free_text', or 'numeric'."""
    for f in spec.extraction_schema.fields:
        if f.name == field_name:
            if f.type == "categorical":
                return "categorical"
            if field_name in ("sample_size",):
                return "numeric"
            return "free_text"
    return "free_text"


def _field_tier(field_name: str, spec) -> int:
    for f in spec.extraction_schema.fields:
        if f.name == field_name:
            return f.tier
    return 0


def _load_paper_info(db_path: str) -> dict[int, dict]:
    """Load paper_id → {title, first_author, year} from papers table."""
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("SELECT id, title, authors, year FROM papers").fetchall()
    conn.close()
    info = {}
    for r in rows:
        authors_raw = r["authors"] or ""
        first_author = ""
        try:
            authors_list = json.loads(authors_raw)
            if authors_list:
                first_author = authors_list[0].split()[-1]
        except (json.JSONDecodeError, TypeError):
            parts = authors_raw.split(";")
            if parts and parts[0].strip():
                first_author = parts[0].strip().split()[-1]
        label = f"{first_author} {r['year']}" if first_author else str(r["id"])
        info[r["id"]] = {
            "title": r["title"] or "",
            "first_author": first_author,
            "year": r["year"],
            "label": label,
        }
    return info


def _pair_key(arm_a: str, arm_b: str) -> str:
    """Readable column prefix for an arm pair."""
    short = {
        "local": "local",
        "openai_o4_mini_high": "o4mini",
        "anthropic_sonnet_4_6": "sonnet",
    }
    return f"{short[arm_a]}_vs_{short[arm_b]}"


# ── Core: build 3-arm disagreement rows ─────────────────────────────


def build_disagreement_rows(
    db_path: str, spec_path: str
) -> tuple[list[dict], dict[str, dict[str, FieldSummary]]]:
    """Build flat disagreement rows across all 3 arms.

    Returns:
        (rows, summaries_by_pair) where each row is a dict with all columns
        and summaries_by_pair maps pair_key → {field_name: FieldSummary}.
    """
    spec = load_review_spec(spec_path)

    # Load arms
    arm_data = {arm: load_arm(db_path, arm) for arm in ARMS}

    # Papers present in all 3 arms
    shared_ids = set(arm_data[ARMS[0]].keys())
    for arm in ARMS[1:]:
        shared_ids &= set(arm_data[arm].keys())
    shared_ids = sorted(shared_ids)
    print(f"Papers shared across all 3 arms: {len(shared_ids)}")

    # All field names across all arms/papers
    all_fields: set[str] = set()
    for arm in ARMS:
        for pid in shared_ids:
            all_fields.update(arm_data[arm].get(pid, {}).keys())
    # Sort: free-text first, then by tier + alpha
    sorted_fields = sorted(
        all_fields,
        key=lambda f: (0 if f in FREE_TEXT_FIELDS else 1, _field_tier(f, spec), f),
    )

    paper_info = _load_paper_info(db_path)

    # Score all pairs for metrics
    scores_by_pair_field: dict[str, dict[str, list[FieldScore]]] = {
        _pair_key(a, b): defaultdict(list) for a, b in ARM_PAIRS
    }

    rows = []
    for pid in shared_ids:
        for fname in sorted_fields:
            values = {arm: arm_data[arm].get(pid, {}).get(fname) for arm in ARMS}

            # Score all 3 pairs
            pair_scores = {}
            any_disagree = False
            for arm_a, arm_b in ARM_PAIRS:
                fs = score_pair(fname, values[arm_a], values[arm_b], spec)
                pk = _pair_key(arm_a, arm_b)
                pair_scores[pk] = fs
                scores_by_pair_field[pk][fname].append(fs)
                if fs.result != "MATCH":
                    any_disagree = True

            if not any_disagree:
                continue

            pinfo = paper_info.get(pid, {"title": "", "label": str(pid)})
            ft = _field_type(fname, spec)
            tier = _field_tier(fname, spec)

            row = {
                "paper_id": pid,
                "paper_label": pinfo["label"],
                "paper_title": pinfo["title"],
                "field_name": fname,
                "field_tier": tier,
                "field_type": ft,
                "local_value": values["local"],
                "o4mini_value": values["openai_o4_mini_high"],
                "sonnet_value": values["anthropic_sonnet_4_6"],
            }
            for pk, fs in pair_scores.items():
                row[f"{pk}_score"] = fs.result
            rows.append(row)

    # Compute summaries
    summaries_by_pair: dict[str, dict[str, FieldSummary]] = {}
    for pk, field_scores in scores_by_pair_field.items():
        summaries_by_pair[pk] = {
            fname: field_summary(fname, scores)
            for fname, scores in sorted(field_scores.items())
        }

    return rows, summaries_by_pair


# ── CSV export ───────────────────────────────────────────────────────


def write_csv(rows: list[dict], path: Path) -> None:
    headers = [
        "paper_id", "paper_label", "paper_title", "field_name", "field_tier",
        "field_type", "local_value", "o4mini_value", "sonnet_value",
        "local_vs_o4mini_score", "local_vs_sonnet_score", "o4mini_vs_sonnet_score",
    ]
    tmp = path.with_suffix(".tmp")
    with open(tmp, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    tmp.replace(path)
    print(f"CSV: {path} ({len(rows)} rows)")


# ── XLSX export ──────────────────────────────────────────────────────


def write_xlsx(
    rows: list[dict],
    summaries_by_pair: dict[str, dict[str, FieldSummary]],
    spec_path: str,
    path: Path,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    spec = load_review_spec(spec_path)

    wb = Workbook()

    # Header style
    hdr_fill = PatternFill(start_color="0A5E56", end_color="0A5E56", fill_type="solid")
    hdr_font = Font(name="IBM Plex Sans", bold=True, color="FFFFFF", size=10)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    top_align = Alignment(vertical="top")

    headers = [
        ("paper_id", 10), ("paper_label", 18), ("paper_title", 45),
        ("field_name", 28), ("field_tier", 8), ("field_type", 12),
        ("local_value", 45), ("o4mini_value", 45), ("sonnet_value", 45),
        ("local_vs_o4mini_score", 18), ("local_vs_sonnet_score", 18),
        ("o4mini_vs_sonnet_score", 18),
    ]

    def _write_sheet(ws, sheet_rows):
        # Headers
        for col_idx, (hdr, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=hdr)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Data
        for row_idx, row in enumerate(sheet_rows, 2):
            for col_idx, (key, _) in enumerate(headers, 1):
                val = row.get(key, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val or "")
                if key in ("local_value", "o4mini_value", "sonnet_value",
                           "paper_title"):
                    cell.alignment = wrap_align
                else:
                    cell.alignment = top_align

        # Freeze header
        ws.freeze_panes = "A2"

    # Sheet 1: Free Text
    ws_ft = wb.active
    ws_ft.title = "Free Text"
    ft_rows = [r for r in rows if r["field_type"] == "free_text"]
    _write_sheet(ws_ft, ft_rows)

    # Sheet 2: Categorical
    ws_cat = wb.create_sheet("Categorical")
    cat_rows = [r for r in rows if r["field_type"] != "free_text"]
    _write_sheet(ws_cat, cat_rows)

    # Sheet 3: Summary
    ws_sum = wb.create_sheet("Summary")
    sum_headers = [
        ("field_name", 28), ("arm_pair", 22), ("field_type", 12), ("field_tier", 8),
        ("n", 6), ("n_match", 8), ("n_mismatch", 10), ("n_ambiguous", 10),
        ("pct_agreement", 12), ("kappa", 8), ("ci_lower", 8), ("ci_upper", 8),
    ]
    for col_idx, (hdr, width) in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=col_idx, value=hdr)
        cell.fill = hdr_fill
        cell.font = hdr_font
        ws_sum.column_dimensions[get_column_letter(col_idx)].width = width

    sum_row = 2
    for fname in sorted(
        {f for sums in summaries_by_pair.values() for f in sums},
        key=lambda f: (0 if f in FREE_TEXT_FIELDS else 1, _field_tier(f, spec), f),
    ):
        for pk, sums in summaries_by_pair.items():
            fs = sums.get(fname)
            if not fs:
                continue
            ft = _field_type(fname, spec)
            tier = _field_tier(fname, spec)
            is_nan = lambda v: isinstance(v, float) and math.isnan(v)
            ws_sum.cell(row=sum_row, column=1, value=fname)
            ws_sum.cell(row=sum_row, column=2, value=pk)
            ws_sum.cell(row=sum_row, column=3, value=ft)
            ws_sum.cell(row=sum_row, column=4, value=tier)
            ws_sum.cell(row=sum_row, column=5, value=fs.n)
            ws_sum.cell(row=sum_row, column=6, value=fs.n_match)
            ws_sum.cell(row=sum_row, column=7, value=fs.n_mismatch)
            ws_sum.cell(row=sum_row, column=8, value=fs.n_ambiguous)
            ws_sum.cell(row=sum_row, column=9,
                        value=round(fs.percent_agreement, 4) if not is_nan(fs.percent_agreement) else "")
            ws_sum.cell(row=sum_row, column=10,
                        value=round(fs.kappa, 4) if not is_nan(fs.kappa) else "")
            ws_sum.cell(row=sum_row, column=11,
                        value=round(fs.ci_lower, 4) if not is_nan(fs.ci_lower) else "")
            ws_sum.cell(row=sum_row, column=12,
                        value=round(fs.ci_upper, 4) if not is_nan(fs.ci_upper) else "")
            sum_row += 1

    ws_sum.freeze_panes = "A2"

    # Sheet 4: All
    ws_all = wb.create_sheet("All")
    _write_sheet(ws_all, rows)

    tmp = path.with_suffix(".tmp.xlsx")
    wb.save(str(tmp))
    tmp.replace(path)
    print(f"XLSX: {path} (Free Text: {len(ft_rows)}, Categorical: {len(cat_rows)}, "
          f"All: {len(rows)}, Summary: {sum_row - 2} rows)")


# ── HTML export ──────────────────────────────────────────────────────


def _esc(text: str | None) -> str:
    return html_mod.escape(str(text)) if text else "&mdash;"


def _score_color(score: str) -> str:
    if score == "MATCH":
        return "#d4edda"
    if score == "AMBIGUOUS":
        return "#fff3cd"
    return "#f8d7da"


def _is_nan(v: float) -> bool:
    return isinstance(v, float) and math.isnan(v)


def write_html(
    rows: list[dict],
    summaries_by_pair: dict[str, dict[str, FieldSummary]],
    spec_path: str,
    path: Path,
) -> None:
    spec = load_review_spec(spec_path)
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    # Compute summary stats
    total = len(rows)
    by_field: dict[str, int] = defaultdict(int)
    by_type: dict[str, int] = defaultdict(int)
    by_tier: dict[int, int] = defaultdict(int)
    for r in rows:
        by_field[r["field_name"]] += 1
        by_type[r["field_type"]] += 1
        by_tier[r["field_tier"]] += 1

    all_fields_sorted = sorted(
        by_field.keys(),
        key=lambda f: (0 if f in FREE_TEXT_FIELDS else 1, _field_tier(f, spec), f),
    )
    unique_fields = list(by_field.keys())

    parts = [f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>3-Arm Disagreement Pairs — PLUM Lab Export</title>
<style>
  :root {{
    --bg: #EEF5F4; --text: #2C2C2C; --teal: #0A5E56; --terra: #B85D3A;
    --border: #c8d8d6; --card-bg: #fff;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'IBM Plex Sans', system-ui, sans-serif; background: var(--bg);
         color: var(--text); padding: 24px; max-width: 1600px; margin: 0 auto; }}
  h1 {{ font-family: 'Fraunces', serif; color: var(--teal); margin-bottom: 8px; font-size: 1.6rem; }}
  h2 {{ font-family: 'Fraunces', serif; color: var(--teal); font-size: 1.2rem; margin: 16px 0 8px; }}
  .meta {{ color: #666; font-size: 0.85rem; margin-bottom: 20px; }}
  .card {{ background: var(--card-bg); border-radius: 8px; padding: 20px; margin-bottom: 20px;
           box-shadow: 0 1px 3px rgba(0,0,0,0.08); }}
  .stats-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)); gap: 12px; margin-bottom: 16px; }}
  .stat-box {{ background: var(--bg); border-radius: 6px; padding: 12px; }}
  .stat-box h3 {{ font-size: 0.85rem; color: var(--teal); margin-bottom: 6px; }}
  .stat-box table {{ font-size: 0.8rem; width: 100%; }}
  .stat-box td {{ padding: 2px 6px; }}
  .stat-box td:last-child {{ text-align: right; font-weight: 600; }}
  .controls {{ margin-bottom: 16px; display: flex; gap: 12px; align-items: center; flex-wrap: wrap; }}
  .controls select, .controls input {{ padding: 6px 10px; border: 1px solid var(--border); border-radius: 4px;
                                       font-size: 0.85rem; }}
  .controls label {{ font-size: 0.85rem; font-weight: 600; color: var(--teal); }}
  table.main {{ border-collapse: collapse; width: 100%; font-size: 0.8rem; }}
  table.main th {{ background: var(--teal); color: #fff; font-weight: 600; position: sticky; top: 0;
                   padding: 8px 6px; text-align: left; cursor: pointer; white-space: nowrap; }}
  table.main th:hover {{ background: #084a44; }}
  table.main td {{ padding: 6px; border-bottom: 1px solid var(--border); vertical-align: top; }}
  table.main tr:hover {{ background: #e8f0ef; }}
  .val-cell {{ max-width: 280px; word-wrap: break-word; white-space: pre-wrap; }}
  .score-cell {{ font-weight: 600; font-size: 0.75rem; padding: 3px 6px; border-radius: 3px; text-align: center; }}
  .ft-badge {{ display: inline-block; background: var(--terra); color: #fff; font-size: 0.65rem;
               padding: 1px 5px; border-radius: 3px; font-weight: 600; }}
  .cat-badge {{ display: inline-block; background: var(--teal); color: #fff; font-size: 0.65rem;
                padding: 1px 5px; border-radius: 3px; font-weight: 600; }}
  .kappa-table {{ border-collapse: collapse; width: 100%; font-size: 0.8rem; }}
  .kappa-table th, .kappa-table td {{ padding: 4px 8px; border-bottom: 1px solid var(--border); text-align: left; }}
  .kappa-table th {{ background: var(--teal); color: #fff; }}
  .k-good {{ color: #1a7a1a; font-weight: 600; }}
  .k-mod {{ color: #b8860b; font-weight: 600; }}
  .k-poor {{ color: #c0392b; font-weight: 600; }}
  .footer {{ text-align: center; color: #999; font-size: 0.8rem; margin-top: 24px; }}
  .hidden {{ display: none; }}
</style>
</head>
<body>
<h1>3-Arm AI Disagreement Pairs</h1>
<p class="meta">Generated {_esc(now)} &middot; Arms: Local (DeepSeek-R1:32b), o4-mini, Sonnet 4.6</p>
"""]

    # ── Summary statistics ───────────────────────────────────────────
    parts.append('<div class="card"><h2>Summary Statistics</h2><div class="stats-grid">')

    # By type
    parts.append('<div class="stat-box"><h3>By Field Type</h3><table>')
    for ft in ("free_text", "categorical", "numeric"):
        if ft in by_type:
            parts.append(f'<tr><td>{_esc(ft)}</td><td>{by_type[ft]}</td></tr>')
    parts.append(f'<tr style="border-top:1px solid var(--border);font-weight:600">'
                 f'<td>Total</td><td>{total}</td></tr></table></div>')

    # By tier
    parts.append('<div class="stat-box"><h3>By Tier</h3><table>')
    for t in sorted(by_tier.keys()):
        parts.append(f'<tr><td>Tier {t}</td><td>{by_tier[t]}</td></tr>')
    parts.append('</table></div>')

    # By field
    parts.append('<div class="stat-box"><h3>By Field (top 10)</h3><table>')
    for fname in sorted(by_field.keys(), key=lambda f: -by_field[f])[:10]:
        parts.append(f'<tr><td>{_esc(fname)}</td><td>{by_field[fname]}</td></tr>')
    parts.append('</table></div>')

    parts.append('</div>')  # stats-grid

    # ── Kappa table ──────────────────────────────────────────────────
    parts.append('<h2>Per-Field Kappa by Arm Pair</h2>')
    parts.append('<table class="kappa-table"><tr><th>Field</th><th>Type</th><th>Tier</th>')
    for pk in sorted(summaries_by_pair.keys()):
        parts.append(f'<th>{_esc(pk)} &kappa;</th><th>%Agr</th>')
    parts.append('</tr>')

    for fname in all_fields_sorted:
        ft = _field_type(fname, spec)
        tier = _field_tier(fname, spec)
        parts.append(f'<tr><td>{_esc(fname)}</td><td>{_esc(ft)}</td><td>{tier}</td>')
        for pk in sorted(summaries_by_pair.keys()):
            fs = summaries_by_pair[pk].get(fname)
            if fs:
                k = fs.kappa
                klass = "k-good" if not _is_nan(k) and k >= 0.8 else (
                    "k-mod" if not _is_nan(k) and k >= 0.6 else "k-poor")
                k_str = f"{k:.3f}" if not _is_nan(k) else "N/A"
                p_str = f"{fs.percent_agreement:.1%}" if not _is_nan(fs.percent_agreement) else "N/A"
                parts.append(f'<td class="{klass}">{k_str}</td><td>{p_str}</td>')
            else:
                parts.append('<td>—</td><td>—</td>')
        parts.append('</tr>')
    parts.append('</table></div>')

    # ── Filter controls ──────────────────────────────────────────────
    parts.append("""
<div class="card">
<h2>Disagreement Pairs</h2>
<div class="controls">
  <label for="filter-field">Field:</label>
  <select id="filter-field" onchange="applyFilters()">
    <option value="">All fields</option>
""")
    for fname in all_fields_sorted:
        parts.append(f'    <option value="{_esc(fname)}">{_esc(fname)}</option>\n')
    parts.append("""  </select>
  <label for="filter-type">Type:</label>
  <select id="filter-type" onchange="applyFilters()">
    <option value="">All types</option>
    <option value="free_text">free_text</option>
    <option value="categorical">categorical</option>
    <option value="numeric">numeric</option>
  </select>
  <label for="filter-score">Score:</label>
  <select id="filter-score" onchange="applyFilters()">
    <option value="">Any disagreement</option>
    <option value="MISMATCH">Has MISMATCH</option>
    <option value="AMBIGUOUS">Has AMBIGUOUS</option>
  </select>
  <label for="search-box">Search:</label>
  <input type="text" id="search-box" placeholder="Search values..." oninput="applyFilters()">
  <span id="row-count" style="font-size:0.85rem;color:var(--teal);font-weight:600;"></span>
</div>
""")

    # ── Data table ───────────────────────────────────────────────────
    parts.append("""<div style="overflow-x:auto;">
<table class="main" id="main-table">
<thead><tr>
  <th onclick="sortTable(0)">Paper</th>
  <th onclick="sortTable(1)">Field</th>
  <th onclick="sortTable(2)">Tier</th>
  <th onclick="sortTable(3)">Type</th>
  <th onclick="sortTable(4)">Local Value</th>
  <th onclick="sortTable(5)">o4-mini Value</th>
  <th onclick="sortTable(6)">Sonnet Value</th>
  <th onclick="sortTable(7)">L vs O</th>
  <th onclick="sortTable(8)">L vs S</th>
  <th onclick="sortTable(9)">O vs S</th>
</tr></thead>
<tbody>
""")

    for r in rows:
        type_badge = ('ft-badge' if r['field_type'] == 'free_text' else 'cat-badge')
        lo_color = _score_color(r["local_vs_o4mini_score"])
        ls_color = _score_color(r["local_vs_sonnet_score"])
        os_color = _score_color(r["o4mini_vs_sonnet_score"])

        parts.append(
            f'<tr data-field="{_esc(r["field_name"])}" '
            f'data-type="{_esc(r["field_type"])}" '
            f'data-scores="{r["local_vs_o4mini_score"]},{r["local_vs_sonnet_score"]},{r["o4mini_vs_sonnet_score"]}">'
            f'<td title="{_esc(r["paper_title"])}">{r["paper_id"]} {_esc(r["paper_label"])}</td>'
            f'<td>{_esc(r["field_name"])}</td>'
            f'<td>{r["field_tier"]}</td>'
            f'<td><span class="{type_badge}">{_esc(r["field_type"])}</span></td>'
            f'<td class="val-cell">{_esc(r["local_value"])}</td>'
            f'<td class="val-cell">{_esc(r["o4mini_value"])}</td>'
            f'<td class="val-cell">{_esc(r["sonnet_value"])}</td>'
            f'<td class="score-cell" style="background:{lo_color}">{r["local_vs_o4mini_score"]}</td>'
            f'<td class="score-cell" style="background:{ls_color}">{r["local_vs_sonnet_score"]}</td>'
            f'<td class="score-cell" style="background:{os_color}">{r["o4mini_vs_sonnet_score"]}</td>'
            f'</tr>\n'
        )

    parts.append('</tbody></table></div></div>')

    # ── JavaScript ───────────────────────────────────────────────────
    parts.append("""
<script>
function applyFilters() {
  const field = document.getElementById('filter-field').value;
  const type = document.getElementById('filter-type').value;
  const score = document.getElementById('filter-score').value;
  const search = document.getElementById('search-box').value.toLowerCase();
  const rows = document.querySelectorAll('#main-table tbody tr');
  let visible = 0;
  rows.forEach(row => {
    let show = true;
    if (field && row.dataset.field !== field) show = false;
    if (type && row.dataset.type !== type) show = false;
    if (score) {
      const scores = row.dataset.scores.split(',');
      if (!scores.includes(score)) show = false;
    }
    if (search) {
      const text = row.textContent.toLowerCase();
      if (!text.includes(search)) show = false;
    }
    row.style.display = show ? '' : 'none';
    if (show) visible++;
  });
  document.getElementById('row-count').textContent = visible + ' / ' + rows.length + ' rows';
}

let sortCol = -1, sortAsc = true;
function sortTable(col) {
  if (sortCol === col) { sortAsc = !sortAsc; } else { sortCol = col; sortAsc = true; }
  const tbody = document.querySelector('#main-table tbody');
  const rows = Array.from(tbody.querySelectorAll('tr'));
  rows.sort((a, b) => {
    let va = a.children[col].textContent.trim();
    let vb = b.children[col].textContent.trim();
    // Try numeric sort
    const na = parseFloat(va), nb = parseFloat(vb);
    if (!isNaN(na) && !isNaN(nb)) return sortAsc ? na - nb : nb - na;
    return sortAsc ? va.localeCompare(vb) : vb.localeCompare(va);
  });
  rows.forEach(r => tbody.appendChild(r));
}

// Initialize count
applyFilters();
</script>
""")

    parts.append("""
<p class="footer">Surgical Evidence Engine &middot; 3-Arm Disagreement Export for PLUM Lab</p>
</body>
</html>
""")

    path.write_text("".join(parts))
    print(f"HTML: {path}")


# ── Main ─────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(
        description="Export 3-arm AI disagreement pairs for PLUM Lab"
    )
    parser.add_argument("--review", default="surgical_autonomy",
                        help="Review name (default: surgical_autonomy)")
    parser.add_argument("--spec", default=None, help="Path to review spec YAML")
    args = parser.parse_args()

    db_path = str(DATA_ROOT / args.review / "review.db")
    spec_path = args.spec or f"review_specs/{args.review}_v1.yaml"
    output_dir = DATA_ROOT / args.review / "exports"
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"Review: {args.review}")
    print(f"DB: {db_path}")
    print(f"Spec: {spec_path}")
    print()

    rows, summaries_by_pair = build_disagreement_rows(db_path, spec_path)

    if not rows:
        print("No disagreement rows found.")
        return

    # Export all 3 formats
    write_csv(rows, output_dir / "disagreement_pairs_3arm.csv")
    write_xlsx(rows, summaries_by_pair, spec_path, output_dir / "disagreement_pairs_3arm.xlsx")
    write_html(rows, summaries_by_pair, spec_path, output_dir / "disagreement_pairs_3arm.html")

    # Terminal summary
    print(f"\n{'=' * 60}")
    print(f"Disagreement Summary")
    print(f"{'=' * 60}")
    print(f"Total disagreement rows: {len(rows)}")

    by_type = defaultdict(int)
    by_field = defaultdict(int)
    for r in rows:
        by_type[r["field_type"]] += 1
        by_field[r["field_name"]] += 1

    print(f"\nBy type:")
    for ft in sorted(by_type.keys()):
        print(f"  {ft}: {by_type[ft]}")

    print(f"\nBy field:")
    for fname in sorted(by_field.keys(), key=lambda f: -by_field[f]):
        ft = "FT" if fname in FREE_TEXT_FIELDS else "CAT"
        print(f"  {fname} [{ft}]: {by_field[fname]}")

    # Spot-check: print 3 free-text disagreement rows
    ft_rows = [r for r in rows if r["field_type"] == "free_text"]
    if ft_rows:
        print(f"\nSpot-check (3 free-text disagreements):")
        for r in ft_rows[:3]:
            print(f"  Paper {r['paper_id']} / {r['field_name']}:")
            print(f"    Local:  {(r['local_value'] or 'None')[:80]}")
            print(f"    o4mini: {(r['o4mini_value'] or 'None')[:80]}")
            print(f"    Sonnet: {(r['sonnet_value'] or 'None')[:80]}")
            print(f"    Scores: L/O={r['local_vs_o4mini_score']} "
                  f"L/S={r['local_vs_sonnet_score']} "
                  f"O/S={r['o4mini_vs_sonnet_score']}")


if __name__ == "__main__":
    main()
