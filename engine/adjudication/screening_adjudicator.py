"""Screening adjudication pipeline — export flagged papers for human review,
import decisions back into the database.

Works with two data sources:
  1. review.db ABSTRACT_SCREEN_FLAGGED papers (standard pipeline)
  2. Expanded search CSV files (pre-ingestion screening)
"""

import csv
import json
import logging
from datetime import datetime, timezone
from pathlib import Path

from engine.adjudication.categorizer import (
    CategoryConfig,
    categorize_paper,
    get_category_descriptions,
    load_config,
)
from engine.adjudication.schema import ensure_adjudication_table
from engine.adjudication.workflow import complete_stage, is_adjudication_complete
from engine.core.database import ReviewDatabase

logger = logging.getLogger(__name__)

EXPANDED_SEARCH_DIR = Path("data/surgical_autonomy/expanded_search")


# ── Data Collection ─────────────────────────────────────────────────


def _collect_db_flagged(db: ReviewDatabase) -> list[dict]:
    """Collect ABSTRACT_SCREEN_FLAGGED papers from review.db with screening rationale."""
    papers = db.get_papers_by_status("ABSTRACT_SCREEN_FLAGGED")
    results = []

    for p in papers:
        pid = p["id"]

        # Get screening decisions (pass 1 and 2)
        rows = db._conn.execute(
            "SELECT pass_number, decision, rationale, model "
            "FROM abstract_screening_decisions WHERE paper_id = ? ORDER BY pass_number",
            (pid,),
        ).fetchall()

        primary_decision = ""
        primary_rationale = ""
        for r in rows:
            if r["pass_number"] == 1:
                primary_decision = r["decision"]
                primary_rationale = r["rationale"] or ""

        # Get verification decision if any
        vrow = db._conn.execute(
            "SELECT decision, rationale, model "
            "FROM abstract_verification_decisions WHERE paper_id = ? "
            "ORDER BY id DESC LIMIT 1",
            (pid,),
        ).fetchone()

        verifier_decision = vrow["decision"] if vrow else ""
        verifier_rationale = vrow["rationale"] if vrow else ""

        results.append({
            "source_type": "db",
            "paper_id": pid,
            "external_key": "",
            "title": p["title"],
            "abstract": p.get("abstract") or "",
            "doi": p.get("doi") or "",
            "pmid": p.get("pmid") or "",
            "year": p.get("year") or "",
            "journal": p.get("journal") or "",
            "data_source": p.get("source") or "",
            "flagged_by": "primary_disagreement" if not vrow else "verifier_exclude",
            "primary_decision": primary_decision,
            "primary_rationale": primary_rationale,
            "verifier_decision": verifier_decision,
            "verifier_rationale": verifier_rationale,
        })

    return results


def _collect_expanded_flagged(expanded_dir: Path) -> list[dict]:
    """Collect flagged papers from expanded search CSV files.

    Pulls from both screening_results.csv (primary disagreements)
    and verification_results.csv (verifier excludes).
    """
    results = []
    abstracts_map: dict[str, str] = {}

    # Load abstracts for lookup
    abstracts_jsonl = expanded_dir / "abstracts.jsonl"
    if abstracts_jsonl.exists():
        with open(abstracts_jsonl) as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    rec = json.loads(line)
                    key = rec.get("key", "")
                    abstracts_map[key] = rec.get("abstract") or ""
                    # Also index by doi and pmid for lookup
                    if rec.get("doi"):
                        abstracts_map[rec["doi"]] = rec.get("abstract") or ""
                    if rec.get("pmid"):
                        abstracts_map[rec["pmid"]] = rec.get("abstract") or ""
                except (json.JSONDecodeError, KeyError):
                    continue

    # Phase 1: flagged from primary screening (pass1 != pass2)
    screening_csv = expanded_dir / "screening_results.csv"
    if screening_csv.exists():
        with open(screening_csv, newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get("screening_decision") != "flagged":
                    continue

                ext_key = row.get("doi") or row.get("pmid") or row["title"]
                abstract = abstracts_map.get(ext_key, "")
                if not abstract:
                    abstract = abstracts_map.get(row.get("doi", ""), "")
                if not abstract:
                    abstract = abstracts_map.get(row.get("pmid", ""), "")

                results.append({
                    "source_type": "expanded_csv",
                    "paper_id": None,
                    "external_key": ext_key,
                    "title": row["title"],
                    "abstract": abstract,
                    "doi": row.get("doi", ""),
                    "pmid": row.get("pmid", ""),
                    "year": row.get("year", ""),
                    "journal": row.get("journal", ""),
                    "data_source": row.get("source", ""),
                    "flagged_by": "primary_disagreement",
                    "primary_decision": f"pass1={row.get('pass1_decision', '')}, pass2={row.get('pass2_decision', '')}",
                    "primary_rationale": f"Pass1: {row.get('pass1_rationale', '')} | Pass2: {row.get('pass2_rationale', '')}",
                    "verifier_decision": "",
                    "verifier_rationale": "",
                })

    # Phase 2: flagged from verification (primary include, verifier exclude)
    verification_csv = expanded_dir / "verification_results.csv"
    if verification_csv.exists():
        with open(verification_csv, newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get("final_decision") != "flagged":
                    continue

                ext_key = row.get("doi") or row.get("pmid") or row["title"]
                abstract = abstracts_map.get(ext_key, "")
                if not abstract:
                    abstract = abstracts_map.get(row.get("doi", ""), "")
                if not abstract:
                    abstract = abstracts_map.get(row.get("pmid", ""), "")

                results.append({
                    "source_type": "expanded_csv",
                    "paper_id": None,
                    "external_key": ext_key,
                    "title": row["title"],
                    "abstract": abstract,
                    "doi": row.get("doi", ""),
                    "pmid": row.get("pmid", ""),
                    "year": row.get("year", ""),
                    "journal": row.get("journal", ""),
                    "data_source": row.get("source", ""),
                    "flagged_by": "verifier_exclude",
                    "primary_decision": row.get("primary_decision", "include"),
                    "primary_rationale": "",
                    "verifier_decision": row.get("verification_decision", "exclude"),
                    "verifier_rationale": row.get("verification_rationale", ""),
                })

    return results


# ── Reference Content Builder ─────────────────────────────────────


def _build_reference_content(spec) -> str:
    """Build reference sheet content from a ReviewSpec object."""
    lines = []

    lines.append("SCREENING ELIGIBILITY CRITERIA")
    lines.append("")

    if hasattr(spec, "screening_criteria") and spec.screening_criteria:
        lines.append("INCLUSION CRITERIA:")
        for criterion in spec.screening_criteria.inclusion:
            lines.append(f"  + {criterion}")
        lines.append("")
        lines.append("EXCLUSION CRITERIA:")
        for criterion in spec.screening_criteria.exclusion:
            lines.append(f"  - {criterion}")
        lines.append("")

    if hasattr(spec, "pico") and spec.pico:
        lines.append("PICO FRAMEWORK:")
        lines.append(f"  Population:   {spec.pico.population}")
        lines.append(f"  Intervention: {spec.pico.intervention}")
        lines.append(f"  Comparator:   {spec.pico.comparator}")
        if isinstance(spec.pico.outcomes, list):
            lines.append(f"  Outcomes:     {'; '.join(spec.pico.outcomes)}")
        else:
            lines.append(f"  Outcomes:     {spec.pico.outcomes}")
        lines.append("")

    if hasattr(spec, "specialty_scope") and spec.specialty_scope:
        lines.append("SPECIALTY SCOPE:")
        lines.append("  Included specialties:")
        for s in spec.specialty_scope.included:
            lines.append(f"    + {s}")
        lines.append("  Excluded specialties:")
        for s in spec.specialty_scope.excluded:
            lines.append(f"    - {s}")
        if spec.specialty_scope.notes:
            lines.append(f"  Notes: {spec.specialty_scope.notes}")

    return "\n".join(lines)


def _build_decision_criteria(spec) -> list[str]:
    """Extract decision criteria from a ReviewSpec for the Instructions sheet."""
    criteria = []
    criteria.append(
        "INCLUDE: The paper describes autonomous or semi-autonomous surgical robot "
        "execution of a physical task. The robot must CONTROL or DIRECT physical "
        "motion — not just perception, planning, or teleoperation."
    )
    criteria.append(
        "EXCLUDE: The paper is about perception-only (CV/ML without robot control), "
        "planning-only, teleoperation-only, reviews/editorials, hardware/sensors, "
        "rehabilitation/exoskeletons, or non-medical robotics."
    )

    if hasattr(spec, "specialty_scope") and spec.specialty_scope:
        included = ", ".join(spec.specialty_scope.included)
        excluded = ", ".join(spec.specialty_scope.excluded)
        criteria.append(f"SPECIALTY SCOPE — Included: {included}")
        criteria.append(f"SPECIALTY SCOPE — Excluded: {excluded}")
        if spec.specialty_scope.notes:
            criteria.append(f"EDGE CASE: {spec.specialty_scope.notes.strip()}")

    return criteria


def _build_edge_case_guidance(spec) -> str:
    """Build edge case guidance string from a ReviewSpec."""
    parts = []
    if hasattr(spec, "specialty_scope") and spec.specialty_scope and spec.specialty_scope.notes:
        parts.append(spec.specialty_scope.notes.strip())
    parts.append(
        "When in doubt between INCLUDE and EXCLUDE, lean toward INCLUDE — "
        "downstream full-text screening will catch false positives."
    )
    return " ".join(parts)


# ── Export ──────────────────────────────────────────────────────────


def export_adjudication_queue(
    review_db: ReviewDatabase,
    output_path: str | Path,
    *,
    expanded_search_dir: Path | None = None,
    review_name: str | None = None,
    category_config: CategoryConfig | None = None,
    format: str = "xlsx",
    review_spec=None,
) -> dict:
    """Export all flagged papers as a human-review Excel queue.

    Pulls ABSTRACT_SCREEN_FLAGGED from review.db plus flagged papers from
    expanded search CSVs (if expanded_search_dir is provided).

    Category config is resolved in priority order:
      1. Explicit category_config parameter
      2. YAML config at data/{review_name}/adjudication_categories.yaml
      3. No config → all papers are 'ambiguous'

    If review_spec is provided, the workbook includes self-documenting
    Instructions and Reference sheets with criteria from the spec.

    Returns summary dict with counts.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Resolve category config
    if category_config is None:
        category_config = load_config(review_name=review_name)

    # Collect flagged papers from all sources
    all_flagged = _collect_db_flagged(review_db)

    if expanded_search_dir and expanded_search_dir.exists():
        all_flagged.extend(_collect_expanded_flagged(expanded_search_dir))

    if not all_flagged:
        logger.warning("No flagged papers found — nothing to export")
        return {"total": 0, "categories": {}}

    # Auto-advance workflow: ABSTRACT_CATEGORIES_CONFIGURED if config has categories
    if category_config and category_config.categories:
        complete_stage(
            review_db._conn, "ABSTRACT_CATEGORIES_CONFIGURED",
            metadata=f"{len(category_config.categories)} categories loaded",
        )

    # Auto-categorize
    for paper in all_flagged:
        paper["auto_category"] = categorize_paper(
            paper["title"], paper["abstract"], config=category_config
        )

    # Sort: ambiguous first, then by category name
    category_order = {"ambiguous": 0}
    all_flagged.sort(key=lambda p: (
        category_order.get(p["auto_category"], 1),
        p["auto_category"],
        p["title"],
    ))

    # Count by category
    cat_counts: dict[str, int] = {}
    for p in all_flagged:
        cat = p["auto_category"]
        cat_counts[cat] = cat_counts.get(cat, 0) + 1

    if format == "html":
        from engine.adjudication.abstract_adjudication_html import (
            generate_abstract_adjudication_html,
        )
        out, html_stats = generate_abstract_adjudication_html(
            review_name=review_name or "unknown",
            output_path=str(output_path),
        )
        output_path = out
    elif format == "xlsx":
        import sys
        print(
            "Warning: xlsx export is deprecated for interactive review. "
            "Use --format html (default). xlsx retained for reference exports.",
            file=sys.stderr,
        )
        _write_xlsx(
            all_flagged, output_path, cat_counts, category_config,
            review_name=review_name or "unknown",
            review_spec=review_spec,
            db_path=str(review_db.db_path),
        )
    else:
        raise ValueError(f"Unsupported format: {format}")

    logger.info(
        "Exported adjudication queue: %d papers to %s",
        len(all_flagged), output_path,
    )
    for cat, count in sorted(cat_counts.items()):
        logger.info("  %s: %d", cat, count)

    # Auto-advance workflow: ABSTRACT_QUEUE_EXPORTED
    complete_stage(
        review_db._conn, "ABSTRACT_QUEUE_EXPORTED",
        metadata=f"{len(all_flagged)} papers exported to {output_path}",
    )

    return {
        "total": len(all_flagged),
        "categories": cat_counts,
        "output_path": str(output_path),
    }


def _write_xlsx(
    papers: list[dict],
    output_path: Path,
    cat_counts: dict,
    category_config: CategoryConfig | None = None,
    review_name: str = "unknown",
    review_spec=None,
    db_path: str = "",
) -> None:
    """Write the adjudication queue as a self-documenting Excel workbook."""
    from openpyxl.styles import PatternFill

    from engine.exporters.review_workbook import (
        ColumnDef,
        DecisionColumnDef,
        FreeTextColumnDef,
        InstructionsConfig,
        create_review_workbook,
    )

    # Category color map
    cat_fills = {
        "ambiguous": PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"),
        "cv_perception": PatternFill(start_color="E0F0FF", end_color="E0F0FF", fill_type="solid"),
        "review_editorial": PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid"),
        "hardware_sensing": PatternFill(start_color="FFF0E0", end_color="FFF0E0", fill_type="solid"),
        "planning_only": PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid"),
        "teleoperation_only": PatternFill(start_color="F0E0FF", end_color="F0E0FF", fill_type="solid"),
        "rehabilitation_prosthetics": PatternFill(start_color="FFE0F0", end_color="FFE0F0", fill_type="solid"),
        "industrial_nonmedical": PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid"),
    }

    # Build row data — assign category fills per row
    rows = []
    for i, paper in enumerate(papers, 1):
        cat = paper["auto_category"]
        rows.append({
            "row_num": i,
            "auto_category": cat,
            "title": paper["title"],
            "abstract": paper["abstract"][:2000] if paper["abstract"] else "",
            "doi": paper["doi"],
            "pmid": paper["pmid"],
            "year": paper.get("year", ""),
            "journal": paper["journal"],
            "data_source": paper["data_source"],
            "flagged_by": paper["flagged_by"],
            "primary_decision": paper["primary_decision"],
            "primary_rationale": paper["primary_rationale"][:1000] if paper["primary_rationale"] else "",
            "verifier_decision": paper["verifier_decision"],
            "verifier_rationale": paper["verifier_rationale"][:1000] if paper["verifier_rationale"] else "",
            "_cat_fill": cat_fills.get(cat),
        })

    columns = [
        ColumnDef(key="row_num", header="Row #", width=6),
        ColumnDef(key="auto_category", header="Auto Category", width=20),
        ColumnDef(key="title", header="Title", width=50, wrap=True),
        ColumnDef(key="abstract", header="Abstract", width=80, wrap=True),
        ColumnDef(key="doi", header="DOI", width=25),
        ColumnDef(key="pmid", header="PMID", width=12),
        ColumnDef(key="year", header="Year", width=6),
        ColumnDef(key="journal", header="Journal", width=30),
        ColumnDef(key="data_source", header="Source", width=10),
        ColumnDef(key="flagged_by", header="Flagged By", width=20),
        ColumnDef(key="primary_decision", header="Primary Decision", width=15),
        ColumnDef(key="primary_rationale", header="Primary Rationale", width=50, wrap=True),
        ColumnDef(key="verifier_decision", header="Verifier Decision", width=15),
        ColumnDef(key="verifier_rationale", header="Verifier Rationale", width=50, wrap=True),
    ]

    decision_columns = [
        DecisionColumnDef(
            key="PI_decision",
            header="PI_decision (INCLUDE/EXCLUDE)",
            valid_values=["INCLUDE", "EXCLUDE"],
            width=25,
        ),
    ]

    # Build instructions
    import_cmd = (
        f"python -c \"\n"
        f"from engine.core.database import ReviewDatabase\n"
        f"from engine.adjudication import import_adjudication_decisions\n"
        f"db = ReviewDatabase('{review_name}')\n"
        f"import_adjudication_decisions(db, '{output_path}')\n"
        f"\""
    )

    # Build decision criteria and edge case guidance from spec if available
    decision_criteria = [
        "INCLUDE: The paper describes autonomous or semi-autonomous surgical robot execution.",
        "EXCLUDE: The paper is about perception-only, planning-only, teleoperation-only, "
        "reviews/editorials, hardware/sensors, rehabilitation, or non-medical robotics.",
    ]
    edge_case = ""
    if review_spec:
        decision_criteria = _build_decision_criteria(review_spec)
        edge_case = _build_edge_case_guidance(review_spec)

    instr = InstructionsConfig(
        review_name=review_name,
        review_spec_id=f"{review_spec.title} v{review_spec.version}" if review_spec else "",
        db_path=db_path,
        export_trigger=(
            f"{len(papers)} papers flagged during abstract screening where "
            f"primary and verification models disagreed"
        ),
        row_count=len(papers),
        decision_column_name="PI_decision (INCLUDE/EXCLUDE)",
        valid_values=["INCLUDE", "EXCLUDE"],
        decision_criteria=decision_criteria,
        edge_case_guidance=edge_case,
        import_command=import_cmd,
        columns_importer_reads=["Title (C)", "DOI (E)", "PMID (F)", "PI_decision (O)", "Notes (P)"],
        columns_importer_ignores=(
            "All other columns (Row #, Auto Category, Abstract, Journal, Source, "
            "Flagged By, Primary/Verifier Decision/Rationale) are read-only context "
            "and are ignored by the importer."
        ),
    )

    # Build reference content from spec if available
    reference_content = None
    if review_spec:
        reference_content = _build_reference_content(review_spec)

    free_text_columns = [
        FreeTextColumnDef(key="PI_notes", header="PI_notes (optional)", width=30),
    ]

    # Create the workbook
    create_review_workbook(
        output_path=output_path,
        rows=rows,
        columns=columns,
        decision_columns=decision_columns,
        free_text_columns=free_text_columns,
        instructions=instr,
        reference_content=reference_content,
        reference_sheet_title="Screening Criteria",
    )

    # Post-process: apply category fills (review_workbook doesn't know about categories)
    from openpyxl import load_workbook as _load_wb
    wb = _load_wb(output_path)
    ws = wb["Review Queue"]
    for row_idx, row_data in enumerate(rows, 2):
        fill = row_data.get("_cat_fill")
        if fill:
            ws.cell(row=row_idx, column=2).fill = fill  # column B = Auto Category
    wb.save(output_path)


# ── Import ─────────────────────────────────────────────────────────


def import_adjudication_decisions(
    review_db: ReviewDatabase,
    input_path: str | Path | None = None,
) -> dict:
    """Read completed adjudication decisions and write to database.

    If input_path is None, auto-discovers the decisions file using the
    naming convention: {review}_abstract_adjudication_decisions.json

    Supports two input formats (detected by file extension):
      - .xlsx  — Excel workbook (from export_adjudication_queue)
      - .json  — JSON array (from abstract_adjudication_html.py HTML tool)

    JSON schema: [{paper_id: int, decision: "ABSTRACT_SCREENED_IN"|"ABSTRACT_SCREENED_OUT", note: str|null}]

    Validates the entire file before making any changes:
      - Rejects if any decision is missing or invalid

    For papers in review.db (with paper_id), updates their status:
      INCLUDE / ABSTRACT_SCREENED_IN → ABSTRACT_SCREENED_IN
      EXCLUDE / ABSTRACT_SCREENED_OUT → ABSTRACT_SCREENED_OUT

    For expanded search papers (no paper_id), records the decision in
    the abstract_screening_adjudication table for later pipeline use.

    Returns summary dict.
    """
    if input_path is None:
        from engine.core.naming import review_artifact_path
        review_name = Path(review_db.db_path).parent.name
        data_dir = Path(review_db.db_path).parent
        input_path = review_artifact_path(
            data_dir, review_name, "abstract_adjudication", "decisions", "json",
        )
        if not input_path.exists():
            raise FileNotFoundError(
                f"Expected decisions file at: {input_path}\n"
                "Export from the HTML review tool first."
            )

    input_path = Path(input_path)

    if input_path.suffix.lower() == ".json":
        return _import_abstract_json(review_db, input_path)

    return _import_abstract_xlsx(review_db, input_path)


def _import_abstract_json(
    review_db: ReviewDatabase,
    input_path: Path,
) -> dict:
    """Import abstract adjudication decisions from JSON (HTML tool output).

    Expected schema: [{paper_id: int, decision: "ABSTRACT_SCREENED_IN"|"ABSTRACT_SCREENED_OUT", note: str|null}]
    """
    try:
        with open(input_path) as f:
            records = json.load(f)
    except (json.JSONDecodeError, OSError) as e:
        error_msg = f"\nIMPORT REJECTED — Cannot read JSON file: {e}\n"
        print(error_msg)
        logger.error(error_msg)
        return {
            "stats": {"include": 0, "exclude": 0, "missing": 0, "invalid": 0, "total": 0},
            "warnings": [error_msg],
        }

    if not isinstance(records, list):
        error_msg = "\nIMPORT REJECTED — JSON must be an array of decision objects.\n"
        print(error_msg)
        logger.error(error_msg)
        return {
            "stats": {"include": 0, "exclude": 0, "missing": 0, "invalid": 0, "total": 0},
            "warnings": [error_msg],
        }

    # Map HTML tool decisions to internal INCLUDE/EXCLUDE
    decision_map = {
        "ABSTRACT_SCREENED_IN": "INCLUDE",
        "ABSTRACT_SCREENED_OUT": "EXCLUDE",
        "INCLUDE": "INCLUDE",
        "EXCLUDE": "EXCLUDE",
    }

    # ── Pass 1: Validate ─────────────────────────────────────────
    parsed_rows = []
    invalid_rows = []

    for i, rec in enumerate(records):
        paper_id = rec.get("paper_id")
        decision_raw = (rec.get("decision") or "").strip().upper()
        note = rec.get("note") or ""

        if not paper_id:
            invalid_rows.append(f"  Record {i}: missing paper_id")
            continue

        decision = decision_map.get(decision_raw)
        if not decision:
            invalid_rows.append(
                f"  Record {i} (paper {paper_id}): '{decision_raw}' "
                f"(must be ABSTRACT_SCREENED_IN/ABSTRACT_SCREENED_OUT or INCLUDE/EXCLUDE)"
            )
            continue

        # Look up paper info from DB
        row = review_db._conn.execute(
            "SELECT title, doi, pmid FROM papers WHERE id = ?", (paper_id,)
        ).fetchone()
        title = row["title"] if row else ""
        doi = row["doi"] if row else ""
        pmid = row["pmid"] if row else ""

        parsed_rows.append({
            "paper_id": paper_id,
            "category": "",
            "title": title,
            "doi": doi or "",
            "pmid": pmid or "",
            "decision": decision,
            "notes": note,
        })

    # ── Reject on validation failure ──────────────────────────────
    if invalid_rows:
        msg_parts = [
            f"\nIMPORT REJECTED — {len(invalid_rows)} validation error(s) found.",
            "No database changes were made.\n",
            f"INVALID RECORDS ({len(invalid_rows)}):",
        ]
        msg_parts.extend(invalid_rows)
        msg_parts.append("\nFix the JSON file and re-run.")

        error_msg = "\n".join(msg_parts)
        print(error_msg)
        logger.error(error_msg)
        return {
            "stats": {
                "include": 0, "exclude": 0,
                "missing": 0, "invalid": len(invalid_rows),
                "total": len(parsed_rows) + len(invalid_rows),
            },
            "warnings": invalid_rows,
        }

    # ── Pass 2: Apply ─────────────────────────────────────────────
    return _apply_abstract_decisions(review_db, parsed_rows)


def _apply_abstract_decisions(
    review_db: ReviewDatabase,
    parsed_rows: list[dict],
) -> dict:
    """Shared logic: write validated abstract decisions to DB and advance workflow."""
    ensure_adjudication_table(review_db._conn)

    now = datetime.now(timezone.utc).isoformat()
    stats = {"include": 0, "exclude": 0, "missing": 0, "invalid": 0, "total": len(parsed_rows)}

    for pr in parsed_rows:
        decision = pr["decision"]
        title = pr["title"]
        doi = pr.get("doi", "")
        pmid = pr.get("pmid", "")
        notes = pr.get("notes", "")
        category = pr.get("category", "")
        paper_id = pr.get("paper_id")

        ext_key = doi or pmid or title

        # Write to adjudication table
        review_db._conn.execute(
            """INSERT INTO abstract_screening_adjudication
               (paper_id, external_key, title, adjudication_decision,
                adjudication_source, adjudication_reason,
                adjudication_category, adjudication_timestamp, created_at)
               VALUES (?, ?, ?, ?, 'human', ?, ?, ?, ?)""",
            (
                paper_id,
                ext_key,
                title,
                decision,
                notes,
                category,
                now,
                now,
            ),
        )

        # Try to find and update paper in review.db
        paper_row = None
        if paper_id:
            paper_row = review_db._conn.execute(
                "SELECT id, status FROM papers WHERE id = ?", (paper_id,)
            ).fetchone()
        if not paper_row and pmid:
            paper_row = review_db._conn.execute(
                "SELECT id, status FROM papers WHERE pmid = ?", (str(pmid),)
            ).fetchone()
        if not paper_row and doi:
            paper_row = review_db._conn.execute(
                "SELECT id, status FROM papers WHERE doi = ?", (doi,)
            ).fetchone()

        if paper_row and paper_row["status"] == "ABSTRACT_SCREEN_FLAGGED":
            pid = paper_row["id"]
            new_status = "ABSTRACT_SCREENED_IN" if decision == "INCLUDE" else "ABSTRACT_SCREENED_OUT"
            review_db.update_status(pid, new_status)

            # Update adjudication record with paper_id if it wasn't set
            if not paper_id:
                review_db._conn.execute(
                    """UPDATE abstract_screening_adjudication
                       SET paper_id = ?
                       WHERE external_key = ? AND adjudication_timestamp = ?""",
                    (pid, ext_key, now),
                )

        if decision == "INCLUDE":
            stats["include"] += 1
        else:
            stats["exclude"] += 1

    review_db._conn.commit()

    # Success summary
    print(
        f"\nIMPORT SUCCESSFUL — {stats['total']} decisions processed.\n"
        f"  INCLUDE: {stats['include']}\n"
        f"  EXCLUDE: {stats['exclude']}\n"
        f"  Database updated."
    )

    logger.info(
        "Import complete: %d include, %d exclude (of %d total)",
        stats["include"], stats["exclude"], stats["total"],
    )

    # Auto-advance workflow: ABSTRACT_ADJUDICATION_COMPLETE
    complete_stage(
        review_db._conn, "ABSTRACT_ADJUDICATION_COMPLETE",
        metadata=(
            f"{stats['include']} included, {stats['exclude']} excluded "
            f"(of {stats['total']} total)"
        ),
    )

    return {"stats": stats, "warnings": []}


def _import_abstract_xlsx(
    review_db: ReviewDatabase,
    input_path: Path,
) -> dict:
    """Import abstract adjudication decisions from Excel workbook."""
    from openpyxl import load_workbook

    wb = load_workbook(input_path)
    ws = wb["Review Queue"]

    # ── Build header index for robust column lookup ───────────
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_index = {}
    for idx, val in enumerate(header_row):
        if val:
            col_index[str(val).strip()] = idx

    def _find_col(candidates: list[str]) -> int | None:
        """Find column index by trying multiple header name patterns."""
        for c in candidates:
            for header, idx in col_index.items():
                if c.lower() in header.lower():
                    return idx
        return None

    title_col = _find_col(["Title"])
    doi_col = _find_col(["DOI"])
    pmid_col = _find_col(["PMID"])
    decision_col = _find_col(["PI_decision", "DECISION"])
    notes_col = _find_col(["PI_notes", "Notes"])
    category_col = _find_col(["Auto Category", "Category"])
    row_num_col = _find_col(["Row #", "Row"])

    if title_col is None or decision_col is None:
        error_msg = (
            "\nIMPORT REJECTED — Required columns not found.\n"
            f"  Found headers: {list(col_index.keys())}\n"
            "  Required: Title, PI_decision (or DECISION)\n"
        )
        print(error_msg)
        logger.error(error_msg)
        return {
            "stats": {"include": 0, "exclude": 0, "missing": 0, "invalid": 0, "total": 0},
            "warnings": [error_msg],
        }

    # ── Pass 1: Validate all rows before any DB writes ──────────
    parsed_rows = []
    blank_rows = []
    invalid_rows = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        title_val = row[title_col].value
        if not title_val:
            continue

        row_num = row[row_num_col].value if row_num_col is not None else row[0].row - 1
        category = row[category_col].value if category_col is not None else ""
        title = row[title_col].value
        doi = (row[doi_col].value or "") if doi_col is not None else ""
        pmid = (row[pmid_col].value or "") if pmid_col is not None else ""
        decision_raw = row[decision_col].value
        notes = (row[notes_col].value or "") if notes_col is not None else ""

        if not decision_raw or str(decision_raw).strip() == "":
            blank_rows.append(
                f"  Row {row_num}: '{title[:60]}...'" if len(title) > 60
                else f"  Row {row_num}: '{title}'"
            )
            continue

        decision = str(decision_raw).strip().upper()
        if decision not in ("INCLUDE", "EXCLUDE"):
            invalid_rows.append(
                f"  Row {row_num}: '{decision_raw}' (must be INCLUDE or EXCLUDE)"
            )
            continue

        parsed_rows.append({
            "row_num": row_num,
            "category": category,
            "title": title,
            "doi": doi,
            "pmid": pmid,
            "decision": decision,
            "notes": notes,
        })

    # ── Reject on validation failure ────────────────────────────
    if blank_rows or invalid_rows:
        total_issues = len(blank_rows) + len(invalid_rows)
        msg_parts = [
            f"\nIMPORT REJECTED — {total_issues} validation error(s) found.",
            "No database changes were made.\n",
        ]
        if blank_rows:
            msg_parts.append(f"BLANK DECISION CELLS ({len(blank_rows)} rows):")
            msg_parts.extend(blank_rows)
            msg_parts.append("")
        if invalid_rows:
            msg_parts.append(f"INVALID DECISION VALUES ({len(invalid_rows)} rows):")
            msg_parts.extend(invalid_rows)
            msg_parts.append("")
        msg_parts.append("Fix the workbook and re-run the import command.")

        error_msg = "\n".join(msg_parts)
        print(error_msg)
        logger.error(error_msg)

        return {
            "stats": {
                "include": 0,
                "exclude": 0,
                "missing": len(blank_rows),
                "invalid": len(invalid_rows),
                "total": len(parsed_rows) + len(blank_rows) + len(invalid_rows),
            },
            "warnings": blank_rows + invalid_rows,
        }

    # ── Pass 2: Apply all validated decisions ───────────────────
    return _apply_abstract_decisions(review_db, parsed_rows)


# ── Pipeline Gate ──────────────────────────────────────────────────


def check_adjudication_gate(review_db: ReviewDatabase) -> int:
    """Check for unresolved ABSTRACT_SCREEN_FLAGGED papers. Returns count."""
    flagged = review_db.get_papers_by_status("ABSTRACT_SCREEN_FLAGGED")
    return len(flagged)
