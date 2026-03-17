"""Tests for analysis.paper1.human_import — human extractor workbook importer."""

import sqlite3
from pathlib import Path
from unittest.mock import patch, MagicMock

import pytest

from analysis.paper1.human_import import (
    _extract_extractor_id,
    _normalize_value,
    parse_workbook,
    validate_workbook,
    store_human_extractions,
    _EXTRACTION_FIELDS,
    _load_codebook_valid_values,
)


# ── Helpers ──────────────────────────────────────────────────────────


CODEBOOK_PATH = Path("data/surgical_autonomy/extraction_codebook.yaml")


def _make_row(overrides: dict | None = None) -> dict:
    """Build a minimal valid row dict."""
    row = {
        "paper_id": "EE-011",
        "extractor_id": "A",
        "sq_key_limitation": "Some quote",
        "sq_clinical_readiness": "Another quote",
        "notes": None,
    }
    for f in _EXTRACTION_FIELDS:
        row[f] = None
    # Set categorical fields to valid values
    row["study_type"] = "Original Research"
    row["key_limitation"] = "Small sample size"
    row["clinical_readiness_assessment"] = "Proof of concept only"
    if overrides:
        row.update(overrides)
    return row


def _make_db(tmp_path: Path, paper_ids: list[int] | None = None) -> Path:
    """Create a minimal papers table."""
    db_path = tmp_path / "test.db"
    conn = sqlite3.connect(str(db_path))
    conn.execute("CREATE TABLE papers (id INTEGER PRIMARY KEY, title TEXT)")
    for pid in (paper_ids or [11]):
        conn.execute("INSERT INTO papers (id, title) VALUES (?, ?)", (pid, f"Paper {pid}"))
    conn.commit()
    conn.close()
    return db_path


# ── Unit: extractor ID extraction ────────────────────────────────────


class TestExtractorId:

    def test_standard_suffix(self):
        assert _extract_extractor_id(Path("Extraction_Workbook_v2_A.xlsx")) == "A"

    def test_lowercase_normalized(self):
        assert _extract_extractor_id(Path("workbook_b.xlsx")) == "B"

    def test_invalid_name_raises(self):
        with pytest.raises(ValueError, match="Cannot derive extractor ID"):
            _extract_extractor_id(Path("workbook.xlsx"))

    def test_multi_char_suffix_raises(self):
        with pytest.raises(ValueError, match="Cannot derive extractor ID"):
            _extract_extractor_id(Path("workbook_AB.xlsx"))


# ── Unit: value normalization ────────────────────────────────────────


class TestNormalize:

    def test_none_returns_none(self):
        assert _normalize_value(None) is None

    def test_empty_string_returns_none(self):
        assert _normalize_value("") is None

    def test_nr_returns_none(self):
        assert _normalize_value("NR") is None
        assert _normalize_value("nr") is None
        assert _normalize_value(" NR ") is None

    def test_whitespace_stripped(self):
        assert _normalize_value("  Original Research  ") == "Original Research"

    def test_number_becomes_string(self):
        assert _normalize_value(42) == "42"


# ── Unit: parse_workbook ─────────────────────────────────────────────


class TestParseWorkbook:

    def test_parse_reference_workbook(self):
        """Parse the actual reference workbook — verifies column mapping."""
        wb_path = Path("data/surgical_autonomy/Extraction_Workbook_v2_A.xlsx")
        if not wb_path.exists():
            pytest.skip("Reference workbook not available")

        rows = parse_workbook(wb_path)
        assert len(rows) == 70
        assert rows[0]["paper_id"] == "EE-011"
        assert rows[0]["extractor_id"] == "A"
        # All 20 fields present as keys
        for f in _EXTRACTION_FIELDS:
            assert f in rows[0], f"Missing field: {f}"

    def test_missing_sheet_raises(self, tmp_path):
        """Workbook without 'Extraction Form' sheet raises ValueError."""
        import openpyxl
        wb = openpyxl.Workbook()
        wb.active.title = "Wrong Sheet"
        path = tmp_path / "bad_A.xlsx"
        wb.save(path)
        wb.close()

        with pytest.raises(ValueError, match="missing 'Extraction Form' sheet"):
            parse_workbook(path)

    def test_missing_columns_raises(self, tmp_path):
        """Workbook with incomplete headers raises ValueError."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extraction Form"
        ws.cell(row=1, column=1, value="Section")
        ws.cell(row=2, column=1, value="Paper ID")
        # Missing all extraction field columns
        path = tmp_path / "incomplete_A.xlsx"
        wb.save(path)
        wb.close()

        with pytest.raises(ValueError, match="Missing extraction columns"):
            parse_workbook(path)

    def test_blank_rows_skipped(self, tmp_path):
        """Rows with no Paper ID are skipped."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extraction Form"
        # Row 1: section headers
        ws.cell(row=1, column=1, value="IDENTIFIERS")
        # Row 2: column headers
        headers = ["Paper ID", "First Author", "Year", "Title"] + _EXTRACTION_FIELDS + [
            "SQ: key_limitation [REQ]", "SQ: clinical_readiness [REQ]", "Extractor Notes"
        ]
        for j, h in enumerate(headers, 1):
            ws.cell(row=2, column=j, value=h)
        # Row 3: valid data
        ws.cell(row=3, column=1, value="EE-001")
        # Row 4: blank
        # Row 5: valid data
        ws.cell(row=5, column=1, value="EE-002")

        path = tmp_path / "sparse_A.xlsx"
        wb.save(path)
        wb.close()

        rows = parse_workbook(path)
        assert len(rows) == 2
        assert rows[0]["paper_id"] == "EE-001"
        assert rows[1]["paper_id"] == "EE-002"


# ── Unit: validate_workbook ──────────────────────────────────────────


class TestValidateWorkbook:

    def test_valid_row_passes(self):
        rows = [_make_row()]
        errors = validate_workbook(rows, CODEBOOK_PATH)
        assert errors == []

    def test_bad_paper_id_format(self):
        rows = [_make_row({"paper_id": "11"})]
        errors = validate_workbook(rows, CODEBOOK_PATH)
        assert any("does not match EE-NNN" in e for e in errors)

    def test_paper_id_not_in_db(self, tmp_path):
        rows = [_make_row({"paper_id": "EE-999"})]
        db_path = _make_db(tmp_path, paper_ids=[11])
        errors = validate_workbook(rows, CODEBOOK_PATH, db_path=db_path)
        assert any("not found in database" in e for e in errors)

    def test_paper_id_in_db_passes(self, tmp_path):
        rows = [_make_row({"paper_id": "EE-011"})]
        db_path = _make_db(tmp_path, paper_ids=[11])
        errors = validate_workbook(rows, CODEBOOK_PATH, db_path=db_path)
        assert errors == []

    def test_invalid_categorical_value(self):
        rows = [_make_row({"study_type": "Made Up Type"})]
        errors = validate_workbook(rows, CODEBOOK_PATH)
        assert any("study_type='Made Up Type' not in codebook" in e for e in errors)

    def test_valid_categorical_case_insensitive(self):
        rows = [_make_row({"study_type": "original research"})]
        errors = validate_workbook(rows, CODEBOOK_PATH)
        assert errors == []

    def test_semicolon_multi_value_validated(self):
        rows = [_make_row({"validation_setting": "Ex vivo; In vivo (animal)"})]
        errors = validate_workbook(rows, CODEBOOK_PATH)
        assert errors == []

    def test_semicolon_multi_value_one_bad(self):
        rows = [_make_row({"validation_setting": "Ex vivo; Outer Space"})]
        errors = validate_workbook(rows, CODEBOOK_PATH)
        assert any("Outer Space" in e for e in errors)

    def test_missing_source_quote_key_limitation(self):
        rows = [_make_row({"key_limitation": "A limitation", "sq_key_limitation": None})]
        errors = validate_workbook(rows, CODEBOOK_PATH)
        assert any("SQ: key_limitation is empty" in e for e in errors)

    def test_missing_source_quote_cra(self):
        rows = [_make_row({
            "clinical_readiness_assessment": "Proof of concept only",
            "sq_clinical_readiness": None,
        })]
        errors = validate_workbook(rows, CODEBOOK_PATH)
        assert any("SQ: clinical_readiness is empty" in e for e in errors)

    def test_null_field_skips_source_quote_check(self):
        """If the extraction field is None, missing source quote is fine."""
        rows = [_make_row({
            "key_limitation": None, "sq_key_limitation": None,
            "clinical_readiness_assessment": None, "sq_clinical_readiness": None,
        })]
        errors = validate_workbook(rows, CODEBOOK_PATH)
        assert errors == []

    def test_empty_rows_detected(self):
        errors = validate_workbook([], CODEBOOK_PATH)
        assert any("No data rows" in e for e in errors)


# ── Unit: codebook loading ───────────────────────────────────────────


class TestCodebookLoading:

    def test_loads_categorical_fields(self):
        valid = _load_codebook_valid_values(CODEBOOK_PATH)
        assert "study_type" in valid
        assert "original research" in valid["study_type"]
        # Free-text fields should NOT be in the dict
        assert "robot_platform" not in valid

    def test_autonomy_level_values(self):
        valid = _load_codebook_valid_values(CODEBOOK_PATH)
        assert "0 (no autonomy)" in valid["autonomy_level"]
        assert "mixed/multiple" in valid["autonomy_level"]


# ── Unit: store_human_extractions ────────────────────────────────────


class TestStore:

    def test_stores_long_format(self, tmp_path):
        db_path = tmp_path / "store_test.db"
        rows = [_make_row()]
        inserted = store_human_extractions(rows, "A", db_path)
        assert inserted == len(_EXTRACTION_FIELDS)  # 20 fields per paper

        conn = sqlite3.connect(str(db_path))
        count = conn.execute("SELECT COUNT(*) FROM human_extractions").fetchone()[0]
        assert count == 20

        # Check a specific row
        row = conn.execute(
            "SELECT value, source_quote FROM human_extractions "
            "WHERE paper_id = 'EE-011' AND field_name = 'key_limitation'"
        ).fetchone()
        assert row[0] == "Small sample size"
        assert row[1] == "Some quote"
        conn.close()

    def test_multiple_papers(self, tmp_path):
        db_path = tmp_path / "multi.db"
        rows = [
            _make_row({"paper_id": "EE-001"}),
            _make_row({"paper_id": "EE-002"}),
        ]
        inserted = store_human_extractions(rows, "A", db_path)
        assert inserted == 40  # 2 papers x 20 fields

    def test_duplicate_import_raises(self, tmp_path):
        db_path = tmp_path / "dup.db"
        rows = [_make_row()]
        store_human_extractions(rows, "A", db_path)

        with pytest.raises(RuntimeError, match="Duplicate import"):
            store_human_extractions(rows, "A", db_path)

    def test_different_extractor_ok(self, tmp_path):
        db_path = tmp_path / "diff.db"
        rows = [_make_row()]
        store_human_extractions(rows, "A", db_path)
        inserted = store_human_extractions(rows, "B", db_path)
        assert inserted == 20

    def test_imported_at_set(self, tmp_path):
        db_path = tmp_path / "ts.db"
        store_human_extractions([_make_row()], "A", db_path)

        conn = sqlite3.connect(str(db_path))
        ts = conn.execute("SELECT imported_at FROM human_extractions LIMIT 1").fetchone()[0]
        assert ts is not None
        assert "T" in ts  # ISO format
        conn.close()

    def test_notes_stored(self, tmp_path):
        db_path = tmp_path / "notes.db"
        rows = [_make_row({"notes": "Check figure 3"})]
        store_human_extractions(rows, "A", db_path)

        conn = sqlite3.connect(str(db_path))
        notes = conn.execute(
            "SELECT DISTINCT notes FROM human_extractions WHERE paper_id = 'EE-011'"
        ).fetchone()[0]
        assert notes == "Check figure 3"
        conn.close()
