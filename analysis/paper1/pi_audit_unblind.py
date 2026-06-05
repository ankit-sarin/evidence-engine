"""PI audit unblinding + judge-reliability scoring.

Joins the completed blinded PI adjudication workbook against the Pass 2
audit key on ``row_id``, runs a fail-fast integrity gate, and computes
judge-reliability metrics for the Pass 2 Gemma verdicts using the PI
adjudication as the gold standard. Emits a Digital-Surgeon-branded
results workbook plus a machine-readable JSON sidecar.

Derived analysis artifact — reads two .xlsx, writes one .xlsx + one
.json. No database access.

The blinded sampler that produced the inputs is
``analysis.paper1.pi_audit_sampler``; this module is its post-adjudication
counterpart and lives alongside it by convention.

Metric layers (see the audit spec):
  A  Confusion matrix              4×3, PI (rows, incl. UNCLEAR) × Gemma
  B  Per-Gemma-class precision     conditional (substantive den) + strict
  C  Fabrication-flag PPV          flag = Gemma==UNSUPPORTED
  D  Weighted Cohen's kappa        substantive rows, linear weights, boot CI
  E  Per-arm agreement + PI dist.  judge reliability across arms
  F  source_window_strategy        UNCLEAR rate, agreement, clean-subset

Statistics are pure Python (Wilson score interval, weighted kappa,
seeded bootstrap). Dependencies: openpyxl + stdlib only.

Usage:
  PYTHONPATH=. python -m analysis.paper1.pi_audit_unblind \\
      --completed artifacts/paper1/pi_audit/<workbook>_COMPLETED.xlsx \\
      --key       artifacts/paper1/pi_audit/pi_audit_key_<ts>.xlsx \\
      [--out-dir  artifacts/paper1/pi_audit]
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import random
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ═════════════════════════════════════════════════════════════════════
# Vocabularies & fixed parameters
# ═════════════════════════════════════════════════════════════════════

PI_VERDICTS = ("SUPPORTED", "PARTIALLY_SUPPORTED", "UNSUPPORTED", "UNCLEAR")
GEMMA_VERDICTS = ("SUPPORTED", "PARTIALLY_SUPPORTED", "UNSUPPORTED")
SUBSTANTIVE = ("SUPPORTED", "PARTIALLY_SUPPORTED", "UNSUPPORTED")
SUBSTANTIVE_SET = frozenset(SUBSTANTIVE)

# Ordinal map for weighted kappa.
ORDINAL = {"UNSUPPORTED": 0, "PARTIALLY_SUPPORTED": 1, "SUPPORTED": 2}

EXPECTED_N = 100

# Bootstrap — the only stochastic component. Seed fixed for reproducibility.
BOOTSTRAP_RESAMPLES = 2000
BOOTSTRAP_SEED = 20260422

WILSON_Z = 1.96

# Strategies whose UNCLEAR is protocol-driven (Instructions told the PI to
# prefer UNCLEAR there — degraded context, no/unlocatable span).
PROTOCOL_UNCLEAR_STRATEGIES = frozenset(
    {"missing_span_fallback_head", "absence_fallback_head"}
)
CLEAN_STRATEGY = "arm_span_window"
CLEAN_PLUS_STRATEGIES = frozenset({"arm_span_window", "full_text"})

SCHEMA_VERSION = "pi_audit_results/1.0"

# Blinding-leak columns that must NOT appear in the completed workbook.
BLINDING_FORBIDDEN_HEADERS = ("arm_name", "gemma_verdict")

# Brand colors — mirrors analysis/paper1/pi_audit_sampler.py.
BRAND_TEAL = "0A5E56"
BRAND_MIST = "EEF5F4"
BRAND_MIST_DARK = "DFEBE9"
BRAND_CHARCOAL = "2C2C2C"
BRAND_WHITE = "FFFFFF"
BRAND_BORDER = "C5CDD6"
BRAND_TERRACOTTA = "B85D3A"

CAVEATS = [
    "Conditional on the Gemma-stratified audit sample — these are NOT "
    "population prevalence estimates. Reweighting to the full Pass 2 "
    "population is deferred to Paper 1.",
    "Single adjudicator (the PI). No inter-adjudicator reliability is "
    "estimable from this sample.",
    "Per-class precision is conditional on the PI rendering a substantive "
    "verdict; UNCLEAR rows are excluded from the conditional denominator "
    "and reported separately as a strict (UNCLEAR-inclusive) sensitivity.",
    "UNCLEAR is protocol-driven in the absence_fallback_head and "
    "missing_span_fallback_head strategies — the Instructions sheet told "
    "the PI to prefer UNCLEAR when context was degraded. Treat UNCLEAR "
    "rates in those strategies as a property of the windowing, not of "
    "true ambiguity.",
]


# ═════════════════════════════════════════════════════════════════════
# Data containers
# ═════════════════════════════════════════════════════════════════════


@dataclass
class JoinedRow:
    row_id: int
    pi: str                      # normalized PI adjudication (uppercase)
    gemma: str                   # gemma_verdict from key
    arm_name: str
    field_name: str
    source_window_strategy: str
    pi_notes: Optional[str]
    pi_adjudicated_at: Optional[str]
    key_cols: dict = field(default_factory=dict)  # full key row, raw


# ═════════════════════════════════════════════════════════════════════
# IO helpers
# ═════════════════════════════════════════════════════════════════════


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def _norm(v) -> str:
    """str → stripped uppercase; None/blank → ''."""
    if v is None:
        return ""
    return str(v).strip().upper()


def _eq_field(a, b) -> bool:
    """Pairing comparison: str+strip, case-sensitive (field/strategy/value
    are emitted verbatim by the sampler so a byte match is expected)."""
    sa = "" if a is None else str(a).strip()
    sb = "" if b is None else str(b).strip()
    return sa == sb


def _read_sheet(path: Path, sheet: str) -> tuple[list[str], list[dict]]:
    """Return (header, list-of-row-dicts keyed by header)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"ABORT: sheet {sheet!r} not in {path.name} "
                         f"(found {wb.sheetnames})")
    ws = wb[sheet]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [(_h if _h is None else str(_h)) for _h in next(rows_iter)]
    except StopIteration:
        wb.close()
        raise SystemExit(f"ABORT: {sheet!r} in {path.name} is empty")
    records: list[dict] = []
    for raw in rows_iter:
        # Skip wholly blank trailing rows (read_only can emit them).
        if raw is None or all(c is None for c in raw):
            continue
        rec = {header[i]: (raw[i] if i < len(raw) else None)
               for i in range(len(header))}
        records.append(rec)
    wb.close()
    return header, records


# ═════════════════════════════════════════════════════════════════════
# Integrity gate (fail-fast)
# ═════════════════════════════════════════════════════════════════════


def _abort(msg: str) -> "None":
    raise SystemExit(f"ABORT (integrity gate): {msg}")


def run_integrity_gate(
    completed_header: list[str],
    completed_rows: list[dict],
    key_header: list[str],
    key_rows: list[dict],
) -> list[JoinedRow]:
    """Validate the join and return the joined rows. Aborts on any failure,
    naming the offending row_ids."""
    pass_lines: list[str] = []

    # --- 0. Required columns present -------------------------------------
    for col in ("row_id", "field_name", "arm_value",
                "source_window_strategy", "adjudication"):
        if col not in completed_header:
            _abort(f"completed workbook missing column {col!r}")
    for col in ("row_id", "field_name", "arm_value",
                "source_window_strategy", "arm_name", "gemma_verdict"):
        if col not in key_header:
            _abort(f"key workbook missing column {col!r}")

    # --- 1. Row counts & 1:1 row_id sets == {1..100} ---------------------
    if len(completed_rows) != EXPECTED_N:
        _abort(f"completed workbook has {len(completed_rows)} data rows, "
               f"expected {EXPECTED_N}")
    if len(key_rows) != EXPECTED_N:
        _abort(f"key workbook has {len(key_rows)} data rows, "
               f"expected {EXPECTED_N}")

    def _row_ids(rows, label):
        ids = []
        for r in rows:
            try:
                ids.append(int(r["row_id"]))
            except (TypeError, ValueError):
                _abort(f"{label}: non-integer row_id {r.get('row_id')!r}")
        return ids

    comp_ids = _row_ids(completed_rows, "completed")
    key_ids = _row_ids(key_rows, "key")
    expected_set = set(range(1, EXPECTED_N + 1))

    comp_dupes = sorted(i for i in set(comp_ids) if comp_ids.count(i) > 1)
    key_dupes = sorted(i for i in set(key_ids) if key_ids.count(i) > 1)
    if comp_dupes:
        _abort(f"completed workbook has duplicate row_ids: {comp_dupes}")
    if key_dupes:
        _abort(f"key workbook has duplicate row_ids: {key_dupes}")
    if set(comp_ids) != expected_set:
        missing = sorted(expected_set - set(comp_ids))
        extra = sorted(set(comp_ids) - expected_set)
        _abort(f"completed row_id set != {{1..{EXPECTED_N}}} "
               f"(missing={missing}, extra={extra})")
    if set(key_ids) != expected_set:
        missing = sorted(expected_set - set(key_ids))
        extra = sorted(set(key_ids) - expected_set)
        _abort(f"key row_id set != {{1..{EXPECTED_N}}} "
               f"(missing={missing}, extra={extra})")
    pass_lines.append(
        f"PASS  row counts: completed={len(completed_rows)}, "
        f"key={len(key_rows)}; row_id sets equal and == {{1..{EXPECTED_N}}}; "
        f"strict 1:1"
    )

    # --- 2. Every adjudication non-null & in PI vocab --------------------
    bad_adj: list[tuple[int, str]] = []
    for r in completed_rows:
        rid = int(r["row_id"])
        norm = _norm(r["adjudication"])
        if norm == "" or norm not in PI_VERDICTS:
            bad_adj.append((rid, repr(r["adjudication"])))
    if bad_adj:
        detail = ", ".join(f"row_id={rid}:{val}" for rid, val in bad_adj[:25])
        _abort(f"{len(bad_adj)} adjudication(s) null or outside PI vocab "
               f"{PI_VERDICTS}: {detail}")
    pass_lines.append(
        f"PASS  all {EXPECTED_N} adjudications non-null and in PI vocab"
    )

    # --- 3. Blinding-absence audit ---------------------------------------
    present = [h for h in BLINDING_FORBIDDEN_HEADERS if h in completed_header]
    if present:
        _abort(f"BLINDING BREACH — completed workbook Adjudication header "
               f"contains {present}; the PI may have seen unblinded data")
    pass_lines.append(
        "PASS  blinding-absence audit: arm_name and gemma_verdict absent "
        "from completed workbook header"
    )

    # --- 4. Pairing check (catastrophic-misjoin guard) -------------------
    comp_by_id = {int(r["row_id"]): r for r in completed_rows}
    key_by_id = {int(r["row_id"]): r for r in key_rows}
    pairing_fail: list[str] = []
    for rid in sorted(expected_set):
        c = comp_by_id[rid]
        k = key_by_id[rid]
        for col in ("field_name", "arm_value", "source_window_strategy"):
            if not _eq_field(c.get(col), k.get(col)):
                pairing_fail.append(
                    f"row_id={rid} {col}: completed={c.get(col)!r} "
                    f"vs key={k.get(col)!r}"
                )
    if pairing_fail:
        detail = "\n  ".join(pairing_fail[:25])
        _abort(f"{len(pairing_fail)} pairing mismatch(es) — wrong key or "
               f"reordered rows:\n  {detail}")
    pass_lines.append(
        "PASS  pairing check: field_name, arm_value, source_window_strategy "
        f"match across all {EXPECTED_N} row_ids"
    )

    # --- Build joined rows -----------------------------------------------
    joined: list[JoinedRow] = []
    for rid in sorted(expected_set):
        c = comp_by_id[rid]
        k = key_by_id[rid]
        gemma = _norm(k.get("gemma_verdict"))
        if gemma not in GEMMA_VERDICTS:
            _abort(f"row_id={rid} key gemma_verdict {k.get('gemma_verdict')!r} "
                   f"outside Gemma vocab {GEMMA_VERDICTS}")
        joined.append(JoinedRow(
            row_id=rid,
            pi=_norm(c.get("adjudication")),
            gemma=gemma,
            arm_name=str(k.get("arm_name")),
            field_name=str(k.get("field_name")),
            source_window_strategy=str(k.get("source_window_strategy")),
            pi_notes=c.get("notes"),
            pi_adjudicated_at=c.get("adjudicated_at"),
            key_cols=dict(k),
        ))

    for line in pass_lines:
        print(line)
    return joined


# ═════════════════════════════════════════════════════════════════════
# Statistics (pure Python)
# ═════════════════════════════════════════════════════════════════════


def wilson(k: int, n: int, z: float = WILSON_Z) -> dict:
    """Wilson score interval. Returns {n, k, p, lo, hi}; lo/hi None if n==0."""
    if n == 0:
        return {"n": 0, "k": k, "p": None, "lo": None, "hi": None}
    p = k / n
    denom = 1.0 + z * z / n
    centre = (p + z * z / (2 * n)) / denom
    half = (z * math.sqrt((p * (1 - p) + z * z / (4 * n)) / n)) / denom
    return {"n": n, "k": k, "p": p,
            "lo": max(0.0, centre - half), "hi": min(1.0, centre + half)}


def _linear_weight(i: int, j: int) -> float:
    """Linear agreement weight w_ij = 1 - |i-j|/2 over categories {0,1,2}."""
    return 1.0 - abs(i - j) / 2.0


def weighted_kappa(pairs: list[tuple[int, int]]) -> Optional[float]:
    """Weighted Cohen's kappa over ordinal categories {0,1,2}, linear
    agreement weights. pairs = [(pi_ord, gemma_ord), ...].

    Returns None if undefined (n==0 or expected weighted agreement == 1)."""
    n = len(pairs)
    if n == 0:
        return None
    cats = (0, 1, 2)
    obs: dict[tuple[int, int], int] = {}
    rowm = {c: 0 for c in cats}
    colm = {c: 0 for c in cats}
    for a, b in pairs:
        obs[(a, b)] = obs.get((a, b), 0) + 1
        rowm[a] += 1
        colm[b] += 1
    po = sum(_linear_weight(i, j) * obs.get((i, j), 0)
             for i in cats for j in cats) / n
    pe = sum(_linear_weight(i, j) * rowm[i] * colm[j]
             for i in cats for j in cats) / (n * n)
    if abs(1.0 - pe) < 1e-12:
        return None
    return (po - pe) / (1.0 - pe)


def _percentile(sorted_vals: list[float], q: float) -> Optional[float]:
    if not sorted_vals:
        return None
    if len(sorted_vals) == 1:
        return sorted_vals[0]
    pos = (q / 100.0) * (len(sorted_vals) - 1)
    lo_i = math.floor(pos)
    hi_i = math.ceil(pos)
    if lo_i == hi_i:
        return sorted_vals[int(lo_i)]
    frac = pos - lo_i
    return sorted_vals[lo_i] * (1 - frac) + sorted_vals[hi_i] * frac


def weighted_kappa_with_ci(pairs: list[tuple[int, int]]) -> dict:
    """Point estimate + seeded bootstrap 95% CI."""
    point = weighted_kappa(pairs)
    n = len(pairs)
    result = {"n": n, "kappa_w": point,
              "ci_lo": None, "ci_hi": None,
              "bootstrap_resamples": BOOTSTRAP_RESAMPLES,
              "bootstrap_valid": 0, "bootstrap_seed": BOOTSTRAP_SEED}
    if point is None or n < 2:
        return result
    rng = random.Random(BOOTSTRAP_SEED)
    ks: list[float] = []
    for _ in range(BOOTSTRAP_RESAMPLES):
        sample = [pairs[rng.randrange(n)] for _ in range(n)]
        k = weighted_kappa(sample)
        if k is not None:
            ks.append(k)
    ks.sort()
    result["bootstrap_valid"] = len(ks)
    result["ci_lo"] = _percentile(ks, 2.5)
    result["ci_hi"] = _percentile(ks, 97.5)
    return result


# ═════════════════════════════════════════════════════════════════════
# Metric layers
# ═════════════════════════════════════════════════════════════════════


def layer_a_confusion(rows: list[JoinedRow]) -> dict:
    """4×3 counts: PI (rows, incl. UNCLEAR) × Gemma (cols)."""
    matrix = {pi: {g: 0 for g in GEMMA_VERDICTS} for pi in PI_VERDICTS}
    for r in rows:
        matrix[r.pi][r.gemma] += 1
    return matrix


def layer_b_per_class_precision(rows: list[JoinedRow]) -> dict:
    """For each Gemma class C: conditional (substantive den) + strict."""
    out: dict[str, dict] = {}
    for C in GEMMA_VERDICTS:
        in_class = [r for r in rows if r.gemma == C]
        num = sum(1 for r in in_class if r.pi == C)
        den_cond = sum(1 for r in in_class if r.pi in SUBSTANTIVE_SET)
        den_strict = len(in_class)
        out[C] = {
            "conditional": wilson(num, den_cond),
            "strict": wilson(num, den_strict),
            "unclear_in_class": den_strict - den_cond,
        }
    return out


def layer_c_fabrication_ppv(rows: list[JoinedRow]) -> dict:
    """Fabrication flag = Gemma==UNSUPPORTED."""
    flagged = [r for r in rows if r.gemma == "UNSUPPORTED"]
    den_cond = sum(1 for r in flagged if r.pi in SUBSTANTIVE_SET)
    den_strict = len(flagged)
    strict_num = sum(1 for r in flagged if r.pi == "UNSUPPORTED")
    lenient_num = sum(
        1 for r in flagged if r.pi in ("UNSUPPORTED", "PARTIALLY_SUPPORTED")
    )
    return {
        "flag_definition": "Gemma == UNSUPPORTED",
        "strict_conditional": wilson(strict_num, den_cond),
        "strict_sensitivity": wilson(strict_num, den_strict),
        "lenient_conditional": wilson(lenient_num, den_cond),
        "lenient_sensitivity": wilson(lenient_num, den_strict),
        "unclear_in_flag": den_strict - den_cond,
    }


def layer_d_weighted_kappa(rows: list[JoinedRow]) -> dict:
    """Weighted kappa on substantive rows (PI substantive; Gemma always is)."""
    pairs = [
        (ORDINAL[r.pi], ORDINAL[r.gemma])
        for r in rows if r.pi in SUBSTANTIVE_SET
    ]
    return weighted_kappa_with_ci(pairs)


def layer_e_per_arm(rows: list[JoinedRow]) -> dict:
    arms = sorted({r.arm_name for r in rows})
    out: dict[str, dict] = {}
    for arm in arms:
        arm_rows = [r for r in rows if r.arm_name == arm]
        subst = [r for r in arm_rows if r.pi in SUBSTANTIVE_SET]
        agree = sum(1 for r in subst if r.pi == r.gemma)
        dist = {pi: sum(1 for r in arm_rows if r.pi == pi) for pi in PI_VERDICTS}
        out[arm] = {
            "n_total": len(arm_rows),
            "exact_agreement_substantive": wilson(agree, len(subst)),
            "pi_verdict_distribution": dist,
        }
    return out


def layer_f_window_strategy(rows: list[JoinedRow]) -> dict:
    strategies = sorted({r.source_window_strategy for r in rows})
    by_strategy: dict[str, dict] = {}
    for s in strategies:
        s_rows = [r for r in rows if r.source_window_strategy == s]
        unclear = sum(1 for r in s_rows if r.pi == "UNCLEAR")
        subst = [r for r in s_rows if r.pi in SUBSTANTIVE_SET]
        agree = sum(1 for r in subst if r.pi == r.gemma)
        by_strategy[s] = {
            "n_total": len(s_rows),
            "n_unclear": unclear,
            "unclear_rate": (unclear / len(s_rows)) if s_rows else None,
            "protocol_driven_unclear": s in PROTOCOL_UNCLEAR_STRATEGIES,
            "substantive_agreement": wilson(agree, len(subst)),
        }

    # strategy × arm cross-tab.
    arms = sorted({r.arm_name for r in rows})
    cross = {s: {a: 0 for a in arms} for s in strategies}
    for r in rows:
        cross[r.source_window_strategy][r.arm_name] += 1

    # Clean-subset re-estimates (repeat Layers B–D).
    clean = [r for r in rows if r.source_window_strategy == CLEAN_STRATEGY]
    clean_plus = [r for r in rows
                  if r.source_window_strategy in CLEAN_PLUS_STRATEGIES]

    def _subset_block(subset: list[JoinedRow]) -> dict:
        return {
            "n": len(subset),
            "per_class_precision": layer_b_per_class_precision(subset),
            "fabrication_ppv": layer_c_fabrication_ppv(subset),
            "weighted_kappa": layer_d_weighted_kappa(subset),
        }

    return {
        "by_strategy": by_strategy,
        "strategy_by_arm": cross,
        "clean_subset": {
            "definition": f"source_window_strategy == '{CLEAN_STRATEGY}'",
            **_subset_block(clean),
        },
        "clean_plus_subset": {
            "definition": "source_window_strategy in "
                          f"{sorted(CLEAN_PLUS_STRATEGIES)}",
            **_subset_block(clean_plus),
        },
    }


def compute_all_layers(rows: list[JoinedRow]) -> dict:
    return {
        "confusion_matrix": layer_a_confusion(rows),
        "per_class_precision": layer_b_per_class_precision(rows),
        "fabrication_ppv": layer_c_fabrication_ppv(rows),
        "weighted_kappa": layer_d_weighted_kappa(rows),
        "per_arm": layer_e_per_arm(rows),
        "window_strategy": layer_f_window_strategy(rows),
    }


# ═════════════════════════════════════════════════════════════════════
# Formatting helpers
# ═════════════════════════════════════════════════════════════════════


def _fmt_pct(x: Optional[float]) -> str:
    return "—" if x is None else f"{100 * x:.1f}%"


def _fmt_ci(w: dict) -> str:
    if w["p"] is None:
        return f"n=0"
    return (f"{_fmt_pct(w['p'])} [{_fmt_pct(w['lo'])}, {_fmt_pct(w['hi'])}] "
            f"({w['k']}/{w['n']})")


def _fmt_kappa(d: dict) -> str:
    if d["kappa_w"] is None:
        return f"undefined (n={d['n']})"
    return (f"{d['kappa_w']:.3f} [{d['ci_lo']:.3f}, {d['ci_hi']:.3f}] "
            f"(n={d['n']})")


# ═════════════════════════════════════════════════════════════════════
# Branded workbook writer
# ═════════════════════════════════════════════════════════════════════


def _sanitize(value):
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def _header_font() -> Font:
    return Font(bold=True, color=BRAND_WHITE, name="Calibri", size=11)


def _header_fill() -> PatternFill:
    return PatternFill(start_color=BRAND_TEAL, end_color=BRAND_TEAL,
                       fill_type="solid")


def _border() -> Border:
    side = Side(style="thin", color=BRAND_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _wrap(vertical: str = "top") -> Alignment:
    return Alignment(horizontal="left", vertical=vertical, wrap_text=True)


def _write_header(ws, headers: list[str], row: int = 1) -> None:
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = _header_font()
        c.fill = _header_fill()
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
        c.border = _border()
    ws.row_dimensions[row].height = 24


def _write_table(ws, headers: list[str], rows: list[list],
                 widths: Optional[list[float]] = None,
                 start_row: int = 1) -> int:
    _write_header(ws, headers, row=start_row)
    mist = PatternFill(start_color=BRAND_MIST, end_color=BRAND_MIST,
                       fill_type="solid")
    mist_dark = PatternFill(start_color=BRAND_MIST_DARK,
                            end_color=BRAND_MIST_DARK, fill_type="solid")
    body = Font(name="Calibri", size=11, color=BRAND_CHARCOAL)
    bd = _border()
    r = start_row
    for ri, values in enumerate(rows):
        r = start_row + 1 + ri
        fill = mist if r % 2 == 0 else mist_dark
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=r, column=ci, value=_sanitize(val))
            c.font = body
            c.alignment = _wrap()
            c.fill = fill
            c.border = bd
    if widths:
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return r


def _write_summary_sheet(ws, layers: dict, meta: dict) -> None:
    ws.title = "Summary"
    teal_h = Font(name="Calibri", size=14, bold=True, color=BRAND_TEAL)
    sub_h = Font(name="Calibri", size=12, bold=True, color=BRAND_TEAL)
    body = Font(name="Calibri", size=11, color=BRAND_CHARCOAL)
    terr = Font(name="Calibri", size=11, bold=True, color=BRAND_TERRACOTTA)

    r = 1

    def line(text, fnt=body, col=1):
        nonlocal r
        c = ws.cell(row=r, column=col, value=text)
        c.font = fnt
        c.alignment = _wrap()
        r += 1

    def kv(label, value, vfont=body):
        nonlocal r
        ws.cell(row=r, column=1, value=label).font = Font(
            bold=True, color=BRAND_CHARCOAL, name="Calibri", size=11)
        c = ws.cell(row=r, column=2, value=value)
        c.font = vfont
        c.alignment = _wrap()
        r += 1

    line("PI Audit — Judge Reliability Results", teal_h)
    r += 1

    # Top caveats block.
    line("Caveats (read first)", sub_h)
    for cav in CAVEATS:
        line("•  " + cav, body)
    r += 1

    line("Headline figures", sub_h)
    fab = layers["fabrication_ppv"]
    kv("Fabrication-flag PPV — strict (headline)",
       _fmt_ci(fab["strict_conditional"]), terr)
    kv("Fabrication-flag PPV — lenient (UNSUP∪PARTIAL)",
       _fmt_ci(fab["lenient_conditional"]))
    kv("Fabrication-flag PPV — strict, UNCLEAR-incl. denom (sensitivity)",
       _fmt_ci(fab["strict_sensitivity"]))
    kv("Weighted Cohen's κ (linear, substantive rows)",
       _fmt_kappa(layers["weighted_kappa"]))
    r += 1

    line("Per-Gemma-class precision (conditional)", sub_h)
    for C in GEMMA_VERDICTS:
        blk = layers["per_class_precision"][C]
        kv(f"  {C}", _fmt_ci(blk["conditional"])
           + f"  (UNCLEAR in class: {blk['unclear_in_class']})")
    r += 1

    line("Clean-subset re-estimate (arm_span_window only)", sub_h)
    cs = layers["window_strategy"]["clean_subset"]
    kv("  n", cs["n"])
    kv("  Fabrication-flag PPV — strict",
       _fmt_ci(cs["fabrication_ppv"]["strict_conditional"]))
    kv("  Weighted κ", _fmt_kappa(cs["weighted_kappa"]))
    line("  Discriminator: if agreement stays low on this span-centered, "
         "low-UNCLEAR subset, the disagreement is judge harshness; if it "
         "climbs, it was window-recovery failure.", body)
    r += 1

    line("Provenance", sub_h)
    kv("completed workbook", meta["completed_filename"])
    kv("completed SHA-256", meta["completed_sha256"])
    kv("key workbook", meta["key_filename"])
    kv("key SHA-256", meta["key_sha256"])
    kv("run timestamp (UTC)", meta["run_timestamp_utc"])
    kv("schema_version", meta["schema_version"])

    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 90


def _write_confusion_sheet(ws, layers: dict) -> None:
    ws.title = "ConfusionMatrix"
    matrix = layers["confusion_matrix"]
    headers = ["PI \\ Gemma"] + list(GEMMA_VERDICTS) + ["row_total"]
    rows = []
    col_totals = {g: 0 for g in GEMMA_VERDICTS}
    for pi in PI_VERDICTS:
        row_total = sum(matrix[pi][g] for g in GEMMA_VERDICTS)
        for g in GEMMA_VERDICTS:
            col_totals[g] += matrix[pi][g]
        rows.append([pi] + [matrix[pi][g] for g in GEMMA_VERDICTS] + [row_total])
    rows.append(["col_total"] + [col_totals[g] for g in GEMMA_VERDICTS]
                + [sum(col_totals.values())])
    _write_table(ws, headers, rows,
                 widths=[22, 20, 20, 20, 12])


def _write_per_class_sheet(ws, layers: dict) -> None:
    ws.title = "PerClassPrecision"
    headers = ["gemma_class", "metric", "k", "n", "p", "ci_lo", "ci_hi",
               "unclear_in_class"]
    rows = []
    for C in GEMMA_VERDICTS:
        blk = layers["per_class_precision"][C]
        for label, key in (("conditional", "conditional"),
                           ("strict (UNCLEAR-incl.)", "strict")):
            w = blk[key]
            rows.append([
                C, label, w["k"], w["n"],
                _fmt_pct(w["p"]), _fmt_pct(w["lo"]), _fmt_pct(w["hi"]),
                blk["unclear_in_class"] if key == "conditional" else "",
            ])
    _write_table(ws, headers, rows,
                 widths=[22, 22, 8, 8, 12, 12, 12, 16])


def _write_fabrication_sheet(ws, layers: dict) -> None:
    ws.title = "FabricationFlag"
    fab = layers["fabrication_ppv"]
    headers = ["metric", "k", "n", "p", "ci_lo", "ci_hi"]
    rows = []
    for label, key in (
        ("strict conditional (headline)", "strict_conditional"),
        ("strict sensitivity (UNCLEAR-incl.)", "strict_sensitivity"),
        ("lenient conditional", "lenient_conditional"),
        ("lenient sensitivity (UNCLEAR-incl.)", "lenient_sensitivity"),
    ):
        w = fab[key]
        rows.append([label, w["k"], w["n"],
                     _fmt_pct(w["p"]), _fmt_pct(w["lo"]), _fmt_pct(w["hi"])])
    r = _write_table(ws, headers, rows, widths=[36, 8, 8, 12, 12, 12])
    note = ws.cell(row=r + 2, column=1,
                   value=f"Flag = {fab['flag_definition']}. "
                         f"UNCLEAR in flag: {fab['unclear_in_flag']}. "
                         "Conditional denominator = flagged rows with a "
                         "substantive PI verdict; strict denominator includes "
                         "PI==UNCLEAR.")
    note.font = Font(name="Calibri", size=10, italic=True,
                     color=BRAND_CHARCOAL)
    note.alignment = _wrap()


def _write_per_arm_sheet(ws, layers: dict) -> None:
    ws.title = "PerArm"
    per_arm = layers["per_arm"]
    headers = (["arm_name", "n_total", "substantive_agreement_%",
                "ci_lo", "ci_hi", "agree_k", "agree_n"]
               + [f"PI={pi}" for pi in PI_VERDICTS])
    rows = []
    for arm in sorted(per_arm):
        blk = per_arm[arm]
        w = blk["exact_agreement_substantive"]
        dist = blk["pi_verdict_distribution"]
        rows.append([
            arm, blk["n_total"], _fmt_pct(w["p"]),
            _fmt_pct(w["lo"]), _fmt_pct(w["hi"]), w["k"], w["n"],
        ] + [dist[pi] for pi in PI_VERDICTS])
    r = _write_table(ws, headers, rows,
                     widths=[24, 8, 18, 10, 10, 8, 8, 14, 20, 14, 12])
    banner = ws.cell(
        row=r + 2, column=1,
        value="CAVEAT: per-arm agreement is conditional on the "
              "Gemma-stratified sample and is NOT population prevalence. "
              "Reweighting is deferred to Paper 1. Agreement is exact "
              "Gemma-vs-PI match on substantive rows.")
    banner.font = Font(name="Calibri", size=10, italic=True,
                       color=BRAND_TERRACOTTA)
    banner.alignment = _wrap()


def _write_window_strategy_sheet(ws, layers: dict) -> None:
    ws.title = "WindowStrategy"
    wf = layers["window_strategy"]
    r = 1
    sub_h = Font(name="Calibri", size=12, bold=True, color=BRAND_TEAL)
    ws.cell(row=r, column=1,
            value="UNCLEAR rate & substantive agreement by strategy").font = sub_h
    r += 1
    headers = ["source_window_strategy", "n_total", "n_unclear",
               "unclear_rate", "protocol_driven_unclear",
               "substantive_agreement_%", "agree_ci_lo", "agree_ci_hi",
               "agree_k", "agree_n"]
    rows = []
    for s in sorted(wf["by_strategy"]):
        b = wf["by_strategy"][s]
        w = b["substantive_agreement"]
        rows.append([
            s, b["n_total"], b["n_unclear"], _fmt_pct(b["unclear_rate"]),
            "YES" if b["protocol_driven_unclear"] else "no",
            _fmt_pct(w["p"]), _fmt_pct(w["lo"]), _fmt_pct(w["hi"]),
            w["k"], w["n"],
        ])
    r = _write_table(ws, headers, rows,
                     widths=[28, 8, 10, 12, 18, 18, 12, 12, 8, 8],
                     start_row=r)

    note = ws.cell(
        row=r + 2, column=1,
        value="absence_fallback_head and missing_span_fallback_head carry "
              "protocol-driven UNCLEAR — the Instructions told the PI to "
              "prefer UNCLEAR under degraded context.")
    note.font = Font(name="Calibri", size=10, italic=True,
                     color=BRAND_TERRACOTTA)
    note.alignment = _wrap()
    r += 4

    # strategy × arm cross-tab.
    ws.cell(row=r, column=1,
            value="strategy × arm cross-tab (counts)").font = sub_h
    r += 1
    cross = wf["strategy_by_arm"]
    arms = sorted({a for s in cross for a in cross[s]})
    headers2 = ["source_window_strategy"] + arms + ["total"]
    rows2 = []
    for s in sorted(cross):
        rc = [cross[s].get(a, 0) for a in arms]
        rows2.append([s] + rc + [sum(rc)])
    r = _write_table(ws, headers2, rows2,
                     widths=[28] + [18] * len(arms) + [10], start_row=r)
    r += 3

    # Clean-subset blocks.
    for key, title in (("clean_subset", "Clean subset (arm_span_window only)"),
                       ("clean_plus_subset",
                        "Secondary subset (arm_span_window ∪ full_text)")):
        cs = wf[key]
        ws.cell(row=r, column=1, value=title).font = sub_h
        r += 1
        ws.cell(row=r, column=1, value=cs["definition"]).font = Font(
            name="Calibri", size=10, italic=True, color=BRAND_CHARCOAL)
        r += 1
        h3 = ["metric", "value"]
        fab = cs["fabrication_ppv"]
        body_rows = [
            ["n", cs["n"]],
            ["Fabrication PPV — strict conditional",
             _fmt_ci(fab["strict_conditional"])],
            ["Fabrication PPV — lenient conditional",
             _fmt_ci(fab["lenient_conditional"])],
            ["Weighted κ", _fmt_kappa(cs["weighted_kappa"])],
        ]
        for C in GEMMA_VERDICTS:
            body_rows.append(
                [f"Per-class precision — {C} (conditional)",
                 _fmt_ci(cs["per_class_precision"][C]["conditional"])])
        r = _write_table(ws, h3, body_rows, widths=[44, 60], start_row=r)
        r += 3


def _write_joined_sheet(ws, joined: list[JoinedRow],
                        key_header: list[str]) -> None:
    ws.title = "JoinedRows"
    # All key columns, then PI fields.
    pi_cols = ["pi_adjudication", "pi_notes", "pi_adjudicated_at"]
    headers = list(key_header) + pi_cols
    rows = []
    for jr in joined:
        base = [jr.key_cols.get(h) for h in key_header]
        rows.append(base + [jr.pi, jr.pi_notes, jr.pi_adjudicated_at])
    widths = [14] * len(key_header) + [18, 40, 20]
    _write_table(ws, headers, rows, widths=widths)
    ws.freeze_panes = "A2"


def write_results_workbook(path: Path, layers: dict, meta: dict,
                           joined: list[JoinedRow],
                           key_header: list[str]) -> None:
    wb = Workbook()
    _write_summary_sheet(wb.active, layers, meta)
    _write_confusion_sheet(wb.create_sheet("ConfusionMatrix"), layers)
    _write_per_class_sheet(wb.create_sheet("PerClassPrecision"), layers)
    _write_fabrication_sheet(wb.create_sheet("FabricationFlag"), layers)
    _write_per_arm_sheet(wb.create_sheet("PerArm"), layers)
    _write_window_strategy_sheet(wb.create_sheet("WindowStrategy"), layers)
    _write_joined_sheet(wb.create_sheet("JoinedRows"), joined, key_header)
    wb.save(path)


# ═════════════════════════════════════════════════════════════════════
# JSON sidecar
# ═════════════════════════════════════════════════════════════════════


def build_json_sidecar(layers: dict, meta: dict,
                       joined: list[JoinedRow]) -> dict:
    return {
        "schema_version": SCHEMA_VERSION,
        "run_timestamp_utc": meta["run_timestamp_utc"],
        "inputs": {
            "completed": {
                "filename": meta["completed_filename"],
                "sha256": meta["completed_sha256"],
            },
            "key": {
                "filename": meta["key_filename"],
                "sha256": meta["key_sha256"],
            },
        },
        "vocabularies": {
            "pi_verdicts": list(PI_VERDICTS),
            "gemma_verdicts": list(GEMMA_VERDICTS),
            "substantive": list(SUBSTANTIVE),
            "ordinal_map": ORDINAL,
        },
        "n_rows": len(joined),
        "metrics": layers,
        "caveats": CAVEATS,
    }


# ═════════════════════════════════════════════════════════════════════
# Stdout report
# ═════════════════════════════════════════════════════════════════════


def print_headline(layers: dict, meta: dict, joined: list[JoinedRow]) -> None:
    fab = layers["fabrication_ppv"]
    print("\n── Headline numbers ─────────────────────────────────────────")
    print(f"  n rows                              : {len(joined)}")
    print(f"  Fabrication-flag PPV (strict, headline): "
          f"{_fmt_ci(fab['strict_conditional'])}")
    print(f"  Fabrication-flag PPV (lenient)      : "
          f"{_fmt_ci(fab['lenient_conditional'])}")
    print(f"  Weighted Cohen's κ (substantive)    : "
          f"{_fmt_kappa(layers['weighted_kappa'])}")
    print("  Per-Gemma-class precision (conditional):")
    for C in GEMMA_VERDICTS:
        blk = layers["per_class_precision"][C]
        print(f"      {C:<22}: {_fmt_ci(blk['conditional'])}"
              f"  (UNCLEAR in class: {blk['unclear_in_class']})")
    cs = layers["window_strategy"]["clean_subset"]
    print(f"  Clean subset (arm_span_window, n={cs['n']}):")
    print(f"      Fab PPV strict : "
          f"{_fmt_ci(cs['fabrication_ppv']['strict_conditional'])}")
    print(f"      Weighted κ     : {_fmt_kappa(cs['weighted_kappa'])}")
    print("─────────────────────────────────────────────────────────────")


# ═════════════════════════════════════════════════════════════════════
# Driver / CLI
# ═════════════════════════════════════════════════════════════════════


_TS_RE = re.compile(r"pi_audit_key_(?P<ts>.+)\.xlsx$")


def _parse_timestamp_token(key_path: Path) -> str:
    m = _TS_RE.search(key_path.name)
    if m:
        return m.group("ts")
    # Fallback: strip extension.
    return key_path.stem


def _discover(audit_dir: Path) -> tuple[Optional[Path], Optional[Path]]:
    completed = sorted(audit_dir.glob("*COMPLETED*.xlsx"))
    keys = sorted(audit_dir.glob("pi_audit_key_*.xlsx"))
    return (completed[-1] if completed else None,
            keys[-1] if keys else None)


def run(argv: Optional[list[str]] = None) -> int:
    p = argparse.ArgumentParser(prog="analysis.paper1.pi_audit_unblind")
    p.add_argument("--completed", type=Path, default=None,
                   help="Path to the PI-adjudicated *_COMPLETED.xlsx")
    p.add_argument("--key", type=Path, default=None,
                   help="Path to pi_audit_key_<ts>.xlsx")
    p.add_argument("--audit-dir", type=Path,
                   default=Path("artifacts/paper1/pi_audit"),
                   help="Directory to auto-discover inputs if --completed / "
                        "--key are omitted.")
    p.add_argument("--out-dir", type=Path, default=None,
                   help="Output directory (default: same dir as --key).")
    args = p.parse_args(argv)

    completed = args.completed
    key = args.key
    if completed is None or key is None:
        d_completed, d_key = _discover(args.audit_dir)
        completed = completed or d_completed
        key = key or d_key

    if completed is None:
        raise SystemExit(
            f"ABORT: no completed workbook found (looked in {args.audit_dir} "
            "for *COMPLETED*.xlsx). The PI-adjudicated file must be copied "
            "from the MacBook (Dropbox/scp) before this can run.")
    if key is None:
        raise SystemExit(
            f"ABORT: no key workbook found (looked in {args.audit_dir} for "
            "pi_audit_key_*.xlsx).")

    completed = completed.resolve()
    key = key.resolve()
    if not completed.exists():
        raise SystemExit(f"ABORT: completed workbook not found: {completed}")
    if not key.exists():
        raise SystemExit(f"ABORT: key workbook not found: {key}")

    out_dir = (args.out_dir or key.parent).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    print("Inputs located:")
    print(f"  completed : {completed}")
    print(f"  key       : {key}")
    print(f"  out-dir   : {out_dir}\n")

    # Load.
    comp_header, comp_rows = _read_sheet(completed, "Adjudication")
    key_header, key_rows = _read_sheet(key, "Key")

    # Integrity gate.
    joined = run_integrity_gate(comp_header, comp_rows, key_header, key_rows)

    # Provenance.
    run_ts = datetime.now(timezone.utc).isoformat()
    ts_token = _parse_timestamp_token(key)
    meta = {
        "schema_version": SCHEMA_VERSION,
        "run_timestamp_utc": run_ts,
        "completed_filename": completed.name,
        "completed_sha256": _sha256(completed),
        "key_filename": key.name,
        "key_sha256": _sha256(key),
        "timestamp_token": ts_token,
    }

    # Compute.
    layers = compute_all_layers(joined)

    # Outputs.
    xlsx_path = out_dir / f"pi_audit_results_{ts_token}.xlsx"
    json_path = out_dir / f"pi_audit_results_{ts_token}.json"
    write_results_workbook(xlsx_path, layers, meta, joined, key_header)
    sidecar = build_json_sidecar(layers, meta, joined)
    with open(json_path, "w") as fh:
        json.dump(sidecar, fh, indent=2)

    print_headline(layers, meta, joined)
    print(f"\nResults workbook : {xlsx_path}")
    print(f"JSON sidecar     : {json_path}")
    return 0


if __name__ == "__main__":
    sys.exit(run())
