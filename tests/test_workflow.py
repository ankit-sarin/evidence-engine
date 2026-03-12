"""Tests for the screening adjudication workflow enforcement."""

import json
from pathlib import Path

import pytest

from engine.adjudication.workflow import (
    WORKFLOW_STAGES,
    advance_stage,
    bypass_stage,
    can_advance_to,
    complete_stage,
    ensure_workflow_table,
    format_workflow_status,
    get_current_blocker,
    get_workflow_status,
    is_adjudication_complete,
    is_stage_done,
    reset_stage,
)
from engine.core.database import ReviewDatabase


@pytest.fixture
def db(tmp_path):
    """Create a ReviewDatabase in a temp directory."""
    d = ReviewDatabase("test_review", data_root=tmp_path)
    yield d
    d.close()


# ── Schema Tests ────────────────────────────────────────────────────


def test_workflow_table_created(db):
    row = db._conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='workflow_state'"
    ).fetchone()
    assert row is not None


def test_workflow_table_seeded(db):
    rows = db._conn.execute(
        "SELECT stage_name, status FROM workflow_state ORDER BY id"
    ).fetchall()
    assert len(rows) == 10
    assert [r["stage_name"] for r in rows] == list(WORKFLOW_STAGES)
    assert all(r["status"] == "pending" for r in rows)


def test_ensure_idempotent(db):
    ensure_workflow_table(db._conn)
    ensure_workflow_table(db._conn)
    rows = db._conn.execute("SELECT COUNT(*) FROM workflow_state").fetchone()
    assert rows[0] == 10


# ── Stage Completion Tests ──────────────────────────────────────────


def test_complete_stage(db):
    complete_stage(db._conn, "SCREENING_COMPLETE", metadata="251 papers screened")
    assert is_stage_done(db._conn, "SCREENING_COMPLETE")
    row = db._conn.execute(
        "SELECT status, metadata FROM workflow_state WHERE stage_name = 'SCREENING_COMPLETE'"
    ).fetchone()
    assert row["status"] == "complete"
    assert "251 papers" in row["metadata"]


def test_complete_unknown_stage_raises(db):
    with pytest.raises(ValueError, match="Unknown stage"):
        complete_stage(db._conn, "FAKE_STAGE")


def test_is_stage_done_false_initially(db):
    assert not is_stage_done(db._conn, "SCREENING_COMPLETE")


# ── Bypass Tests ────────────────────────────────────────────────────


def test_bypass_stage(db):
    bypass_stage(db._conn, "DIAGNOSTIC_SAMPLE_COMPLETE", metadata="skipping for now")
    assert is_stage_done(db._conn, "DIAGNOSTIC_SAMPLE_COMPLETE")
    row = db._conn.execute(
        "SELECT status, metadata FROM workflow_state WHERE stage_name = 'DIAGNOSTIC_SAMPLE_COMPLETE'"
    ).fetchone()
    assert row["status"] == "bypassed"
    assert "bypassed by operator" in row["metadata"]


# ── Reset Tests ─────────────────────────────────────────────────────


def test_reset_stage(db):
    complete_stage(db._conn, "SCREENING_COMPLETE")
    assert is_stage_done(db._conn, "SCREENING_COMPLETE")
    reset_stage(db._conn, "SCREENING_COMPLETE")
    assert not is_stage_done(db._conn, "SCREENING_COMPLETE")


# ── Prerequisite Checks ────────────────────────────────────────────


def test_can_advance_first_stage(db):
    assert can_advance_to(db._conn, "SCREENING_COMPLETE")


def test_cannot_advance_without_prereqs(db):
    assert not can_advance_to(db._conn, "DIAGNOSTIC_SAMPLE_COMPLETE")


def test_can_advance_after_prereqs(db):
    complete_stage(db._conn, "SCREENING_COMPLETE")
    assert can_advance_to(db._conn, "DIAGNOSTIC_SAMPLE_COMPLETE")


def test_cannot_skip_stages(db):
    complete_stage(db._conn, "SCREENING_COMPLETE")
    # Should not be able to advance to stage 3 without stage 2
    assert not can_advance_to(db._conn, "CATEGORIES_CONFIGURED")


# ── Blocker Detection ──────────────────────────────────────────────


def test_get_current_blocker_initial(db):
    blocker = get_current_blocker(db._conn)
    assert blocker is not None
    assert blocker["stage_name"] == "SCREENING_COMPLETE"
    assert blocker["index"] == 0


def test_get_current_blocker_after_first(db):
    complete_stage(db._conn, "SCREENING_COMPLETE")
    blocker = get_current_blocker(db._conn)
    assert blocker["stage_name"] == "DIAGNOSTIC_SAMPLE_COMPLETE"


def test_get_current_blocker_none_when_all_complete(db):
    for stage in WORKFLOW_STAGES:
        complete_stage(db._conn, stage)
    assert get_current_blocker(db._conn) is None


# ── Adjudication Complete Check ─────────────────────────────────────


def test_adjudication_complete_false_initially(db):
    assert not is_adjudication_complete(db._conn)


def test_adjudication_complete_after_all_stages(db):
    for stage in WORKFLOW_STAGES:
        complete_stage(db._conn, stage)
    assert is_adjudication_complete(db._conn)


# ── advance_stage (high-level) ──────────────────────────────────────


def test_advance_stage_success(db):
    result = advance_stage(db._conn, "SCREENING_COMPLETE", note="done")
    assert result["status"] == "complete"
    assert is_stage_done(db._conn, "SCREENING_COMPLETE")


def test_advance_stage_blocked(db):
    result = advance_stage(db._conn, "DIAGNOSTIC_SAMPLE_COMPLETE", note="trying")
    assert result["status"] == "blocked"
    assert "prerequisite" in result["message"]


def test_advance_stage_force(db):
    result = advance_stage(
        db._conn, "DIAGNOSTIC_SAMPLE_COMPLETE", note="skipping", force=True,
    )
    assert result["status"] == "bypassed"
    assert is_stage_done(db._conn, "DIAGNOSTIC_SAMPLE_COMPLETE")


def test_advance_stage_already_complete(db):
    complete_stage(db._conn, "SCREENING_COMPLETE")
    result = advance_stage(db._conn, "SCREENING_COMPLETE", note="again")
    assert result["status"] == "already_complete"


# ── Format Display ──────────────────────────────────────────────────


def test_format_workflow_status_all_pending(db):
    output = format_workflow_status(db._conn, review_name="test_review")
    assert "test_review" in output
    assert "[ ] SCREENING_COMPLETE" in output
    assert "[ ] ADJUDICATION_COMPLETE" in output


def test_format_workflow_status_partial(db):
    complete_stage(db._conn, "SCREENING_COMPLETE")
    output = format_workflow_status(db._conn, review_name="test")
    assert "[✓] SCREENING_COMPLETE" in output
    assert "[ ] DIAGNOSTIC_SAMPLE_COMPLETE" in output


def test_format_workflow_status_bypassed(db):
    bypass_stage(db._conn, "SCREENING_COMPLETE")
    output = format_workflow_status(db._conn, review_name="test")
    assert "[!] SCREENING_COMPLETE" in output
    assert "BYPASSED" in output


def test_format_workflow_status_all_complete(db):
    for stage in WORKFLOW_STAGES:
        complete_stage(db._conn, stage)
    output = format_workflow_status(db._conn)
    assert output.count("[✓]") == 10
    assert "[ ]" not in output


# ── get_workflow_status ─────────────────────────────────────────────


def test_get_workflow_status_returns_all_stages(db):
    statuses = get_workflow_status(db._conn)
    assert len(statuses) == 10
    assert statuses[0]["stage_name"] == "SCREENING_COMPLETE"
    assert statuses[4]["stage_name"] == "ADJUDICATION_COMPLETE"
    assert statuses[5]["stage_name"] == "PDF_ACQUISITION"
    assert statuses[9]["stage_name"] == "AUDIT_REVIEW_COMPLETE"


def test_get_workflow_status_tracks_metadata(db):
    complete_stage(db._conn, "SCREENING_COMPLETE", metadata="test note")
    statuses = get_workflow_status(db._conn)
    assert statuses[0]["metadata"] == "test note"
    assert statuses[0]["completed_at"] is not None


# ── Integration: auto-set from export/import ────────────────────────


def test_export_sets_queue_exported(db, tmp_path):
    """export_adjudication_queue should auto-set QUEUE_EXPORTED."""
    from engine.adjudication.screening_adjudicator import export_adjudication_queue
    from engine.search.models import Citation

    # Add a flagged paper
    db.add_papers([Citation(
        title="Test Paper", abstract="test", pmid="99999999",
        source="pubmed", authors=["A"], journal="J", year=2024,
    )])
    db.add_screening_decision(1, 1, "include", "maybe", "model")
    db.add_screening_decision(1, 2, "exclude", "maybe not", "model")
    db.update_status(1, "SCREEN_FLAGGED")

    out = tmp_path / "queue.xlsx"
    export_adjudication_queue(db, out)

    assert is_stage_done(db._conn, "QUEUE_EXPORTED")


def test_import_sets_adjudication_complete(db, tmp_path):
    """import_adjudication_decisions with zero unresolved should set ADJUDICATION_COMPLETE."""
    from engine.adjudication.screening_adjudicator import (
        export_adjudication_queue,
        import_adjudication_decisions,
    )
    from engine.search.models import Citation

    db.add_papers([Citation(
        title="Test Paper", abstract="test", pmid="88888888",
        source="pubmed", authors=["A"], journal="J", year=2024,
    )])
    db.add_screening_decision(1, 1, "include", "maybe", "model")
    db.add_screening_decision(1, 2, "exclude", "maybe not", "model")
    db.update_status(1, "SCREEN_FLAGGED")

    out = tmp_path / "queue.xlsx"
    export_adjudication_queue(db, out)

    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb["Review Queue"]
    ws.cell(row=2, column=15, value="EXCLUDE")
    wb.save(out)

    import_adjudication_decisions(db, out)

    assert is_stage_done(db._conn, "ADJUDICATION_COMPLETE")


def test_import_does_not_set_complete_with_missing(db, tmp_path):
    """import with missing decisions should NOT set ADJUDICATION_COMPLETE."""
    from engine.adjudication.screening_adjudicator import (
        export_adjudication_queue,
        import_adjudication_decisions,
    )
    from engine.search.models import Citation

    db.add_papers([Citation(
        title="Test Paper", abstract="test", pmid="77777777",
        source="pubmed", authors=["A"], journal="J", year=2024,
    )])
    db.add_screening_decision(1, 1, "include", "maybe", "model")
    db.add_screening_decision(1, 2, "exclude", "maybe not", "model")
    db.update_status(1, "SCREEN_FLAGGED")

    out = tmp_path / "queue.xlsx"
    export_adjudication_queue(db, out)

    # Don't fill in any decisions
    import_adjudication_decisions(db, out)

    assert not is_stage_done(db._conn, "ADJUDICATION_COMPLETE")
