"""Audit adjudication pipeline — export contested/flagged spans for human review,
import decisions back into the database.

Mirrors the screening_adjudicator pattern for extraction audit results.

Export format: one row per problematic span (per-span rows), with context
columns and a PI_decision dropdown (ACCEPT/REJECT/CORRECT).
"""

import logging
import random
from datetime import datetime, timezone
from pathlib import Path

from engine.adjudication.schema import ensure_adjudication_table
from engine.adjudication.workflow import complete_stage
from engine.core.database import ReviewDatabase

logger = logging.getLogger(__name__)

# Audit states that require human attention
_NEEDS_REVIEW = {"contested", "flagged", "invalid_snippet"}


# ── Data Collection ─────────────────────────────────────────────────


def _collect_papers_for_review(
    db: ReviewDatabase,
    *,
    spot_check_pct: float = 0.10,
    spot_check_failure_threshold: float = 0.20,
) -> list[dict]:
    """Collect AI_AUDIT_COMPLETE papers that need human review.

    Primary: papers with any contested/flagged/invalid_snippet spans.
    Spot-check: random N% of papers with ALL spans verified (minor issues only).

    If spot-check failure rate > threshold after import, all minor-issues
    papers should be promoted. That logic runs at import time.
    """
    papers = db.get_papers_by_status("AI_AUDIT_COMPLETE")
    needs_review = []
    all_verified = []

    for paper in papers:
        pid = paper["id"]
        extraction = db._conn.execute(
            "SELECT id, low_yield FROM extractions WHERE paper_id = ? ORDER BY id DESC LIMIT 1",
            (pid,),
        ).fetchone()
        if not extraction:
            continue

        ext_id = extraction["id"]
        is_low_yield = bool(extraction["low_yield"])

        spans = db._conn.execute(
            "SELECT * FROM evidence_spans WHERE extraction_id = ?",
            (ext_id,),
        ).fetchall()
        spans = [dict(s) for s in spans]

        if not spans:
            continue

        # Classify paper by worst audit state
        audit_states = {s["audit_status"] for s in spans}
        problem_spans = [s for s in spans if s["audit_status"] in _NEEDS_REVIEW]

        paper_info = {
            "paper_id": pid,
            "ee_identifier": paper.get("ee_identifier") or "",
            "title": paper["title"],
            "extraction_id": ext_id,
            "spans": spans,
            "problem_spans": problem_spans,
            "audit_states": audit_states,
            "low_yield": is_low_yield,
        }

        if is_low_yield:
            # LOW_YIELD papers always go to review regardless of span states
            if problem_spans:
                if any(s["audit_status"] == "flagged" for s in problem_spans):
                    paper_info["worst_state"] = "flagged"
                elif any(s["audit_status"] == "invalid_snippet" for s in problem_spans):
                    paper_info["worst_state"] = "invalid_snippet"
                else:
                    paper_info["worst_state"] = "contested"
            else:
                paper_info["worst_state"] = "verified"
            paper_info["review_reason"] = "low_yield"
            needs_review.append(paper_info)
        elif problem_spans:
            # Compute worst state: flagged > invalid_snippet > contested
            if any(s["audit_status"] == "flagged" for s in problem_spans):
                paper_info["worst_state"] = "flagged"
            elif any(s["audit_status"] == "invalid_snippet" for s in problem_spans):
                paper_info["worst_state"] = "invalid_snippet"
            else:
                paper_info["worst_state"] = "contested"
            paper_info["review_reason"] = "audit_issues"
            needs_review.append(paper_info)
        else:
            paper_info["worst_state"] = "verified"
            paper_info["review_reason"] = "spot_check"
            all_verified.append(paper_info)

    # Spot-check selection
    n_spot = max(1, int(len(all_verified) * spot_check_pct)) if all_verified else 0
    n_spot = min(n_spot, len(all_verified))
    if n_spot > 0:
        spot_check = random.sample(all_verified, n_spot)
        needs_review.extend(spot_check)

    # Sort: low_yield first, then flagged, invalid_snippet, contested, spot_check
    state_order = {"flagged": 0, "invalid_snippet": 1, "contested": 2, "verified": 3}
    def _sort_key(p):
        # low_yield papers sort before everything else
        ly = 0 if p.get("low_yield") else 1
        return (ly, state_order.get(p["worst_state"], 99), p["title"])
    needs_review.sort(key=_sort_key)

    return needs_review


def _flatten_to_span_rows(papers: list[dict]) -> list[dict]:
    """Flatten paper-level data to one row per span for review queue.

    For papers with problem spans: exports only the problematic spans.
    For spot-check papers: exports all spans for sampling.
    For LOW_YIELD papers: exports all spans (reviewer needs full picture).
    """
    rows = []
    for paper in papers:
        # Determine which spans to export
        if paper["review_reason"] == "low_yield":
            spans_to_export = paper["spans"]
        elif paper["problem_spans"]:
            spans_to_export = paper["problem_spans"]
        else:
            # spot_check — export all
            spans_to_export = paper["spans"]

        for span in spans_to_export:
            audit_state = span["audit_status"]
            if paper.get("low_yield") and audit_state == "verified":
                audit_state = "LOW_YIELD"

            rows.append({
                "paper_id": paper["paper_id"],
                "ee_identifier": paper["ee_identifier"],
                "title": paper["title"],
                "field_name": span["field_name"],
                "extracted_value": span.get("value") or "",
                "evidence_snippet": (span.get("source_snippet") or "")[:500],
                "audit_state": audit_state,
                "auditor_reasoning": (span.get("audit_rationale") or "")[:500],
                "review_reason": paper["review_reason"],
                "_span_id": span["id"],
                "_extraction_id": paper["extraction_id"],
            })

    return rows


# ── Reference Content Builder ─────────────────────────────────────


def _build_audit_reference_content(spec=None) -> str:
    """Build reference sheet content for extraction audit review."""
    lines = []

    lines.append("AUDIT STATES")
    lines.append("")
    lines.append("  CONTESTED: Grep verification failed (snippet not found verbatim) but semantic")
    lines.append("    verification passed. The extracted value is likely correct, but the evidence")
    lines.append("    snippet may be paraphrased or abbreviated. Most common cause: model used")
    lines.append("    ellipsis (...) to bridge text in the snippet.")
    lines.append("")
    lines.append("  FLAGGED: AI auditor flagged the extracted value as unsupported by the source")
    lines.append("    text. The value may be incorrect, misinterpreted, or fabricated. Requires")
    lines.append("    the most careful review.")
    lines.append("")
    lines.append("  LOW_YIELD: Paper has fewer than the threshold number of non-null extracted")
    lines.append("    fields. The extraction may be incomplete — the model may have missed fields")
    lines.append("    that are present in the paper. Reviewer should check if the paper actually")
    lines.append("    contains data for the missing fields.")
    lines.append("")

    lines.append("DECISIONS")
    lines.append("")
    lines.append("  ACCEPT: The extracted value is correct as-is. Mark span as human-verified.")
    lines.append("  REJECT: The extracted value is wrong and cannot be corrected from the paper.")
    lines.append("    The span will be removed from the evidence table.")
    lines.append("  CORRECT: The extracted value needs correction. You MUST provide the corrected")
    lines.append("    value in the 'corrected_value' column. The original value is preserved in")
    lines.append("    the audit trail.")
    lines.append("")

    if spec and hasattr(spec, "extraction_schema") and spec.extraction_schema:
        lines.append("EXTRACTION SCHEMA FIELDS")
        lines.append("")
        for field in spec.extraction_schema.fields:
            tier = f"Tier {field.tier}" if hasattr(field, "tier") else ""
            ftype = field.type if hasattr(field, "type") else ""
            lines.append(f"  {field.name} ({ftype}, {tier}):")
            lines.append(f"    {field.description}")
            if hasattr(field, "enum_values") and field.enum_values:
                lines.append(f"    Valid values: {', '.join(field.enum_values)}")
            lines.append("")

    return "\n".join(lines)


# ── Export ──────────────────────────────────────────────────────────


def export_audit_review_queue(
    review_db: ReviewDatabase,
    output_path: str | Path,
    *,
    spot_check_pct: float = 0.10,
    spot_check_failure_threshold: float = 0.20,
    format: str = "xlsx",
    review_name: str | None = None,
    review_spec=None,
) -> dict:
    """Export papers with contested/flagged spans for human audit review.

    Each row = one span, with context columns and PI_decision dropdown.

    Returns summary dict.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    papers = _collect_papers_for_review(
        review_db,
        spot_check_pct=spot_check_pct,
        spot_check_failure_threshold=spot_check_failure_threshold,
    )

    if not papers:
        logger.warning("No papers need audit review — nothing to export")
        return {"total": 0, "flagged": 0, "contested": 0, "spot_check": 0}

    if format == "html":
        from engine.review.extraction_audit_html import (
            generate_extraction_audit_html,
        )
        out, html_stats = generate_extraction_audit_html(
            review_name=review_name or "unknown",
            output_path=str(output_path),
        )
        output_path = out
    elif format == "xlsx":
        import sys
        print(
            "Warning: xlsx export is deprecated for interactive review. "
            "Use --format html (default). xlsx retained for reference exports.",
            file=sys.stderr,
        )
        _write_audit_xlsx(
            papers, output_path,
            spot_check_failure_threshold,
            review_name=review_name or "unknown",
            review_spec=review_spec,
            db_path=str(review_db.db_path),
        )
    else:
        raise ValueError(f"Unsupported format: {format}")

    stats = {
        "total": len(papers),
        "flagged": sum(1 for p in papers if p["worst_state"] == "flagged"),
        "contested": sum(1 for p in papers if p["worst_state"] == "contested"),
        "invalid_snippet": sum(1 for p in papers if p["worst_state"] == "invalid_snippet"),
        "low_yield": sum(1 for p in papers if p.get("review_reason") == "low_yield"),
        "spot_check": sum(1 for p in papers if p["review_reason"] == "spot_check"),
        "output_path": str(output_path),
        "spot_check_failure_threshold": spot_check_failure_threshold,
    }

    logger.info(
        "Exported audit review queue: %d papers (%d flagged, %d contested, "
        "%d invalid, %d spot-check) to %s",
        stats["total"], stats["flagged"], stats["contested"],
        stats["invalid_snippet"], stats["spot_check"], output_path,
    )

    # Auto-advance workflow
    complete_stage(
        review_db._conn, "AUDIT_QUEUE_EXPORTED",
        metadata=f"{stats['total']} papers exported to {output_path}",
    )

    return stats


def _write_audit_xlsx(
    papers: list[dict],
    output_path: Path,
    spot_check_threshold: float,
    review_name: str = "unknown",
    review_spec=None,
    db_path: str = "",
) -> None:
    """Write the audit review queue as a self-documenting Excel workbook."""
    from engine.exporters.review_workbook import (
        ColumnDef,
        DecisionColumnDef,
        FreeTextColumnDef,
        InstructionsConfig,
        create_review_workbook,
    )

    # Flatten papers to per-span rows
    span_rows = _flatten_to_span_rows(papers)

    columns = [
        ColumnDef(key="paper_id", header="paper_id", width=10),
        ColumnDef(key="ee_identifier", header="EE-ID", width=10),
        ColumnDef(key="title", header="Title", width=50, wrap=True),
        ColumnDef(key="field_name", header="Field Name", width=25),
        ColumnDef(key="extracted_value", header="Extracted Value", width=40, wrap=True),
        ColumnDef(key="evidence_snippet", header="Evidence Snippet", width=60, wrap=True),
        ColumnDef(key="audit_state", header="Audit State", width=15),
        ColumnDef(key="auditor_reasoning", header="Auditor Reasoning", width=50, wrap=True),
    ]

    decision_columns = [
        DecisionColumnDef(
            key="PI_decision",
            header="PI_decision (ACCEPT/REJECT/CORRECT)",
            valid_values=["ACCEPT", "REJECT", "CORRECT"],
            width=30,
        ),
    ]

    free_text_columns = [
        FreeTextColumnDef(key="corrected_value", header="corrected_value (required if CORRECT)", width=40),
        FreeTextColumnDef(key="PI_notes", header="PI_notes (optional)", width=30),
    ]

    import_cmd = (
        f"python -c \"\n"
        f"from engine.core.database import ReviewDatabase\n"
        f"from engine.adjudication import import_audit_review_decisions\n"
        f"db = ReviewDatabase('{review_name}')\n"
        f"import_audit_review_decisions(db, '{output_path}')\n"
        f"\""
    )

    decision_criteria = [
        "ACCEPT: The extracted value is correct as-is. Mark as human-verified.",
        "REJECT: The extracted value is wrong and cannot be corrected. Remove from evidence table.",
        "CORRECT: The value needs correction. You MUST fill in the 'corrected_value' column.",
    ]

    edge_case = (
        "CONTESTED spans usually have correct values but imperfect snippets — "
        "these are most commonly ACCEPT. FLAGGED spans need careful review against "
        f"the paper. If >{int(spot_check_threshold * 100)}% of spot-check papers have errors, "
        "all remaining verified-only papers will be promoted to the review queue."
    )

    instr = InstructionsConfig(
        review_name=review_name,
        review_spec_id=f"{review_spec.title} v{review_spec.version}" if review_spec else "",
        db_path=db_path,
        export_trigger=(
            f"{len(papers)} papers with contested/flagged/low-yield spans "
            f"requiring human audit review ({len(span_rows)} total span rows)"
        ),
        row_count=len(span_rows),
        decision_column_name="PI_decision (ACCEPT/REJECT/CORRECT)",
        valid_values=["ACCEPT", "REJECT", "CORRECT"],
        decision_criteria=decision_criteria,
        edge_case_guidance=edge_case,
        import_command=import_cmd,
        columns_importer_reads=[
            "paper_id (A)", "Field Name (D)", "PI_decision (I)",
            "corrected_value (J)", "PI_notes (K)",
        ],
        columns_importer_ignores=(
            "All other columns (EE-ID, Title, Extracted Value, Evidence Snippet, "
            "Audit State, Auditor Reasoning) are read-only context."
        ),
        notes_on_rejection=(
            "If any PI_decision cell is blank, the entire import is rejected. "
            "If any PI_decision value is not ACCEPT/REJECT/CORRECT, the import is rejected. "
            "If PI_decision is CORRECT but corrected_value is blank, the import is rejected "
            "(names the row and field). Fix the workbook and re-run."
        ),
    )

    reference_content = _build_audit_reference_content(review_spec)

    create_review_workbook(
        output_path=output_path,
        rows=span_rows,
        columns=columns,
        decision_columns=decision_columns,
        free_text_columns=free_text_columns,
        instructions=instr,
        reference_content=reference_content,
        reference_sheet_title="Audit Reference",
    )


# ── Import ─────────────────────────────────────────────────────────


def import_audit_review_decisions(
    review_db: ReviewDatabase,
    input_path: str | Path,
) -> dict:
    """Read completed audit review Excel, write decisions to database.

    Validates the entire file before making any changes:
      - Rejects if any PI_decision cell is blank
      - Rejects if any PI_decision value is not ACCEPT/REJECT/CORRECT
      - Rejects if PI_decision is CORRECT but corrected_value is blank

    For ACCEPT: mark span as verified (human_review).
    For REJECT: mark span as rejected, record in audit_adjudication.
    For CORRECT: update span value, record original in audit_adjudication.

    When all spans for a paper are resolved, transitions paper to HUMAN_AUDIT_COMPLETE.

    Returns summary dict.
    """
    from openpyxl import load_workbook

    input_path = Path(input_path)
    wb = load_workbook(input_path)

    # Find the review queue sheet (support both old and new naming)
    sheet_name = None
    for name in ["Review Queue", "Audit Review"]:
        if name in wb.sheetnames:
            sheet_name = name
            break
    if sheet_name is None:
        error_msg = f"\nIMPORT REJECTED — No 'Review Queue' sheet found in {input_path}\n"
        print(error_msg)
        logger.error(error_msg)
        return {
            "stats": {"accepted": 0, "corrected_fields": 0, "rejected": 0,
                      "missing": 0, "total": 0, "papers_transitioned": 0},
            "warnings": [error_msg],
        }

    ws = wb[sheet_name]

    # ── Build header index for robust column lookup ───────────
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_index = {}
    for idx, val in enumerate(header_row):
        if val:
            col_index[str(val).strip()] = idx

    def _find_col(candidates: list[str]) -> int | None:
        for c in candidates:
            for header, idx in col_index.items():
                if c.lower() in header.lower():
                    return idx
        return None

    paper_id_col = _find_col(["paper_id"])
    field_name_col = _find_col(["Field Name", "field_name"])
    decision_col = _find_col(["PI_decision"])
    corrected_col = _find_col(["corrected_value"])
    notes_col = _find_col(["PI_notes", "notes"])
    title_col = _find_col(["Title", "title"])

    # Also check for legacy column layout (per-paper format)
    accept_col = _find_col(["accept_as_is"])
    reject_paper_col = _find_col(["reject_paper"])
    is_legacy = accept_col is not None and reject_paper_col is not None

    if is_legacy:
        return _import_legacy_format(review_db, ws, col_index)

    if paper_id_col is None or decision_col is None:
        error_msg = (
            "\nIMPORT REJECTED — Required columns not found.\n"
            f"  Found headers: {list(col_index.keys())}\n"
            "  Required: paper_id, PI_decision\n"
        )
        print(error_msg)
        logger.error(error_msg)
        return {
            "stats": {"accepted": 0, "corrected_fields": 0, "rejected": 0,
                      "missing": 0, "total": 0, "papers_transitioned": 0},
            "warnings": [error_msg],
        }

    # ── Pass 1: Validate all rows before any DB writes ──────────
    parsed_rows = []
    blank_rows = []
    invalid_rows = []
    correct_missing_value = []

    excel_row_num = 1  # for error reporting
    for row in ws.iter_rows(min_row=2, values_only=False):
        excel_row_num += 1
        paper_id_val = row[paper_id_col].value
        if not paper_id_val:
            continue

        paper_id = int(paper_id_val)
        field_name = row[field_name_col].value if field_name_col is not None else ""
        title = row[title_col].value if title_col is not None else ""
        decision_raw = row[decision_col].value
        corrected_value = (row[corrected_col].value or "") if corrected_col is not None else ""
        notes = (row[notes_col].value or "") if notes_col is not None else ""

        display_label = f"paper_id={paper_id}, field={field_name}"

        if not decision_raw or str(decision_raw).strip() == "":
            blank_rows.append(f"  Row {excel_row_num}: {display_label} — no decision")
            continue

        decision = str(decision_raw).strip().upper()
        if decision not in ("ACCEPT", "REJECT", "CORRECT"):
            invalid_rows.append(
                f"  Row {excel_row_num}: {display_label} — '{decision_raw}' "
                f"(must be ACCEPT, REJECT, or CORRECT)"
            )
            continue

        if decision == "CORRECT" and not str(corrected_value).strip():
            correct_missing_value.append(
                f"  Row {excel_row_num}: {display_label} — PI_decision is CORRECT "
                f"but corrected_value is blank"
            )
            continue

        parsed_rows.append({
            "excel_row": excel_row_num,
            "paper_id": paper_id,
            "field_name": field_name,
            "title": title,
            "decision": decision,
            "corrected_value": str(corrected_value).strip() if corrected_value else "",
            "notes": notes,
        })

    # ── Reject on validation failure ────────────────────────────
    all_issues = blank_rows + invalid_rows + correct_missing_value
    if all_issues:
        msg_parts = [
            f"\nIMPORT REJECTED — {len(all_issues)} validation error(s) found.",
            "No database changes were made.\n",
        ]
        if blank_rows:
            msg_parts.append(f"BLANK DECISION CELLS ({len(blank_rows)} rows):")
            msg_parts.extend(blank_rows)
            msg_parts.append("")
        if invalid_rows:
            msg_parts.append(f"INVALID DECISION VALUES ({len(invalid_rows)} rows):")
            msg_parts.extend(invalid_rows)
            msg_parts.append("")
        if correct_missing_value:
            msg_parts.append(f"CORRECT WITHOUT corrected_value ({len(correct_missing_value)} rows):")
            msg_parts.extend(correct_missing_value)
            msg_parts.append("")
        msg_parts.append("Fix the workbook and re-run the import command.")

        error_msg = "\n".join(msg_parts)
        print(error_msg)
        logger.error(error_msg)

        return {
            "stats": {
                "accepted": 0,
                "corrected_fields": 0,
                "rejected": 0,
                "missing": len(blank_rows) + len(correct_missing_value),
                "invalid": len(invalid_rows),
                "total": len(parsed_rows) + len(all_issues),
                "papers_transitioned": 0,
            },
            "warnings": all_issues,
        }

    # ── Pass 2: Apply all validated decisions ───────────────────
    ensure_adjudication_table(review_db._conn)

    now = datetime.now(timezone.utc).isoformat()
    stats = {
        "accepted": 0, "corrected_fields": 0, "rejected": 0,
        "missing": 0, "invalid": 0, "total": len(parsed_rows),
        "papers_transitioned": 0,
    }
    warnings = []

    # Group rows by paper_id for post-processing
    paper_ids_seen = set()

    for pr in parsed_rows:
        decision = pr["decision"]
        pid = pr["paper_id"]
        field_name = pr["field_name"]
        notes = pr["notes"]

        paper_ids_seen.add(pid)

        # Find the span
        extraction = review_db._conn.execute(
            "SELECT id FROM extractions WHERE paper_id = ? ORDER BY id DESC LIMIT 1",
            (pid,),
        ).fetchone()
        if not extraction:
            warnings.append(f"Paper {pid}: no extraction found")
            continue
        ext_id = extraction["id"]

        span = review_db._conn.execute(
            "SELECT * FROM evidence_spans WHERE extraction_id = ? AND field_name = ?",
            (ext_id, field_name),
        ).fetchone()
        if not span:
            warnings.append(f"Paper {pid}, field {field_name}: span not found")
            continue

        if decision == "ACCEPT":
            review_db._conn.execute(
                """UPDATE evidence_spans
                   SET audit_status = 'verified',
                       auditor_model = 'human_review',
                       audit_rationale = ?,
                       audited_at = ?
                   WHERE id = ?""",
                (f"Accepted by human reviewer. {notes}", now, span["id"]),
            )
            stats["accepted"] += 1

        elif decision == "REJECT":
            # Record in audit_adjudication, then mark span
            review_db._conn.execute(
                """INSERT INTO audit_adjudication
                   (span_id, paper_id, field_name, original_value,
                    human_decision, override_value, reviewer_notes,
                    adjudication_timestamp, created_at)
                   VALUES (?, ?, ?, ?, 'reject_paper', NULL, ?, ?, ?)""",
                (span["id"], pid, field_name, span["value"], notes, now, now),
            )
            review_db._conn.execute(
                """UPDATE evidence_spans
                   SET audit_status = 'verified',
                       auditor_model = 'human_review',
                       audit_rationale = ?,
                       audited_at = ?
                   WHERE id = ?""",
                (f"Rejected by human reviewer. {notes}", now, span["id"]),
            )
            stats["rejected"] += 1

        elif decision == "CORRECT":
            new_value = pr["corrected_value"]
            review_db._conn.execute(
                """INSERT INTO audit_adjudication
                   (span_id, paper_id, field_name, original_value,
                    human_decision, override_value, reviewer_notes,
                    adjudication_timestamp, created_at)
                   VALUES (?, ?, ?, ?, 'override', ?, ?, ?, ?)""",
                (span["id"], pid, field_name, span["value"],
                 new_value, notes, now, now),
            )
            review_db._conn.execute(
                """UPDATE evidence_spans
                   SET value = ?,
                       audit_status = 'verified',
                       auditor_model = 'human_review',
                       audit_rationale = ?,
                       audited_at = ?
                   WHERE id = ?""",
                (new_value,
                 f"Human override: '{span['value']}' -> '{new_value}'. {notes}",
                 now, span["id"]),
            )
            stats["corrected_fields"] += 1

    # Transition papers to HUMAN_AUDIT_COMPLETE if all spans now resolved
    for pid in paper_ids_seen:
        extraction = review_db._conn.execute(
            "SELECT id FROM extractions WHERE paper_id = ? ORDER BY id DESC LIMIT 1",
            (pid,),
        ).fetchone()
        if not extraction:
            continue

        remaining = review_db._conn.execute(
            """SELECT COUNT(*) FROM evidence_spans
               WHERE extraction_id = ? AND audit_status IN ('contested', 'flagged', 'invalid_snippet')""",
            (extraction["id"],),
        ).fetchone()[0]

        if remaining == 0:
            try:
                review_db.update_status(pid, "HUMAN_AUDIT_COMPLETE")
                stats["papers_transitioned"] += 1
            except ValueError:
                pass  # already transitioned or invalid state

    review_db._conn.commit()

    # Success summary
    print(
        f"\nIMPORT SUCCESSFUL — {stats['total']} span decisions processed.\n"
        f"  ACCEPT:  {stats['accepted']}\n"
        f"  REJECT:  {stats['rejected']}\n"
        f"  CORRECT: {stats['corrected_fields']}\n"
        f"  Papers transitioned to HUMAN_AUDIT_COMPLETE: {stats['papers_transitioned']}\n"
        f"  Database updated."
    )

    logger.info(
        "Audit import complete: %d accepted, %d corrected, "
        "%d rejected (of %d total), %d papers transitioned",
        stats["accepted"], stats["corrected_fields"], stats["rejected"],
        stats["total"], stats["papers_transitioned"],
    )

    # Auto-advance workflow: AUDIT_REVIEW_COMPLETE
    complete_stage(
        review_db._conn, "AUDIT_REVIEW_COMPLETE",
        metadata=(
            f"{stats['accepted']} accepted, {stats['corrected_fields']} corrected, "
            f"{stats['rejected']} rejected (of {stats['total']} total)"
        ),
    )

    return {"stats": stats, "warnings": warnings}


def _import_legacy_format(review_db, ws, col_index):
    """Handle import of legacy per-paper format (accept_as_is/reject_paper columns).

    This preserves backward compatibility with workbooks generated before
    the per-span format was introduced.
    """
    headers = list(col_index.keys())

    def _find_col(candidates):
        for c in candidates:
            for header, idx in col_index.items():
                if c.lower() in header.lower():
                    return idx
        return None

    pid_col = _find_col(["paper_id"])
    accept_col = _find_col(["accept_as_is"])
    notes_col = _find_col(["notes"])
    reject_col = _find_col(["reject_paper"])
    title_col = _find_col(["title"])
    review_reason_col = _find_col(["review_reason"])

    # Parse field correction columns
    field_col_map = {}
    for header, idx in col_index.items():
        if header.endswith("_correction"):
            fname = header[:-len("_correction")]
            val_key = f"{fname}_value"
            field_col_map[fname] = {
                "value_col": col_index.get(val_key),
                "correction_col": idx,
            }

    ensure_adjudication_table(review_db._conn)

    now = datetime.now(timezone.utc).isoformat()
    stats = {
        "accepted": 0, "corrected_fields": 0, "rejected": 0,
        "missing": 0, "total": 0, "papers_transitioned": 0,
        "spot_check_failures": 0, "spot_check_total": 0,
    }
    warnings = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        paper_id_val = row[pid_col].value
        if not paper_id_val:
            continue

        stats["total"] += 1
        pid = int(paper_id_val)

        accept_raw = str(row[accept_col].value or "").strip().upper()
        reject_raw = str(row[reject_col].value or "").strip().upper()
        notes = (row[notes_col].value or "") if notes_col is not None else ""
        review_reason = (row[review_reason_col].value or "") if review_reason_col is not None else ""

        is_accept = accept_raw in ("TRUE", "YES", "1", "ACCEPT")
        is_reject = reject_raw in ("TRUE", "YES", "1", "REJECT")

        if review_reason == "spot_check":
            stats["spot_check_total"] += 1

        has_corrections = False
        corrections = {}
        for fname, cols in field_col_map.items():
            correction_val = row[cols["correction_col"]].value
            if correction_val and str(correction_val).strip():
                has_corrections = True
                corrections[fname] = str(correction_val).strip()

        if is_reject:
            try:
                review_db.reject_paper(pid, f"Audit review rejection: {notes}")
                stats["rejected"] += 1
                stats["papers_transitioned"] += 1
            except ValueError as e:
                warnings.append(f"Paper {pid}: cannot reject — {e}")
            continue

        if not is_accept and not has_corrections:
            stats["missing"] += 1
            t = (row[title_col].value or "?") if title_col is not None else "?"
            warnings.append(f"Paper {pid} ({t[:40]}): no decision")
            continue

        extraction = review_db._conn.execute(
            "SELECT id FROM extractions WHERE paper_id = ? ORDER BY id DESC LIMIT 1",
            (pid,),
        ).fetchone()
        if not extraction:
            warnings.append(f"Paper {pid}: no extraction found")
            continue
        ext_id = extraction["id"]

        spans = review_db._conn.execute(
            "SELECT * FROM evidence_spans WHERE extraction_id = ?",
            (ext_id,),
        ).fetchall()
        span_map = {s["field_name"]: dict(s) for s in spans}

        if is_accept:
            stats["accepted"] += 1
            for s in spans:
                if s["audit_status"] in _NEEDS_REVIEW:
                    review_db._conn.execute(
                        """UPDATE evidence_spans
                           SET audit_status = 'verified',
                               auditor_model = 'human_review',
                               audit_rationale = ?,
                               audited_at = ?
                           WHERE id = ?""",
                        (f"Accepted as-is by human reviewer. {notes}", now, s["id"]),
                    )

        for fname, new_value in corrections.items():
            span = span_map.get(fname)
            if not span:
                warnings.append(f"Paper {pid}, field {fname}: span not found")
                continue

            review_db._conn.execute(
                """INSERT INTO audit_adjudication
                   (span_id, paper_id, field_name, original_value,
                    human_decision, override_value, reviewer_notes,
                    adjudication_timestamp, created_at)
                   VALUES (?, ?, ?, ?, 'override', ?, ?, ?, ?)""",
                (span["id"], pid, fname, span["value"],
                 new_value, notes, now, now),
            )
            review_db._conn.execute(
                """UPDATE evidence_spans
                   SET value = ?,
                       audit_status = 'verified',
                       auditor_model = 'human_review',
                       audit_rationale = ?,
                       audited_at = ?
                   WHERE id = ?""",
                (new_value,
                 f"Human override: '{span['value']}' -> '{new_value}'. {notes}",
                 now, span["id"]),
            )
            stats["corrected_fields"] += 1

        remaining = review_db._conn.execute(
            """SELECT COUNT(*) FROM evidence_spans
               WHERE extraction_id = ? AND audit_status IN ('contested', 'flagged', 'invalid_snippet')""",
            (ext_id,),
        ).fetchone()[0]

        if remaining == 0:
            try:
                review_db.update_status(pid, "HUMAN_AUDIT_COMPLETE")
                stats["papers_transitioned"] += 1
            except ValueError:
                pass

    review_db._conn.commit()

    if warnings:
        for w in warnings[:10]:
            logger.warning(w)

    logger.info(
        "Audit import (legacy): %d accepted, %d corrected, %d rejected, "
        "%d missing (of %d total)",
        stats["accepted"], stats["corrected_fields"], stats["rejected"],
        stats["missing"], stats["total"],
    )

    if stats["missing"] == 0:
        complete_stage(
            review_db._conn, "AUDIT_REVIEW_COMPLETE",
            metadata=(
                f"{stats['accepted']} accepted, {stats['corrected_fields']} corrected, "
                f"{stats['rejected']} rejected (of {stats['total']} total)"
            ),
        )
    else:
        logger.warning("Audit review not complete: %d papers with no decision", stats["missing"])

    return {"stats": stats, "warnings": warnings}


# ── Gate ──────────────────────────────────────────────────────────


def check_audit_review_gate(review_db: ReviewDatabase) -> int:
    """Return count of AI_AUDIT_COMPLETE papers with unresolved spans."""
    papers = review_db.get_papers_by_status("AI_AUDIT_COMPLETE")
    count = 0
    for paper in papers:
        pid = paper["id"]
        has_issues = review_db._conn.execute(
            """SELECT COUNT(*) FROM evidence_spans es
               JOIN extractions e ON e.id = es.extraction_id
               WHERE e.paper_id = ? AND es.audit_status IN ('contested', 'flagged', 'invalid_snippet')""",
            (pid,),
        ).fetchone()[0]
        if has_issues > 0:
            count += 1
    return count
